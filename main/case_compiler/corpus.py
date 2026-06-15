"""idiom 语料库：从镜像的 smoke_test xlsx 提取 (case → DSL 行) 平行语料，供编译器 few-shot。

纯本地构建，吃 knowledge/framework/mirror/。检索=纯查找（关键词/模块匹配最近邻），非向量库
（守 CLAUDE.md 无 RAG/Qdrant 原则）。

每个语料条目 = 一个完整 case 的行序列（含 autoid/标题/步骤行），带可检索的文本签名。
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_MIRROR = Path(__file__).resolve().parents[2] / "knowledge/framework/mirror/smoke_test"


@dataclass
class CorpusCase:
    """语料库里一个完整 case：原始行 + 检索签名。"""

    source_file: str
    autoid: str
    title: str
    rows: list[list[Any]]          # 该 case 的 A-I 行（含首行）
    init_rows: list[list[Any]]     # 所在文件的 C=1 前置行
    tokens: set[str] = field(default_factory=set)
    source_kind: str = "xlsx"      # xlsx=可声明式直转(passthrough); py=imperative,仅作few-shot

    def as_example_text(self) -> str:
        """渲染成 few-shot 文本（紧凑表格）。"""
        lines = []
        lines.append("# 文件级前置 (C=1):")
        for r in self.init_rows:
            lines.append(f"  C=1 | E={r[4]} | F={r[5]} | G={_fmt(r[6])}")
        lines.append(f"# CASE autoid={self.autoid} title={self.title}")
        for r in self.rows:
            c = r[2] if r[2] is not None else "·"
            d = _fmt(r[3]) if r[3] is not None else ""
            h = f" | H={r[7]}" if len(r) > 7 and r[7] else ""
            i = f" | I={r[8]}" if len(r) > 8 and r[8] else ""
            lines.append(f"  C={c} | D={d} | E={r[4]} | F={r[5]} | G={_fmt(r[6])}{h}{i}")
        return "\n".join(lines)


def _fmt(v: Any) -> str:
    if v is None:
        return ""
    return str(v).replace("\n", "\\n")


def _tokenize(text: str) -> set[str]:
    text = (text or "").lower()
    toks = set(re.findall(r"[a-z_]+|[一-鿿]{2,}", text))
    return toks


def _read_grid(path: Path) -> list[list[Any]]:
    wb = load_workbook(path, data_only=True)
    sheet = wb.active
    return [list(r) for r in sheet.iter_rows(values_only=True)]


def _parse_file(path: Path) -> list[CorpusCase]:
    """把一个 xlsx 拆成多个 CorpusCase。"""
    from main.case_compiler.config import get_config
    anchor = get_config().xlsx.header_anchor
    grid = _read_grid(path)
    cases: list[CorpusCase] = []
    init_rows: list[list[Any]] = []
    case_begin = False
    cur: CorpusCase | None = None
    title_hint = ""
    for row in grid:
        # pad to 9
        r = (list(row) + [None] * 9)[:9]
        a, b, c, d, e, f, g, h, i = r
        if a is not None and str(a).strip() == anchor:
            case_begin = True
            continue
        if not case_begin:
            continue
        if e is None and f is None and g is None and a is None:
            continue  # blank separator
        cs = str(c) if c is not None else ""
        if cs == "1":
            init_rows.append(r)
            continue
        if cs == "0":
            continue
        # case 边界判据：A 列有 autoid（与框架引擎 test_xlsx.py:220 `row[0] is not None` 对齐）。
        # 不可用 C==2 判首行——合法 DSL 允许 C 跳号（如 mx_record 首行 C=4）。
        if a is not None:
            if cur:
                cases.append(cur)
            cur = CorpusCase(
                source_file=path.name, autoid=str(a),
                title=str(d or ""), rows=[r], init_rows=[],
            )
        elif cur is not None:
            cur.rows.append(r)
    if cur:
        cases.append(cur)
    # attach init + tokens
    for cc in cases:
        cc.init_rows = init_rows
        sig = cc.title + " " + " ".join(_fmt(x[6]) for x in cc.rows) + " " + path.stem
        cc.tokens = _tokenize(sig)
    return cases


def _parse_py_case(path: Path, normalizer=None) -> "CorpusCase | None":
    """把 host_persistence 风格的 pytest .py 脚本反编译成 CorpusCase（行级 DSL）。

    形态（实证 205271757988517196.py）：
      def test_<autoid>(APV0, test_env, check_point):
          APV0.cmd_config('...')          → E=APV_0 F=cmd_config G=...
          var = APV0.cmd_config('...')    → 同上，var 进 H 列(save_as)  ← R2 修复②
          check_point.found('expect', xx) → E=check_point F=found G=expect
          test_env.routera('dig ...')     → E=test_env F=routera G=...
          Seg0.cmd_config('...')          → 设备别名归一到 E=APV_0       ← R2 修复①
          time.sleep(10)                  → E=time F=sleep G=10
    autoid 来自文件名（[0-9]+.py）。action/expect 注释作 title。

    E 列对象合法集 + 设备别名来自 normalizer（原始 xlsx 下拉菜单的权威源），
    不再写死 {APV0,APV1,...}（根因红线，见 object_normalizer.py）。normalizer=None 时懒加载单例。
    """
    name = path.stem
    if not re.fullmatch(r"\d{7,}", name):
        return None
    try:
        src = path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return None
    if normalizer is None:
        from main.case_compiler.object_normalizer import get_object_normalizer
        normalizer = get_object_normalizer()
    # title: 取注释里的 action 行
    title = ""
    m = re.search(r"#\s*action[:：]\s*(.+)", src)
    if m:
        title = m.group(1).strip()
    rows: list[list[Any]] = []
    # 逐行匹配 [var =] obj.method(args)。
    # R2 修复②：去掉行首 obj 锚定（^\s*\w+\.），改为允许可选 `var =` 前缀，
    # 否则 `result = obj.method(...)` 这类赋值行（占调用行 14.9%）全被丢弃。
    call_re = re.compile(r"^\s*(?:(\w+)\s*=\s*)?(\w+)\.(\w+)\((.*)\)\s*$")
    for line in src.splitlines():
        ls = line.strip()
        if ls.startswith("#") or ls.startswith("def ") or not ls:
            continue
        cm = call_re.match(line)
        if not cm:
            continue
        save_var, obj, method, argstr = cm.group(1), cm.group(2), cm.group(3), cm.group(4)
        # 还原 E 列对象名：经规范源归一（设备别名 Seg0/APV0_C→APV_0），R2 修复①
        eobj = normalizer.canon_object(obj)
        if eobj is None:
            continue
        g = _extract_g(argstr, eobj, method)   # R2 修复③：kwargs/多位置参数
        # E=check_point 的 F=found/abs_found...；E=test_env 的 F=主机名；E=APV_* 的 F=cmd_config 等
        f = method
        # found_times 第二参=times → I 列
        ivar = None
        if eobj == "check_point" and method == "found_times":
            nums = re.findall(r",\s*(\d+)\s*\)?\s*$", argstr)
            if nums:
                ivar = nums[-1]
        # A B C D E F G H I —— 赋值变量进 H 列(save_as)，R2 修复②
        hcol = save_var if save_var else None
        rows.append([None, None, None, None, eobj, f, g, hcol, ivar])
    if not rows:
        return None
    cc = CorpusCase(source_file=path.name, autoid=name, title=title, rows=rows,
                    init_rows=[], source_kind="py")
    sig = title + " " + " ".join(_fmt(r[6]) for r in rows) + " " + path.stem
    cc.tokens = _tokenize(sig)
    return cc


def _extract_g(argstr: str, eobj: str, method: str) -> str:
    """从调用参数串抽 G 列值（R2 修复③：kwargs + 多位置参数）。

    旧实现只 `re.search(['"](.*?)['"])` 抓第一个字符串字面量，对
    `clientc(cmd='...', prompt='...')`（19326 行）会抓错或抓到 prompt。
    规则：
    - time.sleep(n) → 数字原样。
    - kwargs 优先级：cmd > command > config > 第一个字符串字面量。
    - 起始引号配对匹配，避免内层异种引号截断。
    """
    argstrip = argstr.lstrip()
    if eobj == "time" and method == "sleep":
        return argstr.strip().strip("'\"")
    # kwargs：按优先级取语义键的字符串值
    for key in ("cmd", "command", "config", "value"):
        km = re.search(r"\b%s\s*=\s*(['\"])((?:[^\\]|\\.)*?)\1" % key, argstr)
        if km:
            return km.group(2)
    # 起始即字符串字面量：配对收尾匹配（避免内层异种引号截断）
    if argstrip[:1] in ("'", '"'):
        q = argstrip[0]
        mlit = re.match(r"%s((?:[^%s\\]|\\.)*)%s" % (q, q, q), argstrip)
        if mlit:
            return mlit.group(1)
    # 兜底：第一个字符串字面量
    gm = re.search(r"""['"](.*?)['"]""", argstr)
    return gm.group(1) if gm else argstr.strip()


class IdiomCorpus:
    """idiom 语料库（懒加载内存索引）。"""

    def __init__(self, cases: list[CorpusCase]):
        self.cases = cases
        self._by_autoid: dict[str, CorpusCase] = {}
        for cc in cases:
            # 同 autoid 多文件时保留首个（镜像内一般唯一）
            self._by_autoid.setdefault(str(cc.autoid), cc)

    def by_autoid(self, autoid: str, passthrough_only: bool = False) -> "CorpusCase | None":
        """精确 autoid 直取：框架自身已验证的编译（最高保真）。

        passthrough_only=True 时只返回 xlsx 派生（声明式可直转 xlsx 行）的命中；
        .py 派生用的是 imperative 运行时逻辑，不可声明式直转，仅作 few-shot 检索语料。
        """
        cc = self._by_autoid.get(str(autoid))
        if cc is None:
            return None
        if passthrough_only and cc.source_kind != "xlsx":
            return None
        return cc

    @classmethod
    def build(cls, subdirs: list[str] | None = None, include_py: bool = True) -> "IdiomCorpus":
        """从镜像构建。subdirs 限定 smoke_test 下的子目录（相对路径）。

        include_py=True 时同时解析 host_persistence 风格的 pytest .py 脚本为语料。
        """
        roots = []
        if subdirs:
            roots = [_MIRROR / s for s in subdirs]
        else:
            roots = [_MIRROR]
        cases: list[CorpusCase] = []
        seen: set[str] = set()
        # 规范源单例（E×F 合法集 + 设备别名），全 .py 共享一份，避免逐文件重扫语料
        normalizer = None
        if include_py:
            from main.case_compiler.object_normalizer import get_object_normalizer
            normalizer = get_object_normalizer()
        for root in roots:
            if not root.exists():
                continue
            for xlsx in root.rglob("*.xlsx"):
                if xlsx.name.startswith("~$"):
                    continue
                key = str(xlsx)
                if key in seen:
                    continue
                seen.add(key)
                try:
                    cases.extend(_parse_file(xlsx))
                except Exception:
                    continue
            if include_py:
                for pyf in root.rglob("*.py"):
                    if pyf.name.startswith("test_xlsx"):
                        continue
                    key = str(pyf)
                    if key in seen:
                        continue
                    seen.add(key)
                    try:
                        cc = _parse_py_case(pyf, normalizer=normalizer)
                        if cc is not None:
                            cases.append(cc)
                    except Exception:
                        continue
        return cls(cases)

    def nearest(self, query: str, module: str = "", k: int = 3) -> list[CorpusCase]:
        """最近邻检索：token Jaccard + 模块加权。"""
        q = _tokenize(query + " " + module)
        if not q:
            return self.cases[:k]
        scored = []
        for cc in self.cases:
            inter = len(q & cc.tokens)
            if inter == 0:
                continue
            union = len(q | cc.tokens) or 1
            score = inter / union
            scored.append((score, cc))
        scored.sort(key=lambda x: x[0], reverse=True)
        return [cc for _, cc in scored[:k]]
