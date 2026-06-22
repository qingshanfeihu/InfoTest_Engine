"""上机回填核心：把 <RUNTIME> 占位槽位用设备真实值填上并**锁死**。

设计要点（对应用户两条铁律）：
- **不猜留空**：draft 对离线不可知的期望值写 <RUNTIME> 占位（见 provenance_ir）。本模块只负责
  把这些占位用**上机真实输出**回填——值由调用方（ist-fill fork 读设备明细）给定，本模块不编值。
  调用方给空值＝"抽不出"，本模块如实留空、报告，绝不猜。
- **不反复改（锁死）**：回填**只动仍含 <RUNTIME> 的 G 格子**。一旦填上，该格子不再含占位符，
  后续任何回填都定位不到它 → 天然幂等、永不覆盖。锁是结构性的，不靠调用方自律。

xlsx 列布局（与 xlsx_emit 一致）：A=autoid（仅每 case 首行）, C=stmt_type, E=对象, F=方法, G=数据。
数据区从第 29 行起（与 precedent_tools._load_case_rows 同约定），遇 A 以 "999999" 开头的哨兵 case 停。
"""

from __future__ import annotations

import json
import logging
from dataclasses import dataclass, field
from pathlib import Path

from main.case_compiler.provenance_ir import RUNTIME_PLACEHOLDER, parse_provenance

logger = logging.getLogger(__name__)

_DATA_START_ROW = 29
_SENTINEL_PREFIX = "999999"
_COL_A, _COL_C, _COL_E, _COL_F, _COL_G = 1, 3, 5, 6, 7


def _is_observation(e: str, f: str, g: str) -> bool:
    """该步是否产出可被 check_point 文本匹配的回显（test_env 触发 / APV 的 show/统计/dig）。"""
    e = (e or "").strip()
    f = (f or "").strip().lower()
    g = (g or "").lower()
    if e == "test_env":
        return True
    if e.startswith("APV"):
        return any(k in f or k in g for k in ("show", "statistics", "stat", "dig", "nslookup", "get", "display", "list"))
    return False


@dataclass
class RuntimeSlot:
    """一个待回填的 <RUNTIME> 槽位（check_point 的 G 含占位符）。"""
    slot_id: str          # f"{autoid}#{row}"——用**行号**做稳定标识（填别的槽位不会让本 id 漂移）
    autoid: str
    row: int              # sheet 1-based 行号
    current_g: str        # 当前 G 全文（含 <RUNTIME>）
    method: str           # F 列（found / not_found / found_times ...）
    observe_cmd: str      # 紧邻前序观测步的命令（设备据它产出输出；回填值就从这条输出里来）
    observe_obj: str      # 前序观测步的 E（test_env / APV_0）

    def to_dict(self) -> dict:
        return {"slot_id": self.slot_id, "autoid": self.autoid, "row": self.row,
                "current_g": self.current_g, "method": self.method,
                "observe_obj": self.observe_obj, "observe_cmd": self.observe_cmd}


def list_runtime_slots(xlsx_path: str | Path) -> list[RuntimeSlot]:
    """扫描 case.xlsx 数据区，列出所有仍含 <RUNTIME> 占位的 check_point 槽位。

    已回填（不含占位）的格子不返回——天然体现锁：列出的永远只是"还没填的"。
    """
    import openpyxl
    ws = openpyxl.load_workbook(str(xlsx_path), data_only=True).active

    slots: list[RuntimeSlot] = []
    cur_autoid = ""
    last_obs_cmd = ""
    last_obs_obj = ""
    for r in range(_DATA_START_ROW, ws.max_row + 1):
        a = ws.cell(r, _COL_A).value
        if a and str(a).startswith(_SENTINEL_PREFIX):
            break
        if a and str(a).strip():
            cur_autoid = str(a).strip()
            last_obs_cmd = ""  # 新 case 起，前序观测重置
            last_obs_obj = ""
        e = str(ws.cell(r, _COL_E).value or "").strip()
        f = str(ws.cell(r, _COL_F).value or "").strip()
        g = str(ws.cell(r, _COL_G).value or "")
        if not e and not f:
            continue
        if e == "check_point":
            if RUNTIME_PLACEHOLDER in g:
                # slot_id 用行号：填别的槽位不会让本 id 漂移（序号会漂，是锁失效根因）。
                slots.append(RuntimeSlot(
                    slot_id=f"{cur_autoid}#{r}", autoid=cur_autoid, row=r,
                    current_g=g, method=f, observe_cmd=last_obs_cmd, observe_obj=last_obs_obj))
            # check_point 消费上一步输出，不更新观测
        elif _is_observation(e, f, g):
            last_obs_cmd = g.strip()
            last_obs_obj = e
    return slots


@dataclass
class FillResult:
    """一次回填的汇总。"""
    filled: list[str] = field(default_factory=list)        # slot_id 列表（成功填）
    left_blank: list[str] = field(default_factory=list)    # 调用方给空值＝抽不出，如实留空
    not_found: list[str] = field(default_factory=list)     # 定位不到（已填锁死/id 错）
    details: list[str] = field(default_factory=list)
    xlsx_path: str = ""

    def summary(self) -> str:
        return (f"回填[{Path(self.xlsx_path).parent.name}]: 填{len(self.filled)} "
                f"留空{len(self.left_blank)} 未命中{len(self.not_found)}")


def apply_fills(xlsx_path: str | Path, fills: list[dict], *,
                project_root: Path | None = None, run_meta: str = "") -> FillResult:
    """把 fills 里的真实值回填进 xlsx 的 <RUNTIME> 槽位并锁死。

    fills: [{"slot_id": "<autoid>#<n>", "runtime_value": "<设备真实值, 替换 <RUNTIME> 这一段>",
             "evidence": "<可选: 取值依据的设备输出片段>"}]
        runtime_value 空/缺 ＝ 调用方声明"抽不出"→ 该槽位如实留空（不猜），计入 left_blank。
    project_root: 给定则同步更新 outputs/<autoid>/case.provenance.json（device_runtime→device_verified）。
    run_meta: 写入 provenance source.ref 的运行标识（如 "build=...;task=..."）。

    锁：apply 先重扫 slots（只含 <RUNTIME> 的才在列表里）。已填的 slot_id 重扫时不存在 → not_found，
    绝不覆盖已填值。这是"不反复改"的硬保证。
    """
    import openpyxl
    result = FillResult(xlsx_path=str(xlsx_path))

    slot_map = {s.slot_id: s for s in list_runtime_slots(xlsx_path)}
    if not fills:
        return result

    wb = openpyxl.load_workbook(str(xlsx_path))  # 不用 data_only：要写回
    ws = wb.active
    touched_autoids: dict[str, list[tuple[str, str, str]]] = {}  # autoid -> [(old_g,new_g,evidence)]
    dirty = False
    for item in fills:
        if not isinstance(item, dict):
            result.details.append(f"✗ 非 dict fill 项已跳过: {item!r}")
            continue
        sid = str(item.get("slot_id", "")).strip()
        rv = item.get("runtime_value")
        rv = "" if rv is None else str(rv)
        slot = slot_map.get(sid)
        if slot is None:
            result.not_found.append(sid)
            result.details.append(f"✗ {sid} 未命中（已填锁死 / slot_id 不存在）")
            continue
        if not rv.strip():
            result.left_blank.append(sid)
            result.details.append(f"· {sid} 留空（调用方未给值＝抽不出，不猜）")
            continue
        new_g = slot.current_g.replace(RUNTIME_PLACEHOLDER, rv)
        if RUNTIME_PLACEHOLDER in new_g:
            # 极端：rv 自身含占位符 → 拒填，避免假回填
            result.left_blank.append(sid)
            result.details.append(f"· {sid} 留空（回填值仍含占位符，拒绝假回填）")
            continue
        ws.cell(slot.row, _COL_G).value = new_g
        dirty = True
        result.filled.append(sid)
        result.details.append(f"✓ {sid} 行{slot.row}: {slot.current_g[:30]} → {new_g[:40]}")
        touched_autoids.setdefault(slot.autoid, []).append(
            (slot.current_g, new_g, str(item.get("evidence", ""))))

    if dirty:
        wb.save(str(xlsx_path))

    # 同步 provenance（best-effort，失败不挡）：device_runtime → device_verified
    if project_root and touched_autoids:
        for autoid, changes in touched_autoids.items():
            try:
                _sync_provenance(Path(project_root), autoid, changes, run_meta)
            except Exception as e:  # noqa: BLE001
                result.details.append(f"⚠ provenance 同步失败 {autoid}: {e}")

    logger.info(result.summary())
    return result


def _sync_provenance(project_root: Path, autoid: str,
                     changes: list[tuple[str, str, str]], run_meta: str) -> None:
    """把回填后的值同步进 outputs/<autoid>/case.provenance.json，来源转 device_verified。

    按 old_g 精确匹配 check_point 步（draft 写的占位 G 与 xlsx 一致）；命中则改 G + 来源。
    """
    prov_path = project_root / "workspace" / "outputs" / autoid / "case.provenance.json"
    if not prov_path.is_file():
        return
    prov = parse_provenance(prov_path.read_text(encoding="utf-8"))
    if prov is None:
        return
    change_map = {old: (new, ev) for old, new, ev in changes}
    hit = False
    for s in prov.steps:
        if s.E.strip() != "check_point":
            continue
        if s.G in change_map:
            new_g, ev = change_map[s.G]
            s.G = new_g
            s.source.kind = "device_verified"
            ref = run_meta
            if ev:
                ref = (ref + "|" if ref else "") + f"evidence:{ev[:200]}"
            s.source.ref = ref or s.source.ref
            hit = True
    if hit:
        prov_path.write_text(prov.to_json(), encoding="utf-8")
