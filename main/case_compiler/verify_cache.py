"""上机结果缓存：上机通过的 case 不反复跑，只对有问题的在交付前回填/重跑。

用户铁律③：上机 pass 的 case 别反复跑（设备极慢），只回填/重跑有问题的。

缓存按 **case 内容哈希**键控（该 autoid 的 E/F/G 行）——
- case 被重编译/回填后内容变 → 哈希变 → 缓存失效 → 重跑（不会拿旧 pass 蒙混）；
- pass 且无 `<RUNTIME>` 残留才记入缓存（没填完的不算"交付级通过"）。

缓存文件：workspace/outputs/<脑图名>/.verify_cache.json
"""

from __future__ import annotations

import hashlib
import json
import logging
from datetime import datetime, timezone
from pathlib import Path

from main.case_compiler.provenance_ir import RUNTIME_PLACEHOLDER

logger = logging.getLogger(__name__)

_CACHE_NAME = ".verify_cache.json"
_DATA_START_ROW = 29
_SENTINEL_PREFIX = "999999"


def case_rows_by_autoid(xlsx_path: str | Path) -> dict[str, list[dict]]:
    """把合并 xlsx 数据区按 autoid 分组为 {autoid: [{E,F,G}, ...]}（遇哨兵停）。"""
    import openpyxl
    ws = openpyxl.load_workbook(str(xlsx_path), data_only=True).active
    out: dict[str, list[dict]] = {}
    cur = ""
    for r in range(_DATA_START_ROW, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a and str(a).startswith(_SENTINEL_PREFIX):
            break
        if a and str(a).strip():
            cur = str(a).strip()
            out.setdefault(cur, [])
        E = str(ws.cell(r, 5).value or "").strip()
        F = str(ws.cell(r, 6).value or "").strip()
        G = str(ws.cell(r, 7).value or "")
        if cur and (E or F):
            out[cur].append({"E": E, "F": F, "G": G})
    return out


def case_content_hash(rows: list[dict]) -> str:
    """case 内容的稳定哈希（E|F|G 规范化）。回填/重编译改了任一格子哈希就变。"""
    canon = "\n".join(f"{r.get('E','')}|{r.get('F','')}|{r.get('G','')}" for r in rows)
    return hashlib.sha256(canon.encode("utf-8")).hexdigest()[:16]


def has_unfilled_runtime(rows: list[dict]) -> bool:
    """该 case 是否还有未回填的 `<RUNTIME>` 槽位（有则不算交付级通过）。"""
    return any(RUNTIME_PLACEHOLDER in (r.get("G") or "") for r in rows)


def _cache_file(mindmap_dir: str | Path) -> Path:
    return Path(mindmap_dir) / _CACHE_NAME


def load_cache(mindmap_dir: str | Path) -> dict:
    """读缓存 {autoid: {hash, verdict, build, task_id, ts}}；缺失/坏→空 dict。"""
    p = _cache_file(mindmap_dir)
    if not p.is_file():
        return {}
    try:
        d = json.loads(p.read_text(encoding="utf-8"))
        return d if isinstance(d, dict) else {}
    except Exception:  # noqa: BLE001
        return {}


def save_cache(mindmap_dir: str | Path, cache: dict) -> None:
    try:
        _cache_file(mindmap_dir).write_text(
            json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:  # noqa: BLE001
        logger.warning("verify_cache 写入失败: %s", e)


def is_cached_pass(cache: dict, autoid: str, content_hash: str) -> bool:
    """该 autoid 在当前内容下是否已缓存为上机通过（哈希必须匹配，否则视作内容变了）。"""
    e = cache.get(autoid)
    return bool(e and e.get("verdict") == "pass" and e.get("hash") == content_hash)


def record_pass(cache: dict, autoid: str, content_hash: str,
                build: str = "", task_id: str = "") -> None:
    """记一个交付级通过（调用方须自行确认 verdict=pass 且无未填 `<RUNTIME>`）。"""
    cache[autoid] = {"hash": content_hash, "verdict": "pass", "build": build,
                     "task_id": task_id, "ts": datetime.now(timezone.utc).isoformat()}


def invalidate(cache: dict, autoid: str) -> None:
    """显式作废某 autoid 的缓存（如发现旧 pass 实为误判）。"""
    cache.pop(autoid, None)
