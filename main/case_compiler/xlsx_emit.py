"""CaseIR → xlsx：克隆 smoke_test 原生模板，仅重写数据区（R29+），保留说明区/字典区/格式。

实证模板结构（sdns_listener.xlsx）：
  R1-R27 说明区（固定模板文字）；R28 表头（A-I）；R29 起数据区。
  Q-V 列字典区 + Z1:AH48 合并帮助区（克隆文件自动保留）。

数据区结构（实证）：
  - 文件级前置：C=1 行（每 case 执行前必跑），紧跟 author 说明行（C=0）之后。
  - 每 case：首行 A=autoid B=优先级 C=2 D=标题，后续步骤 C 递增；
    同步骤多行仅首行写 C/D。
  - case 之间空一行。

round-trip 闸：emit 后用框架 read_excel_with_openpyxl 回读对账（见 emit 返回的 grid）。
"""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from main.case_compiler.case_ir import FileIR, CaseIR, Row, Step
from main.case_compiler.config import get_config, detect_xlsx_layout

# 模板：跳转机原生样例（已镜像）。克隆它继承全部说明区/字典区/列宽/合并单元格。
_TEMPLATE = (
    Path(__file__).resolve().parents[2]
    / "knowledge/framework/mirror/smoke_test/sdns/listener/sdns_listener.xlsx"
)


def _set_row(sheet, r: int, cols: dict[int, Any]) -> None:
    """写一行；cols 是 {列号(1=A): 值}。未给的列清空。"""
    for c in range(1, 10):  # A..I
        sheet.cell(r, c).value = cols.get(c)


def _row_cols(row: Row, *, stmt_type: Optional[int], description: Optional[str],
              autoid: Optional[str] = None, priority: Optional[str] = None) -> dict[int, Any]:
    """构造一行的列字典。stmt_type/description 为 None 表示该行不写 C/D（同步骤后续行）。"""
    cols: dict[int, Any] = {}
    if autoid is not None:
        cols[1] = autoid          # A
    if priority is not None:
        cols[2] = priority        # B
    if stmt_type is not None:
        cols[3] = stmt_type       # C
    if description is not None:
        cols[4] = description     # D
    cols[5] = row.test_object     # E
    cols[6] = row.method          # F
    cols[7] = row.data            # G
    if row.save_as:
        cols[8] = row.save_as     # H
    if row.input_var:
        cols[9] = row.input_var   # I
    return cols


def emit_xlsx(file_ir: FileIR, out_path: Path) -> dict:
    """把 FileIR 写成 xlsx（克隆模板）。返回 round-trip 对账用的统计 + grid。"""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(_TEMPLATE, out_path)

    wb = load_workbook(out_path)
    sheet = wb.active

    # 动态探测数据起始行（从模板自身的表头锚点定位，不写死 R29）
    grid_probe = [list(row) for row in sheet.iter_rows(values_only=True)]
    layout = detect_xlsx_layout(grid_probe)
    data_start = layout.data_start

    # 清空旧数据区（data_start 到 max_row）
    for r in range(data_start, sheet.max_row + 1):
        _set_row(sheet, r, {})

    r = data_start
    # author 说明行（C=0）
    _set_row(sheet, r, {3: 0, 4: f"Author         : {file_ir.author}\n{file_ir.feature}"})
    r += 1

    # 文件级前置（C=1），同一前置块多行共享 C=1：首行写 C/D，后续留空？
    # 实证 listener：C=1 单行 cmds_config。这里每个 init Row 都写 C=1（前置可多条独立）。
    for row in file_ir.init_rows:
        _set_row(sheet, r, _row_cols(row, stmt_type=1, description="初始化配置"))
        r += 1

    # 各 case
    for case in file_ir.cases:
        first_case_row = True
        for st in case.steps:
            first_row_of_step = True
            for row in st.rows:
                if first_row_of_step:
                    cols = _row_cols(
                        row,
                        stmt_type=st.stmt_type,
                        description=st.description,
                        autoid=(case.autoid if first_case_row else None),
                        priority=(case.priority if first_case_row else None),
                    )
                    first_case_row = False
                    first_row_of_step = False
                else:
                    # 同步骤后续行：不写 C/D/A/B
                    cols = _row_cols(row, stmt_type=None, description=None)
                _set_row(sheet, r, cols)
                r += 1
        # case 间空行
        r += 1

    wb.save(out_path)

    # round-trip：回读对账
    return _readback(out_path)


def _readback(path: Path) -> dict:
    """用纯 openpyxl 回读（同框架 read_excel_with_openpyxl 语义），产出对账统计。"""
    wb = load_workbook(path, data_only=True)
    sheet = wb.active
    grid = [list(row) for row in sheet.iter_rows(values_only=True)]
    anchor = get_config().xlsx.header_anchor
    autoids = []
    case_begin = False
    n_check = 0
    for row in grid:
        a = row[0] if len(row) > 0 else None
        c = row[2] if len(row) > 2 else None
        e = row[4] if len(row) > 4 else None
        if a is not None and str(a).strip() == anchor:
            case_begin = True
            continue
        if not case_begin:
            continue
        if a is not None and str(c) not in ("1", "0", "None"):
            autoids.append(str(a))
        if e == "check_point":
            n_check += 1
    return {
        "path": str(path),
        "rows": len(grid),
        "case_count": len(autoids),
        "autoids": autoids,
        "check_point_count": n_check,
    }
