"""读取测试用例 xlsx 的 D/E/F/G 列，返回行号→{D,E,F,G} 映射 JSON。

用法: python scripts/read_xlsx_rows.py <xlsx_path>
"""
import json
import sys

import openpyxl


def read_rows(xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    header_row = None
    for r in range(1, min(ws.max_row or 0, 100) + 1):
        a, b, c = (str(ws.cell(r, i).value or '').strip() for i in (1, 2, 3))
        # 兼容两种表头格式：标准自动化模板 和 decomposed xlsx
        abc = a + b + c
        if ('自动化ID' in a and '优先级' in b and '语句类型' in c) or \
           ('autoid' in a.lower() and '优先级' in b):
            header_row = r
            break

    if header_row is None:
        print(json.dumps({"error": "header row not found"}, ensure_ascii=False))
        sys.exit(1)

    rows = {}
    for r in range(header_row + 1, (ws.max_row or 0) + 1):
        c_val = str(ws.cell(r, 3).value or '').strip()
        if c_val == '0':
            continue
        d, e, f, g = (str(ws.cell(r, i).value or '').strip() for i in (4, 5, 6, 7))
        if not d and not e and not f:
            continue
        if not e and not f:
            continue
        rows[str(r)] = {'D': d, 'E': e, 'F': f, 'G': g}

    print(json.dumps({
        'header_row': header_row,
        'total_rows': len(rows),
        'rows': rows,
    }, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(json.dumps({"error": "usage: read_xlsx_rows.py <xlsx_path>"}, ensure_ascii=False))
        sys.exit(1)
    read_rows(sys.argv[1])
