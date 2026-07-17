"""批量编译 fan-out 工具：把"逐 case 串行"改成"按阶段批量并行/串行"。

设计依据（见计划 linear-imagining-galaxy.md 决策二）：并行的物理边界落在
**工具内部实现**，不靠 prompt 自律——

- ``compile_fanout``：worker/attributor fork 是纯 LLM + 只读检索/本地写，不碰设备可变态，
  用 ThreadPoolExecutor **并发** fan-out 多个 fork。并发安全已由 P0-S1 验证
  （execute_fork_skill 全局部变量 + LangGraph graph 无状态 + ChatOpenAI/httpx
  并发安全），同一缓存 runnable 可安全被多线程并发 invoke。
- ``dev_run_batch``：上机受跳转机框架全局锁约束（device_mcp_client run_and_wait
  拿不到 task_id 即 "submit failed (lock held?)"），且设备配置/统计是全局共享态，
  **必须单线程串行**。它内部就是一个 for 循环，物理上不可能并发上机。

零硬编码红线：本模块只做"调度 + 结果汇总"，不含任何 APV/sdns 命令、不解析领域
语义、不按关键字分支。命令全程由 draft 子 agent 现场查手册/先例产出。
"""

from __future__ import annotations

import concurrent.futures as cf
import json
import logging
import os
import re as _re
import time
from pathlib import Path

from langchain_core.tools import tool

logger = logging.getLogger(__name__)


def _project_root() -> Path:
    """项目根(workspace 的父)。独立成函数供测试替换沙箱根。"""
    return Path(__file__).resolve().parents[4]

# fan-out 并发上限。draft/grade 是 LLM 调用，并发度受端点限流约束而非 CPU；
# 默认 auto（按待编译数自适应），夹紧到 _MAX_FANOUT 防失控（端点 429 / 把自己打挂）。
_DEFAULT_FANOUT = 4
_MAX_FANOUT = 16

# 单个 fork 的兜底超时（秒）。draft 查手册+emit、grade 判分，给足但不无限挂。
_FORK_TIMEOUT_S = 900

# 429 退避：命中端点限流时指数退避重试，不直接判 fork 失败。
_RATE_LIMIT_MAX_RETRIES = 4
_RATE_LIMIT_BASE_SLEEP_S = 2.0

# 上机互斥(进程内)。起因(2026-07-04 实证):orchestrator 在同一 turn 内把
# dev_run_batch_digest 连发 2-3 次,设备床上多个 pytest 并发互踩配置,产出大片
# 真实但无意义的 fail(三轮结果报废)。框架的全局锁只锁 run 提交窗口,挡不住
# "前一个 client 被 Ctrl-C、设备侧 run 继续跑、新调用又 deliver"的堆积。
# 进程内非阻塞锁 + 设备侧残留探测(下)两层配合,把"同时只有一份在跑"做成不变量。
import threading as _threading
_RUN_MUTEX = _threading.Lock()


def _probe_stale_pytest(env=None) -> str | None:
    """经跳板机 SSH 探测设备床上残留的 ist_staging pytest 进程。

    返回残留描述文本(调用方应拒绝 deliver,附清理指引);无残留返回 None。
    探测本身失败(网络/权限)也返回 None——探测是护栏不是闸门,不能因探测挂掉
    把正常上机也堵死(deliver 后框架自身的锁仍是最后约束)。
    """
    try:
        from main.case_compiler.device_mcp_client import _connect
        c = _connect(env)
        try:
            _, out, _ = c.exec_command(
                "ps -eo pid,lstart,args | grep 'pytest' | grep 'ist_staging' | grep -v grep",
                timeout=15)
            txt = out.read().decode(errors="replace").strip()
        finally:
            try:
                c.close()
            except Exception:  # noqa: BLE001
                pass
        if txt:
            return txt
    except Exception:  # noqa: BLE001
        logger.debug("残留 pytest 探测失败(不阻断)", exc_info=True)
    return None


def _probe_device_reachable(env=None) -> bool | None:
    """经跳板机 ping 被测设备((30) 承载链第零层;run14 实弹:设备批中失联,
    11 案 fail 被 s₀ 配对批量误诊为床污染——污染诊断的前提是设备活着)。

    True=可达 False=不可达 None=探测自身失败(未知,不改判——护栏不是闸门)。
    """
    import os
    ip = os.environ.get("APV_DEVICE_IP", "")
    if not ip:
        return None
    try:
        from main.case_compiler.device_mcp_client import _connect
        c = _connect(env)
        try:
            _, out, _ = c.exec_command(f"ping -c 2 -W 2 {ip} | tail -1", timeout=15)
            txt = out.read().decode(errors="replace")
        finally:
            try:
                c.close()
            except Exception:  # noqa: BLE001
                pass
        if "0% packet loss" in txt and " 0%" in txt:
            return True
        if "100% packet loss" in txt:
            return False
        return None
    except Exception:  # noqa: BLE001
        logger.debug("设备可达性探测失败(未知,不改判)", exc_info=True)
        return None


def _kill_stale_pytest(env=None) -> None:
    """清理设备床上残留的 ist_staging pytest(force_clean=True 的执行动作)。"""
    try:
        from main.case_compiler.device_mcp_client import _connect
        c = _connect(env)
        try:
            c.exec_command("pkill -9 -f 'pytest.*ist_staging'", timeout=15)
            time.sleep(2)
        finally:
            try:
                c.close()
            except Exception:  # noqa: BLE001
                pass
    except Exception:  # noqa: BLE001
        logger.warning("残留 pytest 清理失败", exc_info=True)


def _resolve_concurrency(requested: int, n_items: int = 0) -> int:
    """决定 fan-out 并发度。优先级：env 硬覆盖 > 调用方显式值 > auto(按 item 数自适应)。

    auto（requested<=0）：min(_MAX_FANOUT, max(_DEFAULT_FANOUT, n_items))——
    待编译数少时不空开线程，多时自动铺到上限。墙钟从 Σfork/4 降到 Σfork/min(16,n)。
    """
    env = os.environ.get("IST_FANOUT_CONCURRENCY")
    if env:
        try:
            base = int(env)
        except (TypeError, ValueError):
            base = requested
    elif requested and requested > 0:
        base = requested
    else:
        # auto：按待编译数自适应
        base = max(_DEFAULT_FANOUT, n_items) if n_items > 0 else _DEFAULT_FANOUT
    if base <= 0:
        base = _DEFAULT_FANOUT
    return max(1, min(base, _MAX_FANOUT))


def _is_rate_limit_error(exc: Exception) -> bool:
    """判断异常是否端点限流（429 / rate limit）。靠消息匹配，跨 SDK 兜底。"""
    s = str(exc).lower()
    return "429" in s or "rate limit" in s or "too many requests" in s or "overloaded" in s


def _coerce_json_array(val, name: str):
    """数组参数双收:原生 list(首选,无字符串序列化拖尾暴露面)或 JSON 数组字符串(兼容)。

    返回 (list|None, err|None)。字符串通道经供应商 function-calling 序列化实测
    18-33% 解析失败(autoids_json/briefs_json,2026-07-03 全史 jsonl 取证),原生数组没有这层。
    """
    if isinstance(val, list):
        return val, None
    s = str(val or "").strip()
    if not s:
        return [], None
    try:
        arr = json.loads(s)
    except Exception as exc:  # noqa: BLE001
        return None, f"{name} parse failed: {exc} (pass a native array instead of a JSON string)"
    if not isinstance(arr, list):
        return None, f"{name} must be an array"
    return arr, None


# fanout 单项 output 内联上限(字符)。fork 的机读尾块(STATUS:/ARTIFACT:/VERDICT:)在输出**末尾**,
# 截尾保留 → orchestrator 机读协议不受影响;全文落盘供深挖。
_FANOUT_INLINE_MAX = 2000


def _offload_large_outputs(items: list[dict], skill: str) -> None:
    """出参截断保护:超限的 fork output 全文落 workspace,内联只留末尾+文件指针。

    N 个 fork 的完整输出拼进返回值会随 N×|output| 无界增长——这是"批量出参无
    落盘/摘要"的载荷通道缺口(2026-07-04 评审):入参截断治好了,出参把 orchestrator
    上下文撑爆是同一个病的另一半。key 为 autoid 的落 outputs/<autoid>/(与凭证/冻结
    标记同目录),其余落 outputs/_fanout/。落盘失败时仍截尾(保护不因磁盘失败而失效)。
    """
    import re as _re
    root = _project_root()
    slug = _re.sub(r"[^A-Za-z0-9_.\-]", "_", (skill or "fork"))[:40] or "fork"
    for it in items:
        out = it.get("output")
        if not isinstance(out, str) or len(out) <= _FANOUT_INLINE_MAX:
            continue
        key = str(it.get("key", ""))
        safe = _re.sub(r"[^A-Za-z0-9_.\-]", "_", key)[:60] or "item"
        tail = out[-_FANOUT_INLINE_MAX:]
        try:
            if _re.fullmatch(r"\d{15,}", safe):
                f = root / "workspace" / "outputs" / safe / f"fanout_{slug}.md"
            else:
                f = root / "workspace" / "outputs" / "_fanout" / f"{safe}_{slug}.md"
            f.parent.mkdir(parents=True, exist_ok=True)
            f.write_text(out, encoding="utf-8")
            rel = f.relative_to(root)
            it["output_path"] = str(rel)
            it["output"] = (f"[output of {len(out)} chars exceeds the inline cap; full text at {rel}; "
                            f"below is the trailing {_FANOUT_INLINE_MAX} chars (machine-readable tail block here)]\n…" + tail)
        except Exception:  # noqa: BLE001
            logger.debug("fanout 输出落盘失败(仍截尾保护)", exc_info=True)
            it["output"] = (f"[output of {len(out)} chars exceeds the inline cap and writing to disk failed; "
                            f"keeping only the trailing {_FANOUT_INLINE_MAX} chars]\n…" + tail)


def _xlsx_real_autoids(xlsx_file: str) -> list[str]:
    """扫 xlsx 数据区 A 列的全部真实 case autoid(排除哨兵)。

    与 compile_emit_merged 同款判定(纯数字且 ≥15 位);非数字 id 的特殊卷扫不出
    → 调用方按「全集为空跳过校验」处理,宁漏勿杀。
    """
    out: list[str] = []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_file, read_only=True, data_only=True)
        try:
            for row in wb.active.iter_rows(values_only=True):
                a = str((row[0] if row else "") or "").strip()
                if a.isdigit() and len(a) >= 15 and a != "999999999999999" and a not in out:
                    out.append(a)
        finally:
            wb.close()
    except Exception:  # noqa: BLE001
        logger.debug("xlsx autoid 扫描失败: %s", xlsx_file, exc_info=True)
    return out


@tool(parse_docstring=True)
def compile_fanout(skill: str, briefs_json: list | str = "", briefs_path: str = "",
                   concurrency: int = 0, evidence_from_xlsx: str = "") -> str:
    """Dispatch **one fork skill** concurrently across multiple briefs and collect every sub-agent's output.

    For the parallelizable stages of batch compilation (worker / attributor): one brief per
    case, all run concurrently (bounded by the concurrency cap; excess queue), returning each
    brief's product. ~N× faster than serial invoke_skill (N≈concurrency), with forks fully
    isolated from each other.

    **worker / attributor only** (pure LLM + retrieval/local writes; never touches mutable
    device state). **Never use this for on-device runs** — runs are bound by the framework
    global lock + shared device state and must be serial: use dev_run_batch.

    Each brief is exactly the text you would pass to invoke_skill (requirement + current
    state + rules + pointers + boundaries; no concrete command answers — sub-agents look
    commands up themselves).

    Dual payload channels for briefs (same design as compile_emit steps): **large batches
    (>6 cases) must use the briefs_path file channel** — total brief volume grows with case
    count, and large inline arrays get truncated by vendor serialization (measured: an
    18-case inline batch truncated → forced one-by-one dispatch, concurrency lost).

    Args:
        skill: fork skill to dispatch (e.g. "compile-worker" / "compile-attributor").
        briefs_json: small-batch channel: native array (JSON-array string accepted). Each item
            is a dict with key and brief — key is an identifier (e.g. autoid, only used to map
            outputs back to cases), brief is the full brief text.
        briefs_path: **preferred for large batches**. Path to a JSON file inside workspace
            (e.g. workspace/outputs/<batch>/briefs_wave1.json) holding the same-schema array.
            fs_write / run_python the array to disk first, then pass the path — brief bodies
            bypass inline parameters entirely, zero truncation exposure. A native briefs_json
            array takes precedence if both are given.
        concurrency: concurrency degree. **Default 0=auto** (adaptive to pending count:
            min(16, max(4, N))); pass a positive integer to pin; env IST_FANOUT_CONCURRENCY
            hard-overrides. Clamped to 16 against 429s.
        evidence_from_xlsx: optional. For post-run recompile dispatch pass that xlsx path —
            the tool auto-attaches each key's (=autoid) device_context/causality **verbatim**
            from the sibling last_run.json to the brief tail, eliminating transcription loss
            (a standalone ^ line was measurably lost in retelling, causing misattribution).

    Returns:
        JSON array string. Each item {"key": ..., "ok": bool, "output": "<sub-agent output or
        error>"}, in input order. One fork failing does not affect the others (that item gets
        ok=false); decide what to redo from this. Authoring dispatches (worker/draft,
        key=autoid) also carry "produced": bool — the tool directly checks whether
        outputs/<autoid>/case.xlsx exists on disk; trust it for "was anything produced", not
        prose. When an item's output exceeds the inline cap the full text lands in workspace
        (the item gains "output_path") and only the **tail** stays inline — the fork's
        machine-readable tail block (STATUS:/ARTIFACT:/VERDICT:) sits at the end, so machine
        parsing is unaffected; fs_read the output_path to dig into the full text.
    """
    # briefs 通道优先级(与 compile_emit steps 三通道同款):原生数组 > briefs_path
    # (workspace 文件) > 字符串。文件通道是大批次的主路——briefs 总量 O(N×|brief|),
    # 内联通道在传输层有硬上限,设计上不能假设它吃得下任意批次。
    items: list | None = None
    if isinstance(briefs_json, list) and briefs_json:
        items = briefs_json
    elif (briefs_path or "").strip():
        sp = (briefs_path or "").strip()
        root = _project_root()
        p = Path(sp) if sp.startswith("/") else (root / sp)
        try:
            p = p.resolve()
            ws = (root / "workspace").resolve()
            if not p.is_relative_to(ws):
                return json.dumps({"error": f"briefs_path must be inside workspace/: {sp}"},
                                  ensure_ascii=False)
            if not p.is_file():
                return json.dumps({"error": f"briefs_path file does not exist: {sp}"
                                            " (fs_write the file first, then pass its path)"}, ensure_ascii=False)
            items = json.loads(p.read_text(encoding="utf-8"))
        except Exception as exc:  # noqa: BLE001
            return json.dumps({"error": f"briefs_path read/parse failed: {exc}"}, ensure_ascii=False)
        if not isinstance(items, list):
            return json.dumps({"error": "briefs_path file content must be a JSON array"
                                        " (each item {key, brief})"}, ensure_ascii=False)
    else:
        items, err = _coerce_json_array(briefs_json, "briefs_json")
        if err:
            return json.dumps({"error": err + (
                " Do not inline large batches: write the briefs array to a workspace file"
                " (e.g. workspace/outputs/<batch>/briefs_wave1.json) and pass briefs_path"
                " — the file channel has no truncation exposure.")}, ensure_ascii=False)

    if not items:
        # 空派发是调用错误,不静默成功——orchestrator 漏传参时返回 [] 会被当"派发完成",
        # 清单从此丢失(与"过程事实只存在于散文"同型:错误必须显式,不能靠人看出少了)。
        return json.dumps({"error": "briefs empty: neither briefs_json nor briefs_path carried "
                                    "valid content. Small batches: pass a native array; large "
                                    "batches (>6): write a workspace file and pass briefs_path."},
                          ensure_ascii=False)

    norm: list[dict] = []
    for i, it in enumerate(items):
        if isinstance(it, dict):
            key = str(it.get("key", i))
            brief = str(it.get("brief", ""))
        else:
            key, brief = str(i), str(it)
        norm.append({"key": key, "brief": brief})

    # 上机证据自动注入:key=autoid 时从 last_run.json 取该 case 的原文附到 brief 尾,
    # 替代 LLM 手抄转述(转述会丢独行 ^ 等关键证据)。
    if (evidence_from_xlsx or "").strip():
        try:
            from pathlib import Path as _P
            # 沙箱读闸拒绝(路径越界/黑名单)= 直接放弃注入——注入本是 best-effort,
            # 绝不回退原始路径读盘(2026-07-05 安全评审:except 吞拒绝再裸读是读闸旁路)。
            from main.ist_core.tools.deepagent.file_tools import _resolve_inside_root
            _xp = _resolve_inside_root(evidence_from_xlsx, must_exist=True)
            _lr = _P(_xp).parent / "last_run.json"
            if _lr.is_file():
                _recs = {str(r.get("autoid")): r
                         for r in json.loads(_lr.read_text(encoding="utf-8"))
                         if isinstance(r, dict)}
                for it in norm:
                    r = _recs.get(it["key"])
                    if not r:
                        continue
                    ev = (r.get("device_context") or r.get("causality") or "").strip()
                    if ev:
                        block = ('<device_evidence source="last_run.json" '
                                 'note="verbatim, injected by the tool, not retold">\n'
                                 f"{ev[:6000]}\n</device_evidence>")
                        # 长数据置顶(官方长上下文实践):插在机读信封首行之后——信封保持
                        # 首行供卡片/解析读取,长证据紧随其后,指令留在消息末尾。
                        b = it["brief"]
                        first, sep, rest = b.partition("\n")
                        if first.lstrip().startswith("{"):
                            it["brief"] = first + "\n" + block + (("\n" + rest) if sep else "")
                        else:
                            it["brief"] = block + "\n" + b
        except Exception:  # noqa: BLE001
            logger.debug("fanout 证据注入失败(跳过)", exc_info=True)

    skipped: list[dict] = []

    from main.ist_core.skills.loader import execute_fork_skill

    workers = _resolve_concurrency(concurrency, n_items=len(norm))
    logger.info("compile_fanout skill=%s items=%d concurrency=%d skipped=%d",
                skill, len(norm), workers, len(skipped))

    def _run(item: dict) -> dict:
        # tag=per-item 归属标识(2026-07-06):旧版不传,N worker 的 fastlog 行全是同一个
        # skill 名 label、无法分辨哪行属于哪个 case;格式对齐引擎 `engine:{aid[-6:]}`——
        # `{skill短名}:{autoid尾6}`(非 autoid key 取前 12)。
        _k = str(item.get("key") or "").strip()
        _short = (skill or "fork").replace("ist-compile-", "").replace("compile-", "")
        _tag = f"{_short}:{_k[-6:]}" if len(_k) == 18 and _k.isdigit() else f"{_short}:{_k[:12]}"
        last_exc = None
        for attempt in range(_RATE_LIMIT_MAX_RETRIES + 1):
            try:
                out = execute_fork_skill(skill, item["brief"], tag=_tag)
                ok = not (isinstance(out, str) and out.startswith("ERROR:"))
                return {"key": item["key"], "ok": ok, "output": out}
            except Exception as exc:  # noqa: BLE001
                last_exc = exc
                if not _is_rate_limit_error(exc):
                    logger.exception("fanout fork %s key=%s failed", skill, item["key"])
                    return {"key": item["key"], "ok": False, "output": f"ERROR: {exc}"}
                if attempt < _RATE_LIMIT_MAX_RETRIES:
                    sleep_s = _RATE_LIMIT_BASE_SLEEP_S * (2 ** attempt)
                    logger.warning("fanout fork %s key=%s 命中限流(429)，第%d次退避 %.0fs",
                                   skill, item["key"], attempt + 1, sleep_s)
                    time.sleep(sleep_s)
        return {"key": item["key"], "ok": False,
                "output": f"ERROR: rate-limit retries exhausted ({_RATE_LIMIT_MAX_RETRIES}x): {last_exc}"}

    results: dict[str, dict] = {}
    with cf.ThreadPoolExecutor(max_workers=workers) as ex:
        futs = {ex.submit(_run, it): it["key"] for it in norm}
        for fut in cf.as_completed(futs, timeout=_FORK_TIMEOUT_S * len(norm)):
            key = futs[fut]
            try:
                r = fut.result(timeout=_FORK_TIMEOUT_S)
            except Exception as exc:  # noqa: BLE001
                r = {"key": key, "ok": False, "output": f"ERROR: fork timeout/exception: {exc}"}
            results[key] = r

    # 保持输入顺序
    ordered = [results.get(it["key"], {"key": it["key"], "ok": False,
                                       "output": "ERROR: no result"}) for it in norm]
    ordered += skipped
    # 「产没产出」以落盘为准,工具直接探(编写类派发 key=autoid):worker 说产出了但盘上
    # 没有、或返回是调试残句但盘上有——散文与事实曾张冠李戴,produced 字段是机读事实源。
    if "worker" in (skill or "").lower() or "draft" in (skill or "").lower():
        _root = _project_root()
        for r in ordered:
            aid = str(r.get("key", "")).strip()
            if len(aid) == 18 and aid.isdigit():
                r["produced"] = (_root / "workspace" / "outputs" / aid / "case.xlsx").is_file()
    _offload_large_outputs(ordered, skill)
    return json.dumps(ordered, ensure_ascii=False)


# 上机超时。整份 xlsx **一次**提交即跑完其中全部 case(框架把整份当套件整跑),故超时是**整份级**、
# 随 case 数自适应(单 case 含 sleep/多 dig 约 ~45s),夹紧到总上限,防大文件跑不完被误判超时。
_RUN_DEFAULT_MAX_S = 600
_RUN_MAX_MAX_S = 1200          # 兼容旧签名(单 case 语义)的夹紧上限
_PER_CASE_BUDGET_S = 45        # 整份估时:每 case 预算秒
_RUN_TOTAL_CAP_S = 2400        # 整份总超时硬上限(40min)


@tool(parse_docstring=True)
def dev_run_batch(xlsx_path: str, autoids_json: list | str = "", module: str = "",
                 build: str = "", max_s_each: int = _RUN_DEFAULT_MAX_S,
                 force_clean: bool = False) -> str:
    """Run **one merged xlsx on-device in a single submission**, returning every case's verdict + the framework's real ruling.

    **Single whole-file run (the O(N) fix)**: the framework ``test_xlsx.py`` treats the
    delivered xlsx as one **suite** — submitting **one** autoid executes **every** case in the
    file sequentially, and all inner cases' per-check_point logs land under that submitted
    autoid's staging. So this tool **delivers+runs exactly once**, then reads back all
    autoids' verdicts from that staging; it never re-runs the whole file per autoid (the old
    implementation was O(N²) and large files hit the 600s polling cap with no result — the
    root cause of "ran 20 minutes, nothing back").

    **The serial hard constraint still holds**: the jumphost framework has a global run lock
    and device state is globally shared — one on-device task at a time; this tool submits one
    xlsx per call, physically non-concurrent.

    verdict comes from each case's own log per check_point: no ``#### Fail Num`` and
    ``#### Success Num`` >0 → pass; any Fail → fail; no log → unknown (not reached/skipped).
    Cases containing ``<RUNTIME>`` placeholders always fail the first run (the framework
    searches for the literal "<RUNTIME>") — expected, refilled by ist-verify and rerun.

    Args:
        xlsx_path: local path of the merged case.xlsx (multiple real cases + trailing sentinel).
        autoids_json: **prefer a native array** (JSON-array string accepted; omit = whole
            volume). The autoids to fetch verdicts for — validated against the xlsx's actual
            autoid set, with unknown ids rejected explicitly (guards against hand-copied
            truncated ids silently substring-matching log filenames into fake passes).
        module: staging submodule (default: compiler config staging_module).
        build: target device build (default: compiler config build).
        max_s_each: legacy signature compat — treated as the whole-file budget floor; total
            timeout = clamp(max(it, N×45s), …, 2400s).
        force_clean: when stale pytest processes linger on the device bed, the tool refuses to
            run by default and reports them (concurrent pytest runs trample each other's
            config, producing masses of meaningless fails); pass True only after confirming
            the leftover run is abandoned, to clean up and rerun.

    Returns:
        JSON array string, each item {"autoid", "verdict", "task_id", "causality"
        (check_point ruling lines), "detail_tail", plus "device_context" for non-pass},
        in input order.
    """
    import re as _re
    from pathlib import Path

    arr, err = _coerce_json_array(autoids_json, "autoids_json")
    if err:
        return json.dumps({"error": err}, ensure_ascii=False)
    autoids = [str(a).strip() for a in (arr or []) if str(a).strip()]

    # 复用 dev_run_case 的多根路径解析（agent 沙箱视角）
    p = None
    try:
        from main.ist_core.tools.deepagent.file_tools import _resolve_inside_root
        p = _resolve_inside_root(xlsx_path, must_exist=True)
    except Exception:  # noqa: BLE001
        logger.debug("xlsx 路径解析失败(将回退兜底): %s", xlsx_path, exc_info=True)
        p = None
    if p is None or not Path(p).is_file():
        cands = [Path(xlsx_path)]
        if not Path(xlsx_path).is_absolute():
            root = Path(__file__).resolve().parents[4]
            cands += [root / xlsx_path, root / "knowledge" / "data" / xlsx_path]
        p = next((c for c in cands if c.is_file()), None)
    if p is None or not Path(p).is_file():
        return json.dumps({"error": f"xlsx not found: {xlsx_path}"}, ensure_ascii=False)
    p = Path(p)

    # autoid 与 xlsx 实际卷内全集对账(A 层校验):空=全卷;不在卷内的显式报错——
    # 曾实证 LLM 手抄把 203031753342778012 截成 "778012",旧版靠日志文件名子串
    # 静默误匹配成 pass,错 id 一路进最终报告且跨轮对照失效。
    real = _xlsx_real_autoids(str(p))
    if not autoids:
        if not real:
            return json.dumps({"error": "no autoids given and none found in the xlsx data area; pass a native array explicitly"},
                              ensure_ascii=False)
        autoids = real
    elif real:
        unknown = [a for a in autoids if a not in real]
        if unknown:
            return json.dumps({"error": (
                f"autoids not in this xlsx data area (hand-copy error/truncated?): {', '.join(unknown)}. "
                f"The volume actually holds {len(real)}: {', '.join(real[:6])}{'…' if len(real) > 6 else ''}; "
                "omit the autoids parameter to fetch verdicts for the whole volume.")}, ensure_ascii=False)

    try:
        from main.case_compiler.config import get_config
        cfg = get_config()
        module = (module or cfg.staging_module).strip()
        build = (build or cfg.build).strip()
    except Exception as exc:  # noqa: BLE001
        return json.dumps({"error": f"failed to read compiler config: {exc}"}, ensure_ascii=False)

    # 整份总超时随 case 数自适应（单 case 含 sleep/多 dig ~45s），夹紧到硬上限。
    try:
        floor = int(max_s_each or _RUN_DEFAULT_MAX_S)
    except (TypeError, ValueError):
        floor = _RUN_DEFAULT_MAX_S
    total_max = max(floor, len(autoids) * _PER_CASE_BUDGET_S)
    total_max = max(_RUN_DEFAULT_MAX_S, min(total_max, _RUN_TOTAL_CAP_S))

    try:
        from main.case_compiler.device_mcp_client import FrameworkMCPClient
    except Exception as exc:  # noqa: BLE001
        return json.dumps({"error": f"failed to load FrameworkMCPClient: {exc}"}, ensure_ascii=False)

    # 进程内互斥:同一进程里已有一份上机在跑 → 立即拒绝,绝不排队叠加。
    # (2026-07-04 实证:orchestrator 同 turn 连发 2-3 次 digest,设备床多 pytest 互踩,
    # 三轮结果报废。上机是独占设备床的物理动作,重复调用没有任何正确语义。)
    if not _RUN_MUTEX.acquire(blocking=False):
        return json.dumps({"error": "run_in_progress", "busy": True, "message": (
            "an on-device run is already executing in this process — runs own the device bed "
            "exclusively, one at a time. Do not re-call dev_run_batch/digest; wait for the "
            "current run to return.")},
            ensure_ascii=False)

    submit = autoids[0]   # 整份只用一个 autoid 提交，框架据它建 staging 并整跑全文件
    out: list[dict] = []
    import contextlib as _ctx
    try:
        from main.case_compiler import env_pool as _pool
    except Exception:  # noqa: BLE001
        logger.debug("加载 env_pool 失败，回退单环境模式", exc_info=True)
        _pool = None
    try:
        with _ctx.ExitStack() as _stack:
            _stack.callback(_RUN_MUTEX.release)
            # 环境池：启用时认领一个空闲就绪环境（各自独立设备床→真并行）；
            # 未启用/池异常→ env=None 回退现役单环境（行为同今天）；全忙→device_busy。
            env = None
            if _pool is not None and _pool.is_enabled():
                try:
                    env = _stack.enter_context(_pool.acquire(timeout=300))
                except TimeoutError:
                    return json.dumps({"error": "device_busy", "busy": True,
                                       "message": "all automation environments are busy; retry later."},
                                      ensure_ascii=False)
                except Exception:  # noqa: BLE001
                    logger.warning("环境池 acquire 异常，回退单环境", exc_info=True)
                    env = None
            # 设备侧残留探测:上一次跑批被打断时,client 死了但设备上的 pytest 还活着——
            # 此时 deliver 新卷会两份并发互踩配置,产出大片真实但无意义的 fail,且新
            # digest 可能收割到旧执行的日志(2026-07-04 三轮实证)。有残留默认拒绝;
            # 确认残留是弃跑后用 force_clean=True 清场重跑。
            # (30) 承载链第零层前置:设备不可达时上机=整轮盲跑零信息(run14 实弹:
            # 设备批中失联后若再复跑,只会产一批 broken 烧 12min 设备轮)。拒跑+
            # 人话指引;探测自身失败(None)不拦(护栏不是闸门)
            if _probe_device_reachable(env) is False:
                return json.dumps({"error": "device_unreachable", "busy": False, "message": (
                    "the device under test is unreachable from the jumphost (ping 100% "
                    "loss) — running now would burn a full device round producing only "
                    "broken results. Restore the device (console/power), verify with "
                    "ping, then re-run.")}, ensure_ascii=False)
            stale = _probe_stale_pytest(env)
            if stale and force_clean:
                _kill_stale_pytest(env)
                stale = _probe_stale_pytest(env)
            if stale:
                return json.dumps({"error": "stale_run_on_device", "busy": True, "message": (
                    "stale pytest processes are still running on the device bed (an interrupted "
                    "batch left them alive):\n"
                    + stale[:500]
                    + "\nRunning now would have two suites trampling each other's config — all "
                      "results distorted. Wait for it to finish naturally, or, once confirmed "
                      "abandoned, re-call this tool with force_clean=True to clean up and rerun.")},
                    ensure_ascii=False)
            client = _stack.enter_context(FrameworkMCPClient(env))
            dres = client.deliver(module, submit, str(p))
            if dres.get("error"):
                return json.dumps({"error": f"deliver failed: {dres.get('error')}"}, ensure_ascii=False)
            # run-identity 基线:deliver 时刻的跳板机 epoch。staging 目录跨 run 复用,
            # 上次被打断执行的旧日志会留存;收割时 mtime 早于此基线的日志判 stale,
            # 不产 verdict(2026-07-04 实证:收割旧执行日志 → 0/34、1/34 两轮假结果)。
            # getattr 兼容:旧 client 实现/测试替身没有该方法时基线=0(过滤不启用,行为同旧)。
            deliver_epoch = getattr(client, "jumphost_epoch", lambda: 0.0)()

            # 跑批进度 → evidence fastlog（TUI 300ms tail 同一文件即实时显示，零 TUI 改动）。
            # 降噪：完成数变化或 ≥30s 心跳才写一行；任何异常静默——可观测性不拖垮跑批。
            # 双写(2026-07-06):人读 ▸ 行照旧(tail -f 契约);结构化 progress 事件带稳定
            # key=runbatch:{submit}——TUI 卡片模式同 key 恒一行原地更新,不再 48 条心跳平铺。
            _t0 = time.time()
            _prog_key = f"runbatch:{submit}"
            _prog_state = {"sig": "", "ts": 0.0}
            def _on_poll(st: dict) -> None:
                try:
                    import re
                    from main.ist_core.skills.loader import _fork_emit, _fork_emit_event
                    now = time.time()
                    # 整卷单跑模式下 len(results) 开跑即满(框架一次性建全 per-case 条目),
                    # 旧版拿它当分子显示「34/34」恒满假进度(2026-07-03 实证)。改为诚实
                    # 口径:已跑时长/总预算 + 框架日志尾(真实推进信号在日志里)。
                    _log = st.get("log_tail") or ""
                    tail_lines = _log.strip().splitlines()
                    tail_txt = tail_lines[-1].strip()[-70:] if tail_lines else ""
                    # 当前在跑第几个 case:首选框架的语义行「begin case: <autoid>」
                    # (2026-07-12 #66 实证:日志有时只有 .py 路径行含 autoid,旧版
                    # 「排除 .py 行防恒 1/26」把唯一来源也排掉→进度恒 0);语义行
                    # 缺失时回落通用扫描(仍排路径行防恒 1/26)。
                    _cur_idx = 0
                    _begins = re.findall(r"begin case:\s*(\d{18})", _log)
                    if _begins and _begins[-1] in autoids:
                        _cur_idx = autoids.index(_begins[-1]) + 1
                    else:
                        _scan = "\n".join(ln for ln in _log.splitlines()
                                          if ".py" not in ln)
                        for _id in reversed(re.findall(r"(?<!\d)(\d{18})(?!\d)", _scan)):
                            if _id in autoids:
                                _cur_idx = autoids.index(_id) + 1
                                break
                    _env_host = getattr(client, "host", "") or ""
                    # case 级超时嫌疑(§18.5,xUnit per-case timeout 的采集面近似):
                    # 心跳能定位当前案——同案持续超阈=hang 嫌疑,记入 _prog_state
                    # 供 digest 出口附 timeout_suspect(T1 三症状 hang/污染/真fail
                    # 机械可分的第一信号;框架无 per-case 预算,只能采集侧判)
                    _thresh = int(os.environ.get("IST_CASE_TIMEOUT_SUSPECT", "300"))
                    if _cur_idx and _cur_idx == _prog_state.get("case_idx"):
                        if now - _prog_state.get("case_t0", now) > _thresh:
                            _sus = _prog_state.setdefault("timeout_suspects", {})
                            _aid_cur = autoids[_cur_idx - 1]
                            _sus[_aid_cur] = int(now - _prog_state.get("case_t0", now))
                    elif _cur_idx:
                        _prog_state["case_idx"] = _cur_idx
                        _prog_state["case_t0"] = now
                    sig = f"{tail_txt}|{_cur_idx}"
                    if sig == _prog_state["sig"] and now - _prog_state["ts"] < 30:
                        return
                    _prog_state["sig"], _prog_state["ts"] = sig, now
                    _prog_seg = (f"第{_cur_idx}/{len(autoids)}" if _cur_idx
                                 else f"整卷 {len(autoids)} case 单跑")
                    tail = f" · {tail_txt}" if tail_txt else ""
                    _fork_emit(f"▸ 上机运行 {int(now - _t0)}s/{total_max}s"
                               f"(环境 {_env_host} · {_prog_seg}){tail}")
                    _fork_emit_event({"event": "progress", "key": _prog_key,
                                      "phase": "上机", "elapsed_s": int(now - _t0),
                                      "total_s": int(total_max), "n_cases": len(autoids),
                                      "env": _env_host, "case_idx": _cur_idx,
                                      "detail": tail_txt, "status": "running"})
                except Exception:  # noqa: BLE001
                    pass

            run = client.run_and_wait(module, submit, build, autoids, max_s=total_max,
                                      progress_cb=_on_poll)
            try:
                from main.ist_core.skills.loader import _fork_emit_event as _fee
                _run_err = run.get("error") or ("device_busy" if run.get("busy") else "")
                _fee({"event": "progress", "key": _prog_key, "phase": "上机",
                      "elapsed_s": int(time.time() - _t0), "total_s": int(total_max),
                      "n_cases": len(autoids),
                      "detail": str(_run_err)[:70] if _run_err else "完成",
                      "status": "error" if _run_err else "done"})
            except Exception:  # noqa: BLE001
                pass
            if run.get("busy") or run.get("error") == "device_busy":
                return json.dumps({"error": "device_busy", "busy": True,
                                   "message": run.get("message") or "environment busy: a previous case is still being verified; retry later."},
                                  ensure_ascii=False)
            task_id = run.get("task_id", "")
            run_err = run.get("error")
            # 无论 run 是否 done（可能撞总超时仍 running），都尽量读回已写出的逐 case 日志。
            # min_epoch=基线-3s(容 stat 秒级粒度):早于本次 deliver 的日志是上次执行残留,
            # 值为 STALE_LOG_MARK → 判 unknown 并显式标注,绝不当本次结果采信。
            # TypeError 回退:旧 client/测试替身签名不收 min_epoch 时按旧行为全收。
            try:
                details = client.fetch_batch_details(
                    submit, min_epoch=(deliver_epoch - 3) if deliver_epoch else 0)
            except TypeError:
                details = client.fetch_batch_details(submit)
            stale_mark = getattr(client, "STALE_LOG_MARK", "<<STALE_LOG>>")
            for autoid in autoids:
                d = details.get(autoid, "")
                if d == stale_mark:
                    out.append({"autoid": autoid, "verdict": "unknown", "task_id": run.get("task_id", ""),
                                "causality": "", "detail_tail": (
                                    "stale_log: this case's staging log predates this deliver — it is "
                                    "residue of a previous execution; this run never reached the case "
                                    "(the volume may have crashed/timed out midway). Do not attribute "
                                    "from this log.")})
                    continue
                succ = len(_re.findall(r"#### Success\s*Num", d))
                fail = len(_re.findall(r"#### Fail\s*Num", d))
                if fail > 0:
                    verdict = "fail"
                elif succ > 0:
                    verdict = "pass"
                else:
                    verdict = "unknown"   # 无日志：未执行到/被跳过(如 test_env 主机不支持)/超时未跑到
                causality = [ln.rstrip() for ln in d.splitlines()
                             if _re.search(r"(Success|Fail)\s*Num|fail to find|successed to find",
                                           ln, _re.IGNORECASE)]
                # (43) ok(g) 谓词扫描(§18.5,文法数据 exec_failure_markers,零硬编码):
                # pass 案日志含执行失败标记=空真嫌疑((44)违例,668030 形态)→降 broken;
                # fail 案的标记行提取为 anomaly_lines(独立共因对归因可见,坑#24 平权)。
                # echo-grounding 修(2026-07-13):扫描范围含**主设备完整会话**——失败机理
                # (如 write all 撞文件的 'Failed to execute the command'/'Type YES')在
                # apv_*.txt,不在断言摘要 d 里;只扫 d 会漏掉自身执行失败,让 s₀ 归因把
                # 「自身命令流错位」误判成床污染(668030 实证)。fail 案 device_context 提前
                # 拉一次(含 apv 会话),复用给 anomaly 扫描与 rec——不增 SSH 往返。
                _dev_ctx = ""
                if verdict != "pass":
                    _dev_ctx = client.fetch_device_context_under(submit, autoid)
                _markers = _exec_failure_markers()
                _scan = d + ("\n" + _dev_ctx if _dev_ctx else "")
                _anom = ([ln.strip() for ln in _scan.splitlines()
                          if any(m in ln for m in _markers)][:8] if _markers else [])
                if _anom and verdict == "pass":
                    verdict = "broken"
                # oracle 残差门(§18.10,2026-07-14):框架判定不再是公理——旁路取
                # 原始 inner+apv 全文,按提示符重分段对账。方向矛盾(found 判 fail 但
                # 对齐块含期望值/not_found 判过但块含该值)=采集面失真 → verdict 降
                # broken(第三态:挡写回/挡归因/挡 s₀ 配对——668000 假 PASS 三连写回
                # 投毒的直接防线)。块级 ^ 拒绝(设备解析器标记,闭集)并入 anomaly_lines
                # (G 族自身异常:喂 G6 否决免派 + brief 高亮)。pass 案也审(假 PASS
                # 不呼救,症状驱动的流程结构性漏掉静默类)。旧 client/测试替身无此方法
                # → 审计不可用,rec 显式标注(INV-11:不静默 fail-open)。
                _dist: list = []
                _wa_note = ""
                try:
                    _raw = client.fetch_case_raw(submit, autoid)
                except (AttributeError, TypeError):
                    _raw = None
                if _raw and _raw.get("inner"):
                    _dist, _caret = _window_audit(_raw["inner"], _raw.get("apv") or {})
                    for _a in _caret:
                        if _a not in _anom and len(_anom) < 8:
                            _anom.append(_a)
                else:
                    _wa_note = "unavailable (no raw case files or legacy client)"
                if _dist and verdict in ("pass", "fail"):
                    if verdict == "pass" and not _dev_ctx:
                        _dev_ctx = client.fetch_device_context_under(submit, autoid)
                    verdict = "broken"
                rec = {"autoid": autoid, "verdict": verdict, "task_id": task_id,
                       "causality": "\n".join(causality[-12:]) if causality else "",
                       "detail_tail": d[-2500:]}
                if _anom:
                    rec["anomaly_lines"] = _anom
                    if verdict == "broken" and not _dist:
                        rec["broken_reason"] = (
                            "execution-failure marker in a passing case's log — the "
                            "assertions after the failed step are vacuous ((44)); the "
                            "case's target behavior was not actually verified")
                if _dist:
                    rec["window_distortion"] = _dist[:4]
                    rec["broken_reason"] = (
                        "assertion-window distortion: the framework's check window "
                        "disagrees with the raw device stream (re-segmented by prompt) "
                        "— the verdict direction contradicts the aligned response "
                        "block; neither the pass nor the fail is trustworthy here")
                if _wa_note:
                    rec["window_audit"] = _wa_note
                # pyATS 七码子分类(DESIGN_dongkl_finalization §④):给 broken 打协议级
                # 硬码 broken_subtype——引擎侧(views/reconcile)据此细分处置。这里只认
                # **已机械判定**的协议信号,不做语义猜测(守 (44)):
                #   window-audit 失真(_dist:断言被对齐证据反证,确定性缺陷)
                #   / 执行失败标记(_anom:exec-failure/裸 ^ 语法拒绝)= Errored → reflow 重写;
                #   其余(not_run/stale/协议级分不清)留空 → 引擎落 S_BROKEN 复跑(安全默认);
                #   device_unreachable(承载链探测,下方)覆盖为 blocked → env 呈报。
                if verdict == "broken" and (_dist or _anom):
                    rec["broken_subtype"] = "errored"
                _sus = (_prog_state.get("timeout_suspects") or {}).get(autoid)
                if _sus:
                    rec["timeout_suspect_s"] = _sus   # 心跳同案超阈(§18.5,hang 嫌疑)
                if verdict != "pass":
                    rec["device_context"] = _dev_ctx
                    if run_err and not d:
                        rec["detail_tail"] = (f"(no case log; run state={run_err})\n" + rec["detail_tail"])
                out.append(rec)
            # (30) 承载链第零层(run14 实弹修):批内出现 fail 时探设备可达性——
            # 不可达=fail 全是设备失联的下游症状,非案缺陷非床污染,全部降 broken
            # (device_unreachable);禁 s₀ 配对在死设备批上批量误诊(11 案实证)
            if any(r["verdict"] == "fail" for r in out):
                _reach = _probe_device_reachable()
                if _reach is False:
                    for r in out:
                        if r["verdict"] != "pass":
                            r["verdict"] = "broken"
                            r["device_unreachable"] = True
                            # pyATS Blocked 子类(§④):设备不可达 → env 呈报(复跑救不了
                            # 死设备)。覆盖此前可能打的 errored(承载链第零层优先)。
                            r["broken_subtype"] = "blocked"
                            r["broken_reason"] = (
                                "device unreachable (ping 100% loss from jumphost) — "
                                "this failure is a downstream symptom of device loss, "
                                "not a case defect nor bed pollution; restore the "
                                "device, then resume")
            # 文件级崩溃可见性：有 unknown（某 case 把整份 pytest 搞崩、后续全不跑）→ 取框架 task 日志
            # 的 traceback 附到 unknown 上，让 agent 看到“崩在哪一行/什么异常”，而非只看到一堆无解释的 unknown。
            if task_id and any(r["verdict"] == "unknown" for r in out):
                tb = client.fetch_task_log_errors(task_id)
                if tb:
                    for r in out:
                        if r["verdict"] == "unknown":
                            r["framework_traceback"] = tb
    except Exception as exc:  # noqa: BLE001
        return json.dumps({"error": f"batch on-device run exception: {exc}", "partial": out}, ensure_ascii=False)

    return json.dumps(out, ensure_ascii=False)


def _scan_xlsx_for_check_method(xlsx_path: str, method: str) -> list:
    """扫 xlsx 找用了某 check_point 方法（如 found_times）的行，返回 [(autoid, 行号)]。

    供 digest 在识别到文件级崩溃（某断言崩整份 pytest）时**点名元凶 case**，让归因落到
    具体 autoid，而非笼统"框架崩了"。失败静默返回空（点名是增益、非硬依赖）。
    """
    try:
        import openpyxl
        from main.ist_core.tools.deepagent.file_tools import _resolve_inside_root
        p = _resolve_inside_root(xlsx_path, must_exist=True)
        ws = openpyxl.load_workbook(p, data_only=True).active
        hits, cur = [], None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if cells and cells[0].isdigit() and len(cells[0]) >= 12:
                cur = cells[0]
            if method in cells and "check_point" in cells:
                hits.append((cur, i))
        return hits
    except Exception:  # noqa: BLE001
        return []


def normalize_fail_signature(sig: str) -> str:
    """签名归一化——新旧格式签名做交集比较前，两侧共同过它（A1 迁移条款）。

    2026-07-16 结构化解析上线前的存量签名是裸 grep "fail to find" 产物：真 Fail 项常带
    逐步形态的 `` in: <file>`` 尾（如 ``p2 in: xxx.txt``），与新解析的纯 pattern 文本
    （``p2``）逐字比较交集恒空 → 冻结/跨床反驳在跨界轮静默失效。归一化 = 空白压缩 +
    剥 `` in: …`` 尾 + 重截 60，对新格式幂等。本文件的跨轮交集从原文现场重提取、天然
    同格式；真正的跨格式点在引擎读**存量字段**（facts verdict ``signatures`` /
    last_run ``_fail_signatures``）做交集处——那侧比较前 import 本函数对两侧归一。
    旧侧的节头假行/Success 项归一不出对齐——它们是多余项，不阻碍「真 Fail 项对齐后
    交集非空」；唯一盲区是纯 not_found 失败案（旧格式根本没采到真 Fail 项，无从对齐），
    该形态冻结迟一轮触发（fail-safe 方向），下一轮双侧同格式自愈。
    """
    s = _re.sub(r"\s+", " ", str(sig or "")).strip()
    s = _re.sub(r" in ?: ?\S*\s*$", "", s).strip()
    return s[:60]


def _fail_signatures_legacy(text: str) -> set[str]:
    """旧签名抽取（裸 grep，2026-07-16 前唯一实现）——仅兼容腿/开关回退用。

    已证病灶（dongkl 9 案 4 案签名脏 + 778072 语义反转）：B1 无分隔拼接收节头假行、
    B2 收通过的 not_found 断言（Success 行含 "fail to find" 词面）、且漏收 not_found
    断言失败（``Fail … successed to find``）。"""
    return {m.group(1).strip()[:60]
            for m in _re.finditer(r"fail to find:?\s*([^\r\n]{1,80})", text or "")}


def _fail_signatures(text: str) -> set[str]:
    """从裁决明细抽 fail 签名集合（按框架裁决行 ``#### (Fail|Success) Num`` 结构化解析）。

    跨轮对照用：两轮签名集合**交集非空** = 同签名 fail（同一断言以同样方式不中）。
    仅 group(1)==Fail 的行取 pattern 文本（normalize 后入集合）——含
    ``Fail … successed to find``（not_found 断言失败，旧裸 grep 按词面漏收）；
    Success 行（通过的 not_found，B2 病灶）与一切非裁决行（节头/文件名/RTNETLINK，
    B1 病灶）锚定 ``^####`` 天然出局。裁决行两形态都收：逐步
    ``… find[:]? <pat> in ?: <file>``（_WA_CHECK_RE，pattern 与文件名可分）与案末汇总
    ``… find: <pat>``（无 in 尾，_WA_SUMMARY_RE——dongkl 778072 实证 Fail 汇总行即此形）。
    兼容腿：全文零条结构化裁决行（老日志/异构框架版本改 ``####`` 前缀）回退旧正则并
    warning 留声（签名恒空=冻结门静默失效，比误冻结安全但要可见）；
    ``IST_FAIL_SIG_STRUCTURED=0`` 整体回退（跨版本对照/紧急逃生，默认开）。
    """
    if os.environ.get("IST_FAIL_SIG_STRUCTURED", "1").strip().lower() in ("0", "false", "no"):
        return _fail_signatures_legacy(text)
    out: set[str] = set()
    saw_verdict_line = False
    for raw in (text or "").splitlines():
        ln = _WA_TS_RE.sub("", raw.strip())
        m = _WA_CHECK_RE.match(ln) or _WA_SUMMARY_RE.match(ln)
        if not m:
            continue
        saw_verdict_line = True
        if m.group(1) != "Fail":
            continue
        sig = normalize_fail_signature(m.group(3))
        if sig:
            out.add(sig)
    if not saw_verdict_line:
        legacy = _fail_signatures_legacy(text)
        if legacy:
            logger.warning(
                "fail-signature: 0 structured verdict lines but legacy 'fail to find' text "
                "matched (%d sigs) — old-format log or framework verdict-line format drift",
                len(legacy))
        return legacy
    return out


def _xlsx_apv_lines(xlsx_path) -> dict[str, list[str]]:
    """从(合并)卷机械抽取每个 autoid 的 APV 命令列表(E∈{APV_0,APV_1} 行的 G 列,
    cmds_config 多行逐条拆)。verified_runs 台账的 apv_cmds 字段来源——device_verified
    权威源用它交叉校验"写回的命令真实出现在 PASS 卷面上"。"""
    import openpyxl
    out: dict[str, list[str]] = {}
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            begun = False
            cur = ""
            for row in ws.iter_rows(min_row=2, values_only=True):
                a = str(row[0] or "").strip() if row else ""
                if not begun:
                    begun = a == "自动化ID"
                    continue
                # autoid 行=18 位纯数字(勿按前缀认:曾硬编 "203" 前缀,204 批全体 autoid
                # 不被识别→apv_cmds 恒空→device_verified 写回/行为晋升对 204 批静默失效,
                # 2026-07-08 实测台账 105/105 条空 vs 203 批 26/26 全有)
                if len(a) == 18 and a.isdigit():
                    cur = a
                    out.setdefault(cur, [])
                if not cur or len(row) < 7:
                    continue
                e = str(row[4] or "").strip()
                if e in ("APV_0", "APV_1"):
                    for line in str(row[6] or "").splitlines():
                        line = line.strip()
                        if line:
                            out[cur].append(line)
    finally:
        wb.close()
    return out


def _append_verified_runs(xlsx_path, results: list, cur_round: int, run_ts: float,
                          build: str = "") -> None:
    """运行台账(V6 支柱2a):逐 case 追加 runtime/logs/verified_runs.jsonl。

    build 字段=本次 run 提交用的目标 build(K 锚三元组 (build, run_ts, lineage) 的
    build 位,理论 §5.1)——写回链经 device_run_ref 透传进 footprint 条目的
    evidence.device_run,后续 build 锚差派生 stale 判定靠它;空串=当次未解析出。"""
    apv = _xlsx_apv_lines(xlsx_path)
    ledger = _project_root() / "runtime" / "logs" / "verified_runs.jsonl"
    ledger.parent.mkdir(parents=True, exist_ok=True)
    xstat = Path(str(xlsx_path)).stat()
    with ledger.open("a", encoding="utf-8") as f:
        for rec in results:
            if not isinstance(rec, dict) or not rec.get("autoid"):
                continue
            aid = str(rec["autoid"])
            f.write(json.dumps({
                "autoid": aid, "verdict": str(rec.get("verdict", "")),
                "run_ts": run_ts, "round": cur_round,
                "xlsx": str(xlsx_path), "xlsx_mtime": xstat.st_mtime,
                "build": (build or "").strip(),
                "apv_cmds": apv.get(aid, []),
            }, ensure_ascii=False) + "\n")


_EXEC_MARKERS_CACHE: list | None = None


_WA_TS_RE = _re.compile(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2} ")
_WA_CHECK_RE = _re.compile(
    r"^#### (Fail|Success) Num \d+: (fail to find|successed to find):? (.*?) in ?: ?(.*)$")
# 案末汇总形态的裁决行(无 ` in: <file>` 尾;dongkl 778072 实证 Fail 汇总行即此形:
# `#### Fail Num 1: fail to find: Hit:\s+[1-9]\d*\b`)。_fail_signatures 先试
# _WA_CHECK_RE(逐步形态,pattern/文件名可分),不中再用本式(pattern 取到行尾)。
_WA_SUMMARY_RE = _re.compile(
    r"^#### (Fail|Success) Num \d+: (fail to find|successed to find):? (.*)$")
_WA_SEND_RE = _re.compile(r"^\S+ - sends command in (?:config|enable): (.*)$")
_WA_PROMPT_RE = _re.compile(r"^APV(?:\(config\))?#(.*)$")
_WA_CARET_RE = _re.compile(r"^\s+\^\s*$")


def _apv_blocks(apv_text: str) -> dict:
    """设备会话原始流按提示符旁路重分段 → {命令: [按序响应块]}。

    这是对框架读窗的**异构冗余**:框架 read_until(prompt) 的窗口边界可因异常回显
    (^ 拒绝/慢响应/交互提示)整体错位,但 pty 流的内容与顺序是设备真值——按提示符行
    重新切块得到真实的 命令→响应 映射,与框架窗口对账。"""
    out: dict = {}
    cur_cmd, cur = None, []
    for ln in apv_text.splitlines():
        m = _WA_PROMPT_RE.match(ln.strip())
        if m:
            if cur_cmd is not None:
                out.setdefault(cur_cmd, []).append("\n".join(cur))
            cur_cmd = m.group(1).strip()
            cur = []
        else:
            cur.append(ln)
    if cur_cmd is not None:
        out.setdefault(cur_cmd, []).append("\n".join(cur))
    return out


def _window_audit(inner_txt: str, apv_by_file: dict) -> tuple:
    """oracle 残差门(§18.10,采集面第四扫描器):框架断言窗 vs 旁路对齐响应块对账。

    框架判定不再是公理——(π,R) 忠实投影是三合取的独立一元,投影自身可失真
    (2026-07-14 实证:写保存族 3 run 假 FAIL 7 起/假 PASS 12 起,668000 假 PASS
    三连已写回投毒)。对每个 check:第 k 次针对命令 C 的断言 ↔ C 的第 k 个旁路
    响应块;方向矛盾即失真:
      - found 判 fail 但对齐块含期望值 → false_fail
      - not_found 判过但对齐块含该值   → false_pass
    另做块级 G 拒绝检测:响应块含裸 ^ 行=设备解析拒绝(闭合于设备解析器标记,
    非关键字表)→ 记自身执行异常(喂 anomaly_lines:G6 否决 s₀ 免派 + brief 高亮)。
    范围:只对源命令能在 apv 会话定位的 check 对账(dig/RouterA 源不在,如实跳过)。
    返回 (distortions: list[dict], caret_anomalies: list[str])。"""
    body = [_WA_TS_RE.sub("", ln) for ln in (inner_txt or "").splitlines()]
    blocks: dict = {}
    for txt in (apv_by_file or {}).values():
        for cmd, bl in _apv_blocks(txt).items():
            blocks.setdefault(cmd, []).extend(bl)
    distortions: list = []
    caret: list = []
    for cmd, bl in blocks.items():
        for b in bl:
            if any(_WA_CARET_RE.match(x) for x in b.splitlines()):
                s = f"syntax rejected (^): {cmd}"
                if s not in caret:
                    caret.append(s)
    seen: dict = {}
    for i, ln in enumerate(body):
        m = _WA_CHECK_RE.match(ln)
        if not m:
            continue
        verdict, kind, pat = m.group(1), m.group(2), m.group(3).strip()
        src = ""
        for j in range(i - 1, max(-1, i - 10), -1):
            sm = _WA_SEND_RE.match(body[j])
            if sm:
                src = sm.group(1).strip()
                break
        if not src or src not in blocks:
            continue   # dig/触发端源或会话缺失:审计范围外,如实跳过
        k = seen.get(src, 0)
        seen[src] = k + 1
        bl = blocks[src]
        blk = bl[k] if k < len(bl) else (bl[-1] if bl else "")
        try:
            hit = _re.search(pat, blk) is not None
        except _re.error:
            hit = pat.replace("\\", "") in blk
        if verdict == "Fail" and kind == "fail to find" and hit:
            distortions.append({"kind": "false_fail", "cmd": src, "pattern": pat[:60],
                                "evidence": blk.strip()[:200]})
        elif verdict == "Success" and kind == "fail to find" and hit:
            distortions.append({"kind": "false_pass", "cmd": src, "pattern": pat[:60],
                                "evidence": blk.strip()[:200]})
    return distortions, caret


def _exec_failure_markers() -> list:
    """(43) ok(g) 判据(文法数据 exec_failure_markers.patterns;读失败留声返回空)。"""
    global _EXEC_MARKERS_CACHE
    if _EXEC_MARKERS_CACHE is not None:
        return _EXEC_MARKERS_CACHE
    try:
        from main.case_compiler.domain_grammar import load_grammar
        _EXEC_MARKERS_CACHE = [str(x) for x in
                               (load_grammar().get("exec_failure_markers") or {})
                               .get("patterns") or []]
    except Exception:  # noqa: BLE001
        import logging
        logging.getLogger(__name__).warning(
            "exec_failure_markers 文法读取失败——空真嫌疑扫描本次禁用", exc_info=True)
        _EXEC_MARKERS_CACHE = []
    return _EXEC_MARKERS_CACHE



@tool(parse_docstring=True)
def dev_run_batch_digest(xlsx_path: str, autoids_json: list | str = "", module: str = "",
                         build: str = "", max_s_each: int = _RUN_DEFAULT_MAX_S,
                         force_clean: bool = False) -> str:
    """Run the whole xlsx on-device once + per-case four-layer attribution, returning a **compact readable** digest (never offloaded).

    Same params and run mode as ``dev_run_batch`` (single whole-file run, O(N)), but it
    **digests the large result in-process for you** — the deterministic core of ist-verify
    (first run → split per case → four-layer attribution) condensed into one call:

    - full per-case detail (causality / device_context / framework_traceback) lands in
      ``workspace/outputs/<feature>/last_run.json`` (**indented JSON**: pageable via
      ``fs_read``, locatable via ``fs_grep <autoid>``, loadable via ``run_python``);
    - every non-pass case goes through deterministic four-layer attribution (same classifier
      as ``compile_attribute``: transient > E > G > default V);
    - **only returns** summary counts + a one-line-per-case table + the detail file pointer
      (a few KB → never triggers offload).

    Why it exists: the large JSON ``dev_run_batch`` returns verbatim gets offloaded by
    middleware into a single blob — readable back, but hard to parse per case in place (and
    ``run_python``/``run_shell`` cannot reach the offload location). This tool digests
    **in-process**; you receive a small pre-classified summary, and to dig into one case's
    full device_context, ``fs_read`` / ``fs_grep <autoid>`` the ``last_run.json`` (inside
    workspace, reachable by every tool).

    Cases containing ``<RUNTIME>`` placeholders always fail the first run (the framework
    searches for the literal "<RUNTIME>") — expected, pending backfill: check which layer the
    digest put them in first; backfill still goes through ``compile_runtime_slots`` /
    ``compile_runtime_fill``.

    **When not to use**: to run a **single** case → ``dev_run_case`` (light, no merge); to get
    the full raw JSON and parse it yourself → ``dev_run_batch`` (but it gets offloaded, see above).

    Digest shape (decide next steps from this; do not ask for the full detail inline)::

        === dev_run_batch_digest ===
        <run_summary>
        excel: <path> | total cases: N
        true-pass P:n | fail F:m (G(^ rejected):g unattributed:u) | unknown:k
        full detail: <last_run.json path>
        </run_summary>
        <cross_run_alerts>…same-signature-across-rounds / transient-recurrence alerts (if any)…</cross_run_alerts>
        …verdict table / guidance in their own sections…

    Args:
        xlsx_path: local path of the merged case.xlsx.
        autoids_json: **prefer a native array** (JSON-array string accepted; omit = whole
            volume, recommended). Validated against the xlsx's actual autoid set; unknown ids
            rejected explicitly.
        module: staging submodule (default: compiler config).
        build: target device build (default: compiler config).
        max_s_each: whole-file budget floor (same as ``dev_run_batch``).

    Returns:
        Human-readable digest: summary counts + per-case table (autoid | verdict |
        attribution layer | reflow | causality tail) + the full-detail file path. On-device
        errors / device_busy pass through unchanged.
    """
    from pathlib import Path
    from main.ist_core.tools.device.fail_attribution import attribute_fail

    # 进程内跑 dev_run_batch：拿到的完整 JSON 不经 offload（offload 只在 tool→agent 时发生）
    raw = dev_run_batch.func(xlsx_path, autoids_json, module, build, max_s_each, force_clean)
    try:
        results = json.loads(raw)
    except Exception:  # noqa: BLE001
        return raw
    if not isinstance(results, list):
        return raw   # error / device_busy / partial dict 原样透传

    cnt = {"pass": 0, "fail": 0, "unknown": 0}
    layers = {"G": 0, "undetermined": 0}
    rows: list[tuple] = []
    for rec in results:
        aid = rec.get("autoid", "?")
        verdict = rec.get("verdict", "unknown")
        cnt[verdict] = cnt.get(verdict, 0) + 1
        causal = (rec.get("causality") or "").strip().replace("\n", " ")
        tail = causal[-90:] if causal else ""
        if verdict == "pass":
            rows.append((aid, "pass", "-", "-", tail))
        elif verdict == "unknown":
            tb = (rec.get("framework_traceback") or "").strip()
            note = "crashed/not reached"
            if tb:
                note += f"; tb tail: {tb.splitlines()[-1][:70]}"
            rows.append((aid, "unknown", "?", "-", note))
        else:  # fail → 机械预判只认设备 ^ 拒绝；其余给原文不猜（见 fail_attribution）
            detail = rec.get("device_context") or rec.get("detail_tail") or causal
            ar = attribute_fail(detail)
            layers[ar.layer] = layers.get(ar.layer, 0) + 1
            rec["_digest_layer"] = ar.layer   # 落盘供下一轮跨轮对照
            if ar.layer == "G":
                from main.ist_core.tools.device.fail_attribution import caret_rejected_commands
                cmds = caret_rejected_commands(detail, limit=2)
                # 语法层可判定性接线(理论 §3.2,044572 实证):^ 是设备语法反射的入口——
                # 在 ^ 位置追问 ?,"设备此处实际接受什么"是 O(1) 读表事实,落盘给归因/重编。
                # 三分支判定(用例语法错/文档错/特性缺失)由 attributor 拿这份事实做,不再猜。
                if cmds:
                    try:
                        from main.ist_core.tools.device.run_case import dev_help
                        rec["_device_help"] = str(dev_help.func(cmds[0]))[:2000]
                        from main.ist_core.memory.footprint.signals import emit_signal
                        emit_signal("syntax_help_attached", aid,
                                    source="dev_run_batch_digest", rejected_cmd=cmds[0])
                    except Exception:  # noqa: BLE001 — 追问失败不阻断 digest
                        pass
                note = ("⚠ config rejected (^): " + " ; ".join(c[:60] for c in cmds)) if cmds else tail
                rows.append((aid, "fail", "G(^)", "→G", note))
            else:
                # fail 行表尾展示失败签名(fail to find 前缀),不是 causality 末 90 字——
                # 后者常落在最后一条**成功**裁决行上,失败断言反而不可见(2026-07-03 取证)。
                # \n 分隔(B1):无分隔拼接曾让 causality 尾行+context 节头粘成一行假签名
                sigs = sorted(_fail_signatures(
                    "\n".join(((rec.get("causality") or ""), (detail or "")))))
                note = ("✗ " + " | ".join(s[:55] for s in sigs[:2])) if sigs else tail
                rows.append((aid, "fail", "-", "unattributed", note))

    # 全量明细落 workspace（缩进 JSON，全工具可用）——feature 目录 = xlsx 的父目录
    detail_disp = ""
    repeat_ids: list[str] = []          # 连续两轮同签名 fail（非瞬态实锤）
    transient_recur_ids: list[str] = []  # 上轮归瞬态、本轮复现 fail（误归瞬态）
    try:
        from main.ist_core.tools.deepagent.file_tools import _resolve_inside_root, _WORKSPACE_ROOT
        xp = _resolve_inside_root(xlsx_path, must_exist=True)
        out_file = Path(xp).parent / "last_run.json"
        # 跨轮对照（机械事实）：覆盖写之前读上一轮结果。瞬态定义=不可复现——
        # 同签名连续两轮 fail 即系统性问题（实证 dongkl：5 个"瞬态"下一轮 100% 复现=全部误归）。
        prev_map: dict[str, dict] = {}
        try:
            if out_file.exists():
                for r0 in json.loads(out_file.read_text(encoding="utf-8")):
                    if isinstance(r0, dict) and r0.get("autoid"):
                        prev_map[str(r0["autoid"])] = r0
        except Exception:  # noqa: BLE001
            prev_map = {}
        if prev_map:
            for rec in results:
                if not isinstance(rec, dict) or rec.get("verdict") != "fail":
                    continue
                p = prev_map.get(str(rec.get("autoid")))
                if not p or p.get("verdict") != "fail":
                    continue
                # 瞬态复现检查双源:旧 _digest_layer(机械预判,收缩后不再产 transient)
                # + _attribution.layer(LLM 归因经 submit_attribution 落盘)——后者是
                # 该护栏的活水源,没有它这分支是 dead code(2026-07-03 取证)。
                prev_layers = {p.get("_digest_layer"),
                               (p.get("_attribution") or {}).get("layer")}
                if "transient" in prev_layers:
                    transient_recur_ids.append(str(rec.get("autoid")))
                # 两侧都从原文现场重提取(非读存量字段)——跨版本升级轮天然同格式;
                # \n 分隔防跨段假行(B1)。
                sig_now = _fail_signatures("\n".join(
                    ((rec.get("causality") or ""), (rec.get("device_context") or ""))))
                sig_prev = _fail_signatures("\n".join(
                    ((p.get("causality") or ""), (p.get("device_context") or ""))))
                if sig_now & sig_prev:
                    rec["_repeat_fail_same_signature"] = True
                    repeat_ids.append(str(rec.get("autoid")))
                    # 冻结标记落 per-case 目录(A 层):同签名连续两轮 fail=同法已证无效。
                    # compile_emit 对带此标记的 case 要求 override_frozen_reason(说明换了
                    # 什么法)才放行——文本指引曾被实证绕过(721e:未核环境直接 ad-hoc 重编),
                    # 升为工具闸门;LLM 换法的自由保留,只是"是否换法"必须显式声明。
                    try:
                        import time as _t0
                        from main.ist_core.compile_engine_v8 import _shared as _sh
                        # F-Py-9(A1):frozen 写走 sh.outputs_root()——与引擎写 case.xlsx 同一解析器,
                        # 保证 frozen↔case 恒共址;pytest 下 monkeypatch sh.project_root 隔离,不再硬编码
                        # parents[4]/workspace/outputs(测试传非 autoid 键"R_sig"曾写进生产 outputs/)。
                        _cd = _sh.outputs_root() / str(rec.get("autoid"))
                        _cd.mkdir(parents=True, exist_ok=True)
                        _fz_file = _cd / ".frozen.json"
                        # 重写保留 overrides 历史(emit 门记的换法声明)——曾整文件覆盖,
                        # 换法轨迹跨轮丢失,「换过几次法/各是什么」无据可查。
                        _prev_ov = []
                        if _fz_file.is_file():
                            try:
                                _prev_ov = (json.loads(_fz_file.read_text(encoding="utf-8"))
                                            .get("overrides") or [])
                            except Exception:  # noqa: BLE001
                                pass
                        _fz_file.write_text(json.dumps({
                            "reason": "two consecutive rounds failed with the same signature (same approach proven ineffective)",
                            "signatures": sorted(sig_now & sig_prev)[:4],
                            "ts": _t0.time(),
                            **({"overrides": _prev_ov} if _prev_ov else {}),
                        }, ensure_ascii=False, indent=2), encoding="utf-8")
                        try:
                            from main.ist_core.memory.footprint.signals import emit_signal
                            emit_signal("frozen", str(rec.get("autoid")),
                                        source="dev_run_batch_digest",
                                        signature=str(rec.get("_fail_sig") or "")[:120])
                        except Exception:  # noqa: BLE001
                            pass
                    except Exception:  # noqa: BLE001
                        logger.debug("frozen 标记落盘失败", exc_info=True)
        # 按 autoid merge 写盘(不整文件覆盖):分批跑同一卷时第二批曾覆盖丢第一批
        # 17 条记录+10 个 repeat 标记(2026-07-03 取证 dongkl_final8)。本轮没跑到的
        # autoid 保留上一轮记录;每记录带 _round/_run_ts,fail 记录带 _fail_signatures。
        import time as _time
        prev_round = 0
        for r0 in prev_map.values():
            try:
                prev_round = max(prev_round, int(r0.get("_round") or 0))
            except (TypeError, ValueError):
                pass
        cur_round = prev_round + 1
        now_ts = _time.time()
        merged_map = dict(prev_map)
        for rec in results:
            if not isinstance(rec, dict) or not rec.get("autoid"):
                continue
            rec["_round"] = cur_round
            rec["_run_ts"] = now_ts
            if rec.get("verdict") == "fail":
                rec["_fail_signatures"] = sorted(_fail_signatures("\n".join(
                    ((rec.get("causality") or ""), (rec.get("device_context") or "")))))
                # 归因历史跨轮保留(2026-07-06 588691 收口):新记录整条替换曾把上一轮
                # _attribution(含 fix_direction)丢掉——归因 fork 看不到「上轮开过什么
                # 方子」,无从核对修法生效性,方向错的修法(^ 锚)被按同因重复归因。
                _pv = prev_map.get(str(rec["autoid"])) or {}
                _pa = _pv.get("_attribution")
                if isinstance(_pa, dict) and _pa:
                    rec["_prev_attribution"] = {**_pa, "_round": _pv.get("_round")}
            merged_map[str(rec["autoid"])] = rec
        # 原子写(2026-07-05 竞态类加固):last_run.json 是跨轮对照/归因/写回的事实源,
        # 与意图索引同病类——非原子 write_text 被杀进程/并发进程截断成拼接损坏后,
        # 整链(冻结判定/翻案证据/写回门)全部失灵。tmp+os.replace 根治。
        _tmp = out_file.with_suffix(".json.tmp")
        _tmp.write_text(json.dumps(list(merged_map.values()), ensure_ascii=False, indent=2),
                        encoding="utf-8")
        os.replace(_tmp, out_file)
        # 运行台账(V6 支柱2a):每 case 一条 {autoid, verdict, run_ts, apv_cmds…} 追加
        # runtime/logs/verified_runs.jsonl——footprint 写回的 device_verified 第二权威源。
        # runtime/ 在 agent 文件沙箱黑名单,工具进程写、agent 伪造不了;追加失败不阻断 run。
        try:
            # build 锚:digest 入参可能为空,按 dev_run_batch 同款 cfg 兜底解析生效值
            try:
                from main.case_compiler.config import get_config as _gc
                _eff_build = (build or _gc().build or "").strip()
            except Exception:  # noqa: BLE001
                _eff_build = (build or "").strip()
            _append_verified_runs(xlsx_path, results, cur_round, now_ts, build=_eff_build)
        except Exception:  # noqa: BLE001
            logger.debug("verified_runs 台账追加失败(忽略)", exc_info=True)
        try:
            detail_disp = str(out_file.relative_to(_WORKSPACE_ROOT.parent))
        except Exception:  # noqa: BLE001
            detail_disp = str(out_file)
    except Exception as exc:  # noqa: BLE001
        detail_disp = f"(failed to write detail file: {exc})"

    # 文件级崩溃识别：unknown 常是"某 case 断言崩了整份 pytest → 崩溃点后全 unknown（级联）"。
    # 认已知崩溃签名（如 found_times），扫 xlsx 点名元凶 case，给**正确归因**——编译缺陷、
    # 非框架 bug、非各 case 各自失败、非"逐 case 排查"。这是 bare main 最易归因错的地方。
    crash_note = ""
    tb_joined = "\n".join(
        rec.get("framework_traceback", "") for rec in results
        if rec.get("verdict") == "unknown" and rec.get("framework_traceback")
    )
    if tb_joined:
        from main.ist_core.tools.device.fail_attribution import attribute_file_crash
        hit = attribute_file_crash(tb_joined)
        if hit:
            name, guide = hit
            culprits = _scan_xlsx_for_check_method(xlsx_path, name)
            who = ("culprit case(s): " + ", ".join(f"{a} (row {r})" for a, r in culprits[:8])
                   ) if culprits else "(not located in the xlsx; may sit in a pre-merge single-case draft)"
            crash_note = (
                f"⚠ file-level crash (compilation defect, not a framework bug): a {name} assertion "
                f"crashed the whole pytest → the {cnt.get('unknown', 0)} unknowns after the crash "
                f"point are a **cascade** (later cases never ran), not individual failures.\n"
                f"   cause: {guide}\n   {who}\n"
                f"   → correct disposition: **recompile to remove/replace these cases' {name} "
                f"assertions**; not a framework fix, not per-case debugging — the excel **does need "
                f"to change**."
            )

    # 交互面 XML 分节(2026-07-05):摘要/跨轮警报/崩溃分析/裁决表/指引各自成节——
    # 数据与指引不混排(归因抄证据曾从混排文本里抄出转义失真)。行内文本零改动,
    # 只加节标签;本返回仅 LLM 消费,机读事实源仍是 last_run.json。
    lines = ["=== dev_run_batch_digest ===", "<run_summary>"]
    lines.append(f"excel: {xlsx_path} | total cases: {len(results)}")
    lines.append(
        f"true-pass P:{cnt.get('pass', 0)} | fail F:{cnt.get('fail', 0)} "
        f"(G(^ rejected):{layers['G']} unattributed:{layers['undetermined']}) "
        f"| unknown:{cnt.get('unknown', 0)}"
    )
    lines.append(f"full detail: {detail_disp}")
    lines.append("</run_summary>")
    if repeat_ids or transient_recur_ids:
        lines.append("<cross_run_alerts>")
        if repeat_ids:
            lines.append(
                f"⚠ cross-run comparison: **same-signature** fail two rounds in a row ({len(repeat_ids)}): {', '.join(repeat_ids)}\n"
                f"   → not transient, and last round's fix was ineffective. **Same-approach recompiles are "
                f"frozen** (a third same-approach round will very likely fail again); treat as environment "
                f"blockage / suspected product defect: verify the environment facts first (the real state of "
                f"that IP/config on the device); if the environment is fine, go through kb_bug_search against "
                f"the defect library and produce a defect-candidate record instead of recompiling again."
            )
        if transient_recur_ids:
            lines.append(
                f"⚠ attributed \"transient\" last round, failed again this round ({len(transient_recur_ids)}): {', '.join(transient_recur_ids)}\n"
                f"   → transient means non-reproducible; recurrence means misattribution — re-attribute as a "
                f"systemic problem (G/E/V/product defect)."
            )
        lines.append("</cross_run_alerts>")
    if crash_note:
        lines.append("<crash_analysis>")
        lines.append(crash_note)
        lines.append("</crash_analysis>")
    lines.append("<verdict_rows>")
    lines.append("autoid | verdict | attribution layer | reflow | causality/note (tail)")
    for r in rows:
        lines.append(" | ".join(str(x) for x in r))
    lines.append("</verdict_rows>")
    lines.append("<guidance>")
    lines.append(f"To dig into one case: fs_read {detail_disp} or fs_grep <autoid> in that file for the full device_context.")
    lines.append("Attribution notes: G(^) = device syntax rejection (protocol-level deterministic fact; fix it "
                 "first — later parse/assertion failures in the same case are usually downstream consequences); "
                 "unattributed = no mechanical pre-judgement was made — read that case's device_context verbatim "
                 "in last_run.json and judge E (reachability/environment) / V (assertion expectations) / "
                 "transient (vanishes on a later rerun; same-signature fails two rounds in a row are NOT "
                 "transient) / suspected product defect yourself.")
    lines.append("Attribution step one: check knowledge/data/auto_env/env_capabilities.json known_defects first — "
                 "fails matching a known defect (DC-*) are environment/product boundary; annotate and move on, "
                 "do not recompile them as compilation problems.")
    # 子集复测节流:迭代期整卷重跑,pass 的 case 每轮白跑一遍(dongkl 闭环实测:修 5-8 个
    # fail 反复整卷 34 跑了 7 轮,≈200 次多余 case 执行,每轮多等 5-9 分钟)。fail 占少数时,
    # 修复轮只跑 fail 子集卷;last_run.json 是按 autoid merge 的,子集结果回填不覆盖 pass
    # 记录。终版交付前仍需整卷跑一次确认(合并后整体行为以整卷为准)。
    fail_ids = [str(r2.get("autoid")) for r2 in results
                if isinstance(r2, dict) and r2.get("verdict") == "fail"
                and len(str(r2.get("autoid", ""))) == 18]
    if 0 < len(fail_ids) <= max(3, len(results) // 2):
        lines.append("")
        lines.append(
            f"Throttling hint: only {len(fail_ids)}/{len(results)} cases failed this round — after fixing, "
            f"**run the fail subset only** (a whole-volume rerun makes the {len(results) - len(fail_ids)} "
            f"already-passing cases run again for nothing):\n"
            f"   compile_emit_merged(autoids={json.dumps(fail_ids, ensure_ascii=False)}, "
            f"out_name=\"<batch>_fails\") → dev_run_batch_digest on the subset volume.\n"
            f"   Subset results land in the subset volume dir (per-case .frozen.json still lands by autoid, "
            f"cross-run comparison stays valid); once the subset all passes, **run the whole volume once** "
            f"as delivery confirmation (delivery is judged on the whole-volume result).")
    lines.append("</guidance>")
    return "\n".join(lines)
