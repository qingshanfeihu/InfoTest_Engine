"""确定性编译流水线（V3 approach A）：把 prep→draft→grade→merge 锁成一个工具调用。

为什么存在（颗粒度修正）：
编译流水线本质是**低自由度**任务——固定序列 prep→fanout(draft)→fanout(grade)→筛 PASS→merge，
脆弱、一致性关键、跑偏就废。此前用 inline skill 的散文让主 agent 自己编排这 5 个细粒度工具，
实测主 agent 失控（prep×3、fanout×2、误调 review-verification）。

对照标杆（Cursor create-skill §"Degrees of Freedom"：脆弱操作=低自由度=脚本；
ngs-analysis：重活在 scripts/*.py、skill 只调脚本不叙述步骤），把固定流水线降级成
**一个确定性工具**，主 agent 只调一次。论文口径：确定性机制执行流程，语义模型只在
draft/grade fork 内部（查命令/断言）自由——自由度配在 fork 内，不配在编排层。

brief 五要素从 manifest 字段**机械模板化**生成（autoid/title/step_intents/group_path
都在 manifest 里），不是语义判断——故可确定性产出，无需主 agent 逐条写。
"""

from __future__ import annotations

import json
import logging
import os
import re
from pathlib import Path
from typing import Any

from langchain_core.tools import tool

logger = logging.getLogger(__name__)

import concurrent.futures as _cf_mod
import contextvars as _ctxvars

_MAX_REWORK_ROUNDS = 3
# Guard（防 merge 卡死）：per-case 墙钟截止——churner（footprint/dev_probe 过度探索撞满轮次的
# 难 case）超时即 escalate，保证个别难 case 不会永久 gate 整脑图的 compile_emit_merged。
# 注意：此 deadline 只在轮与轮之间检查（见 _compile_one_case），**不能抢占在飞 fork**——
# 单 fork 的硬墙钟由 _FORK_WALLCLOCK_S 看门狗封顶（Fix A）。
_CASE_DEADLINE_S = float(os.environ.get("IST_CASE_DEADLINE_S") or 540)
# transient 端点错误(丢连接/限流/流中断)退避重试：与"质量重做≤3轮"分开，不消耗质量预算。
_TRANSIENT_RETRIES = 4
_TRANSIENT_BASE_SLEEP = 3.0

# Fix A：单 fork 硬墙钟。实测端点 stall 时单个 draft fork 跑出 810/1004/1056s（流式下
# request_timeout 不触发、ai_rounds=0 死挂），而 _CASE_DEADLINE_S 拦不住在飞 fork。看门狗
# 把 fork 跑在独立线程、用 future.result(timeout) 放弃等待（挂死线程泄漏但数量受并发上限约束）。
_FORK_WALLCLOCK_S = float(os.environ.get("IST_FORK_WALLCLOCK_S") or 300)
_FORK_WATCHDOG = _cf_mod.ThreadPoolExecutor(
    max_workers=int(os.environ.get("IST_FORK_WATCHDOG_WORKERS") or 64),
    thread_name_prefix="fork-wd")
# Fix B：一个 fork（含其所有 transient 重试）的总墙钟上限。防 "Request timed out" 被判 transient
# 后整 fork 重试 4 次 = 4×810s 放大。封顶重试之**和**，与 _FORK_WALLCLOCK_S（封顶单次）互补。
_FORK_TRANSIENT_WALLCLOCK_S = float(os.environ.get("IST_FORK_TRANSIENT_WALLCLOCK_S") or 600)


def _emit_progress(text: str) -> None:
    """把流水线进度推到默认 EventBus → TUI 实时渲染（evidence_added → '· …' 行）。

    compile_pipeline 是单个同步工具，内部跑 prep/draft/grade/CUT 的 fork 不会向主
    EventBus 发事件 → TUI 整段编译期间零输出、spinner 像冻住。这里在关键节点显式 emit，
    让用户看到 prep/draft/grade/CUT/merge 实时进展。fork 在线程池里跑，emit 从 worker
    线程经 TuiSink 跨线程投递到 UI（bus seq 有锁、post 线程安全）。失败一律静默，不挡流水线。
    """
    try:
        from main.ist_core.events import get_default_bus
        get_default_bus().emit("evidence_added", payload={"text": text})
    except Exception:  # noqa: BLE001
        pass


def _project_root() -> Path:
    return Path(__file__).resolve().parents[4]


def _grade_extract_facts(xp: Path, prov: Path) -> dict:
    """缺陷①：grade 前确定性预跑 ist_compile_grade/scripts/grade_extract.py 的 extract(xp, prov)。

    脚本在 skills/ 下（非 main 包内可直接 import 的模块），用 importlib 按文件路径加载。
    脚本由并行 agent 新建——若尚未就绪（文件缺失 / import 失败 / extract 抛错），一律吞掉
    返回 {}，不阻断 grade。返回结构作为 brief 的 extract_facts= 段并入。
    """
    try:
        import importlib.util as _ilu
        script = (_project_root() / "main" / "ist_core" / "skills"
                  / "ist_compile_grade" / "scripts" / "grade_extract.py")
        if not script.is_file():
            return {}
        spec = _ilu.spec_from_file_location("ist_compile_grade_extract", script)
        if spec is None or spec.loader is None:
            return {}
        mod = _ilu.module_from_spec(spec)
        spec.loader.exec_module(mod)
        facts = mod.extract(str(xp), str(prov))
        return facts if isinstance(facts, dict) else {}
    except Exception:  # noqa: BLE001
        return {}


# 低置信先例不内联(提速 + 去误导):召回最佳相似度 < 此阈值 → 视作"无好先例",不塞 7.5K 大块,
# draft 回落到 footprint 自查(footprint 是手册事实,比烂先例更可靠)。算法类(ga/wrr)先例普遍只
# 0.16(匹配差且常带错算法),listener/forward 类 0.4+(好匹配)——0.20 正好分界。env 可调。
import os as _os
_PRECEDENT_MIN_SCORE = float(_os.environ.get("IST_PRECEDENT_MIN_SCORE", "0.20"))
# 分数走 precedent_best_and_text 的**结构化排序分**(hits[0][0])，不再正则抠显示文本——
# 旧正则只对齐「意图X」「配置X+意图Y」、漏「相似度X」(config-only 轴误判 0)，且阈值在不同轴语义不一。


def _preretrieve_precedent(case: dict) -> str:
    """确定性预检索：用 case 意图召回最像的已验证先例（含完整配置基线），供 brief 内联。

    轨迹缩减核心：把"draft 自己在 ReAct 轮里调 compile_precedent"前移到流水线确定性预跑，
    draft 拿到现成先例 → 少几轮往返。检索是客观结构相似召回（非硬编码答案），draft 仍自由改写。

    **低置信不内联**：最佳相似度 < `_PRECEDENT_MIN_SCORE` 时返回 ""——烂先例(尤其算法类常带错变体)
    内联只会白塞 token + 先例支配误导,不如让 draft 回落 footprint 自查(已含手册事实)。
    """
    from main.ist_core.tools.device.precedent_tools import precedent_best_and_text
    intent = " ".join(filter(None, [
        str(case.get("title", "")),
        " ".join(str(s.get("desc", "")) for s in (case.get("step_intents") or [])),
    ])).strip()
    if not intent:
        return ""
    try:
        best, text = precedent_best_and_text(my_config="", intent=intent, limit=2)
    except Exception:  # noqa: BLE001
        return ""
    if best < _PRECEDENT_MIN_SCORE:
        return ""   # 低置信 → 不塞大块,走 _precedent_block 的轻量降级提示
    return text


def _precedent_block(precedent_text: str) -> str:
    """把预检索先例渲染成 brief 末尾的内联块（空召回时给降级提示）。"""
    if not precedent_text or not precedent_text.strip():
        return "\n\n（预检索未召回先例——分布外/新类型，自己 kb_footprint + grep 手册推断。）"
    return ("\n\n=== 预检索先例（确定性结构相似召回；含完整配置基线 + 触发→断言链 + 本测试床网络事实源）"
            "===\n照它改写，别再为它已给的东西重复检索。\n" + precedent_text.strip())


# ── footprint 预检索（Phase 1，低风险主力）──────────────────────────────────────
# 不依赖设备态：把 draft 本会查的 G 段命令文法（规则/行为/缺陷）前移到流水线、内联进 brief，
# 消除 draft 的 kb_footprint 往返轮（每轮 = 1 次 LLM 往返）。机制：footprint Index.search(意图)
# 召回相关命令节点 → 调 kb_footprint 取文法（自带 _FP_CACHE 写穿，draft 若仍查同命令直接命中
# 缓存，不另建 session 缓存——见评估挑战③）。失败/空命中一律降级，draft 兜底自查，不更差。
def _footprint_prefetch_enabled() -> bool:
    """运行时读 env（非模块常量）——支持 A/B：同进程内翻 IST_FOOTPRINT_PREFETCH=0/1 即生效。"""
    return (_os.environ.get("IST_FOOTPRINT_PREFETCH", "1").strip().lower()
            not in ("0", "false", "off", "no"))


_FOOTPRINT_PREFETCH_MAX_QUERIES = int(_os.environ.get("IST_FOOTPRINT_PREFETCH_MAX_QUERIES", "3"))
_FOOTPRINT_PREFETCH_MAX_CHARS = int(_os.environ.get("IST_FOOTPRINT_PREFETCH_MAX_CHARS", "4000"))


def _case_intent_text(case: dict) -> str:
    """case 的意图文本（title + 各 step desc），供 footprint 检索 / 命令推断。"""
    return " ".join(filter(None, [
        str(case.get("title", "")),
        " ".join(str(s.get("desc", "")) for s in (case.get("step_intents") or [])),
    ])).strip()


def _infer_footprint_queries(case: dict, precedent_text: str = "") -> list[str]:
    """推断要预检索的 footprint 命令列表（确定性，不调 LLM）。

    用 footprint Index.search(意图) 召回最相关命令节点（feature_id 按 CLI 前缀组织），转成
    命令串（点→空格）。top_k 受 IST_FOOTPRINT_PREFETCH_MAX_QUERIES 控；失败/未启用返回 []。
    **不做族骨架/配置骨架预取**（红线：实测负收益、骨架层无稳健收益，收益在 grounding）。
    """
    if not _footprint_prefetch_enabled():
        return []
    intent = _case_intent_text(case)
    if not intent:
        return []
    try:
        from main.ist_core.memory.footprint import get_footprint_index
        hits = get_footprint_index().search(intent, top_k=_FOOTPRINT_PREFETCH_MAX_QUERIES)
    except Exception:  # noqa: BLE001
        return []
    queries: list[str] = []
    seen: set[str] = set()
    for fid, _formatted in hits:
        cmd = str(fid or "").replace(".", " ").strip()
        if cmd and cmd not in seen:
            seen.add(cmd)
            queries.append(cmd)
    return queries[:_FOOTPRINT_PREFETCH_MAX_QUERIES]


def _preretrieve_footprint(queries: list[str]) -> str:
    """对推断出的命令逐条调 kb_footprint 取文法，拼成内联块（空命中跳过、总量截断）。

    复用 kb_footprint 自身的 _FP_CACHE（写穿）——draft fork 若仍查同命令直接命中、不打索引，
    无需新建 session 缓存。失败/空一律跳过，绝不塞噪声误导 draft。
    """
    if not queries:
        return ""
    try:
        from main.ist_core.tools.knowledge.footprint_lookup import kb_footprint
    except Exception:  # noqa: BLE001
        return ""
    blocks: list[str] = []
    total = 0
    for q in queries:
        try:
            r = kb_footprint.invoke({"command": q})
        except Exception:  # noqa: BLE001
            continue
        r = (r or "").strip()
        if not r or "未找到 '" in r:        # 空命中：不内联（防噪声 + 防 draft 以为查过）
            continue
        remain = _FOOTPRINT_PREFETCH_MAX_CHARS - total
        if remain <= 0:
            break
        if len(r) > remain:
            r = r[:remain] + " …(截断)"
        blocks.append(f"【{q}】\n{r}")
        total += len(r)
    return "\n\n".join(blocks)


def _footprint_block(footprint_text: str) -> str:
    """把预检索 footprint 渲染成 brief 内联块（空则不出块）。"""
    if not footprint_text or not footprint_text.strip():
        return ""
    return ("\n\n=== 预检索 footprint（G 段命令文法 / 决策规则 / 行为 / 已知缺陷；已为你查好"
            "——这些命令禁止再调 kb_footprint 重复查；只有这里没有的命令才查）===\n"
            + footprint_text.strip())


_SAVE_VARIANT_RE = re.compile(r"write\s+(mem\w*|file|all|net)", re.IGNORECASE)


def _derive_save_variant(case: dict) -> str:
    """从脑图 case 文本派生「配置保存」变体(memory|file|all|net),非保存类返回空。

    确定性解析作者意图(脑图原文写 "执行 write all 后..." → all),供持久化门校验 draft 没偷换
    保存变体。零硬编码(按命令词解析,不按 autoid)。
    """
    parts = [str(case.get("title", ""))]
    for s in (case.get("step_intents") or []):
        parts.append(str(s.get("desc", "")))
        parts.append(str(s.get("expected", "")))
    m = _SAVE_VARIANT_RE.search(" ".join(parts))
    if not m:
        return ""
    t = m.group(1).lower()
    return "memory" if t.startswith("mem") else t


# ── 变体保真(B 层):需求点名某枚举维度的变体,产出必须真用那个,不许偷换同枚举里别的 ──
# _SDNS_METHODS:手册 7190/7566 的算法枚举**快照**(footprint method 节点尚未结构化存枚举,
#   故此处冻结;手册改版需手动同步;TODO 待 footprint 补全后改运行时派生)。长在前(正则左优先)。
_SDNS_METHODS = ("gwrr", "grr", "wrr", "rtt", "rr", "ga", "hi", "snmp", "drop", "topology")
# 全局变体可替基础变体:点名 rr 用 grr(全局rr)、点名 wrr 用 gwrr,是合法严格实现,不算偷换(防误杀)。
_GLOBAL_OF = {"rr": "grr", "wrr": "gwrr"}
# intent 端:算法 token 前加边界(挡 xrr算法 粘连)+ 紧跟"算法"
_METHOD_INTENT_RE = re.compile(r"(?<![A-Za-z])(" + "|".join(_SDNS_METHODS) + r")\s*算法")
# config 端:method 命令行整行判定(跳过 no/show/clear 非生效行)+ 按位置取算法参数(防域名误命中)
_METHOD_LINE_RE = re.compile(r"\bsdns\s+(?:host|pool)\s+method\b", re.IGNORECASE)
_OP_PREFIX_RE = re.compile(r"^\s*(?:no|show|clear)\b", re.IGNORECASE)
_HOST_METHOD_ARG_RE = re.compile(r"\bsdns\s+host\s+method\s+\S+\s+\"?(\w+)\"?", re.IGNORECASE)
_POOL_METHOD_ARG_RE = re.compile(
    r"\bsdns\s+pool\s+method\s+(?:(?:primary|secondary|fallback|default)\s+)?\"?[\w.]+\"?\s+\"?(\w+)\"?",
    re.IGNORECASE)


def _intent_methods(case: dict) -> set:
    """需求点名的算法集合。只从 title + step desc(动作)抽,**不含 expected**——断言文本常有
    "与 rr 算法不同" 类对比叙述会污染。findall 取全部(覆盖多算法用例),不只首个。"""
    parts = [str(case.get("title", ""))]
    for s in (case.get("step_intents") or []):
        parts.append(str(s.get("desc", "")))
    text = " ".join(parts)
    return {m.group(1).lower() for m in _METHOD_INTENT_RE.finditer(text)}


def _actual_methods(config_text: str) -> tuple[set, bool]:
    """返回(实际配的算法集合, 是否出现过 method 配置行)。逐行处理,跳过 no/show/clear(非生效配置)。"""
    fams: set = set()
    seen = False
    for line in (config_text or "").splitlines():
        if not _METHOD_LINE_RE.search(line) or _OP_PREFIX_RE.match(line):
            continue
        seen = True
        for rx in (_HOST_METHOD_ARG_RE, _POOL_METHOD_ARG_RE):
            m = rx.search(line)
            if m and m.group(1).lower() in _SDNS_METHODS:
                fams.add(m.group(1).lower())
    return fams, seen


def _method_satisfied(named: str, actual: set) -> bool:
    """点名 named 是否被产出满足:用本身,或(点名基础 rr/wrr 时)用其全局变体 grr/gwrr 也算。"""
    if named in actual:
        return True
    g = _GLOBAL_OF.get(named)
    return bool(g and g in actual)


def _actual_save_variants(config_text: str) -> tuple[set, bool]:
    out: set = set()
    seen = False
    for line in (config_text or "").splitlines():
        if _OP_PREFIX_RE.match(line):  # no/show/clear 非生效
            continue
        for m in re.finditer(r"\bwrite\s+(memory|mem|file|all|net)\b", line, re.IGNORECASE):
            out.add("memory" if m.group(1).lower() in ("mem", "memory") else m.group(1).lower())
            seen = True
    return out, seen


def _read_xlsx_apv_config(xlsx_path: "Path") -> str:
    """读产出 xlsx 的 APV 配置命令文本(供保真校验抽实际变体)。"""
    try:
        from openpyxl import load_workbook
        ws = load_workbook(str(xlsx_path), data_only=True).active
    except Exception:
        return ""
    lines = []
    for r in ws.iter_rows(values_only=True):
        if str(r[0] or "").startswith("999"):
            break
        e = str(r[4] or "").strip() if len(r) > 4 else ""
        g = str(r[6] or "") if len(r) > 6 else ""
        if e.startswith("APV"):
            lines.append(g)
    return "\n".join(lines)


def _check_variant_fidelity(case: dict, xlsx_path: "Path") -> str:
    """B 层保真校验:需求点名的变体 vs 产出实际用的变体,偷换/缺配则返回回流反馈(无则空串)。

    数据驱动(非 per-case);需求没点名某维度 → 跳过(no-op,零回归)。每维度三态:
    满足(放行)/ 换族(偷换违规)/ 缺配(点名了但产出没配该类命令,也违规,堵"空配静默放行")。
    全局变体(grr/gwrr)算满足基础 rr/wrr(防误杀正确草稿)。多算法用例 findall 全查(required ⊆ actual)。
    """
    cfg = _read_xlsx_apv_config(xlsx_path)
    viol: list[str] = []

    req_m = _intent_methods(case)
    if req_m:
        actual, seen = _actual_methods(cfg)
        missing = sorted(v for v in req_m if not _method_satisfied(v, actual))
        if missing:
            if seen or actual:
                viol.append(
                    f"算法变体不符:需求点名 {missing} 算法,产出 method 配的是 {sorted(actual) or '识别不出'}。"
                    f"算法以需求为准、别照抄先例——改用 sdns host/pool method … {missing[0]}"
                    f"(查 footprint 拿语法+行为;ga=优先级故障切换非轮转,用 sdns pool member priority 设优先级)。")
            else:
                viol.append(
                    f"算法缺配:需求点名 {missing} 算法,但产出没有任何 sdns host/pool method 配置行"
                    f"(疑漏配或用了非常规语法)。必须显式配 sdns host/pool method … {missing[0]}。")

    sv = _derive_save_variant(case)
    if sv:
        actual, seen = _actual_save_variants(cfg)
        if sv not in actual:
            if seen or actual:
                viol.append(
                    f"保存变体不符:需求点名 write {sv},产出却用了 write {sorted(actual)}。"
                    f"保存命令以需求为准,改回 write {sv} + 同族 config {sv} 恢复。")
            else:
                viol.append(
                    f"保存变体缺配:需求点名 write {sv},但产出没有任何 write 保存命令。"
                    f"显式配 write {sv} + 同族 config {sv} 恢复。")

    return "；".join(viol)


def _build_case_brief(case: dict, *, product_version: str, manual_glob: str,
                      precedent_text: str = "", footprint_text: str = "") -> str:
    """从 manifest 的一个 case 机械生成 draft 的五要素 brief（模板，非语义判断）。

    五要素：需求(autoid/title/step_intents) + 现状(模块/版本/分组+组级基线) +
    规则(溯源/不observe-then-assert/自包含) + 指路(先例已预检索内联) + 边界。
    命令/期望值不写进 brief（零硬编码红线）；precedent_text 是**确定性预检索**的已验证先例
    （非硬编码答案，是 compile_precedent 的客观结构相似召回），内联进来让 draft 少跑检索轮次
    （轨迹缩减：把确定性检索移出 ReAct 轮循环，draft 退化成近单发生成）。
    """
    autoid = str(case.get("autoid", ""))
    title = str(case.get("title", ""))
    group_path = case.get("group_path") or []
    group_str = " / ".join(str(g) for g in group_path)
    steps = case.get("step_intents") or []
    step_lines = []
    for i, s in enumerate(steps, 1):
        desc = str(s.get("desc", "")).strip()
        exp = str(s.get("expected", "")).strip()
        line = f"  {i}. {desc}"
        if exp:
            line += f" → 期望：{exp}"
        step_lines.append(line)
    steps_block = "\n".join(step_lines) if step_lines else "  (脑图未列步骤，按标题推断)"

    # 配置保存/持久化类用例:派生意图变体 + 不重启的 clear→恢复范式引导(命令靠 footprint 现查,不写死)
    save_variant = _derive_save_variant(case)
    persist_block = ""
    if save_variant:
        persist_block = f"""

【本用例是「配置保存/持久化」测试,意图保存变体 = write {save_variant}】绝不真重启设备(共享设备 + 框架不重连)。
用不重启的 clear→恢复 范式(查 footprint config {save_variant} / write {save_variant} 的语法,参考先例 log_backup):
  配 listener → show/found → write {save_variant} <参数>(就用 {save_variant} 这个变体,别换成别的)
  → no sdns listener <ip>(或 clear sdns,先清运行配置)→ config {save_variant} <同参数>(同族恢复)→ show → 断言。
基线 init 里【不要】放任何 write 保存命令(会污染恢复快照)。
调 compile_emit 时【必须传 expected_save_variant="{save_variant}"】。"""

    # 算法变体类用例(rr/wrr/ga/…):对冲"先例支配"——算法以需求为准,别照抄先例的算法
    _algos = _intent_methods(case)
    algo = sorted(_algos)[0] if _algos else ""
    algo_block = ""
    if algo:
        algo_block = f"""

【本用例算法 = {algo}】先例可能用的是别的算法(如 wrr),**算法这一项以需求为准,绝不照抄先例的算法**。
配 sdns host/pool method … {algo}(查 footprint「{algo}」拿语法+行为)。注意各算法断言形态不同:
轮询类(rr/wrr)断言轮转命中分布;ga(全局可用性)是优先级故障切换→断言「始终命中最高优先级成员」、
用 sdns pool member priority 设优先级,不是轮转。"""

    # ★ 可达 IP 事实(确定性 env_facts)本被 compile_precedent 捆进先例文本;先例为空(新意图/相似度<0.20
    # 被跳过,如会话保持)时 ★ 列表会一起消失→draft 没 IP 锚点会探设备失败后瞎编(实测 533097 撞此坑)。
    # 解耦:先例不带 ★ 时独立注入 env_facts(永远可用);先例已带则不重复。
    env_block = ""
    if not (precedent_text and precedent_text.strip()):
        try:
            from main.ist_core.tools._shared.env_facts import get_env_facts
            env_block = "\n\n" + get_env_facts().summary_for_agent()
        except Exception:  # noqa: BLE001
            env_block = ""

    return f"""需求：autoid={autoid}，标题={title}
脑图步骤 + 期望：
{steps_block}

现状：目标产品+版本={product_version}；对版本手册 glob={manual_glob}；
本 case 属分组「{group_str}」。同组若有共享基线，纳入本 case 自包含 init。

规则：期望值溯源先例/手册/作者意图/**你自己写的配置**，不许 observe-then-assert（不看设备这次输出啥就抄成期望）；每个 case 自包含（init 自建全部前置）。
**期望值按来源三分诊（决定怎么写，条件于你正在写的这份配置——你配的东西你就知道，不是未知）**：
  ① 配置可推导值（池成员 IP、超时秒数、删除/清除后状态、rr 按配置顺序的命中、协议固定响应）——**写常量**，IP 用 \\b…\\b 加词边界 + 转义点（防 1.1.1.1 误匹配 1.1.1.10）。绝不因"运行时才显示"就当不可知。
  ② 跨观测关系（会话保持/亲和性/同-异成员/前后对比）——断言是"两次观测的**关系**"不是"某个值"，**绝不留 <RUNTIME>**（占位只能填一个值，表达不了关系）。用**捕获+比较**：触发步加 H=v1 捕获首次输出（命中啥存啥、**不用预测是哪个池**），后续 check_point 加 H=v1，found=与首次同/not_found=与首次异；dig 用 +short 去时间戳噪声。
  ③ 设备生成的不透明单值（auto-gen 名/PID/抓包间隔/哈希种子）——能 execute 提取就提取；纯不透明才留 <RUNTIME>（部分模式如 "Hits:\\s*<RUNTIME>"，或整值），标 source.kind=device_runtime（<RUNTIME>⟺device_runtime 自洽，emit 强制），由 ist_verify 上机回填锁死。
红线：会话保持/同-异 **走②捕获、绝不走 <RUNTIME>**；配置可推导走①常量、不留 <RUNTIME>；只有真·设备不透明单值才 <RUNTIME>。绝不凭空编一个对不上来源的值。

指路（先例已为你预检索，见下方「预检索先例」）：**优先照预检索先例的完整配置基线改写**——
启用步如 sdns on、数据中心/池法/监听器等基线步一个不能漏（漏了服务不起、dig 零解析、断言全 fail）；
触发(dig 类型/次数)+断言形态照先例配套写。**只有先例里缺的命令**才 kb_footprint
（你没有手册全文 grep 工具——footprint+先例就是全部命令源；缺的用 dev_probe 探设备或推断，绝不空转）。
IP 取下方「本测试床网络事实源」的可达值，绝不照抄示例 IP。
**listener/VIP 选址(dig/curl 必须够得着)**：listener/VIP 的 IP、以及 check 步骤里 dig/curl 的
目标 IP 必须**一致**，且只能取事实源里标 ★ 的「触发设备够得着的 APV 接口 IP」。标 ⚠ 的接口段
没有路由器/客户端(触发源够不着)，配上去上机必不解析、断言全 fail——绝不能用。
{env_block}{_precedent_block(precedent_text)}{_footprint_block(footprint_text)}{persist_block}{algo_block}

边界：只生成 draft（emit 必须 strict_structural=True + 传 provenance_json），不上机、不自评。"""


def _extract_xlsx_path(fork_output: str, autoid: str, since: float = 0.0) -> Path | None:
    """从 draft fork 的文本输出里定位它产的 case.xlsx（落盘规律 outputs/<autoid>/case.xlsx）。

    **新鲜度校验（治旧草稿污染）**：since>0 时，文件 mtime 必须晚于 since（本次 draft 开工时间），
    否则视作"本轮 draft 没真产出新文件"（沿用了上一轮旧草稿）→ 返回 None，让上层重试/escalate，
    绝不把旧 buggy 草稿当本轮产物合并进去。
    """
    root = _project_root()
    # 落盘规律固定：out_name 默认 autoid
    cand = root / "workspace" / "outputs" / autoid / "case.xlsx"
    if not cand.is_file():
        return None
    if since > 0:
        try:
            if cand.stat().st_mtime < since - 1:   # 容 1s 时钟抖动
                return None
        except OSError:
            return None
    return cand


# grade fork 末尾的结构化裁定标记（机读）：`判定：PASS` / `裁定: CUT` 等。
# 优先认它，治朴素 rfind 的双向误读——尤其 CUT 结论后跟"…改成 X 才能 PASS"的重做意见
# 会把 PASS 顶到最后、被 rfind 误读成通过、把弱断言 case 放进 merge。
_VERDICT_MARKER_RE = re.compile(r"(?:判定|裁定)\s*[:：]\s*(PASS|CUT)", re.IGNORECASE)

# grade CUT 时的根因机读标记：`根因：用例预期冲突` / `根因：可修复`（取最后一个）。
# 「用例预期冲突」= 期望值无手册/先例支撑且与手册/实机矛盾，非 draft 可修；
# 「可修复」= 草稿质量问题，重做有望通过。用于 escalate 时区分是否疑似脑图预期错。
_ROOTCAUSE_RE = re.compile(r"根因\s*[:：]\s*(用例预期冲突|可修复)")


def _parse_rootcause(out: str) -> str | None:
    """解析 grade fork 的 CUT 根因标记（取最后一个，结论在末尾）。无标记返回 None。"""
    if not out:
        return None
    marks = _ROOTCAUSE_RE.findall(out)
    return marks[-1] if marks else None


def _build_escalate_reason(summary: str, rounds: list[dict]) -> dict:
    """缺陷②：把 escalate reason 升级为结构化 dict——如实回带各轮裁定 + 根因 + 完整反馈。

    suspect_spec_conflict 判定：任一轮 rootcause=="用例预期冲突"，或连续 ≥2 轮同一根因
    （重做仍同因，说明非 draft 质量问题）→ 疑似脑图预期与手册/实机冲突，需人工核对。
    """
    rcs = [r.get("rootcause") for r in rounds]
    suspect = ("用例预期冲突" in rcs)
    if not suspect:
        # 连续同因（去 None 后相邻相等）→ 也判疑似冲突
        non_none = [c for c in rcs if c]
        for i in range(1, len(non_none)):
            if non_none[i] == non_none[i - 1]:
                suspect = True
                break
    return {"summary": summary, "rounds": rounds, "suspect_spec_conflict": suspect}


def _parse_grade_verdict(out: str) -> bool:
    """解析 grade fork 的最终裁定是否 PASS。

    **优先认结构化标记** `判定/裁定：PASS|CUT`（取最后一个，结论在末尾）——grade 的 CUT
    重做意见常写"…改成 X 才能 PASS",朴素的"最后出现的 PASS/CUT"会把这种 PASS 误读成通过。
    无结构化标记时回退到"最后出现的裁定词"（向后兼容旧 grade 输出 + fake fork 测试），
    ERROR 一律不通过。
    """
    if not out or out.startswith("ERROR:"):
        return False
    marks = _VERDICT_MARKER_RE.findall(out)
    if marks:
        return marks[-1].upper() == "PASS"
    last_pass = out.rfind("PASS")
    last_cut = out.rfind("CUT")
    if last_pass < 0 and last_cut < 0:
        return False          # 两词都没有 → 视为未明确通过
    return last_pass > last_cut


# ── 可观测性（Phase 0）：度量 draft/grade fork 的 LLM 往返 + 查找/探针调用次数 ──
# 目的：给「预检索是否真减少 LLM 调用/查找而质量不降」提供 baseline 与前后对比的硬数据。
# 关注的"查找/设备探针"工具（预检索要削减的几类，优先展示）。
_OBS_COST_KEYS = ("dev_probe", "kb_footprint", "compile_precedent", "fs_grep")


def _accumulate_obs(obs: dict, summary: dict, *, role: str) -> None:
    """把单次 fork 的 summary（execute_fork_skill 经 summary_sink 回传）累加进 case 级 obs。

    role="draft"|"grade"：分别累加 LLM 往返轮数（ai_rounds）；tool_calls 合并到同一桶
    （draft 才会 dev_probe/kb_footprint，grade 一般不探，但合并展示总成本）。空 summary 跳过。
    """
    if not summary:
        return
    key = f"{role}_llm_rounds"
    obs[key] = int(obs.get(key, 0)) + int(summary.get("ai_rounds", 0) or 0)
    bucket = obs.setdefault("tool_calls", {})
    for k, v in (summary.get("tool_calls") or {}).items():
        bucket[k] = int(bucket.get(k, 0)) + int(v or 0)


def _aggregate_observability(per_case: dict) -> dict:
    """把 per-case obs 聚合成全脑图总量：LLM 往返 + 各工具调用总次数 + 预检索量。"""
    draft_llm = grade_llm = prefetch_fp = 0
    tool_totals: dict[str, int] = {}
    for obs in per_case.values():
        draft_llm += int(obs.get("draft_llm_rounds", 0) or 0)
        grade_llm += int(obs.get("grade_llm_rounds", 0) or 0)
        prefetch_fp += int(obs.get("prefetch_footprint", 0) or 0)
        for k, v in (obs.get("tool_calls") or {}).items():
            tool_totals[k] = tool_totals.get(k, 0) + int(v or 0)
    return {"cases": len(per_case), "draft_llm_rounds": draft_llm,
            "grade_llm_rounds": grade_llm, "tool_calls": tool_totals,
            "total_llm_rounds": draft_llm + grade_llm,
            "prefetch_footprint": prefetch_fp}


def _format_observability(agg: dict) -> str:
    """渲染成一行（TUI evidence + compile_pipeline 返回可见）：要削减的工具排前。"""
    tc = agg.get("tool_calls", {}) or {}
    ordered = [k for k in _OBS_COST_KEYS if k in tc] + \
              [k for k in sorted(tc) if k not in _OBS_COST_KEYS]
    tc_str = " ".join(f"{k}={tc[k]}" for k in ordered) or "(无)"
    return (f"观测(LLM/查找成本基线): {agg.get('cases', 0)} cases | "
            f"LLM往返 draft={agg.get('draft_llm_rounds', 0)} "
            f"grade={agg.get('grade_llm_rounds', 0)} 合计={agg.get('total_llm_rounds', 0)} | "
            f"工具调用[{tc_str}] | 预检索 footprint={agg.get('prefetch_footprint', 0)}")


def _format_observability_delta(before: dict, after: dict) -> str:
    """渲染 A/B 对比（baseline=prefetch OFF → 优化=ON）：LLM 往返 + 查找/探针次数的降幅。

    供 `scripts/debug/measure_prefetch_ab.py` 一键产出「明显减少」的 realized 数字。
    before/after 均为 `_aggregate_observability` 的输出（result["observability"]["total"]）。
    """
    def pct(b: int, a: int) -> str:
        return "n/a" if not b else f"{round((a - b) / b * 100):+d}%"
    lines = ["=== 预检索 A/B 对比（baseline=OFF → 优化=ON）==="]
    for label, key in (("draft LLM 往返", "draft_llm_rounds"),
                       ("总 LLM 往返", "total_llm_rounds")):
        b, a = int(before.get(key, 0)), int(after.get(key, 0))
        lines.append(f"  {label}: {b} → {a}  ({pct(b, a)})")
    tb, ta = (before.get("tool_calls") or {}), (after.get("tool_calls") or {})
    for k in ("kb_footprint", "dev_probe", "compile_precedent"):
        b, a = int(tb.get(k, 0)), int(ta.get(k, 0))
        if b or a:
            lines.append(f"  {k}: {b} → {a}  ({pct(b, a)})")
    return "\n".join(lines)


def _init_device_for_compile(result: dict) -> None:
    """编译入口固化设备初始化（compile_pipeline 开头调一次，整机级）。

    经 FrameworkMCPClient 调部署 server 的 ``init_device``（串口 ``clear config all`` + 配接口 IP）
    → draft 的 dev_probe 探到干净已知态。**永不抛、永不阻断编译**：FrameworkMCPClient 不可用 /
    设备不可达 / 撞单跑锁 → 只 _emit_progress 告警并把摘要写进 result["phases"]，编译继续
    （仍可离线产出 excel）。``IST_COMPILE_INIT_DEVICE=0`` 关闭（默认开）。
    """
    if os.environ.get("IST_COMPILE_INIT_DEVICE", "1") == "0":
        _emit_progress("⏭ init_device 已禁用（IST_COMPILE_INIT_DEVICE=0）")
        return
    try:
        from main.case_compiler.device_mcp_client import FrameworkMCPClient
    except Exception as exc:  # noqa: BLE001
        _emit_progress(f"⚠ init_device 跳过（FrameworkMCPClient 不可用：{exc}）")
        result["phases"].append("init_device: 跳过(client 不可用)")
        return
    try:
        with FrameworkMCPClient() as client:
            res = client.init_device()
    except Exception as exc:  # noqa: BLE001
        _emit_progress(f"⚠ init_device 异常：{exc}（编译继续，draft 探针可能见残留态）")
        result["phases"].append(f"init_device: 异常({exc})")
        return
    if isinstance(res, dict) and res.get("error"):
        _emit_progress(f"⚠ init_device 失败：{res['error']}（编译继续）")
        result["phases"].append(f"init_device: 失败({res['error']})")
        return
    n = res.get("initialized", 0) if isinstance(res, dict) else 0
    tot = res.get("total", 0) if isinstance(res, dict) else 0
    _emit_progress(f"🧹 init_device：清干净 {n}/{tot} 台设备（clear config all + 配 IP）→ draft 探针见干净态")
    result["phases"].append(f"init_device: {n}/{tot} 台清干净")


def _run_pipeline(mindmap_path: str, product_version: str, out_name: str,
                  *, draft_skill: str, grade_skill: str) -> dict:
    """确定性跑完 prep→draft→grade→筛→merge。返回结构化结果 dict（不抛，错误进 result）。"""
    from main.ist_core.tools.device.compile_prep import compile_prep
    from main.ist_core.tools.device.emit_xlsx_tool import compile_emit_merged
    from main.ist_core.tools.device.precedent_tools import _load_case_rows
    from main.ist_core.tools.device.run_case import (
        _new_run_token, _clear_run_cache, _current_run_token)

    run_token = _new_run_token()   # dev_probe single-flight 作用域（run 结束清、跨 run 不复用）
    result: dict[str, Any] = {"mindmap": mindmap_path, "out_name": out_name,
                              "phases": [], "errors": []}
    root = _project_root()
    manual_glob = f"cli_{product_version}_Chapter*.md"

    # 1. prep（一次）
    prep_out = compile_prep.invoke({"mindmap_path": mindmap_path, "out_name": out_name})
    manifest_path = root / "workspace" / "outputs" / out_name / "manifest.json"
    if not manifest_path.is_file():
        result["errors"].append(f"prep 未产 manifest: {prep_out[:200]}")
        return result
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    cases = manifest.get("cases", [])
    result["phases"].append(f"prep: {len(cases)} cases")
    if not cases:
        result["errors"].append("manifest 无 case")
        return result

    # 编译入口固化设备初始化（整机 clear config all + 配接口 IP）：让每个 case 的 draft
    # dev_probe 探到「干净已知态」，不再撞跳板机上别人残留的配置（churn 噪声源之一）。
    # 无条件清一次（非 per-case，clear config all 是整机级）；失败只告警不阻断——编译仍可离线
    # 产出 excel。IST_COMPILE_INIT_DEVICE=0 可关（默认开=无条件）。
    _init_device_for_compile(result)

    # 2. 每 case 独立流水线：draft→grade→(CUT 重做≤N 轮) 串成一个任务，N 个任务并发。
    #    消除 P7 屏障——不再"全部 draft 完才开 grade"；case A 在 grade 时 case B 还能 draft。
    #    并发**自适应**(AIMD)：端点健康缓升、丢连接/限流骤降——替代手动调 IST_FANOUT_CONCURRENCY。
    import time as _time
    from main.ist_core.skills.loader import execute_fork_skill
    from main.ist_core.tools.device.batch_tools import _resolve_concurrency
    from main.ist_core.resilience import AdaptiveLimiter, is_transient_error

    ceiling = _resolve_concurrency(0, n_items=len(cases))   # env 可设上限；默认 auto
    limiter = AdaptiveLimiter(start=max(2, ceiling // 2), min_limit=1, max_limit=ceiling)

    def _fork_call_bounded(skill: str, brief: str, tag: str,
                           summary_sink: dict | None) -> str:
        """Fix A：把单次 fork 跑在看门狗线程，墙钟超 _FORK_WALLCLOCK_S 即放弃等待。

        copy_context() 把当前 context（含 _current_run_token）带进 worker，保证 dev_probe
        single-flight 仍命中本 run。超时返回 `[fork-wallclock]` 标记——**故意不被 is_transient_error
        命中** → 不触发 transient 重试放大，由 _compile_one_case 的 wallclock 分支立即 escalate。
        """
        ctx = _ctxvars.copy_context()
        fut = _FORK_WATCHDOG.submit(
            ctx.run, execute_fork_skill, skill, brief, tag=tag, summary_sink=summary_sink)
        try:
            return fut.result(timeout=_FORK_WALLCLOCK_S)
        except _cf_mod.TimeoutError:
            if summary_sink is not None:
                summary_sink.clear()   # 超墙钟 fork 无可信 summary，清空防污染调用方
            return (f"ERROR: [fork-wallclock] fork skill {skill!r} 超 "
                    f"{int(_FORK_WALLCLOCK_S)}s 墙钟未完成 → 放弃(escalate)")

    def _fork_call(skill: str, brief: str, tag: str = "",
                   summary_sink: dict | None = None) -> str:
        """调 fork：遇 transient 端点错误(丢连接/限流/流中断)→降并发+退避重试，非真失败。
        tag 传给 fork runner,让其内部工具调用实时发到主 bus(TUI 看到 draft/grade 运行过程)。
        summary_sink：可观测性回传桶（最后一次 attempt 的 fork summary，供 obs 聚合）。
        Fix A：每次 attempt 走看门狗封顶单 fork 墙钟。
        Fix B：整个 _fork_call（含所有 transient 重试）总墙钟封顶 _FORK_TRANSIENT_WALLCLOCK_S。"""
        last = ""
        deadline = _time.monotonic() + _FORK_TRANSIENT_WALLCLOCK_S
        for attempt in range(_TRANSIENT_RETRIES + 1):
            out = _fork_call_bounded(skill, brief, tag, summary_sink)
            if isinstance(out, str) and out.startswith("ERROR:") and is_transient_error(out):
                limiter.record_overload()
                last = out
                # Fix B：重试次数未满 **且** 总墙钟未超才退避重试；否则停退（不再 4×810s）。
                if attempt < _TRANSIENT_RETRIES and _time.monotonic() < deadline:
                    _time.sleep(_TRANSIENT_BASE_SLEEP * (2 ** attempt))
                    continue
                return out
            limiter.record_success()
            return out if out is not None else ""
        return last

    def _draft_once(case: dict, aid: str, feedback: str, rnd: int,
                    precedent_text: str = "", footprint_text: str = "",
                    obs: dict | None = None) -> tuple[Path | None, str]:
        """返回 (xlsx 路径或 None, 失败类别)。失败类别（Fix E）：
        `"recursion"`=draft 递归 spin、`"wallclock"`=超单 fork 墙钟、`"fail"`=其它生成失败、
        `""`=成功。让上层对 recursion/wallclock 立即 escalate（等价重做无意义），而非 3 轮重做。"""
        brief = _build_case_brief(case, product_version=product_version,
                                  manual_glob=manual_glob,
                                  precedent_text=precedent_text,
                                  footprint_text=footprint_text)
        if feedback:
            brief += (f"\n\n重做（第{rnd}轮）：上一版 grade 反馈——{feedback}\n"
                      "基于上一版针对问题改，保留正确部分。")
        t0 = _time.time()                       # 开工时间戳，校验产物新鲜度（治旧草稿污染）
        sink: dict = {}
        out = _fork_call(draft_skill, brief, tag=f"{aid[-6:]} draft", summary_sink=sink)
        if obs is not None:
            _accumulate_obs(obs, sink, role="draft")
        if isinstance(out, str) and out.startswith("ERROR:"):
            kind = ("recursion" if "[recursion-limit]" in out
                    else "wallclock" if "[fork-wallclock]" in out
                    else "fail")
            return None, kind
        return _extract_xlsx_path(out, aid, since=t0), ""

    def _grade_once(case: dict, aid: str, xp: Path,
                    obs: dict | None = None) -> tuple[bool, str]:
        prov = xp.parent / "case.provenance.json"
        intents = case.get("step_intents") or []
        need = "; ".join(f"{s.get('desc','')}→{s.get('expected','')}" for s in intents)
        # 缺陷①：grade 前确定性预跑 grade_extract.extract(xp, prov)，把结构化事实并入 brief。
        # 该脚本在 skills/ 下，用 importlib 按路径加载（吞异常返回 {}）——由并行 agent 新建，
        # 若运行时尚未就绪，import 失败被吞、不阻断 grade。
        facts = _grade_extract_facts(xp, prov)
        brief = (f"xlsx_path={xp}\nprovenance_path={prov if prov.exists() else '(无)'}\n"
                 f"原始需求={need}\n"
                 f"extract_facts={json.dumps(facts, ensure_ascii=False)}")
        sink: dict = {}
        out = _fork_call(grade_skill, brief, tag=f"{aid[-6:]} grade", summary_sink=sink) or ""
        if obs is not None:
            _accumulate_obs(obs, sink, role="grade")
        is_pass = _parse_grade_verdict(out)
        # 确定性安全网（论文"期望值必须溯源"）：grade 即使 PASS，若 grade_extract 探出 spec_conflict_suspect
        # （某断言 kind=intent 仅凭脑图意图、无手册溯源，且期望是设备错误回显）——脑图预期与手册/实机冲突，
        # draft 改不动、上机必 fail，绝不能蒙混 done。强制改判 CUT + 注入根因，防 LLM 漏判这类假阳性（589432）。
        if is_pass and isinstance(facts, dict) and facts.get("spec_conflict_suspect"):
            out = (out.rstrip()
                   + "\n\n[确定性安全网] grade_extract 探出 spec_conflict_suspect=True：某断言期望值 kind=intent"
                     "（仅凭脑图意图、无手册/先例溯源）且为设备错误回显——脑图预期与手册/实机冲突，draft 改不动。"
                     "\n根因：用例预期冲突\n判定：CUT")
            is_pass = False
        return is_pass, out

    def _compile_one_case(case: dict) -> dict:
        """一个 case 的完整流水线：draft→grade，CUT 带反馈重做≤N 轮。返回终态。

        用自适应限流器 acquire/release 闸住并发——名额随端点健康自动伸缩。
        """
        # worker 线程：把 run_token 绑到本线程 context，让本 case 内（同线程）draft fork 调的
        # dev_probe 命中本 run 的 single-flight 桶（ThreadPoolExecutor 不自动传 contextvar，须在此 set）。
        _current_run_token.set(run_token)
        aid = str(case["autoid"])
        tag = f"{aid[-6:]} {str(case.get('title','') or '')[:20]}"
        precedent_text = _preretrieve_precedent(case)   # 确定性预检索一次（轨迹缩减）
        # footprint 预检索（Phase 1）：把 G 段命令文法前移内联，省 draft 的 kb_footprint 往返。
        fp_queries = _infer_footprint_queries(case, precedent_text)
        footprint_text = _preretrieve_footprint(fp_queries)
        obs: dict = {}                       # 可观测性：本 case 的 LLM 往返/工具调用累加器
        if footprint_text:
            obs["prefetch_footprint"] = len(fp_queries)
        with limiter:                       # 阻塞直到有名额（名额数=当前自适应 limit）
            feedback = ""
            case_t0 = _time.time()
            rounds: list[dict] = []         # 缺陷②：逐轮如实收集完整 grade 输出 + 根因
            for rnd in range(_MAX_REWORK_ROUNDS):
                # Guard：超墙钟截止就 escalate，不再 redo——churner 不再永久 gate 整脑图 merge。
                if rnd > 0 and (_time.time() - case_t0) > _CASE_DEADLINE_S:
                    _emit_progress(f"⏱ {tag} · 超 {int(_CASE_DEADLINE_S)}s 未过 → escalate")
                    return {"autoid": aid, "case": case, "state": "escalated", "obs": obs,
                            "reason": _build_escalate_reason(
                                f"超 {int(_CASE_DEADLINE_S)}s 仍未通过 → escalate（防个别难 case 永久 gate merge）",
                                rounds)}
                _emit_progress(f"✎ {tag} · draft 第 {rnd+1}/{_MAX_REWORK_ROUNDS} 轮…")
                xp, draft_fail = _draft_once(case, aid, feedback, rnd, precedent_text,
                                             footprint_text, obs=obs)
                if xp is None:
                    # Fix E：递归 spin / 超单 fork 墙钟 → 同 brief 等价重做必然同样卡，立即 escalate。
                    if draft_fail in ("recursion", "wallclock"):
                        why = ("递归上限(在 kb_footprint↔dev_probe 间打转、未 commit compile_emit)"
                               if draft_fail == "recursion" else f"超 {int(_FORK_WALLCLOCK_S)}s 单 fork 墙钟")
                        _emit_progress(f"⏹ {tag} · draft {why} → 立即 escalate，不做等价重做")
                        rounds.append({"round": rnd + 1, "verdict": "DRAFT_RECURSION",
                                       "rootcause": "draft 递归 spin（语义打转，非难度）",
                                       "feedback_full": f"draft {why}"})
                        return {"autoid": aid, "case": case, "state": "escalated", "obs": obs,
                                "reason": _build_escalate_reason(
                                    f"draft fork {why} → 立即 escalate（同 brief 等价重做无意义）",
                                    rounds)}
                    _emit_progress(f"✎ {tag} · draft 第 {rnd+1} 轮未出 xlsx（生成失败），重试")
                    feedback = "draft 未产出 xlsx（生成失败），重试"
                    rounds.append({"round": rnd + 1, "verdict": "DRAFT_FAIL",
                                   "rootcause": None, "feedback_full": feedback})
                    continue
                # B 层保真:需求点名的变体(算法/保存)不许偷换;违规当回流反馈重做
                vfb = _check_variant_fidelity(case, xp)
                if vfb:
                    _emit_progress(f"✎ {tag} · 变体保真不符 → 重做")
                    feedback = vfb
                    rounds.append({"round": rnd + 1, "verdict": "VARIANT_FAIL",
                                   "rootcause": None, "feedback_full": feedback})
                    continue
                _emit_progress(f"⚖ {tag} · grade 评分中…")
                ok, gout = _grade_once(case, aid, xp, obs=obs)
                if ok:
                    _emit_progress(f"✓ {tag} · grade PASS → 交付草稿")
                    return {"autoid": aid, "case": case, "xlsx": xp,
                            "state": "done", "obs": obs}
                _emit_progress(f"✂ {tag} · grade CUT：{gout.strip()[:60]}")
                rc = _parse_rootcause(gout)
                # rounds 存完整 gout（不截断）；feedback 给 draft 用前若干字即可。
                rounds.append({"round": rnd + 1, "verdict": "CUT",
                               "rootcause": rc, "feedback_full": gout})
                # 预期冲突（脑图预期与手册/实机矛盾）draft 改不动——立即 escalate，不浪费后续重做轮。
                if rc == "用例预期冲突":
                    _emit_progress(f"⏹ {tag} · 根因=用例预期冲突（draft 改不动）→ 立即 escalate")
                    return {"autoid": aid, "case": case, "state": "escalated", "obs": obs,
                            "reason": _build_escalate_reason(
                                "脑图预期与手册/实机冲突，draft 改不动（期望值无手册溯源）", rounds)}
                feedback = gout[:600]
            _emit_progress(f"⏹ {tag} · 连续 {_MAX_REWORK_ROUNDS} 轮未过 → escalate")
            return {"autoid": aid, "case": case, "state": "escalated", "obs": obs,
                    "reason": _build_escalate_reason("连续重做仍未通过", rounds)}

    result["phases"].append(
        f"per-case pipeline: {len(cases)} cases, 自适应并发(start={limiter.current} max={ceiling})")
    _emit_progress(f"▶ prep 完成，开始编译 {len(cases)} 个 case（并发 {limiter.current}/{ceiling}）")
    done: dict[str, dict] = {}
    # 缺陷②：escalated 的 reason 为结构化 dict {summary, rounds, suspect_spec_conflict}。
    escalated: dict[str, dict] = {}
    obs_per_case: dict[str, dict] = {}   # 可观测性：每 case 的 LLM 往返/工具调用计数
    import concurrent.futures as _cf
    # 线程池开到上限；实际并发由 limiter 闸住（阻塞在 acquire），随端点健康自适应。
    try:
        with _cf.ThreadPoolExecutor(max_workers=ceiling) as ex:
            futs = {ex.submit(_compile_one_case, c): str(c["autoid"]) for c in cases}
            for fut in _cf.as_completed(futs):
                aid = futs[fut]
                try:
                    r = fut.result()
                except Exception as e:  # noqa: BLE001
                    escalated[aid] = _build_escalate_reason(f"流水线异常: {e}", [])
                    continue
                if r.get("obs"):
                    obs_per_case[r.get("autoid", aid)] = r["obs"]
                if r["state"] == "done":
                    done[r["autoid"]] = {"case": r["case"], "xlsx": r["xlsx"]}
                else:
                    reason = r.get("reason", "")
                    # 向后兼容：万一上游给的是 str，包成结构化 dict，下游统一按 dict 消费。
                    if not isinstance(reason, dict):
                        reason = _build_escalate_reason(str(reason or "连续重做仍未通过"), [])
                    escalated[r["autoid"]] = reason
    finally:
        _clear_run_cache(run_token)   # run 结束清 probe single-flight 桶（跨 run 不复用、防泄漏）
    if limiter.history:
        result["phases"].append(f"并发自适应轨迹: {' '.join(limiter.history[-12:])} (终值={limiter.current})")

    result["done"] = list(done.keys())
    result["escalated"] = escalated

    # 可观测性（Phase 0）：聚合本脑图 LLM 往返/查找成本 → result + evidence 一行。
    # 这是「预检索是否真减少 LLM 调用/查找而质量不降」的 baseline 数据源（前后对比看降幅）。
    obs_agg = _aggregate_observability(obs_per_case)
    result["observability"] = {"total": obs_agg, "per_case": obs_per_case}
    if obs_per_case:
        obs_line = _format_observability(obs_agg)
        result["phases"].append(obs_line)
        _emit_progress("📊 " + obs_line)

    # 3. merge（一次，把 done 的 case 读回 steps 合并）
    if done:
        merged_cases = []
        for aid, info in done.items():
            try:
                rows = _load_case_rows(str(info["xlsx"]))
            except Exception as e:  # noqa: BLE001
                result["errors"].append(f"{aid} 读回 steps 失败: {e}")
                continue
            # _load_case_rows 已含 init 行（C=1 的 APV cmds_config，有 E/F 会被读回）
            # + 各步骤（E/F/G/H/I/desc 全读回）。直接当 steps 喂 merge：init 作为首步保留，
            # 基线不丢，故不另传 init 键（传了反会与读回的 init 行重复）。
            merged_cases.append({"autoid": aid, "title": info["case"].get("title", ""),
                                 "steps": rows})
        if merged_cases:
            _emit_progress(f"⛓ merge：{len(merged_cases)} 个 case 合并打包 → {out_name}/case.xlsx")
            merge_out = compile_emit_merged.invoke(
                {"cases_json": json.dumps(merged_cases, ensure_ascii=False), "out_name": out_name})
            result["phases"].append(f"merge: {len(merged_cases)} cases → {out_name}/case.xlsx")
            result["merge_output"] = merge_out[:300]
    return result


@tool(parse_docstring=True)
def compile_pipeline(mindmap_path: str, product_version: str, out_name: str = "") -> str:
    """【确定性编译流水线】一次跑完一个脑图的 prep→draft→grade→筛→合并，产出 case.xlsx。

    把固定编译序列锁在工具内部——你只调这一次，不要自己逐步调 compile_prep /
    compile_fanout / compile_emit_merged（那会让序列失控：prep 重复、fanout 多发、误调别的 skill）。
    工具内部固定：解析 manifest → 每个 case 独立流水线(draft 产 provenance → grade 验 provenance
    → CUT 带反馈重做≤3 轮)，N 个 case 并发且无屏障(case A 在 grade 时 case B 还能 draft，
    不必等全部 draft 完才开 grade) → grade-PASS 合并成一个 excel。命令/断言全由 draft fork
    现场查（零硬编码），自由度在 fork 内不在编排层。

    不上机（上机走 ist_verify 独立环节）。多脑图请逐个调本工具（每脑图一次）。

    Args:
        mindmap_path: 脑图 txt 路径。
        product_version: 产品版本（如 "10.5"），决定查哪个版本手册。从用户请求提取，缺失则先问用户。
        out_name: 输出子目录名（workspace/outputs/<out_name>/case.xlsx），默认用脑图文件名。

    Returns:
        结构化结果：各阶段进度、done/escalated 的 autoid、合并 excel 路径。
    """
    mp = (mindmap_path or "").strip()
    if not mp:
        return "error: 必须提供 mindmap_path"
    ver = (product_version or "").strip()
    if not ver:
        return "error: 必须提供 product_version（如 10.5）——版本决定查哪个手册，不可臆测，缺失请先 ask_user 问用户"
    if not out_name:
        out_name = Path(mp).stem
    out_name = out_name.strip().replace("/", "_")

    try:
        result = _run_pipeline(mp, ver, out_name,
                               draft_skill="ist_compile_draft", grade_skill="ist_compile_grade")
    except Exception as e:  # noqa: BLE001
        logger.exception("compile_pipeline 失败")
        return f"error: 流水线异常: {e}"

    lines = [f"=== compile_pipeline: {out_name} ==="]
    lines += [f"  {p}" for p in result.get("phases", [])]
    lines.append(f"done: {len(result.get('done', []))} cases")
    escalated = result.get("escalated") or {}
    if escalated:
        # 缺陷②：如实回带每条 escalate 的根因摘要 + 疑似预期冲突标记（reason 为结构化 dict）。
        lines.append(f"escalated(N轮CUT未过): {len(escalated)} — {list(escalated.keys())}")
        for autoid, reason in escalated.items():
            if isinstance(reason, dict):
                summary = reason.get("summary", "")
                lines.append(f"  · {autoid}: {summary}")
                if reason.get("suspect_spec_conflict"):
                    lines.append("    ⚠ 疑似用例预期与手册/实机冲突, 需人工核对脑图")
            else:
                lines.append(f"  · {autoid}: {reason}")
    if result.get("errors"):
        lines.append(f"errors: {result['errors']}")
    if result.get("merge_output"):
        lines.append(f"产出: workspace/outputs/{out_name}/case.xlsx")
    return "\n".join(lines)

