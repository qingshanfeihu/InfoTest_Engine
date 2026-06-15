"""用例编译的检索与评估 tools:把置信函数 f() + 已验证先例检索 包成 agent 可调的 tool。

**red line(goal:不许硬代码)**:本文件不写"某模式该用某断言形态"的规则表、不做关键词意图匹配、
不派魔数分。检索只按**客观配置结构相似度**,判分交给 **LLM 看真实证据**(已验证先例+手册+需求)。

角色对应:
- qa_lookup_pattern(先例检索):给你的用例配置,按"配置结构相似度"返回最像的已验证先例
  (客观距离,不判模式、不派分)。你自己看先例的断言形态、自己归纳该怎么测。
- qa_confidence_score(质量评估判据):把"待判断言+所测配置+已验证先例+原始需求"喂给 LLM 判分,
  给CUT/PASS决策。不替代上机 verdict——verdict 看"能不能跑通",f() 看"是不是那道菜"。

红线:f() 只做快筛+abstain,绝不单独定生死;终判仍上机(qa_run_case)。
"""
from __future__ import annotations

import logging
import re as _re
from pathlib import Path

from langchain_core.tools import tool

logger = logging.getLogger(__name__)

_MIRROR = Path(__file__).resolve().parents[4] / "knowledge" / "framework" / "mirror"


def _load_case_rows(xlsx_path: str) -> list[dict]:
    """读 case.xlsx 数据区为 [{E,F,G}...],遇哨兵 case(999...)停。"""
    import openpyxl
    ws = openpyxl.load_workbook(xlsx_path, data_only=True).active
    rows = []
    for r in range(29, ws.max_row + 1):
        A = ws.cell(r, 1).value
        if A and str(A).startswith("999999"):
            break
        E = str(ws.cell(r, 5).value or "").strip()
        F = str(ws.cell(r, 6).value or "").strip()
        G = str(ws.cell(r, 7).value or "").strip()
        if E or F:
            rows.append({"E": E, "F": F, "G": G})
    return rows


def _cmd_tokens(text: str) -> set:
    """取命令词(字母开头的词),丢具体值(IP/数字/引号串)→ 只比'配了什么命令'。客观,无领域偏好。"""
    toks = set()
    for w in _re.findall(r"[a-zA-Z][a-zA-Z_]+", text or ""):
        toks.add(w.lower())
    return toks


def _resolve_xlsx(xlsx_path: str):
    p = None
    try:
        from main.ist_core.tools.deepagent.file_tools import _resolve_inside_root
        p = _resolve_inside_root(xlsx_path, must_exist=True)
    except Exception:
        p = None
    if p is None or not Path(p).is_file():
        cands = [Path(xlsx_path)]
        if not Path(xlsx_path).is_absolute():
            root = Path(__file__).resolve().parents[4]
            cands += [root / xlsx_path, root / "knowledge" / "data" / xlsx_path]
        p = next((c for c in cands if c.is_file()), None)
    return p if (p and Path(p).is_file()) else None


@tool(parse_docstring=True)
def qa_lookup_pattern(my_config: str, limit: int = 3) -> str:
    """【先例检索】给你这个用例的配置,返回配置结构最像的已验证先例当参考。

    排序只看**客观事实**:已验证先例配置命令 与 你的 my_config 命令词的重叠度(Jaccard)。
    不判"测试模式"、不给断言形态派分、不掺任何领域偏好——纯结构距离。越像的排越前。

    你拿到最像的先例后,**自己看它们的断言形态、自己归纳"这类配置该怎么验"**——
    认证菜单是参考(Voyager:检索不冻结),你仍自走上机验证,期望值溯源到先例+手册,不凭空编。

    Args:
        my_config: 你这个用例的关键配置命令(多行),用于按结构相似度找最像的先例。
        limit: 返回先例数(默认3)。

    Returns:
        配置结构最像的已验证先例 + 它们的完整断言形态(配置→断言的真实写法)。你自己判断哪个适用。
    """
    import glob
    import openpyxl
    my_toks = _cmd_tokens(my_config)
    if not my_toks:
        return "error: my_config 为空或无可识别命令——请传入你这个用例的关键配置命令(配置对象/动作那几行)。"
    cands = []
    for fp in glob.glob(str(_MIRROR / "**" / "*.xlsx"), recursive=True):
        try:
            ws = openpyxl.load_workbook(fp, data_only=True).active
        except Exception:
            continue
        rows = []
        for r in range(29, ws.max_row + 1):
            E = str(ws.cell(r, 5).value or "").strip()
            F = str(ws.cell(r, 6).value or "").strip()
            G = str(ws.cell(r, 7).value or "").strip()
            if E or F:
                rows.append({"E": E, "F": F, "G": G})
        if not rows:
            continue
        cfg = " ".join(r["G"] for r in rows if r["E"].startswith("APV") and r["G"])
        a = _cmd_tokens(cfg)
        if not a:
            continue
        sim = len(a & my_toks) / len(a | my_toks)
        # 返回完整"触发→断言"步骤链(test_env 的 dig/触发 + check_point),不只摘断言——
        # 否则 agent 看不到先例怎么触发的(如 A/AAAA 分查、多次触发),学不全做法。
        seq = [(r["E"], r["F"], r["G"]) for r in rows
               if r["E"] in ("test_env", "check_point", "time")
               or (r["E"].startswith("APV") and ("show" in r["G"] or "method" in r["G"]))]
        cands.append((sim, Path(fp).name, seq[:18]))
    cands.sort(key=lambda x: -x[0])
    hits = [c for c in cands[:limit] if c[0] > 0]
    if not hits:
        return ("=== qa_lookup_pattern ===\n你的配置在先例库里没有结构相近的先例(分布外/新类型)。\n"
                "→ 没有现成范式可抄,你得自己查手册推断该测什么 + 上机验证;f() 会因无锚点给低置信,"
                "做不出就诚实上报(escalate-when-stuck)。")
    out = ["=== qa_lookup_pattern(按与你配置的结构相似度排序;含完整触发→断言链)==="]
    for sim, fn, seq in hits:
        out.append(f"\n先例 {fn}(相似度{sim:.2f})的触发→断言链:")
        for e, f, g in seq:
            out.append(f"  {e} {f}({g[:50]})")
    out.append("\n注意先例'怎么触发'(test_env 的 dig/查询类型、次数、A/AAAA)和'怎么断言'是配套的。"
               "照它的完整链写,别只抄断言漏了触发方式;期望值溯源到先例+手册。")
    # 契约:先例 G 列可能混有不可达示例 IP(1.1.1.1 等历史脏数据)。在同一返回里附上本测试床
    # 真实可达集合,让你写 IP 时取真值,别照抄先例的示例 IP——emit 出口会按此校验,不可达必打回。
    try:
        from main.ist_core.tools._shared.env_facts import get_env_facts
        facts = get_env_facts()
        if facts.devices:
            out.append("\n" + facts.summary_for_agent())
    except Exception:  # noqa: BLE001
        pass
    return "\n".join(out)


@tool(parse_docstring=True)
def qa_confidence_score(xlsx_path: str, need_intent: str = "",
                        anchor_examples: str = "", manual_facts: str = "") -> str:
    """【质量评估判据】判 case.xlsx 的断言"配不配得上它所测的配置行为",给CUT/PASS决策(LLM判,无硬规则)。

    **第二判据,不是 verdict**:qa_run_case 看"能不能跑通"(弱断言 found 一个总在的域名也 pass);
    本工具看"是不是那道菜"(断言有没有咬住需求要测的行为)。

    **判分由 LLM 看真实证据现场判**(原始需求 + 每个断言所测的配置 + 已验证先例 + 手册行为),
    不写"某模式该怎样"的规则、不派魔数。证据越全判得越准:先用 qa_lookup_pattern 拿先例填
    anchor_examples、grep 手册填 manual_facts,再调本工具。

    overall<0.5 → CUT:断言没咬住需求行为,看 reason 换做法重写,别充数。
    overall≥0.5 → 放行候选,终判仍以上机 verdict 为准(f() 只快筛,不单独定生死)。
    最弱 check_point 拖垮全局(防一堆强断言掩护一个弱断言——这是结构规则,非领域硬编码)。

    Args:
        xlsx_path: 本地 case.xlsx 路径。
        need_intent: 原始用例需求/作者意图原文(用于跨层对齐判定)。
        anchor_examples: qa_lookup_pattern 返回的同类已验证先例文本(判分锚点,强烈建议填)。
        manual_facts: grep 手册得到的"该配置该产生什么可观测行为"(判分依据,建议填)。

    Returns:
        overall置信 / CUTor放行 / 每个 check_point 的 score + LLM 理由。
    """
    p = _resolve_xlsx(xlsx_path)
    if p is None:
        return f"error: xlsx 不存在: {xlsx_path}(用 qa_deepagent_ls 看真实落盘路径)"
    try:
        from main.case_compiler.confidence_f import score_case
        rows = _load_case_rows(str(p))
    except Exception as exc:  # noqa: BLE001
        return f"error: 读取失败: {exc}"
    if not rows:
        return "error: xlsx 数据区为空,无 check_point 可判"

    res = score_case(rows, need_intent=need_intent,
                     anchor_examples=anchor_examples, manual_facts=manual_facts)
    import json as _json
    # 统一 JSON 格式化返回(不再拼文本):保留逐 check_point 的 score/理由结构,
    # 下游(agent/driver)可直接 json.loads 取结构化判分,不用正则从文本抠数字。
    if res.get("reason") and not res.get("rows"):
        return _json.dumps({
            "tool": "qa_confidence_score", "abstain": True, "overall": 0.0,
            "decision": "abstain", "reason": res["reason"],
            "hint": "f() 不在缺 LLM 时硬猜分;确保判分模型可用,或直接 qa_run_case 由 verdict 兜底。",
            "checkpoints": [],
        }, ensure_ascii=False, indent=2)
    out = {
        "tool": "qa_confidence_score",
        "note": "第二判据(LLM据证据判,非框架verdict);质量评估判据",
        "overall": round(res["overall"], 2),
        "abstain": res["abstain"],
        "decision": "CUT/abstain(断言没咬住需求行为,该换做法重写)" if res["abstain"]
                    else "放行候选(终判仍以上机 verdict 为准)",
        "checkpoints": [
            {"assertion": rs.cp_g, "score": round(rs.score, 2),
             "reasons": [n for n in rs.notes if n]}
            for rs in res.get("rows", [])
        ],
        "how_to_use": "CUT→看每个checkpoint的reasons,用qa_lookup_pattern查相似先例正确断言形态改写;放行→qa_run_case上机终判。",
    }
    return _json.dumps(out, ensure_ascii=False, indent=2)
