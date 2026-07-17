"""用例编译的检索与评估 tools:把置信函数 f() + 已验证先例检索 包成 agent 可调的 tool。

**red line(goal:不许硬代码)**:本文件不写"某模式该用某断言形态"的规则表、不做关键词意图匹配、
不派魔数分。检索只按**客观配置结构相似度**,判分交给 **LLM 看真实证据**(已验证先例+手册+需求)。

角色对应:
- compile_precedent(先例检索):给你的用例配置,按"配置结构相似度"返回最像的已验证先例
  (客观距离,不判模式、不派分)。你自己看先例的断言形态、自己归纳该怎么测。
- compile_score(质量评估判据):把"待判断言+所测配置+已验证先例+原始需求"喂给 LLM 判分,
  给CUT/PASS决策。不替代上机 verdict——verdict 看"能不能跑通",f() 看"是不是要测的那个行为"。

红线:f() 只做快筛+abstain,绝不单独定生死;终判仍上机(dev_run_case)。
"""
from __future__ import annotations

import json
import logging
import re as _re
import threading
from pathlib import Path

from langchain_core.tools import tool

logger = logging.getLogger(__name__)

_MIRROR = Path(__file__).resolve().parents[4] / "knowledge" / "framework" / "mirror"
_INTENT_INDEX_PATH = (
    Path(__file__).resolve().parents[4]
    / "knowledge" / "framework" / "mirror_intent_index.json"
)
# 写回像记忆(§18.15-A / K (45)):先例的「确认状态」旁挂——provisional=子集轮过、未经整卷终验
# 确认(engine 写回事实里 ctx!=delivery)。与意图索引分文件,不动其热路径/损坏容忍逻辑。
_PROVENANCE_PATH = (
    Path(__file__).resolve().parents[4]
    / "knowledge" / "framework" / "mirror_precedent_provenance.json"
)

_INTENT_INDEX_CACHE: dict | None = None

# mirror 先例语料缓存（per-file 静态解析只做一次）。根因：compile_precedent 原本每次调用都
# openpyxl 全量加载 380+ 个 mirror xlsx，16 并发 draft 下被 GIL 串行化成 CPU 墙（编译慢的主因）。
# 静态部分（cfg_tokens + 触发→断言链 seq）缓存一次；相似度仍每次按 query 实时算，结果不变。
_MIRROR_CORPUS_CACHE: list[dict] | None = None
_MIRROR_CORPUS_LOCK = threading.Lock()


def _load_mirror_corpus() -> list[dict]:
    """解析全部 mirror 先例 xlsx **一次**并缓存：返回 [{fn, cfg_tokens, seq}, ...]。

    线程安全（双检 + 锁）：16 并发 draft 同时首调时只有一个线程解析，其余复用，
    避免 16×380 次重复 openpyxl 加载。mirror 在一次运行内静态，无需失效。
    """
    global _MIRROR_CORPUS_CACHE
    if _MIRROR_CORPUS_CACHE is not None:
        return _MIRROR_CORPUS_CACHE
    with _MIRROR_CORPUS_LOCK:
        if _MIRROR_CORPUS_CACHE is not None:
            return _MIRROR_CORPUS_CACHE
        import glob as _glob
        import openpyxl
        corpus: list[dict] = []
        for fp in _glob.glob(str(_MIRROR / "**" / "*.xlsx"), recursive=True):
            try:
                ws = openpyxl.load_workbook(fp, data_only=True).active
            except Exception:  # noqa: BLE001
                logger.debug("mirror 先例 xlsx 加载失败: %s", fp, exc_info=True)
                continue
            # 读全表：autoid 在 col1（18 位数字），case 标题在 col4，E/F/G 在 col5/6/7。
            raw = []
            for r in range(29, ws.max_row + 1):
                c1 = str(ws.cell(r, 1).value or "").strip()
                E = str(ws.cell(r, 5).value or "").strip()
                F = str(ws.cell(r, 6).value or "").strip()
                G = str(ws.cell(r, 7).value or "").strip()
                title = str(ws.cell(r, 4).value or "").strip()
                aid = c1 if _re.fullmatch(r"\d{15,20}", c1) else ""
                if aid or E or F:
                    raw.append({"aid": aid, "E": E, "F": F, "G": G, "title": title})
            if not raw:
                continue
            # **按 case 拆**：一个 xlsx 常塞多个 case（sdns_method 含 rr/wrr/ga 9 个）。整文件当一个
            # entry + seq[:40] 截断，会把靠后的 case（wrr 在第 64 步、ga 更后）切掉，draft 检索对应
            # 算法时看到的是错位的别的 case 形态（dongkl 算法类全 fail 的真根）。按 autoid 拆 → 每
            # case 一个 entry，检索命中自己那条、内联自己的完整链（含该算法的统计断言）。
            first_aid = next((i for i, x in enumerate(raw) if x["aid"]), len(raw))
            preamble = [x for x in raw[:first_aid] if x["E"] or x["F"]]   # autoid 前的通用前置（初始化基线）
            cases, cur = [], None
            for x in raw[first_aid:]:
                if x["aid"]:
                    if cur:
                        cases.append(cur)
                    cur = {"aid": x["aid"], "title": x["title"], "rows": []}
                if cur and (x["E"] or x["F"]):
                    cur["rows"].append(x)
            if cur:
                cases.append(cur)
            if not cases:   # 无 autoid 标记的老先例 → 整文件当一个 case，向后兼容
                cases = [{"aid": "", "title": "", "rows": [x for x in raw if x["E"] or x["F"]]}]
            # F3 血统((45)/(45b),§18.11):mirror 根下 verified_<autoid>.xlsx=引擎写回
            # (机生,与生成器同环→自指风险);子目录(smoke_test/…)=人源套件。血统由
            # 路径机械可导,不经 LLM 自产。engine_verified 未经窗口审计的按 uncertain
            # 采信(检索可见但降先验)——审计状态字段留待 F1/§18.10 联动填,当前默认 unknown。
            _rel = Path(fp).relative_to(_MIRROR)
            _name = _rel.name
            if _rel.parent == Path(".") and _name.startswith("verified_"):
                _lineage, _audit = "engine_verified", "unaudited"
            else:
                _lineage, _audit = "human_suite", "n/a"
            for c in cases:
                full = preamble + c["rows"]   # 每 case = 通用前置 + 该 case 步骤（配置基线完整、可照抄）
                cfg = " ".join(x["G"] for x in full if x["E"].startswith("APV") and x["G"])
                # 完整链：APV 配置基线(sdns on/启用/池法)+ show + 触发 + 断言，基线必须可见才能照抄。
                seq = [(x["E"], x["F"], x["G"]) for x in full
                       if x["E"] in ("test_env", "check_point", "time") or x["E"].startswith("APV")]
                corpus.append({"fn": _name, "autoid": c["aid"], "intent_self": c["title"],
                               "cfg_tokens": _cmd_tokens(cfg), "seq": seq[:40],
                               "lineage": _lineage, "audit": _audit})
        _MIRROR_CORPUS_CACHE = corpus
        logger.info("mirror 先例语料已缓存: %d 个先例", len(corpus))
        return _MIRROR_CORPUS_CACHE


def _salvage_json_objects(text: str) -> dict:
    """从可能损坏的索引文本里抢救全部合法 JSON 对象并按序合并(后写覆盖先写)。

    损坏形态实证(2026-07-05 v12):非原子 write_text 被杀进程截断/两次写拼接,
    文件=「合法对象 + 拖尾另一个(可能不完整的)对象」——json.loads 报 Extra data,
    旧代码整体放弃 → 28 个 PASS 写回全挂。raw_decode 逐段扫,能救多少救多少。
    """
    import json as _json
    dec = _json.JSONDecoder()
    merged: dict = {}
    i, n = 0, len(text)
    while i < n:
        while i < n and text[i] not in "{[":
            i += 1
        if i >= n:
            break
        try:
            obj, end = dec.raw_decode(text, i)
        except Exception:  # noqa: BLE001
            i += 1
            continue
        if isinstance(obj, dict):
            merged.update(obj)
        i = end
    return merged


def _read_intent_index_file() -> dict:
    """读索引文件(容忍损坏)。损坏时抢救合并、备份原件、原子重写干净版。"""
    import json as _json
    import time as _time
    if not _INTENT_INDEX_PATH.is_file():
        return {}
    text = _INTENT_INDEX_PATH.read_text(encoding="utf-8")
    try:
        obj = _json.loads(text)
        return obj if isinstance(obj, dict) else {}
    except Exception:  # noqa: BLE001
        salvaged = _salvage_json_objects(text)
        logger.warning("意图索引损坏,抢救合并 %d 条并重写干净版(原件备份 .corrupt)", len(salvaged))
        try:
            bak = _INTENT_INDEX_PATH.with_suffix(f".corrupt-{int(_time.time())}.json")
            bak.write_text(text, encoding="utf-8")
            _write_intent_index_atomic(salvaged)
        except Exception:  # noqa: BLE001
            logger.debug("索引损坏备份/重写失败(仍返回抢救结果)", exc_info=True)
        return salvaged


def _write_intent_index_atomic(idx: dict) -> None:
    """原子写索引:tmp + os.replace——被杀进程/并发写不再留下拼接损坏。"""
    import json as _json
    import os as _os
    tmp = _INTENT_INDEX_PATH.with_suffix(".json.tmp")
    tmp.write_text(_json.dumps(idx, ensure_ascii=False, indent=1), encoding="utf-8")
    _os.replace(tmp, _INTENT_INDEX_PATH)


# 写回的读-改-写临界区(fanout 多 worker 同进程并发写回时防交错)
import threading as _idx_threading
_INTENT_INDEX_LOCK = _idx_threading.Lock()


def _load_intent_index() -> dict:
    """懒加载 {xlsx_basename: [intent_path,...]} 意图索引（build_intent_index 产出）。

    索引缺失 → 空 dict(intent 轴退化不可用,config 轴照常);损坏 → 抢救合并
    (见 _read_intent_index_file),不再整体放弃。
    """
    global _INTENT_INDEX_CACHE
    if _INTENT_INDEX_CACHE is not None:
        return _INTENT_INDEX_CACHE
    try:
        _INTENT_INDEX_CACHE = _read_intent_index_file()
    except Exception:  # noqa: BLE001
        logger.debug("意图索引加载失败(退化为不可用): %s", _INTENT_INDEX_PATH, exc_info=True)
        _INTENT_INDEX_CACHE = {}
    return _INTENT_INDEX_CACHE


def _intent_tokens(text: str) -> set:
    """意图文本分词：中英文混排，取 ≥2 字的中文片段 + 字母词，丢标点/分隔符。

    意图相似度初版用词重叠即可（PLAN §八：别一上来上向量库 YAGNI）。
    """
    toks = set()
    for w in _re.findall(r"[a-zA-Z][a-zA-Z_]+", text or ""):
        toks.add(w.lower())
    # 中文按 2-gram 切（无分词器，bigram 足够捕捉"健康检查""通道建立"等共现）
    for seg in _re.findall(r"[一-鿿]+", text or ""):
        for i in range(len(seg) - 1):
            toks.add(seg[i:i + 2])
        if len(seg) == 1:
            toks.add(seg)
    return toks


def _intent_similarity(intent: str, intent_paths: list[str]) -> float:
    """intent 与某先例意图的最大相似度——用**覆盖率（包含度）`|a∩q|/|a|`**，不是 Jaccard。

    数学依据：先例意图 a（case 标题，~8 词）vs 脑图 intent q（~46 词）是**非对称**匹配——要量的是
    "脑图覆盖了多少先例意图"，不是两边互相覆盖。Jaccard `|a∩q|/|a∪q|` 把分母放成并集、被 q 的长度
    稀释（算法类金标准 title 100% 被脑图覆盖、Jaccard 却只 0.15~0.20、卡阈值不内联，是降阈值/回归的根）。
    覆盖率 `|a∩q|/|a|` 直接量先例意图被脑图覆盖的比例（0.875~1.0），稳过阈值、无需脆弱调参。
    """
    q = _intent_tokens(intent)
    if not q:
        return 0.0
    best = 0.0
    for path in intent_paths:
        a = _intent_tokens(path)
        if not a:
            continue
        # 覆盖率：先例意图被脑图覆盖的比例（非对称包含度）。a 过短（<3 词）退回 Jaccard，防"删除"
        # 这类一两个词的短标题碰巧落在脑图里就误判成 1.0。
        sim = (len(a & q) / len(a)) if len(a) >= 3 else (len(a & q) / len(a | q))
        if sim > best:
            best = sim
    return best



def _load_case_rows(xlsx_path: str) -> list[dict]:
    """读 case.xlsx 数据区为 [{E,F,G,H,I?,desc?}...],遇哨兵 case(999...)停。

    H 列(save_as/寄存器引用)必读:捕获+比较关系断言(会话保持等)靠它,grade 据此识别关系断言、
    不误判为弱。非捕获行 H 为空串,下游 .get('H') 向后兼容。
    **I 列(input_var / found_times 次数)必读**:否则 compile_pipeline 的 merge 回读时,
    found_times 的次数 / input_var 引用会被丢成 None——found_times(expect,result,times=None)
    上机 `count==None` 恒 False、该 case **永远 fail**(且 grade 验的是 merge 前版本,完全不可见)。
    D 列(步骤描述)一并读回保真,merge 后步骤描述不丢。I/desc 仅在非空时入字典(向后兼容)。
    """
    import openpyxl
    ws = openpyxl.load_workbook(xlsx_path, data_only=True).active
    rows = []
    for r in range(29, ws.max_row + 1):
        A = ws.cell(r, 1).value
        if A and str(A).startswith("999999"):
            break
        D = str(ws.cell(r, 4).value or "").strip()
        E = str(ws.cell(r, 5).value or "").strip()
        F = str(ws.cell(r, 6).value or "").strip()
        G = str(ws.cell(r, 7).value or "").strip()
        H = str(ws.cell(r, 8).value or "").strip()
        I = str(ws.cell(r, 9).value or "").strip()
        if E or F:
            row = {"E": E, "F": F, "G": G, "H": H}
            if I:
                row["I"] = I
            if D:
                row["desc"] = D
            rows.append(row)
    return rows


def _cmd_tokens(text: str) -> set:
    """取命令词(字母开头的词),丢具体值(IP/数字/引号串)→ 只比'配了什么命令'。客观,无领域偏好。"""
    toks = set()
    for w in _re.findall(r"[a-zA-Z][a-zA-Z_]+", text or ""):
        toks.add(w.lower())
    return toks


def _resolve_xlsx(xlsx_path: str):
    p = None
    try:
        from main.ist_core.tools.deepagent.file_tools import _resolve_inside_root
        p = _resolve_inside_root(xlsx_path, must_exist=True)
    except Exception:
        p = None
    if p is None or not Path(p).is_file():
        cands = [Path(xlsx_path)]
        if not Path(xlsx_path).is_absolute():
            root = Path(__file__).resolve().parents[4]
            cands += [root / xlsx_path, root / "knowledge" / "data" / xlsx_path]
        p = next((c for c in cands if c.is_file()), None)
    return p if (p and Path(p).is_file()) else None


def _retrieve_precedent_hits(my_config: str, intent: str, limit: int) -> tuple[list, set, str]:
    """检索（与渲染解耦）：返回 (hits, my_toks, intent_clean)。

    hits = [(score, cfg_sim, intent_sim, fn, seq)…] 按 score 降序。score = 融合排序分
    （两轴都启用取等权和；只有一轴退化为该轴）——**结构化数值**，供阈值判定，免去从显示文本
    正则抠分（旧 _precedent_best_score 只覆盖「意图X」「配置X+意图Y」，「相似度X」config-only 轴误判 0）。
    两轴皆空 → ([], set(), '')。
    """
    my_toks = _cmd_tokens(my_config)
    intent = (intent or "").strip()
    if not my_toks and not intent:
        return [], my_toks, intent
    intent_index = _load_intent_index() if intent else {}
    cands = []
    # 用缓存语料（静态解析一次），每个先例按本次 query 实时算相似度。
    for entry in _load_mirror_corpus():
        a = entry["cfg_tokens"]
        cfg_sim = (len(a & my_toks) / len(a | my_toks)) if (a and my_toks) else 0.0
        # 意图轴:**case 级意图优先**（entry.intent_self = 该 case 的标题/描述）+ 整文件 intent_index 兜底。
        # 治"整 xlsx 意图被同文件无关 case 稀释"——sdns_method 的 wrr 意图本被 rr/ga 拉低成 0.09，
        # 按 case 拆 + case 级 intent 后，wrr query 对上 wrr case 的标题，intent_sim 不再被稀释。
        intent_sim = 0.0
        if intent:
            paths = ([entry["intent_self"]] if entry.get("intent_self") else []) + intent_index.get(entry["fn"], [])
            if paths:
                intent_sim = _intent_similarity(intent, paths)
        # 融合排序分:两轴都启用时取等权和;只有一轴时退化为该轴。
        if my_toks and intent:
            score = cfg_sim + intent_sim
        elif intent:
            score = intent_sim
        else:
            score = cfg_sim
        if score <= 0:
            continue
        # 返回完整"触发→断言"步骤链(test_env 的 dig/触发 + check_point),不只摘断言——
        # 否则 agent 看不到先例怎么触发的(如 A/AAAA 分查、多次触发),学不全做法。
        cands.append((score, cfg_sim, intent_sim, entry["fn"], entry["seq"],
                      entry.get("autoid", ""), entry.get("lineage", "human_suite")))
    cands.sort(key=lambda x: -x[0])
    top = [c for c in cands[:limit] if c[0] > 0]
    # F3 配额保底((45b) 防自指,§18.11):机生血统在词面相似度上结构性碾压人源
    # (同 autoid 族/同标题措辞)——若 top 全机生但存在有效人源命中,置换末位为最高分
    # 人源命中,保证结果集人源可见性(不改排序哲学,只破"自产血统垄断")。
    if top and all(c[6] == "engine_verified" for c in top):
        human = next((c for c in cands if c[6] == "human_suite" and c[0] > 0), None)
        if human:
            top[-1] = human
    return top, my_toks, intent


def _precedent_sampling_note(seq: list) -> str:
    """采样敏感度事实（写回像记忆，§18.15-A / K (45)）：一条先例若在**分布类算法**（rr/wrr…）下
    断言了**命中计数字段**（Hit/命中/计数…），该计数随 dig 采样窗口变化——同一配置态、不同采样
    给不同计数。把这个**结构事实**随检索结果摆出来，让读的人对照本案意图判断断言形态能不能照抄
    （§0 摆事实不替判：只陈述事实+后果，不在此文件写「该用哪种形态」的规则表——守本文件顶部红线）。

    实证锚：593516（wrr 3:2:1）用 `Hit:\\s+3`/`[1-9]\\d*`/`[4-9]\\d*` 断言，flaky pass 写回后
    live 可检索、断言链可被后继 RR/WRR 案照抄（`docs/forensics/A_oracle.md` §1.4）。
    分布方法词 / 计数字段词从 `domain_grammar` 数据读（单一事实源），不硬编；纯成员归属断言
    （`abs_found <成员IP>`，无计数字段）不触发——那是 h-不变式的正确形态、不该被误标。
    """
    try:
        from main.case_compiler.domain_grammar import (
            count_field_words,
            distribution_methods,
        )
        dmeths = distribution_methods()
        cwords = count_field_words()
    except Exception:  # noqa: BLE001
        logger.debug("domain_grammar 读取失败,跳过采样敏感标注", exc_info=True)
        return ""
    config_text = " ".join((g or "") for e, f, g in seq if e.startswith("APV")).lower()
    has_dist = any(_re.search(rf"\b{_re.escape(m.lower())}\b", config_text) for m in dmeths)
    if not has_dist:
        return ""
    has_count_ck = any(
        e == "check_point" and any(w.lower() in (g or "").lower() for w in cwords)
        for e, f, g in seq
    )
    if not has_count_ck:
        return ""
    return (
        "sampling-sensitive (memory hint, verify first): this precedent asserts a hit-count "
        "field under a distribution algorithm (rr/wrr…); hit counts vary with the dig sample "
        "window (same config, different samples give different counts). Confirm the assertion "
        "form (membership via abs_found / conserving interval on show statistics / exact count) "
        "matches THIS case's intent against the manual before copying — an engine-written "
        "precedent that PASSed once may have passed on sampling luck."
    )


def _load_precedent_annotations() -> dict:
    """先例策展标注(判例化,2026-07-08):不删迁就嫌疑卷(结构仍是金标准),检索返回时附
    警示——pe1 减法实验实证:裸 worker 把迁就卷的断言方向当"产品行为如此"的佐证,
    与污染知识互相印证后放弃了正确的配置探索。标注让先例回到"结构参考"的本位。"""
    try:
        p = _MIRROR / "precedent_annotations.json"
        if p.is_file():
            d = json.loads(p.read_text(encoding="utf-8"))
            return {k: v for k, v in d.items() if isinstance(v, dict)}
    except Exception:  # noqa: BLE001
        logger.debug("precedent_annotations.json 读取失败(忽略)", exc_info=True)
    return {}


def _load_precedent_provenance() -> dict:
    """读先例确认状态旁挂(best-effort，§18.15-A)：{fn: {"provisional": bool}}。
    缺失/损坏 → 空 dict（检索照常、只是不显 provisional，与旧行为一致）。"""
    try:
        if _PROVENANCE_PATH.is_file():
            d = json.loads(_PROVENANCE_PATH.read_text(encoding="utf-8"))
            return {k: v for k, v in d.items() if isinstance(v, dict)} if isinstance(d, dict) else {}
    except Exception:  # noqa: BLE001
        logger.debug("mirror_precedent_provenance.json 读取失败(忽略)", exc_info=True)
    return {}


def _record_precedent_provenance(fn: str, provisional: bool) -> None:
    """写回时如实记 fn 的确认状态（best-effort，进程内锁复用意图索引锁避免并发写撞）。
    失败不阻断写回主流程——旁挂缺失只是少一个提示，不影响先例可用性。"""
    try:
        with _INTENT_INDEX_LOCK:
            cur = _load_precedent_provenance()
            cur[fn] = {"provisional": bool(provisional)}
            tmp = _PROVENANCE_PATH.with_suffix(".json.tmp")
            tmp.write_text(json.dumps(cur, ensure_ascii=False, indent=1), encoding="utf-8")
            tmp.replace(_PROVENANCE_PATH)
    except Exception:  # noqa: BLE001
        logger.debug("mirror_precedent_provenance.json 写入失败(忽略)", exc_info=True)


def _format_precedent_hits(hits: list, my_toks: set, intent: str) -> str:
    """把 hits 渲染成 draft 可读的先例文本（触发→断言链 + 警示 + env_facts）。"""
    axis = "config+intent fused" if (my_toks and intent) else ("intent axis" if intent else "config structure axis")
    anns = _load_precedent_annotations()
    prov = _load_precedent_provenance()
    out = [f"=== compile_precedent (ranked by {axis} similarity; full trigger→assertion chains) ==="]
    for score, cfg_sim, intent_sim, fn, seq, autoid, *_rest in hits:
        lineage = _rest[0] if _rest else "human_suite"
        if my_toks and intent:
            tag = f"config {cfg_sim:.2f} + intent {intent_sim:.2f}"
        elif intent:
            tag = f"intent {intent_sim:.2f}"
        else:
            tag = f"similarity {cfg_sim:.2f}"
        # F3 血统外显((45),§18.11):机生血统标注,提醒 draft 别把自产判例当人源金标准
        lin_tag = " ⟨engine-written; structure only, not an authority⟩" if lineage == "engine_verified" else ""
        fn_show = f"{fn}[{autoid}]" if autoid else fn
        out.append(f"\nPrecedent {fn_show} ({tag}){lin_tag} trigger→assertion chain:")
        # 采样敏感事实(§18.15-A):分布类算法下的命中计数断言随采样变化,随结果摆出来供读者先核。
        samp = _precedent_sampling_note(seq)
        if samp:
            out.append(f"  ⚠ {samp}")
        # 确认状态旁挂(写回像记忆,K (45)):子集轮过、未经整卷终验确认的先例摆出来供先核。
        if prov.get(fn, {}).get("provisional") is True:
            out.append("  ⚠ provisional (memory hint, verify first): this precedent passed only "
                       "on a subset re-run, not yet confirmed by a full delivery verify.")
        ann = anns.get(str(autoid)) if autoid else None
        if ann and ann.get("note"):
            out.append(f"  ⚠ curation note [{ann.get('flag', 'curated')}]: {str(ann['note'])[:260]}")
        for e, f, g in seq:
            # 多行 cmds_config 用**真换行**展示(缩进续行),draft 照抄即得真 \n——
            # 切勿用 ⏎ 等替身字符:draft 会把字面替身抄进配置,框架按 \n 拆命令时整串变一条废命令。
            g_show = g[:300].replace("\n", "\n        ")
            out.append(f"  {e} {f}: {g_show}")
    out.append("\n⚠ Copy the precedent's **complete config baseline**, never truncate: enable/activation "
               "steps and every baseline step in the precedent must all be present — miss one and the "
               "device service never comes up, dig resolves nothing, every assertion fails. A precedent's "
               "'how to configure (full baseline) + how to trigger (dig types/counts) + how to assert' is a "
               "matched set; write the whole chain as it does. Copy the precedent for **config form** "
               "(commands/quoting/baseline); an expected value's polarity (found vs not_found) and target "
               "trace to THIS case's intent + the manual, NOT to the precedent — a precedent testing a "
               "different intent can assert the opposite direction (precedent-then-assert is a fake-PASS "
               "path, the twin of observe-then-assert). Runtime values unknowable offline stay <RUNTIME>.")
    out.append("⚠ **Copy the closest precedent's command format verbatim; change only values (IP/domain/name), "
               "never the format**: the precedent's **quoting, parameter count and order** are what actually ran "
               "on-device — if it double-quotes a parameter (e.g. `\"24\"`) you quote it too; if it does not, "
               "neither do you; **never 'normalize' quoting yourself, never invent parameter combinations the "
               "precedent does not have** (if the precedent stops at the mask, do not add query_type out of thin "
               "air). For shapes no precedent covers, settle the format from footprint/manual; when unsure leave "
               "it for on-device verify to backstop via the device `^` error / `show` echo — do not guess a "
               "spelling the device will reject.")
    # 契约:先例 G 列可能混有不可达示例 IP(1.1.1.1 等历史脏数据)。在同一返回里附上本测试床
    # 真实可达集合,让你写 IP 时取真值,别照抄先例的示例 IP——emit 出口会按此校验,不可达必打回。
    try:
        from main.ist_core.tools._shared.env_facts import get_env_facts
        facts = get_env_facts()
        if facts.devices:
            out.append("\n" + facts.summary_for_agent())
    except Exception:  # noqa: BLE001
        pass
    return "\n".join(out)


def precedent_best_and_text(my_config: str = "", intent: str = "", limit: int = 2) -> tuple[float, str]:
    """供确定性预检索：返回 (最像先例的结构化排序分 = hits[0][0], 渲染文本)。无召回 → (0.0, "")。

    替代从 compile_precedent 显示文本正则抠分——直接读结构化分，消掉 parser/formatter 跨文件
    只对齐 2/3 格式 + 阈值在不同轴语义不一的脆弱点。
    """
    hits, my_toks, intent_clean = _retrieve_precedent_hits(my_config, intent, limit)
    if not hits:
        return 0.0, ""
    return hits[0][0], _format_precedent_hits(hits, my_toks, intent_clean)


@tool(parse_docstring=True)
def compile_precedent(my_config: str, limit: int = 3, intent: str = "") -> str:
    """Precedent retrieval: given this case's config/intent, return the most similar verified precedents as reference.

    Ranking uses **objective facts** on two (fused) axes:
    - config axis: token overlap (Jaccard) between verified precedent config commands and your my_config;
    - intent axis (optional): text overlap between your intent description and the precedent's
      mindmap intent chain (intent_path).
    No "test mode" judgement, no assertion-form scoring, no domain preference — pure
    structural/semantic distance. The most similar rank first.

    **The intent axis cures "out-of-distribution cases unreachable when config can't be
    guessed"**: when the requirement is one sentence and no commands are settled yet, pass
    intent (the raw requirement / author intent verbatim) and skeleton precedents are still
    reachable through the intent chain. With my_config empty but intent non-empty, retrieval
    switches to the pure intent axis (no longer an error). With both, ranking is fused.

    Once you have the closest precedents, **inspect their assertion forms yourself and
    generalize "how this class of config gets verified"** — the verified-precedent list is a
    reference (retrieval never freezes your choices); you still verify on-device, and
    expected values trace to precedent+manual, never invented.

    Args:
        my_config: this case's key config commands (multi-line), for structural-similarity
            retrieval. May be empty (pure intent axis).
        limit: number of precedents to return (default 3).
        intent: raw case requirement / author intent verbatim (optional). Enables the intent
            axis so out-of-distribution cases can still retrieve a skeleton.

    Returns:
        The most similar verified precedents + their full assertion forms (the real
        config→assertion write-ups). You judge which applies.
    """
    # 消融实验 Arm-E(基线裸生成):不提供先例约束(G 段),模拟业界默认的无先例自由生成。
    # 生产默认 Arm-L 不进此分支。
    from main.ist_core.tools._shared.ablation import is_baseline
    if is_baseline():
        return ("=== compile_precedent (baseline arm: no precedents provided) ===\n"
                "This arm does not retrieve verified precedents (ablation Arm-E). Generate "
                "directly from the requirement and general knowledge, without relying on "
                "similar precedents' trigger→assertion forms.")
    hits, my_toks, intent_clean = _retrieve_precedent_hits(my_config, intent, limit)
    # 向后兼容:config 与 intent 都空才报错;只要有一轴可用就检索（intent 轴治分布外）。
    if not my_toks and not intent_clean:
        return ("error: both my_config and intent are empty — pass this case's key config "
                "commands (the object/action lines), or pass intent (the raw requirement) "
                "for intent-axis retrieval.")
    if not hits:
        return ("=== compile_precedent ===\nNo structurally similar precedent for your config/intent "
                "(out-of-distribution / new type).\n"
                "→ No ready-made paradigm to copy: derive what to test from the manual yourself and "
                "verify on-device; if it cannot be done, report honestly (escalate-when-stuck).")
    return _format_precedent_hits(hits, my_toks, intent_clean)


@tool(parse_docstring=True)
def compile_writeback(autoid: str, last_run_path: str, intent_path: str = "",
                      provisional: bool | None = None) -> str:
    """Write a **true on-device PASS** volume back into the precedent store (mirror + intent index) so later compilations can retrieve it.

    Closed-loop self-evolution: growing the precedent store raises retrieval hit rate,
    driving G/E-segment errors toward their intrinsic floor. Measured counterexample: with
    the writeback chain broken, next-round workers re-stepped on pits the previous round had
    already verified through.

    Mechanical gates: ① the autoid's verdict in last_run.json must be pass (on-device
    oracle, no retellings trusted); ② the volume credential must be fresh (what is written
    back is exactly what ran on-device, not something edited afterwards). This tool is the
    precedent-store channel only; footprint fact writeback goes through the router path.

    Args:
        autoid: the full autoid of the on-device-PASSed case.
        last_run_path: path to the last_run.json recording that run (written by dev_run_batch_digest).
        intent_path: intent path (e.g. "<feature> > <algorithm> > <case intent>"); when empty
            it is best-effort assembled from the sibling batch manifest (falling back to the title).
        provisional: confirmation status of this pass, recorded alongside the precedent like a
            memory annotation per K 2.9.4 axiom 45. True means it passed only on a subset re-run
            and is not yet confirmed by a full delivery verify. False means delivery-confirmed.
            None means the caller did not say so nothing is recorded. compile_precedent surfaces
            a provisional precedent as a hint to verify rather than ground truth to copy.

    Returns:
        Writeback result (mirror filename + index entry); on gate failure an error stating
        what is missing.
    """
    import json as _json
    import re as _re
    aid = (autoid or "").strip()
    # F-Py-9b-1b(写侧补口):outputs 读(src_dir 602)+ workspace 沙箱(_ws 607)+ last_run 解析(608)
    # 走 _sh.project_root() 单一根——生产 == parents[4](字节等价),pytest monkeypatch 后同步落 tmp,
    # 保 602/607/608 同源(_ws relative_to 校验一致,同 batch_tools 委托)。
    from main.ist_core.compile_engine_v8 import _shared as _sh
    root = _sh.project_root()
    # 安全:autoid 白名单(路径分量净化)——它拼进写 mirror 的文件名与读 outputs 的目录。
    # 未净化时 aid="../.." 可写穿沙箱(安全评审高危项:工具进程内直写不经 file_tools 四闸)。
    if not _re.fullmatch(r"[A-Za-z0-9_.\-]+", aid) or ".." in aid:
        return f"error: illegal autoid (only alphanumerics ._- allowed, no ..): {aid!r}"
    src_dir = root / "workspace" / "outputs" / aid
    xp = src_dir / "case.xlsx"
    if not xp.is_file():
        return f"error: case.xlsx for {aid} does not exist"
    # 门①:上机 oracle。last_run_path 收敛到 workspace 内(禁绝对路径读任意文件)。
    _ws = (root / "workspace").resolve()
    lrp = Path(last_run_path) if Path(last_run_path).is_absolute() else root / last_run_path
    try:
        if not lrp.resolve().is_relative_to(_ws):
            return f"error: last_run_path must be inside workspace/: {last_run_path}"
    except Exception:  # noqa: BLE001
        return f"error: last_run_path cannot be resolved: {last_run_path}"
    if not lrp.is_file():
        return f"error: last_run does not exist: {last_run_path}"
    try:
        recs = _json.loads(lrp.read_text(encoding="utf-8"))
        rec = next((r for r in recs if str(r.get("autoid")) == aid), None)
    except Exception as exc:  # noqa: BLE001
        return f"error: last_run parse failed: {exc}"
    if rec is None:
        return f"error: {aid} is not in this last_run — pass the one recording this case's run"
    if str(rec.get("verdict")) != "pass":
        return (f"error: {aid} ran with verdict={rec.get('verdict')} — only true-PASS volumes are "
                "written back; writing back fail/unknown poisons the precedent store.")
    # <RUNTIME> 护栏(红线评审提示):含未回填运行时占位的卷进先例库,环境态值与
    # 溯源值对消费方不可分辨。真 PASS 卷本不该还带 <RUNTIME>(它会使断言失败),
    # 但直改/半回填的边角情况防一手——检出即拒,让先例库只收干净的确定性卷。
    try:
        import openpyxl as _ox
        _wsx = _ox.load_workbook(xp).active
        for _r in _wsx.iter_rows(min_row=2):
            if "<RUNTIME>" in str(_r[6].value or ""):
                return (f"error: {aid} volume still contains <RUNTIME> placeholders — volumes with "
                        "unfilled runtime values are not written back (environment state would "
                        "pollute provenance). compile_runtime_fill first, then re-verify.")
    except Exception:  # noqa: BLE001
        pass
    # 门②:凭证新鲜(写回的是上机那份,不是之后改过的)。凭证缺失=没走过 emit 门,
    # 拒绝而非静默跳过(红线评审弱门:缺凭证跳过新鲜校验会放进未经门的卷)。
    credf = src_dir / ".grade_credential.json"
    if not credf.is_file():
        return (f"error: {aid} has no credential (never passed the emit gate) — only gated and "
                "on-device-verified volumes are written back.")
    try:
        cred = _json.loads(credf.read_text(encoding="utf-8"))
        if abs(float(cred.get("xlsx_mtime", -1)) - xp.stat().st_mtime) >= 1e-6:
            return (f"error: {aid} volume was modified after its credential (mtime mismatch) — what "
                    "was verified on-device is not this file; re-pass the gate first, then write back.")
    except Exception as exc:  # noqa: BLE001
        return f"error: {aid} credential parse failed: {exc}"
    # 意图路径
    ip = (intent_path or "").strip()
    if not ip:
        title = ""
        for mani in src_dir.parent.glob("*/manifest.json"):
            try:
                m = _json.loads(mani.read_text(encoding="utf-8"))
                c = next((c for c in m.get("cases", []) if c.get("autoid") == aid), None)
                if c:
                    title = " > ".join([*(c.get("group_path") or []), c.get("title") or ""])
                    break
            except Exception:  # noqa: BLE001
                continue
        ip = title or aid
    # 写 mirror + 索引(同名覆盖=同 case 新版本)
    fn = f"verified_{aid}.xlsx"
    try:
        import shutil as _sh
        _MIRROR.mkdir(parents=True, exist_ok=True)
        _dst = (_MIRROR / fn).resolve()
        if not _dst.is_relative_to(_MIRROR.resolve()):   # 双保险:白名单已挡,这里兜底
            return f"error: write target escapes the mirror dir (anomalous autoid): {aid!r}"
        _sh.copyfile(xp, _dst)
        # 索引更新走「容损读 + 原子写 + 进程内锁」——2026-07-05 实证:旧版裸
        # read+write_text 被杀进程截断成拼接损坏后,28 个 PASS 写回全挂在这一行。
        with _INTENT_INDEX_LOCK:
            idx = _read_intent_index_file()
            idx[fn] = [ip]
            _write_intent_index_atomic(idx)
    except Exception as exc:  # noqa: BLE001
        return f"error: writeback failed: {exc}"
    # 写回像记忆:如实记确认状态(caller 说了才记;None=不记,与旧行为一致)。检索时旁挂摆出。
    if provisional is not None:
        _record_precedent_provenance(fn, provisional)
    # 失效进程内缓存:同 run 内后续 compile_precedent 立即能检索到(ρ_k 增长可观测)
    global _INTENT_INDEX_CACHE, _MIRROR_CORPUS_CACHE
    _INTENT_INDEX_CACHE = None
    _MIRROR_CORPUS_CACHE = None
    return (f"written back to the precedent store: {fn}\nintent index: {ip}\n"
            f"compile_precedent in later compilations can retrieve this verified precedent "
            f"immediately (in-process caches invalidated and rebuilt).")
