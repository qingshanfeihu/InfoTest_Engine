"""v2 结构约束门（correct-by-construction，论文命题3.18 / PLAN_footprint_v2_compile.md §2.2）。

与 grade(LLM 语义自评)**独立**的确定性强制——执行**结构约束**，不碰**骨架选择**：
- ① 命令头∈手册 allowlist（footprint 已有的命令集）——挡幻觉/越界命令；
- ② 断言对象挂在某前序观测算子（show/dig 产出回显的步骤）的值域上——根治 655203 悬空断言；
- ③ 绑定 IP∈env_facts 可达表（已由 emit_xlsx_tool._gate_unreachable_ips 做，这里不重复）。

红线（§七）：只执行**与意图无关的类型规则**（命令合法性、断言非悬空、IP 可达），
绝不替 LLM 决定"测什么/什么命令序列/什么断言形态"（H_G≠0 的语义决策永远 LLM）。

只在 v2 编译链启用（compile_emit(strict_structural=True)）；v1 默认 False，行为零变化。
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)

# 命令操作前缀（与 footprint extractor._OP_PREFIXES 对齐）：剥掉再取命令主体 token
_OP_PREFIXES = ("no", "show", "clear")
# 参数记法 / 具体值 token：<x> [x] {a|b}、纯数字、IP、引号串 → 不是命令主体 token
_VALUE_TOK_RE = re.compile(r"^([<\[{].*|.*[|].*|\d.*|[\"'].*|.*\.\d+.*)$")
# 观测算子：check_point 校验的是“上一个非 check_point 步骤”的回显输出。
# 产出可匹配回显的算子：test_env 触发（dig/query）、APV 的 show/统计类命令。
# 用词边界匹配（避免 "listener" 命中 "list"、"display" 误伤等子串假阳性）。
_OBSERVE_RE = re.compile(
    r"\b(show|statistics|stat|display|dig|nslookup|get|list)\b", re.IGNORECASE
)

# ---- 分发闭集:一律从 mirror 框架源码**解析**,不手抄(权威=源码,手抄=漂移) ----
# 失败语义(源码事实):E 不在 devices 表 → locals().get(E)=None → 框架 continue,
# **整步静默跳过**(test_xlsx.py 分发循环);F 不在对象方法集 → getattr 无默认值
# AttributeError → **崩整份文件**。mirror 缺失/解析失败 → 对应闭集为空 → 相关
# 白名单检查自动跳过(fail-open,门宁可少管不误杀)。
from functools import lru_cache


def _mirror_src(rel: str) -> str:
    from pathlib import Path
    p = Path(__file__).resolve().parents[4] / "knowledge" / "framework" / "mirror" / rel
    try:
        return p.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return ""


@lru_cache(maxsize=1)
def _devices_keys() -> frozenset:
    """框架 devices dict 的键集(test_xlsx.py `devices = {...}` 字面解析)。"""
    src = _mirror_src("lib/test_xlsx.py")
    m = re.search(r"devices\s*=\s*\{(.*?)\}", src, re.S)
    return frozenset(re.findall(r"'([^']+)'\s*:", m.group(1))) if m else frozenset()


@lru_cache(maxsize=1)
def _host_slot_es() -> frozenset:
    """直连槽 = devices 表里经 request.getfixturevalue 拿 ssh_server 对象的键
    (conftest 里 `ssh_server(configer, logger, ...)` 构造的 fixture 名)。"""
    conftest = _mirror_src("smoke_test/conftest.py")
    slots = set()
    for m in re.finditer(r"def (\w+)\([^)]*\):(.*?)(?=\n@|\ndef |\Z)", conftest, re.S):
        if "ssh_server(" in m.group(2) and "class " not in m.group(2):
            slots.add(m.group(1))
    return frozenset(slots & _devices_keys())


def _public_methods(rel: str, cls: str) -> frozenset:
    """解析 mirror 某类的 def 方法名(用例面白名单的原料;私有/框架生命周期方法由
    调用方按源码事实剔除)。"""
    src = _mirror_src(rel)
    m = re.search(rf"class {cls}\b.*?(?=\nclass |\Z)", src, re.S)
    return frozenset(re.findall(r"\n    def (\w+)\(", m.group(0))) if m else frozenset()


# 框架生命周期方法(test_xlsx/conftest 自动调用,用例步骤调它会破坏 per-case 自动
# 恢复/日志切分——ssh_server.py:63-131 与 check_point.py:83+ 的调用点是依据):
_LIFECYCLE_FS = frozenset({
    "__init__", "xlsx_begin", "read_until", "close", "soft_close", "clear",
    "delete_ip", "delete_route", "load_synonyms", "check_direct_match",
    "get_same", "get_similar_function",
})


@lru_cache(maxsize=1)
def _valid_es() -> frozenset:
    keys = _devices_keys()
    # check_point/time 不在 devices 表但被框架特殊分发(fixture 参数/内置模块)
    return (keys | {"check_point", "time"}) if keys else frozenset()


@lru_cache(maxsize=1)
def _valid_fs_by_e() -> dict:
    env_hosts = _public_methods("lib/env.py", "Env") - _LIFECYCLE_FS
    cp = _public_methods("lib/check_point.py", "Check_Point") - _LIFECYCLE_FS
    slot_fs = _public_methods("lib/ssh_server.py", "ssh_server") - _LIFECYCLE_FS
    out: dict = {}
    if env_hosts:
        out["test_env"] = env_hosts
    if cp:
        out["check_point"] = cp
    out["time"] = frozenset({"sleep"})   # E=time 实为 python time 模块,框架仅对 sleep 有参数特判
    if slot_fs:
        for slot in _host_slot_es():
            out[slot] = slot_fs
    return out


@lru_cache(maxsize=4)
def _execute_returning_actions(src_rel: str) -> frozenset:
    """解析 mirror 动作注册表源码,返回**有 return 值**的动作名闭集。"""
    src = _mirror_src(src_rel)
    if not src:
        return frozenset()
    mapping = dict(re.findall(r"'([^']+)':\s*self\.(func_\d+)", src))
    returning: set[str] = set()
    for m in re.finditer(r"\n    def (func_\d+)\(self[^)]*\):(.*?)(?=\n    def |\Z)", src, re.S):
        if re.search(r"\n\s+return\s+\S", m.group(2)):
            returning.add(m.group(1))
    return frozenset(name for name, fn in mapping.items() if fn in returning)


# 注册表源码位置(E=APV_* 走设备侧 action 混入,E=直连槽走客户端侧 client_action)
_APV_ACTION_SRC = "lib/apv/apv_action.py"
_CLIENT_ACTION_SRC = "lib/client_action.py"


@dataclass
class StructuralViolation:
    code: str          # cmd_not_in_allowlist | dangling_assertion
    detail: str
    step_index: int = -1


@dataclass
class StructuralResult:
    ok: bool = True
    violations: list[StructuralViolation] = field(default_factory=list)

    def add(self, code: str, detail: str, step_index: int = -1) -> None:
        self.ok = False
        self.violations.append(StructuralViolation(code, detail, step_index))

    def render(self, autoid: str) -> str:
        lines = [f"case {autoid} 违反结构约束（correct-by-construction 门，与 grade 独立）："]
        for v in self.violations:
            loc = f"step[{v.step_index}] " if v.step_index >= 0 else ""
            lines.append(f"  - [{v.code}] {loc}{v.detail}")
        lines.append(
            "\n这些是与意图无关的**结构**错误（命令合法性 / 断言是否悬空），"
            "确定性可判、必须改对——不是骨架选择问题。修正后重新 emit。"
        )
        return "\n".join(lines)


def _command_head_tokens(cmd: str) -> list[str]:
    """从一条 CLI 命令取命令主体 token 序列（剥 no/show/clear 前缀、剥参数值）。

    "sdns listener 172.16.34.70 53" → ["sdns","listener"]
    "no slb real http rs1"          → ["slb","real","http"]
    "show sdns service status"      → ["sdns","service","status"]
    """
    toks = cmd.strip().split()
    while toks and toks[0].lower() in _OP_PREFIXES:
        toks = toks[1:]
    out: list[str] = []
    for t in toks:
        tl = t.strip().lower()
        if not tl:
            continue
        # 命令主体只取连续的字母 token；遇到第一个参数值 token 即停
        if not re.match(r"^[a-z][a-z0-9_-]*$", tl) or _VALUE_TOK_RE.match(tl):
            break
        out.append(tl)
    return out


def _load_allowlist_prefixes() -> tuple[set[str], set[str]]:
    """从 footprint 知识树取命令 allowlist。

    返回 (full_command_keys, module_roots)：
    - full_command_keys: 所有节点 feature_id 的点分形式（如 "sdns.listener"），
      以及每个节点 cli.commands 里命令的主体 token 路径；
    - module_roots: 一级模块名（feature_id 首 token，如 "sdns"/"slb"）。

    footprint 缺失/空 → 返回空集（调用方据此降级为不拦，避免误杀）。
    """
    full: set[str] = set()
    roots: set[str] = set()
    try:
        from main.ist_core.memory.footprint import get_footprint_index
        idx = get_footprint_index()
        for fid in idx.list_nodes():
            full.add(fid)
            roots.add(fid.split(".")[0])
    except Exception as exc:  # noqa: BLE001
        logger.debug("structural_gate 读 footprint 失败: %s", exc)
    return full, roots


def _check_command_allowlist(steps: list, init: str, result: StructuralResult) -> None:
    """命令头∈allowlist：APV 配置步骤 + init 的命令主体 token 路径须命中 footprint。

    分级判定（防 footprint 文法不全误杀，§八风险）：
    - footprint 空 → 整体跳过（不拦）；
    - 命令所属**一级模块**（首 token）不在 footprint 任何模块里 → 判违规（明显越界/幻觉）；
    - 模块在、但具体命令路径不在 → 只 logger 记录、不拦（footprint 可能没覆盖到该子命令）。
    """
    full, roots = _load_allowlist_prefixes()
    if not roots:
        return  # footprint 不可用 → 降级不拦

    def _check_one(cmd: str, idx: int) -> None:
        head = _command_head_tokens(cmd)
        if not head:
            return
        module = head[0]
        if module not in roots:
            result.add(
                "cmd_not_in_allowlist",
                f"命令 {cmd!r} 的模块 {module!r} 不在手册命令树任何模块中"
                f"（已知模块: {', '.join(sorted(roots))}）——疑似越界/幻觉命令。",
                idx,
            )

    # init 可能多行
    for line in (init or "").splitlines():
        line = line.strip()
        if line:
            _check_one(line, -1)
    for i, s in enumerate(steps):
        if not isinstance(s, dict):
            continue
        e = str(s.get("E", "")).strip()
        f = str(s.get("F", "")).strip()
        if e.startswith("APV") and f in ("cmd_config", "cmds_config"):
            for line in str(s.get("G", "") or "").splitlines():
                line = line.strip()
                if line:
                    _check_one(line, i)


def _is_observation_step(step: dict) -> bool:
    """该步骤是否让框架 `result` 持有非 None 回显（可被 check_point `re.search`，不抛 TypeError）。

    判据是**框架的结构化契约（F 列方法名）**，不是 grep G 列命令关键字——后者会把"失败的单条配置
    命令"误判成无回显（994957：`sdns host pool … p11` 不含 show/dig，却真回显错误文字 `A maximum
    of 10…`，逼 draft 多插一步 show 把 result 冲掉、断言读错缓冲）。框架事实：
      - test_env 触发（dig/clientc/routera 等）→ 回显。
      - APV `cmd_config`（**单条**，apv_ssh.py:151 `return output`）→ 回显；失败命令回显错误文字、
        成功命令回显 prompt，均非 None、可被断言。
      - APV `cmds_config`（**多条**，apv_ssh.py:166-169 遍历调 cmd_config 但**不收集返回值**）→ None。
    （test_xlsx.py:309 `result = func()`，func 即该 F 列方法：cmd_config 返回 output、cmds_config 返回 None。）
    """
    e = str(step.get("E", "")).strip()
    f = str(step.get("F", "")).strip()
    if e == "test_env":
        return True  # dig/clientc/routera 等触发都回显
    if e in _host_slot_es():
        # 直连槽(fixture 返回 ssh_server 对象;mirror lib/test_xlsx.py:176):
        # cmd 返回整段回显(ssh_server.py:93-106);execute 按动作定(部分动作无 return → None)。
        if f == "cmd":
            return True
        if f == "execute":
            g = str(step.get("G", "")).strip()
            action = g.split("：", 1)[0].strip()
            return action in _execute_returning_actions(_CLIENT_ACTION_SRC)
        return False
    if e.startswith("APV"):
        if f == "execute":
            # APV 类混入设备侧动作注册表(apv.py:23 class APV(...,action,...));按动作定
            g = str(step.get("G", "")).strip()
            action = g.split("：", 1)[0].strip()
            return action in _execute_returning_actions(_APV_ACTION_SRC)
        return f == "cmd_config"   # 单条 cmd_config 产 result 回显；cmds_config（多条）返回 None
    return False


def _check_dangling_assertions(steps: list, result: StructuralResult) -> None:
    """断言的 result 来源必须有效——否则框架 ``found(None)`` 抛 TypeError、**整份文件崩溃**
    （pytest 用例异常退出，该 case 之后的 case 全不执行；实证一个坏断言拖垮 31→只跑 4）。

    框架契约（实证 lib/test_xlsx.py + check_point.py）：
    - 非 check_point 步**仅当不带 H(save_as)** 时 `result = func(...)`；带 H 的步只把输出存进
      寄存器、**不更新 result**。
    - 配置步 `cmds_config/cmd_config` 的 func() 返回 None（不产回显）。
    - check_point 若 I 列(row[8])空，被测值 = 框架 `result`；`re.search(expect, None)` → TypeError 崩溃。
    ⟹ check_point(I 空) 的 result 非 None ⟺ 它**之前最近的「不带 H 的步」是观测算子**(dig/show 产回显)。
      若最近的不带 H 步是配置步(result=None)、或前面的观测步都带了 H(没更新 result) → 崩溃。

    与意图无关、机械可判。捕获比较的正确三步式：dig(H=v1 捕获基线) → dig(**无 H**，设 result) →
    check_point(H=v1)；draft 漏掉中间「无 H 观测步」时本门拦下。
    """
    result_is_observe = False   # 框架 `result` 当前是否持有观测回显（非 None 可被 re.search）
    for i, s in enumerate(steps):
        if not isinstance(s, dict):
            continue
        e = str(s.get("E", "")).strip()
        h = str(s.get("H", "") or "").strip()
        i_col = str(s.get("I", "") or "").strip()
        if e == "check_point":
            # I 列非空时被测值取 I 寄存器（另一路径，此处不判）；I 空时取框架 result。
            if not i_col and not result_is_observe:
                result.add(
                    "dangling_assertion",
                    "该断言取的 result 无有效观测回显 → 框架 result=None、found(None) 抛 TypeError "
                    "**崩溃整份文件**(该 case 之后全不跑)。成因:断言紧前最近的「不带 H 步」是配置步"
                    "(cmds_config 返回 None)、或前面的观测步都带了 H(save_as 不更新 result)。"
                    "修法:断言**紧前**放一个**不带 H** 的观测步(dig/show)让其回显成为 result;"
                    "捕获比较用三步式 dig(H=v1) → dig(无H) → check_point(H=v1)。",
                    i,
                )
            # check_point 不改变 result（它消费 result，不产出新回显）
        else:
            if not h:
                # 不带 H → 框架 `result = func()`;观测步→字符串(有效),配置步→None(无效)
                result_is_observe = _is_observation_step(s)
            # 带 H → 仅存寄存器、不更新 result → result_is_observe 不变


def _check_dead_capture(steps: list, result: StructuralResult) -> None:
    """寄存器写而不读（死捕获）：观测/动作步用 H(save_as) 把回显捕获进寄存器 v，但**没有任何
    check_point 引用 v**——写了一个永不读的寄存器，是废动作/残缺断言。

    典型（GA 断错缓冲）：dig 带 H 捕获命中 IP 进 v1/v3，但 check_point 去读框架 result（紧前常是
    `show host pool` 配置回显，里面没有该 IP）→ 既没验到 dig 行为(寄存器没人读)、断言又读错缓冲。
    这正是 grade 的 `is_genuine_v_assertion` 修 H 后仍不主动报警的那类（无 mutating、IP token 空），
    需本确定性门兜底。

    与意图无关、纯数据流（写入集 − 读取集），机械可判。**天然区分合法三步式**：dig(H=v1)→dig(无H)→
    check_point(H=v1)，v1 被 check_point 的 H 读了 → 不死；残缺混合体里 v1 无人读 → 死。
    读取集保守地收 check_point 的 H(作 expect 寄存器) + **所有步**的 I(作被查文本/输入引用)——
    宁可漏报、绝不误杀（对齐本门「静态约束宁漏报不误杀」原则，见 check_structural_constraints 注）。
    """
    captured: dict[str, int] = {}   # 寄存器名 → 捕获它的步 idx（非 check_point 步带 H）
    referenced: set[str] = set()    # 被引用（消费）的寄存器名
    for i, s in enumerate(steps):
        if not isinstance(s, dict):
            continue
        e = str(s.get("E", "")).strip()
        h = str(s.get("H", "") or "").strip()
        i_col = str(s.get("I", "") or "").strip()
        if i_col:
            referenced.add(i_col)            # 任何步的 I = 从寄存器取被查文本/输入
        if e == "check_point":
            if h:
                referenced.add(h)            # check_point 带 H = expect 从寄存器取（abs_found 比较）
        elif h:
            captured[h] = i                  # 观测/动作步带 H = 把回显捕获进寄存器
    for reg, idx in captured.items():
        if reg not in referenced:
            result.add(
                "dead_capture",
                f"寄存器 '{reg}' 被 save_as 捕获但**无任何 check_point 引用**（既不作 expect 的 H、"
                f"也不作被查文本的 I）= 写了永不读的寄存器、废动作/残缺断言。典型：dig 带 H 捕获进 "
                f"'{reg}' 却用 found 读 result(紧前常是 show 配置回显→断错缓冲)，没 abs_found(H='{reg}') "
                f"消费它 → 既没验到该 dig 行为、断言又读错缓冲。修法：用 check_point abs_found 引用 "
                f"'{reg}'（三步式 dig(H={reg})→dig(无H)→check_point(H={reg})），或删掉这个没用的捕获。",
                idx,
            )


def check_structural_constraints(autoid: str, steps: list, init: str = "") -> StructuralResult:
    """v2 结构约束门入口：命令 allowlist + 断言非悬空（IP 可达由 emit 既有门做）。

    返回 StructuralResult；ok=False 时 emit 应拒绝并把 render() 反馈给 draft 让它改。

    注（为何不做「引号/参数格式门」）：实证 `sdns host persistence` 先例**逐位置且不一致**
    （域名从不加引号、掩码总加引号），且先例集**不完整**（不覆盖 query_type 的 arity-5 合法形态）。
    任何静态格式门要么漏报、要么误杀合法形态。命令的参数格式正确性靠：①照最像先例**逐字复制**
    （compile_precedent 已强调）；②查 footprint/手册；③上机 verify 用设备 `^`/show 反馈兜底回流。
    """
    result = StructuralResult()
    if not isinstance(steps, list):
        return result
    _check_command_allowlist(steps, init, result)
    _check_dangling_assertions(steps, result)
    _check_no_found_times(steps, result)
    _check_dead_capture(steps, result)
    return result


def _check_no_found_times(steps: list, result: StructuralResult) -> None:
    """拒绝 found_times 断言——框架 xlsx 分派**不支持**它,必崩整份文件(A 层机械事实,非语义判断)。

    实证(lib/test_xlsx.py 无任何 found_times/times 特殊处理):check_point 一律走 generic 2 参分派
    `func(expect, result)`,而 `check_point.found_times(expect, result, times)` 需 3 参 →
    `TypeError: found_times() missing 1 required positional argument: 'times'` → pytest 崩、
    该 case 之后全不跑(实证 yzg 10 pass 后崩、14 不执行)。
    改用 `found`(出现即可)或 `abs_found`(字面);要"恰好 N 次"语义,框架表达不了,换断言。
    注:这是**机械崩溃**判定(等同语法错),与"某 claim 可不可证伪"的语义判断(归 verifiability 工具 +
    LLM)是两码事——本门只认 F==found_times 这个必崩形态,不在这里算可证伪性。
    """
    for i, s in enumerate(steps):
        if not isinstance(s, dict):
            continue
        if str(s.get("E", "")).strip() == "check_point" and str(s.get("F", "")).strip() == "found_times":
            result.add(
                "found_times_unsupported",
                "found_times 框架 xlsx 流不支持(分派只传 2 参、缺 times)→ TypeError 崩整份文件。"
                "改用 found(出现即可)/abs_found(字面匹配);'恰好 N 次'语义本框架表达不了。",
                i,
            )


def check_no_found_times_mandatory(steps: list) -> StructuralResult:
    """found_times **无条件**拒绝门——与 strict_structural opt-in 解耦(A 层机械崩溃门)。

    found_times 保证崩框架(分派只传 2 参必 TypeError,见 _check_no_found_times);其拒绝绝不该可选。
    实证:draft agent 偶尔漏传 `strict_structural=True`(tool 默认 False 的 prompt 指令) → 整个结构门
    跳过 → found_times 漏进 excel → 上机崩、崩溃点后全 case unknown(dongkl 实测 31 unknown 根因)。
    故 emit **无条件**先跑这一条(其余启发式门仍留 strict_structural opt-in)。等同 SWE-agent 的
    linter"语法不对不让 edit 过"——机械可判、误判即真错,不误杀好制品。
    """
    result = StructuralResult()
    if isinstance(steps, list):
        _check_no_found_times(steps, result)
    return result


def _check_no_manual_ip_cleanup(steps: list, result: StructuralResult) -> None:
    """test_env / 主机直连槽步禁止 ``ip addr add/del``(主机网络状态变更)——框架契约的机械事实。

    框架 ``lib/ssh_server.py:93-106`` 的记账在 ``ssh_server.cmd`` 里(``addr add`` → ip_list,
    不去重),恢复在每 case 开头 ``xlsx_begin → soft_close → delete_ip``——test_env 转发形态
    与直连槽形态走的是**同一个类的同一条 cmd 路径**,契约对两者同样成立。两条实证的必炸路径:
    - case 内自行 del → 框架恢复对已删 IP delete → RTNETLINK 报错 → 崩整份文件
      (dongkl_final: 1 pass + 22 unknown 级联);
    - 多个 case add 同一 IP → 记账两条 → 恢复时第二条 delete 必失败 → 失败输出残留
      触发机 SSH 通道,污染后续 case 的 dig 回显 → 成批假 fail(第五轮 8/9、第七轮 7/8)。
    触发机网络状态是框架管理的基础设施;rr/wrr 轮转按请求轮转、不按源,不需要多源 IP——
    删掉 ip addr 步,直接用拓扑既有触发机发请求。
    """
    for i, s in enumerate(steps):
        if not isinstance(s, dict):
            continue
        e = str(s.get("E", "")).strip()
        g = " ".join(str(s.get("G", "") or "").split())
        if (e == "test_env" or e in _host_slot_es()) and (
                "ip addr " in g or "ip address " in g
                or "ip route " in g or "ip -6 route " in g) and (
                " add" in g or " del" in g):
            # 同一契约的两半:addr(ip_list,:99-100)与 route(route_list,:95-98)都被
            # ssh_server.cmd 记账、soft_close 时 delete_ip/delete_route 恢复。
            result.add(
                "manual_ip_cleanup",
                "变更测试环境主机 IP/路由(ip addr / ip route 的 add/del)——框架对 add "
                "自动记账(不去重)并在下一 case 开头 delete 恢复:自行 del 或跨 case 重复 "
                "add 都会让恢复失败,**崩整份文件或以 RTNETLINK 残留污染后续 case 的回显"
                "(成批假 fail)**。删掉这个步骤——主机网络状态由框架管理;要多源就用"
                "拓扑既有的多台触发机,要新路径先查拓扑既有可达性。",
                i,
            )


def _check_dispatch_targets(steps: list, result: StructuralResult) -> None:
    """E/F 必须在框架分发闭集内(闭集从 mirror 源码解析,见 _valid_es/_valid_fs_by_e)。

    两种失败模式,都机械可判:
    - E 不在 devices 表 → 框架静默跳过整步(无异常无日志)——步骤蒸发,后续断言拿错缓冲;
    - F 不在该对象方法闭集 → getattr AttributeError → 崩整份文件。
    E=test_env 的 F 先做 emit 同款小写归一再判(大小写是可修形态,emit 会归一,不误杀)。
    """
    for i, s in enumerate(steps):
        if not isinstance(s, dict):
            continue
        e = str(s.get("E", "")).strip()
        f = str(s.get("F", "")).strip()
        if not e:
            continue   # 缺 E 由 emit 入口拒,不重复报
        _es = _valid_es()
        if _es and e not in _es:
            result.add(
                "unknown_dispatch_target",
                f"E={e!r} 不在框架 devices 表(合法: {', '.join(sorted(_es))})——"
                "框架对未知 E **静默跳过整步**(不执行、无日志异常),后续断言拿错缓冲跑。"
                "对照 EXCEL_FUNCTIONS.md 的 E 表改对。",
                i,
            )
            continue
        allowed = _valid_fs_by_e().get(e)
        if allowed is None:
            continue   # APV*/Seg* 方法集未全量取证,不设白名单
        f_norm = f.lower() if e == "test_env" else f
        if f_norm not in allowed:
            result.add(
                "unknown_dispatch_method",
                f"F={f!r} 不是 E={e} 对象的合法方法(合法: {', '.join(sorted(allowed))})——"
                "框架 getattr 无默认值,方法名拼错 **AttributeError 崩整份文件**。"
                "对照 EXCEL_FUNCTIONS.md 该 E 的 F 说明改对。",
                i,
            )


def _check_command_payload_sanity(steps: list, result: StructuralResult) -> None:
    """命令载荷完整性——G 列空/None、或含**字面** ``\\n`` 转义(非真实换行)必拒。

    两个实证的必拒形态(2026-07-02 第八轮:LLM 重编 steps_json 质量问题批量漏进 xlsx,
    11 个 case 全部 G(^) 被设备拒):
    - G 为空/None:框架把单元格 str 化后原样发送,设备收到 "None" → ``^`` 拒;
    - G 含字面反斜杠 n(``\\\\n`` 两字符,非换行符):多条命令被拼成一行发送,设备在第二条
      命令词处 ``^`` 拒。多命令的正确形态是**真实换行**分隔(cmds_config 逐行发送)。
    机械可判、与命令内容无关——不看命令是什么,只看载荷形态。
    """
    for i, s in enumerate(steps):
        if not isinstance(s, dict):
            continue
        e = str(s.get("E", "")).strip()
        # 会把 G 当命令发出去的步全查:APV_0/1/2、test_env 转发、主机直连槽。
        # (旧版漏 APV_1/直连槽——2026-07-05 框架取证后对齐;check_point 的 G 是断言表达式不查。)
        if not (e.startswith("APV") or e == "test_env" or e in _host_slot_es()):
            continue
        g_raw = s.get("G")
        g = str(g_raw) if g_raw is not None else ""
        # 精确对齐框架行为:G 为**空串**时 cmds_config 对 splitlines() 零循环、什么也不发
        # ——无害(实证 merge 空 init 的占位步,case 照常 PASS),放行。
        # 危险形态是 None(json null,下游 str 化成 "None")与字面 "None" 字符串——发给设备必 ^ 拒。
        if g_raw is None or g.strip().lower() == "none":
            result.add(
                "empty_command_payload",
                "命令步 G 列为 None/字面\"None\"——框架 str 化后原样发送,设备收到 \"None\" 必被"
                " ^ 拒。补上真实命令或删掉该步(纯空串占位步无害,不在此列)。",
                i,
            )
        elif str(s.get("F", "")).strip() == "cmd_config" and ("\n" in g.strip() or "\r" in g.strip()):
            result.add(
                "cmd_config_multiline",
                "cmd_config 的 G 含换行——框架对 cmd_config 先 replace 删光换行"
                "(test_xlsx.py:307 single_line_parameter),多条命令**无分隔直接粘连**成"
                "一条发送,设备必 ^ 拒。多条命令改用 cmds_config(逐行发送)。",
                i,
            )
        elif "\\n" in g:
            result.add(
                "literal_backslash_n",
                "命令 G 列含**字面** \\\\n(反斜杠+n 两个字符,不是换行符)——多条命令会被拼成"
                "一行发送,设备在第二条命令处必 ^ 拒。多命令用**真实换行**分隔(JSON 里写 \\n "
                "转义会由解析还原为换行;若你在字符串里写了 \\\\\\\\n 就成了字面反斜杠,改掉)。",
                i,
            )


def check_crash_gates_mandatory(steps: list) -> StructuralResult:
    """必崩形态**无条件**拒绝门集合——与 strict_structural opt-in 解耦(A 层机械崩溃门)。

    收录标准(严进):该形态上机**保证**崩整份 pytest 文件(崩溃点之后全不跑)——
    误判即真错、不存在误杀好制品的可能。当前三条:
    - found_times:框架分派只传 2 参必崩(_check_no_found_times;dongkl 首跑 31 unknown 根因)。
    - 悬空断言:check_point(I 空)前无 result 生产步 → found(None) 必崩(_check_dangling_assertions;
      实证 dongkl 778012 重编版:配置步后直接断言,worker 漏传 strict_structural 使 opt-in 门被
      跳过而漏网 → 第三轮上机 1 pass + 33 unknown)。
    - test_env 自行 ip addr del:与框架自动恢复双重清理,崩下一 case(_check_no_manual_ip_cleanup;
      实证 dongkl_final 22 unknown 级联)。
    - 载荷完整性:G 为 None/字面 \\n → 设备必 ^ 拒(_check_command_payload_sanity)。
    - 分发白名单:E 不在 devices 表=整步静默蒸发、F 不在方法闭集=getattr AttributeError
      崩整卷(_check_dispatch_targets;2026-07-05 框架源码取证,白名单全为源码闭集)。
    教训同源:必崩类检查躲在 opt-in 开关后面就等于没有——draft agent 会漏传参数。启发式/
    allowlist 类门仍留 strict_structural(可能误杀,须可选);必崩类一律进本门。
    """
    result = StructuralResult()
    if isinstance(steps, list):
        _check_no_found_times(steps, result)
        _check_dangling_assertions(steps, result)
        _check_no_manual_ip_cleanup(steps, result)
        _check_command_payload_sanity(steps, result)
        _check_dispatch_targets(steps, result)
        _check_line_anchor_assertions(steps, result)
        _check_assertion_matches_command_echo(steps, result)
        _check_has_assertion(steps, result)
        # 2026-07-06 从 lint 前移:未定义 I 引用=NameError 崩卷、I 注入 format 结构坏=
        # ValueError/KeyError/IndexError 崩卷、H 撞框架名字空间=覆盖执行器状态——全部
        # 必崩/必污染类,按本门收录标准(必崩类不躲后置卡点)进 emit 无条件门。
        _check_capture_refs_defined(steps, result)
    return result


# ---------------------------------------------------------------------------
# xlsx 级 lint —— 把必崩门挂到"成品卷"本身,堵死绕过 emit 的编辑路径
# ---------------------------------------------------------------------------
# 起因(dongkl 34-case 闭环实证,2026-07-04):orchestrator 用 run_python 直改
# case.xlsx 修断言,不经 compile_emit → check_crash_gates_mandatory 完全失效;
# 直改版带"dig(H=x)后直接断言"形态,上机 result=None 抛 TypeError,整份 pytest
# 39 秒崩掉、34 case 只跑 1 个,连续两轮。门放在编辑入口挡不住绕行,放在
# **凭证/合并的必经之路**才是不变量:任何来源的卷面(emit 产 / 直改 / 手工)
# 要拿 grade 凭证、要进合并终卷,都必须过同一套 lint。

_AUTOID_RE = re.compile(r"^\d{18}$")
# +short 模式 dig 的输出只有记录值(无 header/status/ANSWER 段)——对这些文本断言必 fail
_SHORT_INCOMPATIBLE_RE = re.compile(r"status:|->>HEADER<<-|ANSWER SECTION|QUESTION SECTION")


def steps_from_xlsx(xlsx_path) -> tuple[str, list[dict]]:
    """反解单 case 卷(框架模板 xlsx)为 (autoid, steps dict 列表)。

    列序与 emit 写入一致:A=autoid B=priority C=步号 D=desc E=对象 F=方法 G=参数
    H=save_as I=引用寄存器。只取首个 autoid 的连续步骤区(单卷=一个 case)。
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    autoid = ""
    steps: list[dict] = []
    # 对齐框架执行语义(test_xlsx.py `case_begin`):A 列="自动化ID" 表头行**之后**的行
    # 才是步骤;之前是模板说明区/字典区(框架不执行,反解也不该当步骤收——旧版全扫
    # 恰好没炸,分发白名单门一严说明行就现形误报)。
    begun = False
    for row in ws.iter_rows(min_row=2):
        a = str(row[0].value or "").strip()
        if not begun:
            if a == "自动化ID":
                begun = True
            continue
        e = str(row[4].value or "").strip()
        if a and _AUTOID_RE.match(a) is None and a.startswith("203"):
            # 位数异常的 autoid 也要带出去让上层报——用原文记录
            autoid = autoid or a
        elif a.startswith("203"):
            autoid = autoid or a
        if not e:
            continue
        steps.append({
            "D": str(row[3].value or ""),
            "E": e,
            "F": str(row[5].value or ""),
            "G": str(row[6].value or ""),
            "H": str(row[7].value or ""),
            "I": str(row[8].value or "") if len(row) > 8 else "",
        })
    return autoid, steps


def _check_line_anchor_assertions(steps: list, result: StructuralResult) -> None:
    """found/not_found 的 ^/\\A 行首锚必假门(2026-07-06 588691 三轮实证)。

    框架 check_point.found/not_found 用 ``re.compile(expect, re.DOTALL)``——**无
    MULTILINE**,``^`` 只匹配字符串开头,而 result 开头永远是命令回显行:
    - found ^X  → 永不匹配 → 恒 fail;
    - not_found ^X → 永不匹配 → **恒真假 PASS**(什么都没验证,比恒 fail 危险)。
    行级匹配用 ``\\n`` 前缀或直接去锚(DOTALL 下 re.search 全文扫)。
    """
    for i, s in enumerate(steps):
        if not isinstance(s, dict):
            continue
        if str(s.get("E", "")).strip() != "check_point":
            continue
        f = str(s.get("F", "")).strip()
        g = str(s.get("G", "") or "")
        if re.match(r"\(\?[aiLmsux]*m[aiLmsux]*\)", g):
            continue   # 内联 (?m) 开 MULTILINE,行锚合法生效——恰是知道该语义后的正确修法
        if f in ("found", "not_found") and (g.endswith("$") and not g.endswith("\\$")
                                            or g.endswith("\\Z")):
            result.add(
                "line_anchor_never_matches",
                f"断言 {f} 的模式以结尾锚收尾——无 MULTILINE 下 $/\\Z 只匹配字符串末尾,"
                "而窗口末尾永远是设备提示符(read_until(prompt)),数据行的结尾锚永不匹配:"
                + ("该断言**恒真**(永远找不到=永远通过),什么都没验证。"
                   if f == "not_found" else "该断言**恒 fail**。")
                + "去掉结尾锚,或用 \\n 界定数据行边界。",
                i,
            )
            continue
        if f in ("found", "not_found") and (g.startswith("^") or g.startswith("\\A")):
            result.add(
                "line_anchor_never_matches",
                f"断言 {f} 的模式以 {g[:2]!r} 开头——框架 found/not_found 是 DOTALL"
                "、**无 MULTILINE**,行首锚只匹配字符串开头(开头是命令回显行),"
                + ("该断言**恒真**(永远找不到=永远通过),什么都没验证。"
                   if f == "not_found" else "该断言**恒 fail**。")
                + "要匹配数据行,用 \\n 前缀锚定换行后,或去掉锚(全文搜索)。",
                i,
            )


def _echo_src_cmd(e: str, f: str) -> bool:
    """该步的 G 是否会以命令原文形式回显在其输出窗口开头(send(cmd)+read_until(prompt),
    ssh_server.py:cmd / env.py 各主机方法 / APV cmd_config 同一机制)。execute 的 G 是
    动作调用式非命令原文、cmds_config 无 return(result 被置 None)——都不算。"""
    if e == "test_env":
        return True
    if e in _host_slot_es() and f == "cmd":
        return True
    if e.startswith("APV") and f == "cmd_config":
        return True
    return False


def _check_assertion_matches_command_echo(steps: list, result: StructuralResult) -> None:
    """断言模式命中「被测窗口来源步的命令原文」必假门(2026-07-06 588691 round1 同族收口)。

    框架窗口 = 命令回显 + 输出 + 提示符(send(cmd)+read_until(prompt))——模式若能在命令
    原文上匹配,则不论设备输出什么都必中:
    - found/abs_found → **恒真假 PASS**(什么都没验证,最危险);
    - not_found → 恒 fail(想验「输出无 X」但 X 就在命令里,如 not_found 查询目标 IP)。
    典型:dig @<服务IP> 后断言该服务 IP——回显必含 @IP。断言应匹配**数据行**,与命令原文
    区分开(加词边界仍撞命令时,换能定位数据行的上下文,如 \\n 前缀+相邻字段)。
    """
    last_src_g: str | None = None      # 最近更新框架 result 的步:命令原文(可回显)或 None
    reg_src: dict[str, str] = {}       # 寄存器名 → 捕获步命令原文(仅可回显步)
    for i, s in enumerate(steps):
        if not isinstance(s, dict):
            continue
        e = str(s.get("E", "")).strip()
        f = str(s.get("F", "")).strip()
        g = str(s.get("G", "") or "")
        h = str(s.get("H", "") or "").strip()
        i_col = str(s.get("I", "") or "").strip()
        if e == "check_point":
            if h or not g.strip() or f not in ("found", "not_found", "abs_found"):
                continue   # 期望取寄存器(运行时值)/空模式/found_times 另有门
            src = reg_src.get(i_col) if i_col else last_src_g
            if not src:
                continue
            try:
                hit = (src.find(g) >= 0 if f == "abs_found"
                       else re.compile(g, re.DOTALL).search(src) is not None)
            except re.error:
                continue   # 编译失败由 regex 门报
            if hit:
                result.add(
                    "assertion_matches_command_echo",
                    f"断言 {f} 的模式在其被测窗口的**命令原文**上就能匹配({src[:60]!r}…)——"
                    "窗口开头永远是命令回显,该断言"
                    + ("**恒 fail**(想验「输出无 X」但 X 在命令里)。"
                       if f == "not_found" else
                       "**恒真假 PASS**(不论设备输出什么都通过,什么都没验证)。")
                    + "改成匹配数据行的形态,与命令原文区分开。",
                    i,
                )
            continue
        if h:
            if _echo_src_cmd(e, f):
                reg_src[h] = g
            else:
                reg_src.pop(h, None)
        else:
            last_src_g = g if _echo_src_cmd(e, f) else None


def _check_has_assertion(steps: list, result: StructuralResult) -> None:
    """零断言 case 恒 fail 门。框架 per-case 结算(check_point.py:124-128,case 边界
    test_xlsx.py:253/344 调 close()):fail>0 → FAIL;**success==0 也 FAIL**——一个
    check_point 都没有的 case 无从产生 success,上机必 FAIL,且日志无任何断言明细可归因。
    """
    for s in steps:
        if isinstance(s, dict) and str(s.get("E", "")).strip() == "check_point":
            return
    if steps:
        result.add(
            "no_assertion_in_case",
            "该 case 没有任何 check_point 步——框架结算对 success==0 判 FAIL"
            "(check_point.py:126),纯配置/纯观测卷上机恒 fail。补至少一条断言;"
            "只想执行不验证的步不构成测试用例。",
            0,
        )


def _check_assertion_regex_compiles(steps: list, result: StructuralResult) -> None:
    """check_point 的 G 列按 Python 正则编译——编译失败(如未闭合 `[^`)则框架
    re.compile 处抛异常崩整份文件。实证:593545 一轮直改版带 `[^` 进卷被 grade 抓到,
    若漏网上机即崩。abs_found 语义是字面匹配(框架 escape),不判。"""
    for i, s in enumerate(steps):
        if not isinstance(s, dict) or str(s.get("E", "")).strip() != "check_point":
            continue
        f = str(s.get("F", "")).strip()
        g = str(s.get("G", "") or "")
        if f not in ("found", "not_found") or not g:
            continue
        try:
            re.compile(g)
        except re.error as exc:
            result.add(
                "assertion_regex_invalid",
                f"断言正则无法编译({exc}): {g[:80]!r} —— 框架 re.compile 处抛异常,"
                "整份文件崩溃。修正正则语法(常见:未闭合的字符类 [^ 应写 [^\\n])。",
                i,
            )


def _check_short_mode_assertions(steps: list, result: StructuralResult) -> None:
    """dig +short 的输出只有记录值——紧随其后的断言若匹配 status:/HEADER/SECTION
    类文本必 fail(不崩但恒假)。实证:211027 两轮 fail 的直接根因。"""
    last_obs_short = False
    for i, s in enumerate(steps):
        if not isinstance(s, dict):
            continue
        e = str(s.get("E", "")).strip()
        if e == "check_point":
            g = str(s.get("G", "") or "")
            if last_obs_short and _SHORT_INCOMPATIBLE_RE.search(g):
                result.add(
                    "short_mode_status_assertion",
                    "断言要匹配 dig 的 status/HEADER/SECTION 文本,但它消费的观测步用了 "
                    "+short(输出只有记录值,无这些段)——恒 fail。去掉该 dig 的 +short,"
                    "或改断言为记录值形态。",
                    i,
                )
            continue
        h = str(s.get("H", "") or "").strip()
        if not h:
            g = str(s.get("G", "") or "")
            last_obs_short = ("dig" in g and "+short" in g)


def _check_capture_refs_defined(steps: list, result: StructuralResult) -> None:
    """H/I 变量引用必须先捕获(框架源码 test_xlsx.py:293-336):
    - check_point 的 H/I 引用未捕获变量 → locals().get() 得 None → 断言失真或 TypeError;
    - **非 check_point 步的 I** 引用未捕获变量 → 框架 ``raise NameError``(:319-324)**崩整卷**;
    - 非 check_point 步带 I 但 G 无 ``{}`` 占位 → format 无操作,注入静默不发生(值没用上,
      断言表面跑通实际验错东西)。反向(G 有 {} 无 I)不判:shell 命令里 {} 有合法用法(awk 等)。
    """
    defined: set[str] = set()
    for i, s in enumerate(steps):
        if not isinstance(s, dict):
            continue
        e = str(s.get("E", "")).strip()
        h = str(s.get("H", "") or "").strip()
        i_col = str(s.get("I", "") or "").strip()
        if e == "check_point":
            for ref in (h, i_col):
                if ref and ref not in defined:
                    result.add(
                        "undefined_capture_ref",
                        f"断言引用寄存器 {ref!r},但之前没有任何步骤用 H={ref} 捕获过——"
                        "框架取到 None,断言失真。补捕获步或改引用名。",
                        i,
                    )
        else:
            if i_col:
                # I 支持 obj.attr(框架 :315-318);obj 可为已捕获变量或 devices 对象
                base = i_col.split(".", 1)[0]
                if base not in defined and base not in _valid_es():
                    result.add(
                        "undefined_capture_ref",
                        f"步骤 I={i_col!r} 引用的变量没有被之前任何步骤用 H 捕获过——"
                        "框架对非断言步的未定义 I **raise NameError,崩整份文件**。"
                        "先用 H 捕获,或改引用名。",
                        i,
                    )
                g = str(s.get("G", "") or "")
                # format 结构安全(框架 :309 parameters[0].format(单值)):裸 {/} →
                # ValueError、命名占位 → KeyError、匿名占位 ≥2 → IndexError,全部崩整卷。
                from string import Formatter
                _fmt_err = ""
                try:
                    _fields = [fx for _, fx, _, _ in Formatter().parse(g) if fx is not None]
                    _bad_named = [fx for fx in _fields if fx not in ("", "0")]
                    if _bad_named:
                        _fmt_err = (f"G 含命名/多位置占位 {{{_bad_named[0]}}} ——框架只传一个"
                                    "值,format 抛 KeyError/IndexError")
                    elif len(set(_fields)) > 1:
                        # {} 与 {0} 混用 → ValueError(cannot switch numbering)
                        _fmt_err = "G 混用自动 {} 与手动 {0} 占位——format 抛 ValueError"
                    elif _fields.count("") > 1:
                        # 纯 {0} 重复引用合法({0} {0}.format(单值) 不抛);纯 {} 多个才 IndexError
                        _fmt_err = f"G 含 {len(_fields)} 个自动占位符——框架只传一个值,format 抛 IndexError"
                except ValueError:
                    _fmt_err = ("G 含未配对的 {{ 或 }}(JSON 载荷等字面大括号)——I 注入时框架"
                                " format 抛 ValueError;字面大括号写成 {{{{ }}}}")
                if _fmt_err:
                    result.add(
                        "injection_format_crash",
                        f"步骤带 I={i_col!r} 时 G 经框架 str.format 注入:{_fmt_err},"
                        "**崩整份文件**。",
                        i,
                    )
                elif "{}" not in g:
                    result.add(
                        "injection_without_placeholder",
                        f"步骤带 I={i_col!r} 但 G 里没有 {{}} 占位符——框架把变量 format 进"
                        " G 的 {},没有占位注入就静默不发生,这一步跑的还是原文。"
                        "在 G 里要用变量值的位置写 {}。",
                        i,
                    )
            if h:
                if h in _framework_reserved_names():
                    result.add(
                        "register_shadows_framework_name",
                        f"H={h!r} 与框架执行器的名字冲突(test_xlsx.py 执行帧的参数/局部变量"
                        "或 devices 槽)——同名下后续 locals().get(H) 读回的是**框架对象而非"
                        "你捕获的回显**(如 result/value/设备槽),断言/I 注入拿错值必错。"
                        "换一个普通寄存器名(v1/ip1 这类)。",
                        i,
                    )
                defined.add(h)


@lru_cache(maxsize=1)
def _framework_reserved_names() -> frozenset:
    """框架执行器名字空间闭集——从 mirror lib/test_xlsx.py 用 ast 解析(函数参数+全部
    Store 上下文的 Name),并集 devices 槽(_valid_es)。不手抄:框架升级重解析即准。"""
    names: set[str] = set(_valid_es())
    src = _mirror_src("lib/test_xlsx.py")
    if src:
        import ast as _ast
        try:
            tree = _ast.parse(src)
        except SyntaxError:
            tree = None
        if tree is not None:
            # 只收执行帧真实存在的名字:模块级 Store + 执行函数(test_xlsx*)的参数与
            # 局部 Store——曾 walk 全文件所有函数,把 get_parameter 等辅助函数的局部名
            # (i/m/key/name 这类常用短名)也收进来,误杀无害寄存器名(评审建议收窄)。
            for node in tree.body:
                if isinstance(node, _ast.Assign):
                    for tgt in node.targets:
                        for n in _ast.walk(tgt):
                            if isinstance(n, _ast.Name):
                                names.add(n.id)
                elif isinstance(node, (_ast.FunctionDef, _ast.AsyncFunctionDef))                         and node.name.startswith("test_xlsx"):
                    for a in list(node.args.args) + list(node.args.kwonlyargs):
                        names.add(a.arg)
                    for sub in _ast.walk(node):
                        if isinstance(sub, _ast.Name) and isinstance(sub.ctx, _ast.Store):
                            names.add(sub.id)
    return frozenset(n for n in names if n)


# 框架 get_parameter 的切分正则(test_xlsx.py:57,原样复刻):单行 G 按**引号外**逗号切段
_PARAM_SPLIT_RE = re.compile(r'((?:(?:"(?:\\.|[^\\"])*"|\'(?:\\.|[^\\\'])*\'|[^,])+))')
_KWARG_SEG_RE = re.compile(r"^\s*(timeout|prompt)\s*=")


def _check_parameter_splitting(steps: list, result: StructuralResult) -> None:
    """test_env/直连槽步的 G 含引号外英文逗号 → 框架 get_parameter 把 G 切成多个位置
    参数(test_xlsx.py:48-80),第二段起会错位传给主机方法的 prompt/timeout 形参——
    read_until 等一个不存在的提示符,恒等满 timeout、输出可能截断,**静默失真不报错**。
    G 含真实换行时框架整段单参(不切),安全;``timeout=``/``prompt=`` 形态的段是框架
    支持的具名参数惯用法,放行。APV 的 cmd_config 有单行拼接特判、check_point 不走
    get_parameter,均不受影响。
    """
    for i, s in enumerate(steps):
        if not isinstance(s, dict):
            continue
        e = str(s.get("E", "")).strip()
        if not (e == "test_env" or e in _host_slot_es()):
            continue
        f = str(s.get("F", "")).strip()
        if f == "execute":
            continue   # execute 的 G 整段传动作函数,中文逗号是其参数语法,不切英文段判定
        g = str(s.get("G", "") or "")
        if "\n" in g or "," not in g:
            continue
        segs = [p.strip() for p in _PARAM_SPLIT_RE.findall(g) if p.strip()]
        stray = [p for p in segs[1:] if not _KWARG_SEG_RE.match(p)]
        if stray:
            result.add(
                "comma_splits_parameters",
                f"G 含引号外英文逗号,框架会把它切成 {len(segs)} 个参数——"
                f"{stray[0]!r} 会错位传给主机方法的 prompt/timeout 形参,该步恒等满超时、"
                "输出失真。逗号是命令一部分时给该段加引号,或改写命令避开逗号;"
                "要传超时用 timeout=N 形态(框架具名参数)。",
                i,
            )


_DOMAIN_TOKEN_RE = re.compile(r"\b([A-Za-z0-9][A-Za-z0-9.-]{10,}\.(?:com|net|org|cn|test))\b")


def _check_dns_label_limit(steps: list, result: StructuralResult) -> None:
    """DNS 单标签最长 63 字符(RFC 1035)——超限域名 dig 直接报 'not a legal IDNA2008
    name'、查询永远发不出去,该 case 必 fail。实证:994838 的"128字符域名"需求被写成
    单标签 120 字符,三轮上机 fail 后才定位到协议物理约束。长域名需求的合法形态是
    多标签拼总长(每段 ≤63)。"""
    seen: set[str] = set()
    for i, s in enumerate(steps):
        if not isinstance(s, dict):
            continue
        g = str(s.get("G", "") or "")
        for dom in _DOMAIN_TOKEN_RE.findall(g):
            if dom in seen:
                continue
            seen.add(dom)
            too_long = [lab for lab in dom.split(".") if len(lab) > 63]
            if too_long:
                result.add(
                    "dns_label_over_63",
                    f"域名 {dom[:60]}… 含超过 63 字符的标签(长 {len(too_long[0])})——"
                    "违反 DNS 单标签上限,dig 侧 IDNA 直接拒绝、查询永远失败。"
                    "长域名需求用多标签拼总长(每段≤63,如 www.<61字符>.<58字符>.com)。",
                    i,
                )


def lint_xlsx_case(xlsx_path) -> StructuralResult:
    """成品卷必崩/必假 lint:凭证(submit_verdict/compile_score)与合并(emit_merged)
    的**必经门**。规则全部与意图无关、机械可判:
    - autoid 18 位(截断 id 曾混入终卷成 35 case);
    - check_crash_gates_mandatory 全集(悬空断言/found_times/手动 ip del/载荷);
    - 断言正则可编译;
    - +short 观测 与 status 类断言互斥;
    - 寄存器引用必先捕获;
    - G 逗号切参 / autoid 行 E 列非空(框架 ifrun:E 空整 case 静默不跑)。
    """
    result = StructuralResult()
    try:
        autoid, steps = steps_from_xlsx(xlsx_path)
    except Exception as exc:  # noqa: BLE001
        result.add("xlsx_unreadable", f"卷面无法读取: {exc}")
        return result
    if autoid and not _AUTOID_RE.match(autoid):
        result.add(
            "autoid_malformed",
            f"卷面 autoid {autoid!r} 不是 18 位数字——手抄截断 id 会静默生成垃圾目录并"
            "混入终卷(实证曾致终卷 35 case)。以 last_run.json/manifest 的机读全名为准。",
        )
    mand = check_crash_gates_mandatory(steps)
    if not mand.ok:
        result.ok = False
        result.violations.extend(mand.violations)
    _check_assertion_regex_compiles(steps, result)
    _check_short_mode_assertions(steps, result)
    _check_dns_label_limit(steps, result)
    _check_parameter_splitting(steps, result)
    _check_autoid_rows_runnable(xlsx_path, result)
    return result


def _check_autoid_rows_runnable(xlsx_path, result: StructuralResult) -> None:
    """autoid 行(每 case 首行)的 E 列必须非空——框架 ifrun(test_xlsx.py:88-90)对
    E 为 None 的 autoid 行直接 return False,**整个 case 静默不跑**(无日志无 fail,
    执行统计里凭空少一个 case)。emit 把 autoid 与第一步同行写入天然满足;人工卷/直改
    卷常把 autoid 写成独占行,即触发。卷面级检查(steps 反解已丢行绑定信息)。
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
    except Exception:  # noqa: BLE001 — 卷面不可读已由上游报
        return
    begun = False
    for row in ws.iter_rows(min_row=2):
        a = str(row[0].value or "").strip()
        if not begun:
            begun = a == "自动化ID"
            continue
        if a.startswith("203") and not str(row[4].value or "").strip():
            result.add(
                "autoid_row_not_runnable",
                f"autoid 行({a})的 E 列为空——框架 ifrun 对此整 case **静默跳过**"
                "(不执行、不计 fail、无日志)。autoid 必须与第一个步骤同行(E 列非空)。",
            )
