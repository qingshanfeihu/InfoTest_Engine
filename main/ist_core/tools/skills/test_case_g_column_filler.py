"""qa_fill_g_column: 生成 xlsx + LLM 驱动 G 列填充.

将 qa_decompose_test_cases 产出的 decomposed JSON → 生成标准 xlsx →
调用 fork skill g-column-filler (LLM 子代理) 填充 G 列 → 输出 filled_*.xlsx.

管线：
1. 按 group 生成 xlsx（复用 _build_xlsx_for_group）
2. 逐 xlsx 读取 D/E/F 列 + 网络拓扑
3. 构建 structured brief，调 execute_fork_skill("g-column-filler", brief)
4. 解析 g_updates JSON，写入 G 列
"""

from __future__ import annotations

import json
import re
from collections import OrderedDict
from pathlib import Path
from typing import Any

from langchain_core.tools import tool

_PROJECT_ROOT = Path(__file__).resolve().parents[4]
_OUTPUT_BASE = _PROJECT_ROOT / "workspace" / "outputs"
_TOPO_PATH = _PROJECT_ROOT / "knowledge" / "data" / "auto_env" / "network_topology_rag.md"

# 跨模块依赖检测关键词
_CROSS_MODULE_KW: dict[str, list[str]] = {
    "slb": ["slb vip", "slb virtual", "vip", "port-"],
    "ssl": ["ssl", "https"],
    "fw": ["fw", "acl"],
}


def _parse_topology() -> dict[str, Any]:
    """从 network_topology_rag.md 提取设备→IP 映射和网段信息."""
    text = _TOPO_PATH.read_text(encoding="utf-8") if _TOPO_PATH.exists() else ""
    device_map: dict[str, dict[str, Any]] = {}

    # 匹配表格行：设备名 | IP信息 | ...
    table_rows = re.findall(
        r'\|\s*(\w+(?:\d+)?)\s*\|(.+?)\|',
        text,
    )

    for name, rest in table_rows:
        name = name.strip()
        if name.lower() in ("设备名称", "vlan/子网"):
            continue
        ips_v4 = re.findall(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}(?:/\d{1,2})?)', rest)
        ips_v6 = re.findall(r'([0-9a-fA-F:]+(?::[0-9a-fA-F]+)*(?:/\d{1,3})?)', rest)
        # 过滤短的 v6 片段
        ips_v6 = [ip for ip in ips_v6 if ip.count(":") >= 2]
        if ips_v4 or ips_v6:
            device_map[name] = {
                "ipv4": [ip.split("/")[0] for ip in ips_v4],
                "ipv6": [ip.split("/")[0] for ip in ips_v6],
            }

    # 子网信息
    subnets = re.findall(
        r'\|\s*(\d{1,3}\.\d{1,3}\.\d{1,3}\.x)\s*\|(.+?)\|',
        text,
    )
    subnet_info = {s.strip(): desc.strip() for s, desc in subnets}

    return {"device_ip_map": device_map, "subnet_info": subnet_info}


def _read_xlsx_rows(xlsx_path: Path) -> dict[str, Any]:
    """读取 xlsx 的数据行，返回 rows_map + header_row + base_config_row."""
    import openpyxl

    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active

    # 找表头行
    header_row = None
    for r in range(1, min((ws.max_row or 0), 5) + 1):
        a = str(ws.cell(r, 1).value or "").lower()
        if "autoid" in a:
            header_row = r
            break

    if header_row is None:
        return {"error": "header row not found", "rows_map": {}, "header_row": 0, "base_config_row": 0}

    rows_map: dict[str, dict[str, str]] = {}
    base_config_row = 0
    e_column_types: dict[str, str] = {}

    for r in range(header_row + 1, (ws.max_row or 0) + 1):
        c_val = str(ws.cell(r, 3).value or "").strip()
        if c_val == "0":
            continue
        d = str(ws.cell(r, 4).value or "").strip()
        e = str(ws.cell(r, 5).value or "").strip()
        f = str(ws.cell(r, 6).value or "").strip()
        g = str(ws.cell(r, 7).value or "").strip()

        if not d and not e and not f:
            continue
        if not e and not f:
            continue

        rows_map[str(r)] = {"D": d, "E": e, "F": f, "G": g}
        e_column_types[str(r)] = e

        if base_config_row == 0 and c_val == "1" and "APV" in e and "cmds_config" in f:
            base_config_row = r

    return {
        "header_row": header_row,
        "rows_map": rows_map,
        "base_config_row": base_config_row,
        "e_column_types": e_column_types,
    }


def _extract_module_keywords(filename: str) -> str:
    """从 xlsx 文件名提取模块关键词."""
    stem = Path(filename).stem
    # 去掉 _test_cases 后缀
    stem = stem.replace("_test_cases", "")
    return stem


def _detect_cross_module_deps(rows_map: dict[str, dict[str, str]], module_keywords: str) -> list[str]:
    """检测跨模块依赖."""
    deps: list[str] = []
    module_lower = module_keywords.lower()

    for row_str, row in rows_map.items():
        d = row.get("D", "").lower()
        for mod, keywords in _CROSS_MODULE_KW.items():
            if mod not in module_lower:
                if any(kw in d for kw in keywords):
                    deps.append(f"Row {row_str}: 跨模块依赖 {mod} ({', '.join(k for k in keywords if k in d)})")

    return deps


def _build_brief(xlsx_path: Path, rows_info: dict[str, Any], topo: dict[str, Any]) -> str:
    """构建传给 g-column-filler 的 structured brief."""
    filename = xlsx_path.name
    module_kw = _extract_module_keywords(filename)
    cross_deps = _detect_cross_module_deps(rows_info["rows_map"], module_kw)

    brief_parts = [
        f"xlsx_path: {xlsx_path}",
        f"base_config_row: {rows_info.get('base_config_row', 0)}",
        f"module_keywords: {module_kw}",
        f"cross_module_deps: {json.dumps(cross_deps, ensure_ascii=False)}",
        f"device_ip_map: {json.dumps(topo.get('device_ip_map', {}), ensure_ascii=False, indent=2)}",
        f"subnet_info: {json.dumps(topo.get('subnet_info', {}), ensure_ascii=False, indent=2)}",
        f"rows_map: {json.dumps(rows_info.get('rows_map', {}), ensure_ascii=False, indent=2)}",
        f"e_column_types: {json.dumps(rows_info.get('e_column_types', {}), ensure_ascii=False, indent=2)}",
    ]
    return "\n".join(brief_parts)


def _parse_g_updates(fork_output: str) -> dict[str, Any]:
    """从 fork skill 输出中提取 g_updates JSON."""
    # 尝试找 ```json ... ``` 代码块
    json_block = re.search(r'```json\s*(\{.*?\})\s*```', fork_output, re.DOTALL)
    if json_block:
        try:
            return json.loads(json_block.group(1))
        except json.JSONDecodeError:
            pass

    # 尝试找裸 JSON 对象
    json_match = re.search(r'\{[^{}]*"g_updates"\s*:\s*\{[^}]*\}', fork_output, re.DOTALL)
    if json_match:
        try:
            return json.loads(json_match.group())
        except json.JSONDecodeError:
            pass

    # 更宽松的匹配
    json_match = re.search(r'\{[\s\S]*"g_updates"[\s\S]*\}', fork_output, re.DOTALL)
    if json_match:
        try:
            return json.loads(json_match.group())
        except json.JSONDecodeError:
            pass

    return {"g_updates": {}, "unfilled": [{"row": "N/A", "reason": "failed to parse fork output"}]}


def _write_g_column(xlsx_path: Path, g_updates: dict[str, str]) -> Path:
    """将 G 列内容写入 xlsx，输出 filled_ 文件."""
    import openpyxl

    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active

    written = 0
    for row_str, value in g_updates.items():
        try:
            row_num = int(row_str)
        except ValueError:
            continue
        cell = ws.cell(row=row_num, column=7)
        cell.value = value
        written += 1

    out_dir = _OUTPUT_BASE
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"filled_{xlsx_path.name}"
    wb.save(str(out_path))

    return out_path


@tool
def qa_fill_g_column(decomposed_json_path: str, project: str = "") -> str:
    """Generate xlsx from decomposed JSON and fill G column via LLM-driven fork skill.

    Pipeline:
    1. Generate standard automation xlsx per group
    2. Read each xlsx's D/E/F columns + network topology
    3. Invoke fork skill ``g-column-filler`` to generate G column content
       (LLM sub-agent queries CLI docs and produces precise commands)
    4. Write filled xlsx to ``workspace/outputs/filled_*.xlsx``

    Args:
        decomposed_json_path: Path to the decomposed JSON file
            (e.g. "workspace/outputs/yzg/yzg_decomposed.json").
        project: Optional project name. Defaults to parent folder of JSON file.

    Returns:
        JSON string with status, filled files list, and per-file fill summary.
    """
    # Resolve input path
    p = Path(decomposed_json_path)
    candidates = [p, _PROJECT_ROOT / decomposed_json_path]
    resolved = None
    for c in candidates:
        if c.exists():
            resolved = c
            break
    if resolved is None:
        return json.dumps({
            "status": "error",
            "error": f"File not found: {decomposed_json_path}",
        }, indent=2, ensure_ascii=False)

    try:
        data = json.loads(resolved.read_text(encoding="utf-8"))
    except Exception as exc:
        return json.dumps({
            "status": "error",
            "error": f"Failed to parse JSON: {exc}",
        }, indent=2, ensure_ascii=False)

    if data.get("status") != "success":
        return json.dumps({
            "status": "error",
            "error": f"Input is not a valid decomposed JSON (status={data.get('status')})",
        }, indent=2, ensure_ascii=False)

    cases = data.get("cases", [])
    if not cases:
        return json.dumps({
            "status": "error",
            "error": "No cases found in decomposed JSON",
        }, indent=2, ensure_ascii=False)

    # Step 1: Generate xlsx files (same logic as qa_generate_test_case_xlsx)
    from main.ist_core.tools.skills.test_case_xlsx_generator import _build_xlsx_for_group

    groups: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for case in cases:
        g = case.get("group", "default")
        if g not in groups:
            groups[g] = []
        groups[g].append(case)

    proj = project or resolved.parent.name
    output_dir = _OUTPUT_BASE / proj

    # Generate all xlsx files first
    xlsx_paths: dict[str, Path] = {}
    for group_name, group_cases in groups.items():
        out_path = _build_xlsx_for_group(group_name, group_cases, output_dir)
        xlsx_paths[group_name] = out_path

    # Step 2: Read network topology (once, shared across all xlsx)
    topo = _parse_topology()

    # Load env for LLM access (needed by fork skill subagent)
    from main.langchain_env import langchain_load_dotenv_if_present
    langchain_load_dotenv_if_present()

    from main.ist_core.skills.loader import clear_subagent_cache, execute_fork_skill
    clear_subagent_cache()

    # Step 3-5: For each xlsx, read rows, invoke fork skill, write G column
    results: list[dict[str, Any]] = []
    total_filled = 0
    total_skipped = 0
    total_unfilled = 0

    for group_name, xlsx_path in xlsx_paths.items():
        rows_info = _read_xlsx_rows(xlsx_path)
        if "error" in rows_info:
            results.append({
                "group": group_name,
                "xlsx": str(xlsx_path.relative_to(_PROJECT_ROOT)),
                "status": "error",
                "error": rows_info["error"],
            })
            continue

        if not rows_info["rows_map"]:
            results.append({
                "group": group_name,
                "xlsx": str(xlsx_path.relative_to(_PROJECT_ROOT)),
                "status": "skipped",
                "reason": "no data rows found",
            })
            continue

        # Build brief and invoke fork skill
        brief = _build_brief(xlsx_path, rows_info, topo)

        try:
            fork_output = execute_fork_skill("g-column-filler", brief)
        except Exception as exc:
            results.append({
                "group": group_name,
                "xlsx": str(xlsx_path.relative_to(_PROJECT_ROOT)),
                "status": "error",
                "error": f"fork skill failed: {exc}",
            })
            continue

        # Parse g_updates from fork output
        parsed = _parse_g_updates(fork_output)
        g_updates = parsed.get("g_updates", {})
        unfilled = parsed.get("unfilled", [])

        if not g_updates:
            results.append({
                "group": group_name,
                "xlsx": str(xlsx_path.relative_to(_PROJECT_ROOT)),
                "status": "error",
                "error": "no g_updates found in fork output",
                "fork_output_snippet": fork_output[:500],
            })
            continue

        # Write G column
        try:
            filled_path = _write_g_column(xlsx_path, g_updates)
        except Exception as exc:
            results.append({
                "group": group_name,
                "xlsx": str(xlsx_path.relative_to(_PROJECT_ROOT)),
                "status": "error",
                "error": f"failed to write filled xlsx: {exc}",
            })
            continue

        row_count = len(rows_info["rows_map"])
        results.append({
            "group": group_name,
            "xlsx": str(xlsx_path.relative_to(_PROJECT_ROOT)),
            "filled_xlsx": str(filled_path.relative_to(_PROJECT_ROOT)),
            "status": "success",
            "total_rows": row_count,
            "g_filled": len(g_updates),
            "g_unfilled": len(unfilled),
            "unfilled": unfilled,
        })
        total_filled += len(g_updates)
        total_skipped += row_count - len(g_updates) - len(unfilled)
        total_unfilled += len(unfilled)

    return json.dumps({
        "status": "success",
        "project": proj,
        "total_groups": len(xlsx_paths),
        "total_filled_rows": total_filled,
        "total_skipped_rows": total_skipped,
        "total_unfilled_rows": total_unfilled,
        "files": results,
    }, indent=2, ensure_ascii=False)
