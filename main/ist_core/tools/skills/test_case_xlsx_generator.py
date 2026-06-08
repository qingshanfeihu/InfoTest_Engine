"""qa_generate_test_case_xlsx: decomposed JSON → 标准自动化 xlsx.

将 qa_decompose_test_cases 产出的 decomposed JSON 转为自动化执行引擎
可识别的 xlsx 文件。纯机械操作，不调 LLM。

规则：
- 按 group 拆分，一个 group → 一个 xlsx（同名 group 合并在一个文件）
- 列映射严格遵循标准模板：
  A=autoid, B=优先级, C=步骤号, D=步骤描述, E=操作对象(actor),
  F=操作方法(action), G=数据(data/hint)
- C=1 为文件级共享前置（A/B 列为空）：
  整个 xlsx 只有一行 C=1，不属于任何 case。
  每个 case 执行前都会先跑 C=1 前置，case 执行完后配置自动清除，
  下个 case 再重新跑 C=1。
- 同一个 case 的多行仅首行填 autoid + 优先级
- case 之间空一行分隔
"""

from __future__ import annotations

import json
from collections import OrderedDict
from pathlib import Path
from typing import Any

from langchain_core.tools import tool

_PROJECT_ROOT = Path(__file__).resolve().parents[4]
_OUTPUT_BASE = _PROJECT_ROOT / "workspace" / "outputs"

# 标准列宽（字符宽度）
_COL_WIDTHS: dict[str, int] = {
    "A": 22,  # autoid
    "B": 8,   # 优先级
    "C": 6,   # 步骤
    "D": 60,  # 步骤描述
    "E": 14,  # 操作对象
    "F": 14,  # 操作方法
    "G": 70,  # 数据
    "H": 20,  # 临时保存期望结果
    "I": 20,  # 输入变量
}


def _resolve_g_value(step: dict[str, Any]) -> str:
    """从 step 中提取 G 列内容：优先 data → hint → infer_from → 空。"""
    if step.get("data"):
        return str(step["data"])
    if step.get("hint"):
        return str(step["hint"])
    infer = step.get("infer_from")
    if isinstance(infer, dict):
        return infer.get("expected_meaning", "") or infer.get("concrete_example", "") or ""
    if infer:
        return str(infer)
    return ""


def _build_xlsx_for_group(
    group_name: str,
    cases: list[dict[str, Any]],
    output_dir: Path,
) -> Path:
    """为一个 group 的 cases 生成一个 xlsx，返回输出路径。"""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    safe_name = group_name.replace("/", "_").replace("\\", "_").replace(":", "_")
    ws.title = safe_name[:31]  # sheet name max 31 chars

    # ── 表头（第 1 行）──
    headers = ["autoid", "优先级", "步骤", "步骤描述", "操作对象", "操作方法", "数据", "临时保存期望结果", "输入变量"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # ── 列宽 ──
    for col_letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── 数据行（从第 2 行开始）──
    row = 2
    wrap_align = Alignment(vertical="top", wrap_text=True)

    # ── C=1 共享前置 — 整个 xlsx 文件的公共初始化步骤 ──
    # 从第一个 case 中提取 C=1 步骤，作为文件级前置（A/B 列为空）。
    # 每个 case 执行前都会先跑 C=1，case 执行完后配置清除，下个 case 再重新跑。
    if cases:
        first_steps = cases[0].get("steps", [])
        shared_init_steps = [s for s in first_steps if str(s.get("c", "")) == "1"]
        for init_step in shared_init_steps:
            c_val = init_step.get("c", "")
            actor = init_step.get("actor", "")
            action = init_step.get("action", "")
            describe = init_step.get("describe", "")
            values = ["", "", c_val, describe, actor, action, "", "", ""]
            for ci, v in enumerate(values, 1):
                cell = ws.cell(row, ci, v)
                cell.alignment = wrap_align
                cell.border = thin_border
            row += 1

    for case in cases:
        autoid = str(case.get("autoid", ""))
        priority = str(case.get("priority", ""))
        steps = case.get("steps", [])

        # 跳过 C=1 共享前置（已在上面单独写入）
        case_steps = [s for s in steps if str(s.get("c", "")) != "1"]
        if not case_steps:
            continue

        for si, step in enumerate(case_steps):
            c_val = step.get("c", "")
            actor = step.get("actor", "")
            action = step.get("action", "")
            describe = step.get("describe", "")

            # 仅每个 case 的首行填 autoid + 优先级
            a_val = autoid if si == 0 else ""
            p_val = priority if si == 0 else ""

            # G 列留空，由 g-column-filler fork skill 后续填充
            values = [a_val, p_val, c_val, describe, actor, action, "", "", ""]
            for ci, v in enumerate(values, 1):
                cell = ws.cell(row, ci, v)
                cell.alignment = wrap_align
                cell.border = thin_border
            row += 1

        # case 之间空一行
        row += 1

    # ── 冻结表头 ──
    ws.freeze_panes = "A2"

    # ── 保存 ──
    out_path = output_dir / f"{safe_name}_test_cases.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


@tool
def qa_generate_test_case_xlsx(decomposed_json_path: str, project: str = "") -> str:
    """Generate standard automation xlsx from a decomposed test case JSON.

    Reads the output of qa_decompose_test_cases, groups cases by their
    ``group`` field, and produces one xlsx per group.  Column layout follows
    the automation-engine standard:

        A=autoid  B=优先级  C=步骤  D=步骤描述  E=操作对象(actor)  F=操作方法(action)  G=数据

    C 列约定：
        C=1 — 文件级共享前置，A/B 列为空。在执行每个 case 之前自动运行，
               case 结束后配置清除，下一个 case 再重新执行。
        C>=2 — 普通测试步骤。每个 case 的 A 列首行填 autoid，B 列填优先级。

    Args:
        decomposed_json_path: Path to the decomposed JSON file
            (e.g. "workspace/inputs/yzg/zhaiyq_decomposed.json").
        project: Optional project name for the output sub-directory.
            Defaults to the parent folder name of the JSON file.

    Returns:
        JSON string with status, project, output_dir, and a list of
        generated file paths (one per group).
    """
    p = Path(decomposed_json_path)
    candidates = [p, _PROJECT_ROOT / decomposed_json_path]
    resolved = None
    for c in candidates:
        if c.exists():
            resolved = c
            break
    if resolved is None:
        return json.dumps({
            "status": "error",
            "error": f"File not found: {decomposed_json_path}",
        }, indent=2, ensure_ascii=False)

    try:
        data = json.loads(resolved.read_text(encoding="utf-8"))
    except Exception as exc:
        return json.dumps({
            "status": "error",
            "error": f"Failed to parse JSON: {exc}",
        }, indent=2, ensure_ascii=False)

    if data.get("status") != "success":
        return json.dumps({
            "status": "error",
            "error": f"Input is not a valid decomposed JSON (status={data.get('status')})",
        }, indent=2, ensure_ascii=False)

    cases = data.get("cases", [])
    if not cases:
        return json.dumps({
            "status": "error",
            "error": "No cases found in decomposed JSON",
        }, indent=2, ensure_ascii=False)

    # 按 group 分组（保持出现顺序）
    groups: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
    for case in cases:
        g = case.get("group", "default")
        if g not in groups:
            groups[g] = []
        groups[g].append(case)

    # 确定输出目录
    proj = project or resolved.parent.name
    output_dir = _OUTPUT_BASE / proj

    # 每个 group 生成一个 xlsx
    generated: list[str] = []
    for group_name, group_cases in groups.items():
        out_path = _build_xlsx_for_group(group_name, group_cases, output_dir)
        generated.append(str(out_path.relative_to(_PROJECT_ROOT)))

    return json.dumps({
        "status": "success",
        "project": proj,
        "output_dir": str(output_dir.relative_to(_PROJECT_ROOT)),
        "total_cases": len(cases),
        "total_groups": len(groups),
        "groups": list(groups.keys()),
        "files": generated,
    }, indent=2, ensure_ascii=False)
