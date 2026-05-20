"""orgin 文档分类器 v2 —— LLM 判定 + 文件级缓存 + 用户覆盖层。

把 orgin/ 下文件分到三类管线：

- ``product``       → mineru 链（spec / cli.pdf / app.pdf / design / pptx / 其他产品文档）
- ``test_case_list``→ qa 链（"Test List ..." xlsx / 测试用例 / Bug→功能 映射）
- ``test_strategy`` → qa 链（"Test Strategy ..."）

设计要点（来自 ``memory/feedback_kms_asset_classifier.md``）：

1. **不靠文件名硬编码正则**。LLM 直接判，给每个文件返回 ``category + confidence + reason``。
2. **文件级缓存**：按 ``(filename, mtime, size)`` 复用结果，没改的文件不再调 LLM。
   缓存落 ``knowledge/.intermediate/.classifier_cache.json``。
3. **用户覆盖层**：``knowledge/.classifier_overrides.json`` 是用户级白名单，
   有 entry 直接用，跳过 LLM 和 cache。
4. **xlsx 提示**：用 openpyxl 读 sheet 名 + 第一行表头，喂给 LLM 提高判得准。
   doc/docx/pdf 只用文件名 + 扩展名 + 字节大小（足够区分测试用例 vs 产品规格）。
5. **fallback**：LLM 不可达时返回 ``category="unclassified"``，让 ``/kms`` 命令显式报警，
   而不是默认归 product 污染 mineru 链。
"""

from __future__ import annotations

import json
import logging
import os
from pathlib import Path
from typing import Literal, TypedDict

logger = logging.getLogger(__name__)

OrginCategory = Literal["product", "test_case_list", "test_strategy", "unclassified"]


class ClassifierResult(TypedDict):
    category: OrginCategory
    confidence: float          # 0.0-1.0；用户 override 视为 1.0；LLM 失败 0.0
    reason: str                # 单行人类可读理由
    source: Literal["llm", "cache", "override", "fallback"]


_VALID_CATEGORIES = {"product", "test_case_list", "test_strategy", "unclassified"}

_SYSTEM_PROMPT = """\
你是测试评审平台 IST-Core 的知识库分类器。任务：把一个文件归到下面三类之一：

- product        : 产品规格 / 设计 / 用户手册 / CLI/APP 参考 / 架构白皮书 / 协议方案。
                  特征：spec / 设计文档 / 需求 / 概要设计 / API 手册 / 协议规范 /
                  PRD / 白皮书 / 增强要求 / 配置指南 / pptx 方法说明 / cli.pdf / app.pdf。
- test_case_list : 测试用例清单 / 测试范围列表 / Bug→Function 映射表 / 用例 xlsx。
                  特征：Test List / 测试用例 / Test Scope / Priority 列表 /
                  BugID-功能 映射 / xlsx 表头出现 "用例ID/Pre-condition/预期结果"。
- test_strategy  : 测试策略文档（不是用例本身，而是怎么测的方针）。
                  特征：Test Strategy / 测试策略 / 测试方针。

只输出 JSON，schema：

  {
    "category": "product" | "test_case_list" | "test_strategy" | "unclassified",
    "confidence": 0.0-1.0 的浮点数，
    "reason": "一行简短理由（≤80 字符），用中文。"
  }

不要解释，不要 markdown，不要前后文，只返回一个 JSON 对象。
"""


# ---------------------------------------------------------------------------
# cache + overrides
# ---------------------------------------------------------------------------


def _cache_path() -> Path:
    from main import knowledge_paths as kp
    kp.ensure_intermediate_dirs()
    return kp.KNOWLEDGE_INTERMEDIATE / ".classifier_cache.json"


def _overrides_path() -> Path:
    from main import knowledge_paths as kp
    return kp.KNOWLEDGE_ROOT / ".classifier_overrides.json"


def _load_json(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001
        logger.warning("classifier load %s failed: %s", path, exc)
        return {}


def _save_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _file_signature(file_path: Path) -> tuple[float, int]:
    st = file_path.stat()
    return (round(st.st_mtime, 3), st.st_size)


# ---------------------------------------------------------------------------
# xlsx hint extractor
# ---------------------------------------------------------------------------


def _extract_xlsx_hints(file_path: Path) -> str:
    """读 xlsx 的 sheet 名 + 每个 sheet 前 2 行内容，最多前 3 个 sheet。

    返回紧凑文本，供 LLM 当上下文。失败时返回空串（让 LLM 只用文件名判）。
    """
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError:
        return ""
    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return f"(open failed: {exc})"
    sheets = wb.sheetnames[:3]
    parts: list[str] = [f"sheets({len(wb.sheetnames)}): {', '.join(wb.sheetnames[:5])}"]
    for sn in sheets:
        try:
            ws = wb[sn]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 2:
                    break
                cells = ["" if c is None else str(c)[:40] for c in row[:10]]
                rows.append(" | ".join(cells))
            parts.append(f"[{sn}] " + " || ".join(rows))
        except Exception as exc:  # noqa: BLE001
            parts.append(f"[{sn}] (read failed: {exc})")
    try:
        wb.close()
    except Exception:  # noqa: BLE001
        pass
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# LLM call
# ---------------------------------------------------------------------------


def _call_llm(filename: str, hints: str) -> ClassifierResult:
    api_key = (os.environ.get("DASHSCOPE_API_KEY")
               or os.environ.get("BAILIAN_API_KEY")
               or "").strip()
    if not api_key:
        return {
            "category": "unclassified",
            "confidence": 0.0,
            "reason": "DASHSCOPE_API_KEY missing",
            "source": "fallback",
        }

    try:
        import requests  # noqa: PLC0415
        from main.function_llm import chat_completion  # type: ignore
    except Exception as exc:  # noqa: BLE001
        return {
            "category": "unclassified",
            "confidence": 0.0,
            "reason": f"llm import failed: {exc}",
            "source": "fallback",
        }

    user_prompt = (
        f"文件名: {filename}\n"
        f"扩展名: {Path(filename).suffix.lower()}\n"
    )
    if hints:
        user_prompt += f"\nxlsx 提示:\n{hints}\n"

    try:
        with requests.Session() as session:
            data = chat_completion(
                session, api_key,
                system_prompt=_SYSTEM_PROMPT,
                user_prompt=user_prompt,
                max_tokens=256,
                temperature=0.0,
                top_p=0.1,
                max_retries=2,
            )
    except Exception as exc:  # noqa: BLE001
        logger.warning("classifier LLM call failed for %s: %s", filename, exc)
        return {
            "category": "unclassified",
            "confidence": 0.0,
            "reason": f"llm call failed: {type(exc).__name__}",
            "source": "fallback",
        }

    cat = str(data.get("category", "")).strip().lower()
    if cat not in _VALID_CATEGORIES:
        return {
            "category": "unclassified",
            "confidence": 0.0,
            "reason": f"llm returned invalid category: {cat!r}",
            "source": "fallback",
        }
    try:
        conf = float(data.get("confidence", 0.0))
    except Exception:  # noqa: BLE001
        conf = 0.0
    reason = str(data.get("reason", "")).strip()[:120]
    return {
        "category": cat,  # type: ignore[typeddict-item]
        "confidence": conf,
        "reason": reason or "(no reason)",
        "source": "llm",
    }


# ---------------------------------------------------------------------------
# public API
# ---------------------------------------------------------------------------


def classify_file(file_path: Path | str) -> ClassifierResult:
    """对单个文件做分类。三层兜底：override → cache → LLM。"""
    p = Path(file_path)

    overrides = _load_json(_overrides_path())
    if p.name in overrides:
        ov = overrides[p.name]
        cat = str(ov.get("category", "")).strip().lower()
        if cat in _VALID_CATEGORIES:
            return {
                "category": cat,  # type: ignore[typeddict-item]
                "confidence": 1.0,
                "reason": str(ov.get("reason", "user override"))[:120],
                "source": "override",
            }

    if not p.exists():
        return {"category": "unclassified", "confidence": 0.0,
                "reason": "file not found", "source": "fallback"}

    sig = _file_signature(p)
    cache = _load_json(_cache_path())
    cached = cache.get(p.name)
    if cached and tuple(cached.get("sig", [])) == sig:
        return {
            "category": cached["category"],
            "confidence": float(cached.get("confidence", 0.0)),
            "reason": cached.get("reason", ""),
            "source": "cache",
        }

    hints = ""
    if p.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        hints = _extract_xlsx_hints(p)

    result = _call_llm(p.name, hints)
    if result["source"] == "llm":
        cache[p.name] = {
            "sig": list(sig),
            "category": result["category"],
            "confidence": result["confidence"],
            "reason": result["reason"],
        }
        try:
            _save_json(_cache_path(), cache)
        except Exception as exc:  # noqa: BLE001
            logger.warning("save classifier cache failed: %s", exc)
    return result


def bucketize_orgin_dir(orgin_dir: Path | str) -> dict[OrginCategory, list[str]]:
    """对 orgin_dir 下所有文件分桶。返回 ``{category: [filename, ...]}``。"""
    root = Path(orgin_dir)
    buckets: dict[OrginCategory, list[str]] = {
        "product": [],
        "test_case_list": [],
        "test_strategy": [],
        "unclassified": [],
    }
    if not root.exists():
        return buckets
    for p in sorted(root.iterdir()):
        if not p.is_file():
            continue
        r = classify_file(p)
        buckets[r["category"]].append(p.name)
    return buckets


def classify_orgin_file(filename: str) -> OrginCategory:
    """旧 API 兼容：只返回 category 字符串。

    如果 ``filename`` 是绝对路径或可在 ``orgin/`` 下找到，做完整分类；
    否则只跑 LLM 仅基于文件名。
    """
    from main import knowledge_paths as kp
    p = Path(filename)
    candidate = p if p.is_absolute() else (kp.KNOWLEDGE_ORGIN / filename)
    if candidate.exists():
        return classify_file(candidate)["category"]
    return _call_llm(p.name, "")["category"]


def list_orgin_with_reasons(orgin_dir: Path | str) -> list[dict]:
    """对 orgin_dir 下每个文件返回 ``{name, category, confidence, reason, source}``。

    给 ``/kms product status`` / ``/kms qa status`` 显示理由用。
    """
    root = Path(orgin_dir)
    rows: list[dict] = []
    if not root.exists():
        return rows
    for p in sorted(root.iterdir()):
        if not p.is_file():
            continue
        r = classify_file(p)
        rows.append({"name": p.name, **r})
    return rows
