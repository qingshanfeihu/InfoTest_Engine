"""Read-only generic filesystem tools for the phase-one QA agent surface."""

from __future__ import annotations

import html
import os
import re
import tempfile
import zipfile
from pathlib import Path
from typing import Iterable

from langchain_core.tools import tool

from main.qa_agent.tools.deepagent._range_reader import read_text_range
from main.qa_agent.tools.deepagent._rg import RipgrepResult, run_ripgrep


_PROJECT_ROOT = Path(__file__).resolve().parents[4]
# ⚠ 安全边界：IST-Core agent 的文件视野白名单根。
# 与 CLAUDE.md "agent 视野：agent 只看 knowledge/data/" 对齐。
# 修改此值即修改 agent 沙箱范围，需经安全评审。
_AGENT_ROOT = (_PROJECT_ROOT / "knowledge" / "data").resolve()
_WORKSPACE_ROOT = (_PROJECT_ROOT / "workspace").resolve()


def _agent_roots() -> tuple[Path, ...]:
    """返回 agent 可访问的根目录列表（按优先级排序）。

    优先级：knowledge/data > workspace > IST_SESSION_DIR > IST_USER_DIR。
    不存在的目录自动跳过，保证向后兼容。
    """
    roots: list[Path] = [_AGENT_ROOT]
    if _WORKSPACE_ROOT.is_dir():
        roots.append(_WORKSPACE_ROOT)
    session = os.environ.get("IST_SESSION_DIR")
    if session:
        p = Path(session).resolve()
        if p.is_dir():
            roots.append(p)
    user_dir = os.environ.get("IST_USER_DIR")
    if user_dir:
        p = Path(user_dir).resolve()
        if p.is_dir():
            roots.append(p)
    return tuple(roots)

# Defense-in-depth：仓库根下平台自身资产；即使白名单失守也必须挡。
_PLATFORM_DENIED_TOP_LEVEL = {
    # 平台代码
    "main",
    "tests",
    "scripts",
    "agent-chat-ui",
    # 虚拟环境与三方包源码
    ".venv",
    ".venv311",
    # VCS / 运行时状态 / 缓存
    ".git",
    ".langgraph_api",
    ".langgraph_data",
    ".pytest_cache",
    "memory",
    # 运行时产物（统一收纳到 runtime/）
    "runtime",
    # 历史归档
    "backup",
    "conversation_history",
    "large_tool_results",
    "postgres_storage",
    "qdrant_storage",
    # 凭据
    ".env",
    "environment",
    # IDE / 工程元
    ".claude",
    ".github",
    ".vscode",
    ".idea",
}
# 仓库根下"平台元信息"文件（用户确认这些也算平台东西，不暴露给 agent）。
_PLATFORM_DENIED_FILES = {
    "CLAUDE.md",
    "todolist.md",
    "ARCHITECTURE.md",
    "README.md",
    "README_zh.md",
    "requirements.txt",
    "pyproject.toml",
    "package.json",
    "package-lock.json",
    "environment.example",
}
# rg defense-in-depth：即便 rg 被锁在 knowledge/data/ 子树下，也额外排除内部产物路径。
_RG_DEFENSE_PREFIXES = (
    "knowledge/.intermediate",
    "knowledge/.cache",
    "knowledge/.cache.json",
    "knowledge/.index_cache.json",
    "knowledge/.schema_gaps.jsonl",
    "knowledge/_ingest_quality_report.json",
)
_TEXT_SUFFIXES = {
    ".cfg",
    ".css",
    ".csv",
    ".html",
    ".ini",
    ".json",
    ".jsonc",
    ".log",
    ".md",
    ".py",
    ".rst",
    ".sh",
    ".text",
    ".toml",
    ".ts",
    ".tsx",
    ".txt",
    ".yaml",
    ".yml",
}
_SPREADSHEET_SUFFIXES = {".xlsx", ".xlsm"}
_DOCX_SUFFIXES = {".docx"}
_VCS_DIRECTORIES_TO_EXCLUDE = {".git", ".svn", ".hg", ".bzr", ".jj", ".sl"}
_RG_MAX_COLUMNS = 500
_TYPE_SUFFIXES = {
    "css": {".css"},
    "go": {".go"},
    "html": {".html", ".htm"},
    "java": {".java"},
    "js": {".js", ".jsx", ".mjs", ".cjs"},
    "json": {".json", ".jsonc"},
    "md": {".md", ".markdown"},
    "py": {".py"},
    "python": {".py"},
    "rs": {".rs"},
    "rust": {".rs"},
    "sh": {".sh", ".bash", ".zsh"},
    "toml": {".toml"},
    "ts": {".ts", ".tsx"},
    "tsx": {".tsx"},
    "txt": {".txt", ".text"},
    "yaml": {".yaml", ".yml"},
    "yml": {".yaml", ".yml"},
}


def _project_rel(path: Path) -> str:
    try:
        return path.resolve().relative_to(_PROJECT_ROOT).as_posix()
    except ValueError:
        return path.as_posix()


def _resolve_inside_root(raw_path: str | None, *, must_exist: bool = False) -> Path:
    """Resolve a user-supplied path against the agent sandbox.

    Three gates (defense-in-depth):
      1. Traversal: reject ``..`` and ``~`` components outright.
      2. Platform deny-list: reject any repo-root subpath that lives in
         ``_PLATFORM_DENIED_TOP_LEVEL`` or is a ``_PLATFORM_DENIED_FILES`` entry.
      3. Sandbox white-list: resolved path must live under one of
         ``_agent_roots()`` (knowledge/data, workspace, session, user).
    """
    text = (raw_path or ".").strip() or "."

    # Gate 1: refuse traversal tokens explicitly so the agent gets a clear error.
    parts = Path(text).parts
    if ".." in parts or text.startswith("~") or "~" in parts:
        raise PermissionError(
            "path traversal not allowed; agent sandbox is rooted at knowledge/data/"
        )

    path = Path(text)
    if path.is_absolute():
        resolved = path.resolve()
    else:
        # Try each root in priority order; first existing match wins.
        # Also accept "knowledge/data/..." and "workspace/..." prefixes
        # resolved under _PROJECT_ROOT for legacy compatibility.
        resolved = None
        for root in _agent_roots():
            candidate = (root / path).resolve()
            if candidate.exists():
                resolved = candidate
                break
        if resolved is None:
            # Fallback: try under _PROJECT_ROOT (legacy "knowledge/data/x" form)
            project_candidate = (_PROJECT_ROOT / path).resolve()
            if project_candidate.exists():
                resolved = project_candidate
            else:
                # Default to first root for non-existent relative paths
                resolved = (_agent_roots()[0] / path).resolve()

    # Gate 2: platform deny-list.
    try:
        rel_to_project = resolved.relative_to(_PROJECT_ROOT)
    except ValueError:
        rel_to_project = None
    if rel_to_project is not None:
        rel_parts = rel_to_project.parts
        if rel_parts:
            top = rel_parts[0]
            if top in _PLATFORM_DENIED_TOP_LEVEL:
                raise PermissionError(
                    f"path is in platform-denied directory: {top}/"
                )
            if len(rel_parts) == 1 and top in _PLATFORM_DENIED_FILES:
                raise PermissionError(
                    f"path is a platform metadata file: {top}"
                )

    # Gate 3: white-list. Must live under one of _agent_roots().
    for root in _agent_roots():
        try:
            resolved.relative_to(root)
            break
        except ValueError:
            continue
    else:
        raise PermissionError(
            "path outside agent sandbox; agent can only read paths under "
            f"knowledge/data/ or workspace/. requested: {raw_path}"
        )

    if must_exist and not resolved.exists():
        raise FileNotFoundError(f"path not found: {raw_path}")
    return resolved


def _coerce_limit(value: int | None, default: int, *, upper: int = 1000) -> int:
    return max(1, min(int(value or default), upper))


def _coerce_offset(value: int | None) -> int:
    return max(0, int(value or 0))


def _normalise_pattern(pattern: str, base: Path) -> str:
    pattern = (pattern or "*").strip() or "*"
    if ".." in Path(pattern).parts or pattern.startswith("~"):
        raise PermissionError(
            "path traversal not allowed; agent sandbox is rooted at knowledge/data/"
        )
    pattern_path = Path(pattern)
    if pattern_path.is_absolute():
        try:
            return pattern_path.resolve().relative_to(base).as_posix()
        except ValueError:
            return pattern_path.name
    return pattern


def _expand_brace_patterns(pattern: str) -> list[str]:
    match = re.search(r"\{([^{}]+)\}", pattern)
    if not match:
        return [pattern]
    out: list[str] = []
    for part in match.group(1).split(","):
        out.extend(_expand_brace_patterns(pattern[: match.start()] + part + pattern[match.end() :]))
    return out


def _is_probably_binary(path: Path) -> bool:
    try:
        chunk = path.read_bytes()[:2048]
    except Exception:
        return True
    return b"\x00" in chunk


def _rg_target(path: Path) -> str:
    rel = _project_rel(path)
    return rel or "."


def _rg_exclusion_args() -> list[str]:
    args: list[str] = []
    # VCS dirs are universal noise — keep them excluded even though _AGENT_ROOT
    # is already below knowledge/data/ and shouldn't contain them in practice.
    for name in sorted(_VCS_DIRECTORIES_TO_EXCLUDE):
        args.extend(["--glob", f"!{name}"])
        args.extend(["--glob", f"!{name}/**"])
    for prefix in _RG_DEFENSE_PREFIXES:
        args.extend(["--glob", f"!{prefix}"])
        args.extend(["--glob", f"!{prefix}/**"])
    return args


def _resolve_rg_path(raw_path: str) -> Path | None:
    text = raw_path.strip()
    if not text:
        return None
    path = Path(text)
    if not path.is_absolute():
        path = _PROJECT_ROOT / path
    try:
        return _resolve_inside_root(path.as_posix(), must_exist=True)
    except Exception:
        return None


_CONTENT_PATH_RE = re.compile(r"^(?P<path>.*?)(?::(?P<line>\d+):|-(?P<context_line>\d+)-)")


def _content_line_path(raw_line: str) -> str | None:
    match = _CONTENT_PATH_RE.match(raw_line)
    if not match:
        return None
    return match.group("path")


def _filter_rg_file_lines(lines: list[str], *, output_mode: str) -> list[str]:
    out: list[str] = []
    for line in lines:
        if output_mode == "count":
            path_text, sep, count_text = line.rpartition(":")
            if not sep:
                continue
            resolved = _resolve_rg_path(path_text)
            if resolved is not None:
                out.append(f"{_project_rel(resolved)}:{count_text}")
            continue
        resolved = _resolve_rg_path(line)
        if resolved is not None:
            out.append(_project_rel(resolved))
    return out


def _filter_rg_content_lines(lines: list[str]) -> list[str]:
    out: list[str] = []
    for line in lines:
        if line == "--":
            if out and out[-1] != "--":
                out.append(line)
            continue
        path_text = _content_line_path(line)
        if path_text is None:
            out.append(line)
            continue
        resolved = _resolve_rg_path(path_text)
        if resolved is None:
            continue
        out.append(_project_rel(resolved) + line[len(path_text) :])
    return out


def _apply_window(lines: list[str], *, offset: int, head_limit: int | None) -> tuple[list[str], bool]:
    offset = _coerce_offset(offset)
    if head_limit == 0:
        return lines[offset:], False
    limit = _coerce_limit(head_limit, 250, upper=10000)
    end = offset + limit
    return lines[offset:end], len(lines) > end


def _rg_warning(result: RipgrepResult) -> str | None:
    if result.timed_out:
        return "... rg timed out; partial results shown. Narrow path/glob or reduce the query."
    if result.truncated:
        return "... rg output exceeded 20MB; partial results shown. Narrow path/glob or reduce the query."
    return None


def _format_page_from_range(lines: list[str], *, total: int, offset: int, limit: int) -> str:
    header = f"total_lines={total}, offset={offset}, returned={len(lines)}"
    if offset + limit < total:
        header += f", next_offset={offset + limit}"
    return header + "\n" + "\n".join(lines)


def _iter_candidate_files(base: Path, glob_pattern: str, *, max_files: int = 5000) -> Iterable[Path]:
    pattern = _normalise_pattern(glob_pattern, base)
    count = 0
    seen: set[Path] = set()
    for expanded_pattern in _expand_brace_patterns(pattern):
        for path in base.glob(expanded_pattern):
            try:
                resolved = _resolve_inside_root(path.as_posix(), must_exist=True)
            except Exception:
                continue
            if not resolved.is_file() or resolved in seen:
                continue
            seen.add(resolved)
            count += 1
            if count > max_files:
                return
            yield resolved


def _format_page(lines: list[str], *, offset: int, limit: int) -> str:
    offset = max(0, int(offset or 0))
    limit = max(1, min(int(limit or 200), 1000))
    total = len(lines)
    page = lines[offset : offset + limit]
    header = f"total_lines={total}, offset={offset}, returned={len(page)}"
    if offset + limit < total:
        header += f", next_offset={offset + limit}"
    return header + "\n" + "\n".join(page)


def _read_text_file(path: Path, *, offset: int, limit: int) -> str:
    page = read_text_range(path, offset=offset, limit=limit)
    return _format_page_from_range(page.lines, total=page.total_lines, offset=page.offset, limit=page.limit)


def _clean_xml_text(raw: str) -> str:
    text = re.sub(r"<[^>]+>", " ", raw)
    text = html.unescape(text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _read_docx(path: Path, *, offset: int, limit: int) -> str:
    try:
        with zipfile.ZipFile(path) as zf:
            names = [n for n in zf.namelist() if n.startswith("word/") and n.endswith(".xml")]
            parts: list[str] = []
            for name in names:
                if name != "word/document.xml" and not name.startswith("word/header") and not name.startswith("word/footer"):
                    continue
                raw = zf.read(name).decode("utf-8", errors="replace")
                cleaned = _clean_xml_text(raw)
                if cleaned:
                    parts.append(f"=== {name} ===\n{cleaned}")
    except Exception as exc:  # noqa: BLE001
        return f"error: unable to read docx as generic document: {exc}"
    lines = "\n\n".join(parts).splitlines()
    numbered = [f"{i + 1}: {line}" for i, line in enumerate(lines)]
    return _format_page(numbered, offset=offset, limit=limit)


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", text)


def _read_spreadsheet(path: Path, *, offset: int, limit: int) -> str:
    try:
        import openpyxl  # type: ignore[import-not-found]
    except Exception as exc:  # noqa: BLE001
        return f"error: openpyxl is required to read spreadsheet files: {exc}"

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        return f"error: unable to read spreadsheet: {exc}"

    lines: list[str] = [f"workbook={_project_rel(path)} sheets={len(wb.sheetnames)}"]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(
            f"=== Sheet: {sheet_name} dims={ws.calculate_dimension()} "
            f"max_row={ws.max_row} max_col={ws.max_column} ==="
        )
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [_cell_text(v) for v in row]
            if not any(values):
                continue
            trimmed = values[:20]
            suffix = " || ..." if len(values) > len(trimmed) else ""
            lines.append(f"{sheet_name}!{row_idx}: " + " || ".join(trimmed) + suffix)
    return _format_page(lines, offset=offset, limit=limit)


@tool(parse_docstring=True)
def qa_deepagent_ls(path: str = ".", max_entries: int = 200) -> str:
    """List files and directories under the project root.

    This is a generic DeepAgents-style ``ls`` tool for phase-one static review.
    It is read-only, concurrency-safe, and does not run project code. Use it
    to inspect the repository layout or a known evidence directory before
    narrowing to ``qa_deepagent_read_file`` / ``qa_deepagent_grep``.

    Boundaries:
    - Generic read-only project exploration tool.
    - Does not write files, start processes, access vector stores, or call
      language models.
    - Safe to call concurrently with other generic read tools.
    - Denies secret/runtime directories such as environment, .env, .git,
      virtualenvs, local vector storage, and internal run logs.

    Args:
        path: Project-relative or absolute path inside this repository.
        max_entries: Maximum directory entries to return.

    Returns:
        Text lines with type, relative path, size, and child count when useful.
    """
    try:
        target = _resolve_inside_root(path, must_exist=True)
        if target.is_file():
            return f"file {_project_rel(target)} size={target.stat().st_size}"
        entries = sorted(target.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower()))
        max_entries = max(1, min(int(max_entries or 200), 1000))
        out = [f"path={_project_rel(target) or '.'} entries={len(entries)} returned={min(len(entries), max_entries)}"]
        for child in entries[:max_entries]:
            try:
                _resolve_inside_root(child.as_posix(), must_exist=True)
            except Exception:
                continue
            stat = child.stat()
            if child.is_dir():
                try:
                    child_count = len(list(child.iterdir()))
                except Exception:
                    child_count = -1
                out.append(f"dir  {_project_rel(child)}/ children={child_count}")
            else:
                out.append(f"file {_project_rel(child)} size={stat.st_size}")
        if len(entries) > max_entries:
            out.append(f"... truncated; increase max_entries or narrow path")
        return "\n".join(out)
    except Exception as exc:  # noqa: BLE001
        return f"error: {exc}"


def _python_glob(base: Path, pattern: str, *, max_results: int, offset: int) -> tuple[list[str], bool]:
    matches: list[Path] = []
    seen: set[Path] = set()
    for expanded_pattern in _expand_brace_patterns(pattern):
        for match in base.glob(expanded_pattern):
            try:
                resolved = _resolve_inside_root(match.as_posix(), must_exist=True)
            except Exception:
                continue
            if resolved in seen:
                continue
            seen.add(resolved)
            matches.append(resolved)
    matches.sort(key=lambda p: p.name.lower())
    window = matches[offset : offset + max_results]
    return [_project_rel(p) + ("/" if p.is_dir() else "") for p in window], len(matches) > offset + max_results


def _rg_glob(base: Path, pattern: str, *, max_results: int, offset: int) -> tuple[list[str], bool] | None:
    args = [
        "--files",
        "--glob",
        pattern,
        "--sort=modified",
        "--no-ignore",
        "--hidden",
        *_rg_exclusion_args(),
    ]
    result = run_ripgrep(args, _rg_target(base), cwd=_PROJECT_ROOT)
    if result.unavailable:
        return None
    if result.timed_out and not result.lines:
        return None
    if not result.ok and not result.lines:
        return None

    paths = _filter_rg_file_lines(result.lines, output_mode="files_with_matches")
    window = paths[offset : offset + max_results]
    truncated = len(paths) > offset + max_results or result.timed_out or result.truncated
    return window, truncated


@tool(parse_docstring=True)
def qa_deepagent_glob(pattern: str, path: str = ".", max_results: int = 200, offset: int = 0) -> str:
    """Find files by glob pattern under the project root.

    This is a generic DeepAgents-style ``glob`` tool for locating candidate
    source documents, test lists, JSON features, markdown notes, or code files.

    Boundaries:
    - Generic read-only tool for locating candidate files.
    - Concurrency-safe.
    - Does not read file contents; use ``qa_deepagent_read_file`` for that.
    - Denies secret/runtime directories.

    Args:
        pattern: Glob pattern such as ``knowledge/orgin/*.xlsx`` or
            ``**/*Cookie*.md``.
        path: Optional base directory inside the project.
        max_results: Maximum paths returned.
        offset: Number of matches to skip before returning results.

    Returns:
        Matching project-relative paths, one per line.
    """
    try:
        base = _resolve_inside_root(path, must_exist=True)
        if base.is_file():
            base = base.parent
        pattern = _normalise_pattern(pattern, base)
        max_results = _coerce_limit(max_results, 200)
        offset = _coerce_offset(offset)
        rg_result = _rg_glob(base, pattern, max_results=max_results, offset=offset)
        if rg_result is None:
            matches, truncated = _python_glob(base, pattern, max_results=max_results, offset=offset)
        else:
            matches, truncated = rg_result
        if not matches:
            return "(no matches)"
        if truncated:
            matches.append("... truncated; increase max_results/offset or narrow path/pattern")
        return "\n".join(matches)
    except Exception as exc:  # noqa: BLE001
        return f"error: {exc}"


def _normalise_output_mode(output_mode: str | None) -> str:
    mode = (output_mode or "content").strip().lower()
    if mode not in {"content", "files_with_matches", "count"}:
        raise ValueError("output_mode must be one of: content, files_with_matches, count")
    return mode


def _effective_head_limit(head_limit: int | None, max_results: int) -> int:
    if head_limit is None:
        return _coerce_limit(max_results, 100)
    return max(0, min(int(head_limit), 10000))


def _type_suffixes(type_name: str | None) -> set[str] | None:
    if not type_name:
        return None
    key = type_name.strip().lower()
    return _TYPE_SUFFIXES.get(key, {f".{key}"})


def _python_grep(
    *,
    pattern: str,
    base: Path,
    glob_pattern: str,
    case_sensitive: bool,
    max_results: int,
    output_mode: str,
    offset: int,
    type_name: str | None,
    context: int | None,
) -> tuple[list[str], bool]:
    flags = 0 if case_sensitive else re.IGNORECASE
    try:
        rx = re.compile(pattern, flags)
    except re.error:
        rx = re.compile(re.escape(pattern), flags)

    files = [base] if base.is_file() else list(_iter_candidate_files(base, glob_pattern, max_files=20000))
    suffixes = _type_suffixes(type_name)
    matched_files: list[str] = []
    count_lines: list[str] = []
    content_lines: list[str] = []

    for file_path in files:
        if suffixes is not None and file_path.suffix.lower() not in suffixes:
            continue
        if file_path.suffix.lower() not in _TEXT_SUFFIXES and _is_probably_binary(file_path):
            continue
        try:
            lines = file_path.read_text(encoding="utf-8", errors="replace").splitlines()
        except Exception:
            continue

        match_indexes = [idx for idx, line in enumerate(lines) if rx.search(line)]
        if not match_indexes:
            continue
        rel_path = _project_rel(file_path)
        if output_mode == "files_with_matches":
            matched_files.append(rel_path)
            continue
        if output_mode == "count":
            count_lines.append(f"{rel_path}:{len(match_indexes)}")
            continue

        selected_indexes: set[int] = set()
        context_size = max(0, int(context or 0))
        for idx in match_indexes:
            start = max(0, idx - context_size)
            end = min(len(lines), idx + context_size + 1)
            selected_indexes.update(range(start, end))
        for idx in sorted(selected_indexes):
            snippet = lines[idx].strip()
            if len(snippet) > 500:
                snippet = snippet[:497] + "..."
            content_lines.append(f"{rel_path}:{idx + 1}: {snippet}")

    if output_mode == "files_with_matches":
        matched_files.sort()
        return _apply_window(matched_files, offset=offset, head_limit=max_results)
    if output_mode == "count":
        count_lines.sort()
        return _apply_window(count_lines, offset=offset, head_limit=max_results)
    return _apply_window(content_lines, offset=offset, head_limit=max_results)


def _rg_grep(
    *,
    pattern: str,
    base: Path,
    glob_pattern: str,
    case_sensitive: bool,
    output_mode: str,
    offset: int,
    head_limit: int,
    type_name: str | None,
    context: int | None,
) -> tuple[list[str], bool, str | None] | None:
    args = [
        "--hidden",
        "--color=never",
        "--with-filename",
        "--max-columns",
        str(_RG_MAX_COLUMNS),
        *_rg_exclusion_args(),
    ]
    if glob_pattern and base.is_dir():
        args.extend(["--glob", glob_pattern])
    if not case_sensitive:
        args.append("-i")
    if type_name:
        args.extend(["--type", type_name])
    if output_mode == "files_with_matches":
        args.append("-l")
    elif output_mode == "count":
        args.append("-c")
    else:
        args.extend(["--line-number", "--no-heading"])
        if context is not None:
            args.extend(["-C", str(max(0, int(context)))])

    try:
        re.compile(pattern)
        fixed_string = False
    except re.error:
        fixed_string = True
    if fixed_string:
        args.append("-F")
    if pattern.startswith("-"):
        args.extend(["-e", pattern])
    else:
        args.append(pattern)

    result = run_ripgrep(args, _rg_target(base), cwd=_PROJECT_ROOT)
    if result.unavailable:
        return None
    if result.timed_out and not result.lines:
        return [], False, "error: rg timed out before returning matches; narrow path/glob or reduce the query"
    if not result.ok and not result.lines:
        return None

    if output_mode == "content":
        lines = _filter_rg_content_lines(result.lines)
    else:
        lines = _filter_rg_file_lines(result.lines, output_mode=output_mode)
        if output_mode == "files_with_matches":
            resolved = [_resolve_rg_path(p) for p in lines]
            paths = [p for p in resolved if p is not None]
            paths.sort(key=lambda p: (p.stat().st_mtime, _project_rel(p)), reverse=True)
            lines = [_project_rel(p) for p in paths]
    window, truncated_by_window = _apply_window(lines, offset=offset, head_limit=head_limit)
    warning = _rg_warning(result)
    truncated = truncated_by_window or warning is not None
    if warning:
        window.append(warning)
    return window, truncated, None


@tool(parse_docstring=True)
def qa_deepagent_grep(
    pattern: str,
    path: str = ".",
    glob: str = "**/*",
    case_sensitive: bool = False,
    max_results: int = 100,
    output_mode: str = "content",
    head_limit: int | None = None,
    offset: int = 0,
    type: str | None = None,
    context: int | None = None,
) -> str:
    """Search text patterns across project files.

    This is a generic DeepAgents-style ``grep`` tool. Use it for broad local
    evidence discovery, such as finding product terms, CLI strings, bug ids,
    or existing review documents. It is intentionally simple and read-only.

    Boundaries:
    - Generic read-only search tool.
    - Concurrency-safe.
    - Searches text-like files only; binary files are skipped.
    - Does not infer conclusions. The agent must reason from matched
      lines and cited files.

    Args:
        pattern: Regex pattern. Invalid regex is treated as literal text.
        path: Project-relative directory or file to search.
        glob: Glob filter under path, for example ``**/*.md``.
        case_sensitive: Whether matching is case-sensitive.
        max_results: Backward-compatible maximum matching lines returned.
        output_mode: One of ``content``, ``files_with_matches``, or ``count``.
        head_limit: Maximum returned lines or entries. ``0`` means unlimited.
        offset: Number of result lines or entries to skip.
        type: Optional ripgrep file type filter, for example ``py`` or ``ts``.
        context: Optional number of context lines around content matches.

    Returns:
        Matches formatted as ``path:line:text``, paths, or ``path:count``.
    """
    try:
        base = _resolve_inside_root(path, must_exist=True)
        mode = _normalise_output_mode(output_mode)
        offset = _coerce_offset(offset)
        effective_head_limit = _effective_head_limit(head_limit, max_results)
        rg_result = _rg_grep(
            pattern=pattern,
            base=base,
            glob_pattern=glob,
            case_sensitive=case_sensitive,
            output_mode=mode,
            offset=offset,
            head_limit=effective_head_limit,
            type_name=type,
            context=context,
        )
        if rg_result is None:
            out, truncated = _python_grep(
                pattern=pattern,
                base=base,
                glob_pattern=glob,
                case_sensitive=case_sensitive,
                max_results=effective_head_limit,
                output_mode=mode,
                offset=offset,
                type_name=type,
                context=context,
            )
        else:
            out, truncated, error = rg_result
            if error:
                return error
        if not out:
            return "(no matches)"
        if truncated and (not out[-1].startswith("... rg ")):
            out.append("... truncated; increase head_limit/offset or narrow path/glob")
        return "\n".join(out)
    except Exception as exc:  # noqa: BLE001
        return f"error: {exc}"


_WRITABLE_SUBDIRS: frozenset[str] = frozenset({
    "outputs",
})
_MAX_WRITE_BYTES = 1 * 1024 * 1024  # 1 MiB
_WRITABLE_SUFFIXES = _TEXT_SUFFIXES | {".json", ".jsonl"}


def _resolve_writable_path(raw_path: str | None) -> Path:
    """Resolve a path for write operations (four gates).

    Writes are restricted to workspace/outputs/ only.
    """
    text = (raw_path or ".").strip() or "."

    # Gate 1: refuse traversal tokens
    parts = Path(text).parts
    if ".." in parts or text.startswith("~") or "~" in parts:
        raise PermissionError(
            "path traversal not allowed; agent sandbox is rooted at knowledge/data/"
        )

    path = Path(text)
    if path.is_absolute():
        resolved = path.resolve()
    else:
        # For write: resolve under workspace/ (file may not exist yet).
        # Accept "workspace/outputs/..." or bare "outputs/..." prefix.
        if text.startswith("workspace/"):
            resolved = (_PROJECT_ROOT / path).resolve()
        elif text.startswith("outputs/"):
            resolved = (_WORKSPACE_ROOT / path).resolve()
        else:
            resolved = (_WORKSPACE_ROOT / "outputs" / path).resolve()

    # Gate 2: platform deny-list
    try:
        rel_to_project = resolved.relative_to(_PROJECT_ROOT)
    except ValueError:
        rel_to_project = None
    if rel_to_project is not None:
        rel_parts = rel_to_project.parts
        if rel_parts:
            top = rel_parts[0]
            if top in _PLATFORM_DENIED_TOP_LEVEL:
                raise PermissionError(
                    f"path is in platform-denied directory: {top}/"
                )
            if len(rel_parts) == 1 and top in _PLATFORM_DENIED_FILES:
                raise PermissionError(
                    f"path is a platform metadata file: {top}"
                )

    # Gate 3: must live under workspace/
    try:
        rel = resolved.relative_to(_WORKSPACE_ROOT)
    except ValueError as exc:
        raise PermissionError(
            "write only allowed under workspace/; "
            f"requested: {raw_path}"
        ) from exc

    # Gate 4: writable subdirectory whitelist (only outputs/)
    ws_parts = rel.parts
    if not ws_parts:
        raise PermissionError(
            "cannot write directly to workspace/ root; "
            f"must target: {sorted(_WRITABLE_SUBDIRS)}"
        )
    top_subdir = ws_parts[0]
    if top_subdir not in _WRITABLE_SUBDIRS:
        raise PermissionError(
            f"write not allowed to workspace/{top_subdir}/; "
            f"writable: {sorted(_WRITABLE_SUBDIRS)}"
        )
    return resolved


def _atomic_write(target: Path, data: bytes) -> None:
    """Write data to target atomically via tmpfile + rename."""
    fd, tmp_path = tempfile.mkstemp(
        dir=str(target.parent), suffix=".tmp", prefix=".qa_write_"
    )
    try:
        os.write(fd, data)
        os.close(fd)
        os.replace(tmp_path, str(target))
    except BaseException:
        try:
            os.close(fd)
        except OSError:
            pass
        Path(tmp_path).unlink(missing_ok=True)
        raise


@tool(parse_docstring=True)
def qa_deepagent_read_file(path: str, offset: int = 0, limit: int = 200) -> str:
    """Read a project file with line-oriented pagination.

    This is the phase-one generic equivalent of DeepAgents ``read_file``. It
    reads text files directly and renders ``.xlsx`` / ``.xlsm`` spreadsheets as
    plain text sheet rows so an analysis task can inspect a test list without
    invoking any non-read-only flow.

    Boundaries:
    - Generic read-only file inspection tool.
    - Concurrency-safe.
    - Does not write derived files, update caches, run commands, call language
      models, or index anything.
    - Spreadsheet reading is structural only: it returns workbook/sheet/row
      content for analysis.

    Args:
        path: Project-relative or absolute path inside this repository.
        offset: Zero-based line offset for pagination.
        limit: Number of rendered lines to return, capped to 1000.

    Returns:
        Text with ``total_lines`` metadata and numbered/paged content.
    """
    try:
        target = _resolve_inside_root(path, must_exist=True)
        if target.is_dir():
            return qa_deepagent_ls.invoke({"path": _project_rel(target), "max_entries": limit})
        suffix = target.suffix.lower()
        if suffix in _SPREADSHEET_SUFFIXES:
            return _read_spreadsheet(target, offset=offset, limit=limit)
        if suffix in _DOCX_SUFFIXES:
            return _read_docx(target, offset=offset, limit=limit)
        if suffix not in _TEXT_SUFFIXES and _is_probably_binary(target):
            return f"error: unsupported binary file for generic read_file: {_project_rel(target)}"
        return _read_text_file(target, offset=offset, limit=limit)
    except Exception as exc:  # noqa: BLE001
        return f"error: {exc}"


@tool(parse_docstring=True)
def qa_deepagent_write_file(path: str, content: str, overwrite: bool = False) -> str:
    """Write content to a file inside the agent sandbox.

    Creates a new file under the writable subdirectories of knowledge/data/.
    If the file already exists, set overwrite=True to replace it.

    Boundaries:
    - Write-only to allowed subdirectories (defects, markdown, baselines,
      reports).
    - Refuses to overwrite existing files unless overwrite=True.
    - Parent directory must already exist.
    - Content capped at 1 MiB; only text file suffixes allowed.
    - Subject to the four-gate security model.

    Args:
        path: Project-relative or absolute path inside this repository.
        content: Text content to write.
        overwrite: If True, allows overwriting an existing file.

    Returns:
        Success message with the written path and byte count, or error string.
    """
    try:
        target = _resolve_writable_path(path)
        suffix = target.suffix.lower()
        if suffix and suffix not in _WRITABLE_SUFFIXES:
            return f"error: suffix '{suffix}' not allowed for write; use text files"
        if not suffix:
            return "error: file must have a text extension (e.g. .md, .json, .txt)"
        encoded = content.encode("utf-8")
        if len(encoded) > _MAX_WRITE_BYTES:
            return f"error: content too large ({len(encoded)} bytes); max {_MAX_WRITE_BYTES}"
        if target.exists() and not overwrite:
            return f"error: file already exists: {_project_rel(target)}; set overwrite=True"
        if not target.parent.exists():
            return f"error: parent directory does not exist: {_project_rel(target.parent)}"
        _atomic_write(target, encoded)
        return f"wrote {len(encoded)} bytes to {_project_rel(target)}"
    except PermissionError as exc:
        return f"error: {exc}"
    except Exception as exc:  # noqa: BLE001
        return f"error: {exc}"


@tool(parse_docstring=True)
def qa_deepagent_edit_file(
    path: str, old_string: str, new_string: str, replace_all: bool = False
) -> str:
    """Edit an existing file by replacing exact string matches.

    Performs in-place string replacement on a file within the agent sandbox.
    The file must already exist and be in a writable subdirectory.

    Boundaries:
    - Target file must exist and be within the writable sandbox.
    - old_string must differ from new_string.
    - Unless replace_all=True, old_string must appear exactly once.
    - Subject to the four-gate security model.

    Args:
        path: Project-relative or absolute path to the file to edit.
        old_string: The exact text to find. Must be unique unless replace_all.
        new_string: Replacement text. Must differ from old_string.
        replace_all: If True, replaces all occurrences. Default False.

    Returns:
        Success message with replacement count, or error string.
    """
    try:
        target = _resolve_writable_path(path)
        if not target.exists():
            return f"error: file not found: {_project_rel(target)}"
        if not target.is_file():
            return f"error: not a file: {_project_rel(target)}"
        if old_string == new_string:
            return "error: old_string and new_string are identical"
        if not old_string:
            return "error: old_string must not be empty"
        content = target.read_text(encoding="utf-8")
        count = content.count(old_string)
        if count == 0:
            return "error: old_string not found in file"
        if count > 1 and not replace_all:
            return (
                f"error: old_string found {count} times; "
                "set replace_all=True or provide more context"
            )
        new_content = content.replace(old_string, new_string, -1 if replace_all else 1)
        encoded = new_content.encode("utf-8")
        if len(encoded) > _MAX_WRITE_BYTES:
            return f"error: result too large ({len(encoded)} bytes); max {_MAX_WRITE_BYTES}"
        _atomic_write(target, encoded)
        return f"replaced {count} occurrence(s) in {_project_rel(target)}"
    except PermissionError as exc:
        return f"error: {exc}"
    except Exception as exc:  # noqa: BLE001
        return f"error: {exc}"
