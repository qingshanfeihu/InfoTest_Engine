"""把 xlsx 测试用例文件直接转成 markdown 表格，供 IST-Core agent 直读。

无 LLM、无网络，只依赖 ``openpyxl``。每个 sheet 转一段 GFM 表格，
落到 ``knowledge/data/markdown/qa/{stem}.md``。

CLI::

    python -m main.xlsx_to_markdown <xlsx_path> [--out <md_path>]

库函数::

    from main.xlsx_to_markdown import convert_xlsx_to_markdown
    md = convert_xlsx_to_markdown(Path("xxx.xlsx"))
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Iterable


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("|", "\\|").replace("\n", "<br>")
    return text.strip()


def _row_to_md(cells: Iterable[object]) -> str:
    return "| " + " | ".join(_cell_to_str(c) for c in cells) + " |"


def _sheet_to_markdown(sheet) -> str:  # type: ignore[no-untyped-def]
    rows: list[list[object]] = []
    max_width = 0
    for row in sheet.iter_rows(values_only=True):
        cells = list(row)
        while cells and (cells[-1] is None or str(cells[-1]).strip() == ""):
            cells.pop()
        if not cells:
            continue
        rows.append(cells)
        max_width = max(max_width, len(cells))

    if not rows:
        return f"## {sheet.title}\n\n_(空 sheet)_\n"

    padded = [list(r) + [None] * (max_width - len(r)) for r in rows]

    header = padded[0]
    body = padded[1:]
    sep = "| " + " | ".join(["---"] * max_width) + " |"

    lines = [f"## {sheet.title}", "", _row_to_md(header), sep]
    for r in body:
        lines.append(_row_to_md(r))
    lines.append("")
    return "\n".join(lines)


def convert_xlsx_to_markdown(xlsx_path: Path) -> str:
    """读 xlsx/xls 所有 sheet，返回拼好的 markdown 文本。"""
    suffix = xlsx_path.suffix.lower()
    if suffix == ".xls":
        return _convert_xls_to_markdown(xlsx_path)

    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as exc:
        raise RuntimeError("xlsx_to_markdown 需要 openpyxl，请 pip install openpyxl") from exc

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    parts: list[str] = [f"# {xlsx_path.stem}", "", f"_来源：{xlsx_path.name}_", ""]
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(_sheet_to_markdown(ws))
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass
    return "\n".join(parts).rstrip() + "\n"


def _convert_xls_to_markdown(xls_path: Path) -> str:
    """旧 .xls 用 xlrd 读（xlrd 2.0+ 只支持 .xls）。"""
    try:
        import xlrd  # type: ignore[import-not-found]
    except ImportError as exc:
        raise RuntimeError(
            ".xls 文件需要 xlrd（pip install 'xlrd>=2.0,<3'），或先另存为 .xlsx"
        ) from exc

    book = xlrd.open_workbook(str(xls_path))
    parts: list[str] = [f"# {xls_path.stem}", "", f"_来源：{xls_path.name}_", ""]
    for sheet_name in book.sheet_names():
        sh = book.sheet_by_name(sheet_name)
        rows: list[list[object]] = []
        max_width = 0
        for r in range(sh.nrows):
            cells = list(sh.row_values(r))
            while cells and (cells[-1] is None or str(cells[-1]).strip() == ""):
                cells.pop()
            if not cells:
                continue
            rows.append(cells)
            max_width = max(max_width, len(cells))
        if not rows:
            parts.append(f"## {sheet_name}\n\n_(空 sheet)_\n")
            continue
        padded = [list(r) + [None] * (max_width - len(r)) for r in rows]
        sep = "| " + " | ".join(["---"] * max_width) + " |"
        lines = [f"## {sheet_name}", "", _row_to_md(padded[0]), sep]
        for row in padded[1:]:
            lines.append(_row_to_md(row))
        lines.append("")
        parts.append("\n".join(lines))
    return "\n".join(parts).rstrip() + "\n"


def write_markdown(xlsx_path: Path, out_path: Path) -> Path:
    md = convert_xlsx_to_markdown(xlsx_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(md, encoding="utf-8")
    return out_path


def _default_out_path(xlsx_path: Path) -> Path:
    from main import knowledge_paths as kp
    return kp.KNOWLEDGE_MARKDOWN_QA / f"{xlsx_path.stem}.md"


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="xlsx → markdown (GFM 表格)")
    parser.add_argument("xlsx", type=Path, help="输入 xlsx 路径")
    parser.add_argument("--out", type=Path, default=None,
                        help="输出 md 路径（默认 knowledge/data/markdown/qa/{stem}.md）")
    args = parser.parse_args(argv)

    if not args.xlsx.exists():
        print(f"错误：{args.xlsx} 不存在", file=sys.stderr)
        return 2

    out = args.out or _default_out_path(args.xlsx)
    written = write_markdown(args.xlsx, out)
    print(f"已写出 {written}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
