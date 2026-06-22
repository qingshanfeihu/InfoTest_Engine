"""Arm D 最终：全接地引擎。

红线：引擎对任何 case 一视同仁——从配置文本解析 service-ip 类型 / pool 资源类型 /
host 池集+权重+算法 / 清零事件，全部按【通用规则】驱动确定性计数器。
无任何 per-autoid 分支、无写死的池名/IP/期望值。

接地来源：
- 路由按记录类型分池：手册 SDNS_Pool Extension_Spec.md:89/177（一池一类型，A/AAAA 各自独立轮转）
- IPv6 字面（含 ::）判 AAAA 型，点分四段判 A 型
- 清零规则：重设 host method / 重设 host pool 权重 → 该 host 统计归零（通用假设，可证伪）
"""
from __future__ import annotations

import glob
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path("/Users/jiangyongze/Library/CloudStorage/SynologyDrive-home/Project/InfoTest_Engine")
CORPUS = Path("/tmp/real_cases")
HIT = re.compile(r"Hit:.*?(\d+)")
V6 = re.compile(r"[0-9a-fA-F]*::[0-9a-fA-F:]*")
V4 = re.compile(r"\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b")


def ip_type(ip: str) -> str:
    if V6.search(ip):
        return "AAAA"
    if V4.search(ip):
        return "A"
    return "?"


def parse_rows(xlsx):
    ws = openpyxl.load_workbook(xlsx, data_only=True).active
    rows = []
    for r in range(28, ws.max_row + 1):
        e = str(ws.cell(r, 5).value or "").strip()
        f = str(ws.cell(r, 6).value or "").strip()
        g = str(ws.cell(r, 7).value or "").strip()
        if e or g:
            rows.append((r, e, f, g))
    return rows


class Device:
    """通用 sdns 选池状态机。规则驱动，无 per-case 知识。"""
    def __init__(self):
        self.svc_type = {}          # service-id -> 'A'|'AAAA'
        self.pool_svcs = {}         # pool -> [service-id]
        self.host_pools = {}        # host -> {pool: weight}
        self.method = {}            # host -> 'rr'|'wrr'|'ga'|...
        self.hits = {}              # pool -> 累计命中
        self.rr_idx = {}            # host -> 下一个轮转位置(按记录类型分别维护)

    def pool_type(self, pool):
        """池含哪些记录类型(可A可AAAA可混)。"""
        ts = set()
        for s in self.pool_svcs.get(pool, []):
            t = self.svc_type.get(s)
            if t:
                ts.add(t)
        return ts

    def apply_cfg(self, g):
        for line in g.split("\n"):
            ln = line.strip().replace('"', "")
            if not ln:
                continue
            m = re.match(r"sdns service ip (\S+)\s+(\S+)", ln)
            if m:
                self.svc_type[m.group(1)] = ip_type(m.group(2)); continue
            m = re.match(r"sdns pool service (\S+)\s+(\S+)", ln)
            if m:
                self.pool_svcs.setdefault(m.group(1), [])
                if m.group(2) not in self.pool_svcs[m.group(1)]:
                    self.pool_svcs[m.group(1)].append(m.group(2))
                continue
            m = re.match(r"no sdns pool name (\S+)", ln)
            if m:
                self.pool_svcs.pop(m.group(1), None)
                for h in self.host_pools:
                    self.host_pools[h].pop(m.group(1), None)
                continue
            m = re.match(r"no sdns host pool (\S+)\s+(\S+)", ln)
            if m:
                self.host_pools.get(m.group(1), {}).pop(m.group(2), None)
                self._reset_host(m.group(1)); continue
            m = re.match(r"sdns host pool (\S+)\s+(\S+)(?:\s+(\d+))?", ln)
            if m:
                h, p, w = m.group(1), m.group(2), m.group(3)
                self.host_pools.setdefault(h, {})[p] = int(w) if w else 1
                # 重设权重(带数字)视为重配 -> 清零该 host 相关池统计
                if w:
                    self._reset_host(h)
                continue
            m = re.match(r"sdns host method (\S+)\s+(\S+)", ln)
            if m:
                self.method[m.group(1)] = m.group(2)
                self._reset_host(m.group(1)); continue

    def _reset_host(self, host):
        for p in self.host_pools.get(host, {}):
            self.hits[p] = 0
        self.rr_idx[host] = {"A": 0, "AAAA": 0}

    def query(self, host, qtype):
        """一次查询：在该 host 下、含 qtype 资源的池里按 method 轮转，命中池 +1。"""
        pools = [p for p in self.host_pools.get(host, {}) if qtype in self.pool_type(p)]
        if not pools:
            return None
        meth = self.method.get(host, "rr")
        idxmap = self.rr_idx.setdefault(host, {"A": 0, "AAAA": 0})
        if meth == "wrr":
            seq = []
            for p in pools:
                seq += [p] * max(1, self.host_pools[host].get(p, 1))
        else:  # rr / 其它默认按池序轮转
            seq = pools
        pick = seq[idxmap[qtype] % len(seq)]
        idxmap[qtype] += 1
        self.hits[pick] = self.hits.get(pick, 0) + 1
        return pick


def run_case(xlsx):
    rows = parse_rows(xlsx)
    dev = Device()
    host = None
    ok = tot = 0
    for i, (r, e, f, g) in enumerate(rows):
        gl = g.lower()
        if e == "APV_0" and "host name" in gl:
            m = re.search(r"host name (\S+)", g)
            if m:
                host = m.group(1)
        if e == "APV_0":
            dev.apply_cfg(g)
        elif e == "test_env" and "dig" in gl:
            qtype = "AAAA" if "aaaa" in gl else "A"
            dev.query(host, qtype)
        if "show statistics sdns pool" in gl:
            m = re.search(r"pool\s+(\w+)", gl)
            pool = m.group(1) if m else "?"
            pred = dev.hits.get(pool, 0)
            for (r2, e2, f2, g2) in rows[i + 1:i + 4]:
                if e2 == "check_point":
                    mm = HIT.search(g2)
                    if mm:
                        hv = int(mm.group(1)); tot += 1; good = (hv == pred); ok += good
                        yield (r, pool, pred, hv, good)
                        break


def main():
    files = sys.argv[1:] or [
        "smoke_test/sdns/sdns_method/sdns_method.xlsx",
    ]
    g_ok = g_tot = 0
    for rel in files:
        xp = CORPUS / rel
        if not xp.exists():
            xp = Path(rel)
        if not xp.exists():
            print(f"skip {rel}"); continue
        ok = tot = 0
        details = []
        for (r, pool, pred, hv, good) in run_case(xp):
            ok += good; tot += 1
            details.append(f"  r{r} {pool}: 引擎={pred} 真人={hv} {'OK' if good else 'X'}")
        g_ok += ok; g_tot += tot
        acc = f"{100*ok//tot}%" if tot else "—"
        print(f"\n=== {rel.split('/')[-1]}  {ok}/{tot} = {acc} ===")
        for d in details:
            print(d)
    if g_tot:
        print(f"\n##### Arm D 全接地引擎 总计: {g_ok}/{g_tot} = {100*g_ok//g_tot}% #####")
        print("对照: A一把梭39% / B状态机23% / C oracle轮转43%")


if __name__ == "__main__":
    main()
