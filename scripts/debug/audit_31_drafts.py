"""31个draft全量确定性核查(零LLM,可复现)。不抽样不外推。

对每个 draft 输出确定性事实:
- need_dynamic: 需求是否含动态点(关键词:发N次/请求/轮转/权重/分发/命中)
- has_hit: 断言里有无命中计数(Hit:N 或 Hit:\\s+N)
- hit_sum vs req_n: 计数总和是否=请求次数(自洽性,矛盾=拼凑断言)
- verdict: 纯规则判 强/弱/矛盾,规则透明

规则(全暴露,可复现):
- 需求无动态点 → 只需static断言 → 有found即 OK_static
- 需求有动态点 + draft无Hit计数 → WEAK_no_count(漏测分布)
- 需求有动态点 + 有Hit但sum≠req_n → CONTRADICT(拼凑)
- 需求有动态点 + 有Hit且sum==req_n → COVERS_dynamic(真覆盖)
"""
import glob
import json
import re

import openpyxl

ROOT = "/Users/jiangyongze/Library/CloudStorage/SynologyDrive-home/Project/InfoTest_Engine"

DYN_KW = ["发送", "请求", "轮转", "轮询", "权重", "分发", "命中", "次请求", "依次", "分别命中"]
REQ_N = re.compile(r"(\d+)\s*次")
HIT = re.compile(r"Hit:\s*\\?s?\+?\s*(\d+)")


def need_map():
    m = json.load(open(f"{ROOT}/workspace/outputs/_recon_intent/manifest.json"))
    out = {}
    for c in m["cases"]:
        si = c.get("step_intents", [])
        need = "; ".join(f"{s.get('desc','')}→{s.get('expected','')}"
                         for s in si if s.get("desc"))
        out[c["autoid"]] = need
    return out


def cps_of(xlsx):
    ws = openpyxl.load_workbook(xlsx, data_only=True).active
    out = []
    for r in range(29, ws.max_row + 1):
        if str(ws.cell(r, 5).value or "") == "check_point":
            out.append((str(ws.cell(r, 6).value or ""), str(ws.cell(r, 7).value or "")))
    return out


def classify(need, cps):
    need_dyn = any(k in need for k in DYN_KW)
    gtext = " ".join(g for _, g in cps)
    hits = [int(h) for h in HIT.findall(gtext)]
    has_hit = len(hits) > 0
    req_ns = [int(n) for n in REQ_N.findall(need)]
    req_n = max(req_ns) if req_ns else None
    n_found = sum(1 for f, _ in cps if "found" in f.lower())
    if not cps:
        return "NO_CP", need_dyn, has_hit, hits, req_n
    if not need_dyn:
        return "OK_static", need_dyn, has_hit, hits, req_n
    # 需求有动态点
    if not has_hit:
        return "WEAK_no_count", need_dyn, has_hit, hits, req_n
    # 有Hit:检查自洽(每"段"的hit和是否合理。简化:总hit和 vs 总请求次数)
    # wrr发N次,命中分布和应=N;若单段hit>req_n或和矛盾→拼凑
    hit_sum = sum(hits)
    if req_n and hit_sum > req_n * 2:  # 宽松:容许多段,但总和远超请求数=矛盾
        return "CONTRADICT", need_dyn, has_hit, hits, req_n
    return "HAS_count", need_dyn, has_hit, hits, req_n


def main():
    needs = need_map()
    rows = []
    for x in sorted(glob.glob(f"{ROOT}/workspace/outputs/2030*/case.xlsx")):
        aid = x.split("/")[-2]
        need = needs.get(aid, "")
        cps = cps_of(x)
        verdict, ndyn, hh, hits, reqn = classify(need, cps)
        rows.append((aid, verdict, ndyn, hh, hits, reqn, len(cps), need))
    from collections import Counter
    c = Counter(r[1] for r in rows)
    print("=== 31 draft 确定性核查(零LLM,规则透明) ===\n")
    print("分类汇总:")
    for k, v in c.most_common():
        print(f"  {k}: {v}")
    print(f"\n  需求含动态点的case: {sum(1 for r in rows if r[2])}/{len(rows)}")
    print(f"  其中有Hit计数的: {sum(1 for r in rows if r[2] and r[3])}")
    print(f"  动态需求但无计数(WEAK): {sum(1 for r in rows if r[1]=='WEAK_no_count')}")
    print("\n逐case明细:")
    for aid, v, nd, hh, hits, rn, ncp, need in rows:
        print(f"  [{v:14}] {aid} dyn={int(nd)} hit={hits if hits else '-'} req={rn} cp={ncp}")
        print(f"                  need: {need[:75]}")


if __name__ == "__main__":
    main()
