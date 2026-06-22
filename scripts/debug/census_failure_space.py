"""失败空间普查:137 动态路由 case 到底有几类动态行为。
不预设"漏清零是唯一病"——按测试族分组 + 抽结构特征,看 LLM 要现推的
规则空间有多大。特征(全是结构观测,不含算法知识):
  - mid_cfg: 请求序列中途是否插了 APV_0 配置(状态会变:重配/disable/enable)
  - qtype: 是否出现多种查询类型(A/AAAA/MX/PTR...) → 可能按类型路由
  - ip_kinds: found 的不同 IP 数
  - rotates: found 序列是否像定长轮转(周期性)
"""
import glob
import os
import re
from collections import defaultdict

import openpyxl

IP = re.compile(r"\b(\d{1,3}(?:\.\d{1,3}){3})\b")
QT = re.compile(r"\b(A|AAAA|MX|PTR|CNAME|NS|SOA|TXT|SRV)\b")


def signature(ws):
    rows = []
    for r in range(1, ws.max_row + 1):
        E = str(ws.cell(r, 5).value or "").strip()
        F = str(ws.cell(r, 6).value or "").strip()
        G = str(ws.cell(r, 7).value or "").strip()
        if E or F or G:
            rows.append((E, F, G))
    found_ips, qtypes, mid_cfg = [], set(), 0
    seen_req = False
    for E, F, G in rows:
        if E == "test_env" and ("dig" in G or "curl" in G):
            seen_req = True
            for q in QT.findall(G):
                qtypes.add(q)
        if E == "APV_0" and seen_req and "show" not in G:
            mid_cfg += 1            # 请求开始后又插配置 → 状态中途变
        if "found" in F.lower():
            m = IP.search(G)
            if m:
                found_ips.append(m.group(1))
    # 轮转检测:找最小周期 p(2..4),序列是否近似周期
    rot = False
    for p in (2, 3, 4):
        if len(found_ips) >= p * 2 and all(
                found_ips[i] == found_ips[i % p] for i in range(len(found_ips))):
            rot = True
            break
    return len(found_ips), len(set(found_ips)), len(qtypes), mid_cfg, rot


fam = defaultdict(list)
for f in glob.glob("/tmp/real_cases/**/*.xlsx", recursive=True):
    try:
        ws = openpyxl.load_workbook(f, data_only=True).active
    except Exception:
        continue
    nf, nip, nqt, mid, rot = signature(ws)
    if nf >= 3 and nip >= 2:
        family = os.path.relpath(f, "/tmp/real_cases").split("/")[2] \
            if len(os.path.relpath(f, "/tmp/real_cases").split("/")) > 2 else "?"
        fam[family].append((nf, nip, nqt, mid, rot))

print(f"=== 动态 case 按测试族普查 (族数={len(fam)}) ===\n")
print(f"{'族':24} {'数':>3} {'多查询型':>6} {'中途改配':>6} {'纯轮转':>5}")
for k in sorted(fam, key=lambda x: -len(fam[x])):
    v = fam[k]
    n = len(v)
    multiqt = sum(1 for x in v if x[2] >= 2)
    midcfg = sum(1 for x in v if x[3] >= 1)
    rot = sum(1 for x in v if x[4])
    print(f"{k:24} {n:>3} {multiqt:>6} {midcfg:>6} {rot:>5}")
print(f"\n总动态 case: {sum(len(v) for v in fam.values())}")
print(f"纯轮转(rotate能解释): {sum(1 for v in fam.values() for x in v if x[4])}")
print(f"中途改配(状态变,非纯rr): {sum(1 for v in fam.values() for x in v if x[3] >= 1)}")
print(f"多查询类型(可能按型路由): {sum(1 for v in fam.values() for x in v if x[2] >= 2)}")
