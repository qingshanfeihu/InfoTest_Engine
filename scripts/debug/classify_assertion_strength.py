"""零LLM确定性结构分类:31个draft的check_point断言强弱。

按 v4 计划点名的客观弱断言标记分类,不调任何 LLM:
- strong: 有 Hit 计数断言(咬住命中次数,能区分轮转)
- mid:    多个不同 IP 的 found 序列(逐跳但无 Hit 计数)
- weak:   只单个 IP found / 裸字面量(无法区分真轮转)
"""
import glob
import re
import openpyxl
from collections import Counter


def classify(xlsx):
    ws = openpyxl.load_workbook(xlsx, data_only=True).active
    cps = []
    for r in range(29, ws.max_row + 1):
        E = str(ws.cell(r, 5).value or "")
        F = str(ws.cell(r, 6).value or "")
        G = str(ws.cell(r, 7).value or "")
        if E == "check_point":
            cps.append((F, G))
    if not cps:
        return "no_cp", cps
    gtext = " ".join(g for _, g in cps)
    has_hit = "Hit" in gtext or "hit" in gtext
    ip_re = re.compile(r"\d+\.\d+\.\d+\.\d+")
    n_ip_found = sum(1 for f, g in cps if "found" in f.lower() and ip_re.search(g))
    distinct_ips = len({m for _, g in cps for m in ip_re.findall(g)})
    if has_hit:
        return "strong_Hit计数", cps
    if distinct_ips >= 2:
        return "mid_多IP序列无Hit", cps
    if n_ip_found >= 1:
        return "weak_单IP_found无Hit", cps
    return "other_无IP断言", cps


def main():
    c = Counter()
    samples = {}
    for x in sorted(glob.glob("workspace/outputs/2030*/case.xlsx")):
        cls, cps = classify(x)
        c[cls] += 1
        samples.setdefault(cls, (x.split("/")[-2], cps[:4]))
    print("31个draft结构分类(确定性零LLM):")
    for k, v in c.most_common():
        print(f"  {k}: {v}")
    print("\n各类样本(autoid + check_point前4条):")
    for k, (aid, cps) in samples.items():
        print(f"  [{k}] {aid}")
        for f, g in cps:
            print(f"      {f}: {g[:55]}")


if __name__ == "__main__":
    main()
