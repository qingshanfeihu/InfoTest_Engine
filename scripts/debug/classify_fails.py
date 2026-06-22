"""把 yzg 上机失败按 device_context 归类(只读已有日志,不重跑设备)。

类别:①配置命令被拒(设备 Failed/Invalid)②dig不解析(ANSWER:0/NXDOMAIN/无ANSWER)
③断言值不匹配(dig有答案但≠期望)④SSH掉线(瞬态)⑤超时/无结果 ⑥其它待人工。
"""
import json
import re

import openpyxl

from main.langchain_env import langchain_load_dotenv_if_present

XLSX = "workspace/outputs/yzg/case.xlsx"


def all_autoids():
    ws = openpyxl.load_workbook(XLSX, data_only=True).active
    ids = []
    for r in range(29, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a and str(a).startswith("999999"):
            break
        if a and str(a).strip():
            ids.append(str(a).strip())
    return ids


def classify(ctx, detail):
    low = (ctx + "\n" + detail).lower()
    # ① 配置命令被设备拒
    if re.search(r"failed to execute|invalid input|incomplete command|unknown command|syntax error", low):
        # 排除全局 init 那条无害的 synconfig 警告
        bad = [l for l in (ctx + detail).splitlines()
               if re.search(r"failed to execute|invalid input|incomplete command|unknown command", l, re.I)
               and "synconfig" not in l.lower()]
        if bad:
            return "①配置命令被拒", bad[0].strip()[:80]
    # ④ SSH 掉线(瞬态)
    if "socket is closed" in low or "connection reset" in low or "connection aborted" in low:
        return "④SSH掉线(瞬态)", ""
    # dig 相关
    has_answer = bool(re.search(r"answer:\s*[1-9]", low) or re.search(r"\bin\s+a\b|\bin\s+aaaa\b", low))
    if "nxdomain" in low:
        return "②dig不解析(NXDOMAIN)", ""
    if "servfail" in low:
        return "②dig不解析(SERVFAIL)", ""
    if re.search(r"answer:\s*0", low) or ("dig" in low and not has_answer and "answer section" not in low):
        return "②dig不解析(ANSWER:0/无答案)", ""
    # ③ dig 有答案但断言不匹配
    if has_answer and re.search(r"fail to find", low):
        miss = [l for l in detail.splitlines() if re.search(r"fail to find", l, re.I)]
        return "③断言值不匹配(有答案≠期望)", (miss[0].strip()[:80] if miss else "")
    return "⑥其它/待人工", ""


def main():
    langchain_load_dotenv_if_present()
    from main.case_compiler.device_mcp_client import FrameworkMCPClient
    try:
        passed = set(json.load(open("workspace/outputs/yzg/.verify_cache.json")))
    except Exception:
        passed = set()
    fails = [a for a in all_autoids() if a not in passed]
    print(f"待归类 fail: {len(fails)} 个\n")
    from collections import Counter
    tally = Counter()
    with FrameworkMCPClient() as c:
        for aid in fails:
            try:
                ctx = c.fetch_device_context(aid, max_chars=4000)
                detail = c.fetch_case_detail(aid, max_chars=3000)
            except Exception as ex:
                ctx, detail = "", str(ex)
            cat, ev = classify(ctx, detail)
            tally[cat] += 1
            print(f"  {aid[-6:]}: {cat}  {ev}")
    print("\n=== 归类汇总 ===")
    for cat, n in tally.most_common():
        print(f"  {cat}: {n}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
