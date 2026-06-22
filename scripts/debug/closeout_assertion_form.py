"""收口实验：真人断言点能否被"性质⊔因果⊔配置回显"三类确定性断言覆盖。

红线：判定全部用透明代码,零 LLM 赌具体值。
- ECHO        : 断言值的 IP/token 在该行之前的配置里字面出现 → 可换"输出应含配置X"
- CONSERVATION: 断言是 Hit:\\s+N 计数 → 归"命中总和=请求数"性质断言,不赌N
- DEFER       : 既非回显又非计数的具体观测值 → 编译期不可定,老实转上机

覆盖率 =(ECHO+CONSERVATION)/总断言点。决定"改性质/因果断言"方向成不成立。
"""
from __future__ import annotations

import glob
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl

CORPUS = Path("/tmp/real_cases")
ROOT = Path("/Users/jiangyongze/Library/CloudStorage/SynologyDrive-home/Project/InfoTest_Engine")
IP = re.compile(r"\b(\d{1,3}(?:\.\d{1,3}){3})\b")
IP6 = re.compile(r"\b([0-9a-fA-F]{0,4}(?::[0-9a-fA-F]{0,4}){2,})\b")
HIT = re.compile(r"(?i)\bhit\b:?\\?\s*\+?\s*(\d+)")
TOK = re.compile(r"[A-Za-z][\w.-]{2,}")


def parse_rows(xlsx: Path):
    ws = openpyxl.load_workbook(xlsx, data_only=True).active
    rows = []
    for r in range(29, ws.max_row + 1):
        e = str(ws.cell(r, 5).value or "").strip()
        f = str(ws.cell(r, 6).value or "").strip()
        g = str(ws.cell(r, 7).value or "").strip()
        if e or g:
            rows.append((r, e, f, g))
    return rows


def classify(g_assert: str, prior_cfg_text: str, prior_cfg_toks: set) -> set:
    """对单个断言值,返回它命中的确定性形态标签集合(可多标签)。
    一条复合断言常同时含回显(IP配过)+守恒(hit计数)两种性质,不该单选一桶。"""
    labels: set = set()
    # CONSERVATION: 含 hit 计数(大小写均认) → 归守恒性质,不赌具体数
    if HIT.search(g_assert):
        labels.add("CONSERVATION")
    # ECHO: 断言里的 IP(v4/v6)在此前配置里字面出现 → 配置回显
    ips = IP.findall(g_assert) + [m for m in IP6.findall(g_assert) if ":" in m]
    if ips:
        if all(ip in prior_cfg_text for ip in ips):
            labels.add("ECHO")
        else:
            labels.add("DEFER")  # 有具体IP但配置没配过 → 观测来的,编译期不可定
    elif not labels:
        # 无IP:看关键 token 是否来自配置(仅在还没任何标签时判)
        toks = [t for t in TOK.findall(g_assert) if len(t) >= 3]
        if toks and any(t in prior_cfg_toks for t in toks):
            labels.add("ECHO")
    if not labels:
        labels.add("DEFER")  # 纯数字/符号,无可溯源
    return labels


def run_case(xlsx: Path):
    rows = parse_rows(xlsx)
    cfg_lines = []
    cats = []
    for i, (r, e, f, g) in enumerate(rows):
        if e in ("APV_0", "test_env"):
            cfg_lines.append(g)
        if e == "check_point":
            prior_text = "\n".join(cfg_lines)
            prior_toks = set(t for t in TOK.findall(prior_text) if len(t) >= 3)
            cats.append(classify(g, prior_text, prior_toks))
    return cats


def main():
    out = open(ROOT / "runtime/logs/closeout_assertion_form.jsonl", "w")
    sample = json.load(open(ROOT / "runtime/logs/crossfamily_sample.json"))
    agg = Counter()
    per_fam = []
    for rel in sample:
        xp = CORPUS / rel
        if not xp.exists():
            continue
        fam = rel.split("/")[-2]
        cats = run_case(xp)
        if not cats:
            continue
        c = Counter(cats)
        cov = (c["ECHO"] + c["CONSERVATION"]) / len(cats)
        agg.update(cats)
        per_fam.append((fam, len(cats), c["ECHO"], c["CONSERVATION"], c["DEFER"], cov))
        out.write(json.dumps({"family": fam, "n": len(cats), "echo": c["ECHO"],
                              "conservation": c["CONSERVATION"], "defer": c["DEFER"],
                              "coverage": round(cov, 2)}, ensure_ascii=False) + "\n")
    out.close()
    tot = sum(agg.values())
    cov_tot = (agg["ECHO"] + agg["CONSERVATION"]) / tot if tot else 0
    print(f"=== 收口: {len(per_fam)}族 {tot}个真人断言点 三类形态覆盖(零LLM,代码判) ===\n")
    print(f"  ECHO(配置回显,可换'输出含配置X'): {agg['ECHO']:4} ({100*agg['ECHO']//tot}%)")
    print(f"  CONSERVATION(Hit计数,归守恒性质): {agg['CONSERVATION']:4} ({100*agg['CONSERVATION']//tot}%)")
    print(f"  DEFER(具体观测值,编译期不可定):   {agg['DEFER']:4} ({100*agg['DEFER']//tot}%)")
    print(f"\n  ★ 性质⊔因果⊔回显 覆盖率: {100*cov_tot:.0f}%  (高→方向成立; 低→编译期定值整体不通)")
    print(f"\n逐族(n断言 | echo/conserv/defer | 覆盖率):")
    for fam, n, e, cs, d, cov in sorted(per_fam, key=lambda x: -x[5]):
        print(f"  [{fam[:26]:26}] {n:3} | {e:2}/{cs:2}/{d:2} | {100*cov:3.0f}%")
    print(f"\n证据落: runtime/logs/closeout_assertion_form.jsonl")


if __name__ == "__main__":
    main()
