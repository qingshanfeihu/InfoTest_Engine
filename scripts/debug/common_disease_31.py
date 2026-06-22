"""通病数据:31个机器产物 vs 同autoid真人gold,量化"断言值来源"病灶。

每个机器 check_point 值归类:
- MATCH_GOLD: 命中真人 gold 真值(对)
- SELF_ECHO: 是本case自己配置命令里出现的token(observe-then-assert:断言自己配的回显)
- OTHER: 既非gold也非自配(脑补/错值)

红线:不硬编码任何设备值,只做"机器值 vs gold值 / 机器值 vs 自配token"的集合关系判定。
"""
from __future__ import annotations
import glob
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path("/Users/jiangyongze/Library/CloudStorage/SynologyDrive-home/Project/InfoTest_Engine")
CORPUS = Path("/tmp/real_cases")
IP = re.compile(r"[0-9a-fA-F]+(?:[.:][0-9a-fA-F:]+)+")
TOK = re.compile(r"[A-Za-z0-9_.:]+")


def _cps_and_cfg(ws, lo, hi):
    """返回 (check_point值列表, 配置命令拼接文本)。行区间 [lo,hi)。"""
    cps, cfg = [], []
    for r in range(lo, hi):
        e = str(ws.cell(r, 5).value or "").strip()
        f = str(ws.cell(r, 6).value or "").strip()
        g = str(ws.cell(r, 7).value or "").strip()
        if e == "check_point":
            cps.append((f, g))
        elif e in ("APV_0", "test_env") and g:
            cfg.append(g)
    return cps, "\n".join(cfg)


def _machine_block(aid):
    ws = openpyxl.load_workbook(ROOT / f"workspace/outputs/{aid}/case.xlsx", data_only=True).active
    return _cps_and_cfg(ws, 28, ws.max_row + 1)


def _human_block(aid, human_file):
    """真人语料里同 autoid 的行块 [起,下个autoid)。"""
    ws = openpyxl.load_workbook(human_file, data_only=True).active
    lo = hi = None
    for r in range(28, ws.max_row + 1):
        a = str(ws.cell(r, 1).value or "").strip()
        if a == aid and lo is None:
            lo = r
        elif lo is not None and a.isdigit() and len(a) >= 15 and a != aid:
            hi = r
            break
    if lo is None:
        return [], ""
    return _cps_and_cfg(ws, lo, hi or ws.max_row + 1)


def _toks(text):
    """抽 token 集合(IP 整体 + 普通 token),供"自配回显"判定。"""
    s = set(IP.findall(text))
    s |= {t for t in TOK.findall(text) if len(t) >= 2}
    return s


def _gold_keys(gold_cps):
    """真人 gold 的关键标识集合(每个 gold 值的 IP 或末位 token)。"""
    keys = set()
    for _, g in gold_cps:
        ips = IP.findall(g)
        if ips:
            keys.add(max(ips, key=len))
        else:
            toks = [t for t in TOK.findall(g) if len(t) >= 2]
            if toks:
                keys.add(toks[-1])
    return keys


def classify(mval, gold_keys, cfg_toks):
    """机器断言值归类。"""
    ips = IP.findall(mval)
    key = max(ips, key=len) if ips else None
    if key is None:
        toks = [t for t in TOK.findall(mval) if len(t) >= 2]
        key = toks[-1] if toks else mval.strip()
    if any(key in gk or gk in key for gk in gold_keys):
        return "MATCH_GOLD"
    if key in cfg_toks:
        return "SELF_ECHO"
    return "OTHER"


def main():
    import glob as _g
    idx = json.load(open(ROOT / "runtime/logs/human_autoid_index.json"))
    prods = sorted(set(x.split("/")[-2] for x in _g.glob(str(ROOT / "workspace/outputs/2030*/case.xlsx"))))
    out = open(ROOT / "runtime/logs/common_disease_31.jsonl", "w")
    agg = {"MATCH_GOLD": 0, "SELF_ECHO": 0, "OTHER": 0}
    n_case = 0
    rows = []
    for aid in prods:
        hf = idx.get(aid)
        if not hf:
            continue
        mcps, mcfg = _machine_block(aid)
        gcps, gcfg = _human_block(aid, hf)
        if not mcps or not gcps:
            continue
        n_case += 1
        gkeys = _gold_keys(gcps)
        cfgtoks = _toks(mcfg)
        per = {"MATCH_GOLD": 0, "SELF_ECHO": 0, "OTHER": 0}
        for f, g in mcps:
            c = classify(g, gkeys, cfgtoks)
            per[c] += 1
            agg[c] += 1
        rows.append((aid, len(mcps), len(gcps), per))
        rec = {"autoid": aid, "n_machine_cp": len(mcps), "n_gold_cp": len(gcps),
               "per_class": per, "gold_keys": sorted(gkeys)[:8]}
        out.write(json.dumps(rec, ensure_ascii=False) + "\n")
    out.close()
    tot = sum(agg.values())
    print(f"=== 31产物断言值来源分类(共{n_case}case {tot}个机器断言点) ===\n")
    for k in ("MATCH_GOLD", "SELF_ECHO", "OTHER"):
        print(f"  {k:12}: {agg[k]:4}  ({100*agg[k]/tot:.0f}%)")
    print(f"\n  命中真人gold真值率: {100*agg['MATCH_GOLD']/tot:.0f}%  ← 越低=病越重")
    print(f"  自配回显(observe-then-assert): {100*agg['SELF_ECHO']/tot:.0f}%")
    print("\n逐case(机器cp/gold cp | 命中/自配/其它):")
    for aid, nm, ng, per in rows:
        print(f"  {aid} m{nm:2}/g{ng:2} | {per['MATCH_GOLD']:2}/{per['SELF_ECHO']:2}/{per['OTHER']:2}")
    print(f"\n证据落: runtime/logs/common_disease_31.jsonl")


if __name__ == "__main__":
    main()
