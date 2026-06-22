"""实验2(枢纽):给LLM完整配置(init前置+case本体)+请求序列,挖空check_point期望值,
让它模拟路由推每个期望值,跟gold比。推得对→"编译期不可知"彻底死,模拟路线成立。
红线:不喂任何答案;不按算法写if;只给配置+请求,要LLM自己模拟。"""
import json, sys, re
from pathlib import Path
ROOT = Path("/Users/jiangyongze/Library/CloudStorage/SynologyDrive-home/Project/InfoTest_Engine")
CORPUS = Path("/tmp/real_cases")
sys.path.insert(0, str(ROOT))
from dotenv import load_dotenv
load_dotenv(str(ROOT/"environment"), override=False)
import openpyxl
IP = re.compile(r"[0-9a-fA-F]+(?:[.:][0-9a-fA-F:]+)+")
NUM = re.compile(r"\d+")

def parse_cases(xlsx):
    """返回 (init_cfg, [case dict])。case: {aid, steps:[(E,F,G)], golds:[(slot_row_idx,G)]}"""
    ws = openpyxl.load_workbook(xlsx, data_only=True).active
    init_cfg = ""
    cases = []; cur = None
    for r in range(29, ws.max_row+1):
        a = str(ws.cell(r,1).value or "").strip()
        c = str(ws.cell(r,3).value or "").strip()
        e = str(ws.cell(r,5).value or "").strip()
        f = str(ws.cell(r,6).value or "").strip()
        g = str(ws.cell(r,7).value or "").strip()
        desc = str(ws.cell(r,4).value or "").strip()  # D列:步骤/断言描述(验什么)
        if c == "1" and g:   # init块
            init_cfg = g
            continue
        if a.isdigit() and len(a) >= 15:  # 新case起点
            if cur: cases.append(cur)
            cur = {"aid": a, "steps": [], "golds": []}
        if cur is None: continue
        if e or f or g:
            cur["steps"].append((e,f,g,desc))
    if cur: cases.append(cur)
    return init_cfg, cases

def build_prompt_blocks(init_cfg, case):
    """构造给LLM的步骤视图:check_point行挖空标#slot,其余原样。返回(blocks文本, gold列表)"""
    lines = ["【通用前置配置(init,所有case共享)】", init_cfg, "", "【本case步骤】"]
    golds = []
    for step in case["steps"]:
        e, f, g, desc = step
        if e == "check_point":
            slot = len(golds); golds.append(g)
            dtag = f"（验:{desc}）" if desc else ""
            lines.append(f"  [断言点#{slot}] {f}{dtag} 期望=____(待你推)")
        else:
            dtag = f"  # {desc}" if desc else ""
            lines.append(f"  [{e}|{f}] {g}{dtag}")
    return "\n".join(lines), golds


_SYS = """你是APV sdns设备的精确模拟器。给你一台设备的【完整配置(init前置+case配置)】和
【请求序列】,其中若干【断言点】的期望值被挖空。你要像设备真的执行一样,在脑中维护状态,
推出每个断言点此刻的期望值。

可用的确定性事实(全在配置里,不需外部知识):
- service ip <名> <IP>:服务名→IP映射
- pool service <pool> <服务名>:池→服务成员
- host pool <域名> <pool> [权重]:域名挂哪些池(及wrr权重)
- host method <域名> <算法>:选池算法(rr轮转/wrr加权)
- dig查询→按算法选池→池的service→该service的IP,就是dig应返回的值
- show statistics sdns pool <p>的Hit:从该池被命中计数推(rr:第k次请求命中pool[k%N])

要求:
- 每个断言点给出predicted(你推的期望值)。dig类给IP;Hit类给数字(写成设备显示格式如"Hit:\\s+1"也可,或纯数字)。
- 能唯一推出就推;若配置信息不足以唯一确定,标underdetermined=true别硬猜。
- 严格按配置推,不要用任何配置外的假设IP。
只输出JSON: {"slots":[{"slot":0,"predicted":"1.1.1.1","why":"rr第1次→p1→s1→1.1.1.1"},...]}"""

def _ask(model, blocks, timeout=150):
    from langchain_core.messages import SystemMessage, HumanMessage
    import concurrent.futures as cf
    msgs = [SystemMessage(content=_SYS), HumanMessage(content=blocks+"\n\n按要求推每个断言点,输出JSON。")]
    with cf.ThreadPoolExecutor(max_workers=1) as ex:
        resp = ex.submit(model.invoke, msgs).result(timeout=timeout)
    m = re.search(r"\{.*\}", str(resp.content), re.S)
    if not m:
        return {"slots": []}
    try:
        return json.loads(m.group(0))
    except json.JSONDecodeError:
        # LLM 偶发返回坏 JSON(尾逗号/截断)。不崩,当本 case 全欠定处理。
        return {"slots": [], "_bad_json": True}

def _norm(s):
    """宽松比对:有IP取最长IP;否则若含数字取末位数字(Hit:\\s+1 与 1 同归一为'1')。"""
    s = str(s)
    ips = IP.findall(s)
    if ips: return max(ips, key=len)
    n = NUM.findall(s)
    if n: return n[-1]
    return s.strip()

def main():
    from main.ist_core.agents._llm import build_agent_chat_model
    model = build_agent_chat_model()
    files = sys.argv[1:] or ["smoke_test/sdns/sdns_method/sdns_method.xlsx"]
    out = open(ROOT/"runtime/logs/exp2_simulate.jsonl","w")
    N_ok = N_under = N_wrong = N_tot = 0
    for rel in files:
        xp = CORPUS/rel
        if not xp.exists(): print(f"skip {rel}"); continue
        init_cfg, cases = parse_cases(xp)
        fam = rel.split("/")[-2]
        for case in cases:
            blocks, golds = build_prompt_blocks(init_cfg, case)
            if not golds: continue
            try: d = _ask(model, blocks)
            except Exception as e: print(f"[{case['aid'][:16]}] 异常{str(e)[:40]}"); continue
            slots = {s["slot"]:s for s in d.get("slots",[]) if "slot" in s}
            co=cu=cw=0
            detail=[]
            for i,gold in enumerate(golds):
                s=slots.get(i,{}); N_tot+=1
                if s.get("underdetermined"): cu+=1; N_under+=1; verdict="under"
                elif s.get("predicted") and _norm(s["predicted"])==_norm(gold): co+=1; N_ok+=1; verdict="ok"
                else: cw+=1; N_wrong+=1; verdict="wrong"
                detail.append({"slot":i,"gold":_norm(gold),"pred":_norm(s.get("predicted")) if s.get("predicted") else None,"v":verdict})
            rec={"aid":case["aid"],"family":fam,"n":len(golds),"ok":co,"under":cu,"wrong":cw,"detail":detail}
            out.write(json.dumps(rec,ensure_ascii=False)+"\n"); out.flush()
            print(f"  [{fam[:16]:16}] {case['aid'][:16]} 断言{len(golds):2} 推对{co:2} 欠定{cu} 推错{cw}")
    out.close()
    print(f"\n=== 实验2: {N_tot}断言点 完整配置模拟推导 ===")
    print(f"  推对gold:   {N_ok} ({100*N_ok/N_tot if N_tot else 0:.0f}%)")
    print(f"  诚实欠定:   {N_under} ({100*N_under/N_tot if N_tot else 0:.0f}%)")
    print(f"  推错:       {N_wrong} ({100*N_wrong/N_tot if N_tot else 0:.0f}%)")
    print(f"  判决: 推对率高→'编译期不可知'死,模拟路线成立; 推错高→真推理病")
    print(f"  证据: runtime/logs/exp2_simulate.jsonl")

if __name__ == "__main__":
    main()
