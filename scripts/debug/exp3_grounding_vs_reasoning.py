"""实验3:定位真病——甲(grounding)还是乙(推理)。零 LLM,纯确定性。

对每个机器产物(31 个 case.xlsx),并排算两个数:
  自洽率 = 断言里的关键值(IP/service token)能在机器自己配置 G 列里找到来源的比例
           (机器配了 service ip s1 X、断言 dig 返回 X → 自洽:它在自己配置世界里推得对)
  可达率 = 断言里的 IPv4 中 env_facts.is_reachable 为真的比例(用没用测试床真值)

判读:
  自洽高 + 可达高 → 接地良好(无病)
  自洽高 + 可达低 → 甲 grounding 病(自洽但没用测试床真值)
  自洽低           → 乙 推理病(连自己配置都对不上)
"""
from __future__ import annotations

import glob
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path("/Users/jiangyongze/Library/CloudStorage/SynologyDrive-home/Project/InfoTest_Engine")
sys.path.insert(0, str(ROOT))
from main.ist_core.tools._shared.env_facts import get_env_facts  # noqa: E402

IP = re.compile(r"\b(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})\b")
TOK = re.compile(r"[A-Za-z][\w.:-]{1,}")
# 动态/运行时断言:设备运行后才吐的值(命中计数、权重分布正则),不该出现在静态配置里。
# 它们的对错归实验2 的模拟 oracle 判,不归"静态自洽"。判自洽时必须剔除,否则把
# 正确的动态断言误判成"推理病"(与 self-echo 同类尺子错)。
_DYNAMIC_RE = re.compile(r"Hits?\s*:|\.\*\d|命中|统计")


def is_dynamic_assert(val: str) -> bool:
    return bool(_DYNAMIC_RE.search(val or ""))



def load_blocks(xlsx):
    """返回 (cfg_text 机器自己配置全文, asserts 断言值列表)。"""
    ws = openpyxl.load_workbook(xlsx, data_only=True).active
    cfg, asserts = [], []
    for r in range(28, ws.max_row + 1):
        e = str(ws.cell(r, 5).value or "").strip()
        g = str(ws.cell(r, 7).value or "").strip()
        if not g:
            continue
        if e in ("APV_0", "test_env"):
            cfg.append(g)
        elif e == "check_point":
            asserts.append(g)
    return "\n".join(cfg), asserts


def selfconsistent(val, cfg_text, cfg_toks):
    """断言值是否能在机器自己配置里找到来源(IP 字面出现 / 关键 token 出现)。"""
    ips = IP.findall(val)
    if ips:  # IP 类断言:该 IP 是否被自己配过
        return all(ip in cfg_text for ip in ips)
    toks = [t for t in TOK.findall(val) if len(t) >= 2]
    if not toks:
        return True  # 无可比 token(如 Hit:1 纯计数)→ 不计入自洽判定
    return any(t in cfg_toks for t in toks)


def main():
    ef = get_env_facts()
    out = open(ROOT / "runtime/logs/exp3_grounding.jsonl", "w")
    rows = []
    for xlsx in sorted(glob.glob(str(ROOT / "workspace/outputs/2030*/case.xlsx"))):
        aid = xlsx.split("/")[-2]
        cfg_text, asserts = load_blocks(xlsx)
        cfg_toks = set(t for t in TOK.findall(cfg_text) if len(t) >= 2)
        # 自洽率:有可比 token 的断言里,来源能在自配里找到的比例
        n_sc = n_sc_ok = 0
        # 可达率:断言里所有 IPv4 中可达的比例
        all_ips, reach_ips = [], []
        for a in asserts:
            ips = IP.findall(a)
            for ip in ips:
                all_ips.append(ip)
                if ef.is_reachable(ip):
                    reach_ips.append(ip)
            # 自洽只看"静态、有可比内容"的断言;动态(Hit计数/权重分布)归实验2 oracle,剔除
            if is_dynamic_assert(a):
                continue
            if IP.findall(a) or [t for t in TOK.findall(a) if len(t) >= 2]:
                n_sc += 1
                if selfconsistent(a, cfg_text, cfg_toks):
                    n_sc_ok += 1
        sc_rate = n_sc_ok / n_sc if n_sc else None
        reach_rate = len(reach_ips) / len(all_ips) if all_ips else None
        rows.append((aid, sc_rate, reach_rate, n_sc, len(all_ips)))
        out.write(json.dumps({"aid": aid, "selfconsist": sc_rate, "reach": reach_rate,
                              "n_assert": n_sc, "n_ip": len(all_ips)}, ensure_ascii=False) + "\n")
    out.close()
    # 汇总
    scs = [r[1] for r in rows if r[1] is not None]
    rcs = [r[2] for r in rows if r[2] is not None]
    print(f"=== 实验3: {len(rows)} 机器产物 甲/乙判定(零LLM) ===\n")
    print(f"  自洽率(断言来自机器自己配置): 均值 {sum(scs)/len(scs):.0%}  (n={len(scs)})")
    print(f"  可达率(断言IP是测试床真值):   均值 {sum(rcs)/len(rcs):.0%}  (n={len(rcs)})")
    # 分桶
    def bucket(sc, rc):
        if sc is None:
            return "无可判断言"
        if sc < 0.6:
            return "乙_推理病(连自配都对不上)"
        if rc is not None and rc < 0.6:
            return "甲_grounding病(自洽但IP不可达)"
        return "接地良好"
    from collections import Counter
    bc = Counter(bucket(r[1], r[2]) for r in rows)
    print("\n分桶:")
    for k, v in bc.most_common():
        print(f"  {k:30}: {v}")
    print("\n逐case(自洽/可达 | n断言/n_IP):")
    for aid, sc, rc, na, ni in rows:
        scs_ = f"{sc:.0%}" if sc is not None else "—"
        rcs_ = f"{rc:.0%}" if rc is not None else "—"
        print(f"  {aid} 自洽{scs_:>4} 可达{rcs_:>4} | {na:2}断言 {ni:2}IP  {bucket(sc,rc)}")
    print(f"\n证据落: runtime/logs/exp3_grounding.jsonl")


if __name__ == "__main__":
    main()
