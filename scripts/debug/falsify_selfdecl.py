"""证伪实验：自声明约束 + 领域无关算术校验 能否在动态 Hit 断言上
做到「危险桶(自洽但错)≈0、弃权诚实」。

真值：真人 sdns_method / mx_record 每个 show statistics → Hit:N（直接读,无歧义）。
机制：把该点之前整条命令+请求序列喂 LLM（全 case 上下文）→ LLM 产 预测Hit + 自声明约束。
校验：领域无关算术——代码只 eval LLM 声明的约束表达式,不含 rr/守恒公式。
三桶：自洽且==真人 / 诚实弃权 / 自洽但≠真人(危险)。

红线：本脚本的校验器对 rr 一无所知；约束由 LLM 声明（数据,样本）,代码只做算术。
"""
from __future__ import annotations
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path("/Users/jiangyongze/Library/CloudStorage/SynologyDrive-home/Project/InfoTest_Engine")
CORPUS = Path("/tmp/real_cases/smoke_test/sdns")
HIT = re.compile(r"Hit:\\s\+(\d+)|Hit:\s*(\d+)")


def _hit_val(g: str):
    m = HIT.search(g)
    if not m:
        return None
    return int(m.group(1) or m.group(2))


def extract_points(xlsx: Path) -> list[dict]:
    """每个 Hit 真值点：{idx, human_hit, replay(该点前整条命令+请求+断言文本), pool_line}。"""
    ws = openpyxl.load_workbook(xlsx, data_only=True).active
    rows = []
    for r in range(28, ws.max_row + 1):
        E = str(ws.cell(r, 5).value or ""); F = str(ws.cell(r, 6).value or "")
        G = str(ws.cell(r, 7).value or "")
        if E or F or G:
            rows.append((r, E.strip(), F.strip(), G.strip()))
    points = []
    for i, (r, E, F, G) in enumerate(rows):
        hv = _hit_val(G) if E == "check_point" else None
        if hv is None:
            continue
        # 重放上下文 = 该点之前的 配置 + 请求 序列（不含任何断言行——
        # 真实编译时 draft 没有真人答案,喂答案会让实验失真）。
        replay = []
        for (_, e2, f2, g2) in rows[:i]:
            g2s = g2.replace("\n", " ; ")
            if e2 == "APV_0":
                replay.append(f"[配置] {g2s}")
            elif e2 == "test_env":
                replay.append(f"[请求/{f2}] {g2s}")
        # 找这个 Hit 点对应的 show 命令（往前找最近的 APV show statistics）
        pool_cmd = ""
        for (_, e2, f2, g2) in reversed(rows[:i]):
            if e2 == "APV_0" and "show statistics" in g2:
                pool_cmd = g2.strip(); break
        points.append({"idx": len(points), "row": r, "human_hit": hv,
                       "pool_cmd": pool_cmd, "replay": "\n".join(replay)})
    return points


_SYS = """你是专注的测试断言设计师。现在只做一件事：给定一台 APV 设备此前执行过的
【配置+请求序列】,预测此刻执行某条 `show statistics sdns pool <X>` 命令时,该 pool 的
累计命中次数 Hit 应该是多少。

要点：
- Hit 是【全序列累计状态】：从序列开头读到现在,该 pool 每被一次请求命中就 +1,遇到重配
  /重置该 pool 统计的命令则归零。你要像人通读整条用例那样在脑中维护这个计数器。
- 选池算法语义(rr 轮转/wrr 加权/按查询类型路由等)你自己依据序列判断——这是通用知识。
- 关键诚实要求：如果根据给定序列无法【唯一确定】这个 Hit 值（比如请求次数不足以体现
  分布、算法或路由有歧义、信息缺失）,必须标 underdetermined=true 并说明,不要硬猜一个数。

同时,你要【自己声明】这个预测值必须满足的约束,用 Python 布尔表达式写,变量只能用：
  hit  = 你预测的该 pool 累计命中(int)
  reqs = 你判断的、自上次该 pool 清零以来命中该 pool 的请求次数(int)
例如若你认为"该 pool 被命中 reqs 次、每次 +1、期间没清零",就声明 ["hit == reqs"]。
约束是你对自己推理的自检,必须由上面两个变量算得出真假。

只输出 JSON,不要解释：
{"underdetermined": false, "reason": "",
 "hit": <int>, "reqs": <int>,
 "constraints": ["hit == reqs", ...]}"""


def _model():
    sys.path.insert(0, str(ROOT))
    from dotenv import load_dotenv
    load_dotenv(str(ROOT / "environment"), override=False)
    from main.ist_core.agents._llm import build_agent_chat_model
    return build_agent_chat_model()


def ask_llm(model, pool_cmd: str, replay: str, timeout=120):
    from langchain_core.messages import SystemMessage, HumanMessage
    import concurrent.futures as cf
    user = (f"【此前配置+请求序列】\n{replay}\n\n"
            f"【此刻命令】{pool_cmd}\n\n该 pool 累计 Hit 应是多少?按要求输出 JSON。")
    msgs = [SystemMessage(content=_SYS), HumanMessage(content=user)]
    with cf.ThreadPoolExecutor(max_workers=1) as ex:
        resp = ex.submit(model.invoke, msgs).result(timeout=timeout)
    txt = str(resp.content)
    m = re.search(r"\{.*\}", txt, re.S)
    return (json.loads(m.group(0)) if m else {}), txt


def check_constraints(decl: dict) -> tuple[bool, str]:
    """领域无关算术校验：只 eval LLM 自己声明的约束表达式。
    命名空间仅含 LLM 声明的 hit/reqs + 安全算术。代码不含任何 rr/守恒知识。"""
    ns = {"hit": decl.get("hit"), "reqs": decl.get("reqs"),
          "abs": abs, "min": min, "max": max}
    if not isinstance(ns["hit"], int) or not isinstance(ns["reqs"], int):
        return False, "hit/reqs 非整数,无法校验"
    for c in decl.get("constraints") or []:
        try:
            ok = eval(c, {"__builtins__": {}}, ns)  # noqa: S307 — 受限命名空间,实验用
        except Exception as e:  # noqa: BLE001
            return False, f"约束无法求值: {c} ({e})"
        if not ok:
            return False, f"自相矛盾: 声明值不满足自声明约束 {c} (hit={ns['hit']},reqs={ns['reqs']})"
    return True, "自洽"


def main():
    model = _model()
    out = open(ROOT / "runtime/logs/falsify_selfdecl.jsonl", "w")
    buckets = {"consistent_correct": 0, "abstain": 0, "danger_consistent_wrong": 0,
               "contradict": 0, "error": 0}
    total = 0
    for fn in ["sdns_method/sdns_method.xlsx", "sdns_DynamicDomain_MX/mx_record.xlsx"]:
        pts = extract_points(CORPUS / fn)
        for p in pts:
            total += 1
            try:
                decl, raw = ask_llm(model, p["pool_cmd"], p["replay"])
            except Exception as e:  # noqa: BLE001
                buckets["error"] += 1
                print(f"[{fn.split('/')[-1]} row{p['row']}] LLM异常: {e}")
                continue
            if decl.get("underdetermined"):
                bucket = "abstain"
            else:
                consistent, why = check_constraints(decl)
                if not consistent:
                    bucket = "contradict"
                elif decl.get("hit") == p["human_hit"]:
                    bucket = "consistent_correct"
                else:
                    bucket = "danger_consistent_wrong"
            buckets[bucket] += 1
            rec = {"file": fn, "row": p["row"], "human_hit": p["human_hit"],
                   "llm_hit": decl.get("hit"), "reqs": decl.get("reqs"),
                   "underdetermined": decl.get("underdetermined"),
                   "constraints": decl.get("constraints"), "bucket": bucket}
            out.write(json.dumps(rec, ensure_ascii=False) + "\n"); out.flush()
            flag = {"consistent_correct": "✓对", "abstain": "弃权",
                    "danger_consistent_wrong": "✗危险", "contradict": "矛盾"}.get(bucket, bucket)
            print(f"[{flag}] {fn.split('/')[-1][:12]} row{p['row']:3} "
                  f"真人={p['human_hit']} LLM={decl.get('hit')} reqs={decl.get('reqs')} "
                  f"{('UNDER:'+str(decl.get('reason',''))[:30]) if decl.get('underdetermined') else ''}")
    out.close()
    print(f"\n=== {total} 个真值点 三桶判决 ===")
    print(f"  ✓ 自洽且对真人:      {buckets['consistent_correct']}")
    print(f"  弃权(诚实标欠定):    {buckets['abstain']}")
    print(f"  ✗ 危险(自洽但错):    {buckets['danger_consistent_wrong']}")
    print(f"  矛盾(被算术闸抓):    {buckets['contradict']}")
    print(f"  error:               {buckets['error']}")
    print(f"\n判决: 危险桶={buckets['danger_consistent_wrong']} (要≈0)  证据: runtime/logs/falsify_selfdecl.jsonl")


if __name__ == "__main__":
    main()
