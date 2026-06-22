"""Arm B：设备状态抽象为独立哑账本,LLM 处理每个事件前注入【当前状态】,
只输出该事件的 delta(命中哪些池 / 清零哪些池),账本做确定性累加。

对照 Arm A(falsify_selfdecl：LLM 脑内一次性追踪整条用例累计 Hit)。
判决：Arm B 危险桶(自洽但错)是否塌到≈0 —— 若是,实锤"失败根因是缺状态机/
全局账扛不住",而非 LLM 不懂算法。

红线：账本(state_machine)对 rr 一无所知,只执行 LLM 喂的 delta(+1/归零),
无任何算法 if 分支。算法语义(谁命中)由 LLM 逐事件判,是数据不是代码。
"""
from __future__ import annotations
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path("/Users/jiangyongze/Library/CloudStorage/SynologyDrive-home/Project/InfoTest_Engine")
CORPUS = Path("/tmp/real_cases/smoke_test/sdns")
HIT = re.compile(r"Hit:\\s\+(\d+)|Hit:\s*(\d+)")


def _hit_val(g: str):
    m = HIT.search(g)
    return int(m.group(1) or m.group(2)) if m else None


def parse_events(xlsx: Path) -> list[dict]:
    """整条 case 解析成有序事件流。kind ∈ CFG/REQ/SHOW；SHOW 带 pool + human_hit。"""
    ws = openpyxl.load_workbook(xlsx, data_only=True).active
    raw = []
    for r in range(28, ws.max_row + 1):
        E = str(ws.cell(r, 5).value or "").strip()
        F = str(ws.cell(r, 6).value or "").strip()
        G = str(ws.cell(r, 7).value or "").strip()
        if E or F or G:
            raw.append((r, E, F, G))
    events = []
    for i, (r, E, F, G) in enumerate(raw):
        if E == "APV_0" and "show statistics" in G:
            # 找紧随其后的 check_point Hit 真值
            hv = None
            for (_, e2, f2, g2) in raw[i + 1:i + 4]:
                if e2 == "check_point":
                    v = _hit_val(g2)
                    if v is not None:
                        hv = v; break
            pool = G.split("pool")[-1].strip() if "pool" in G else ""
            events.append({"kind": "SHOW", "row": r, "cmd": G, "pool": pool, "human_hit": hv})
        elif E == "APV_0":
            events.append({"kind": "CFG", "row": r, "cmd": G.replace("\n", " ; ")})
        elif E == "test_env":
            events.append({"kind": "REQ", "row": r, "who": F, "cmd": G})
    return events


def _model():
    sys.path.insert(0, str(ROOT))
    from dotenv import load_dotenv
    load_dotenv(str(ROOT / "environment"), override=False)
    from main.ist_core.agents._llm import build_agent_chat_model
    return build_agent_chat_model()


def _ask(model, sysp, user, timeout=90):
    from langchain_core.messages import SystemMessage, HumanMessage
    import concurrent.futures as cf
    msgs = [SystemMessage(content=sysp), HumanMessage(content=user)]
    with cf.ThreadPoolExecutor(max_workers=1) as ex:
        resp = ex.submit(model.invoke, msgs).result(timeout=timeout)
    txt = str(resp.content)
    m = re.search(r"\{.*\}", txt, re.S)
    if not m:
        return {}, txt
    raw = m.group(0)
    try:
        return json.loads(raw), txt
    except json.JSONDecodeError:
        # LLM 偶发尾随注释/单引号/多余逗号 → 容错重试一次
        fixed = re.sub(r",\s*([}\]])", r"\1", raw).replace("'", '"')
        try:
            return json.loads(fixed), txt
        except json.JSONDecodeError:
            return {}, txt


# REQ 原子判断：当前配置 + 账本下,这一次请求命中哪个池(或不命中)
_REQ_SYS = """你判断 APV sdns 设备上【这一次操作】是否命中某个 sdns pool,命中哪个。
注入了【当前配置】和【各池当前累计命中】。只看这一次操作：
- 若是 dig/DNS 查询且会被某 pool 应答 → 给出该 pool 名(依据当前算法/路由语义,如
  rr 轮转看各池已命中数推下一个、按查询类型 A/AAAA 路由到对应池等,你自己判断)。
- 若是环境准备(如 ip addr add/del)、与选池无关的操作 → hits=null。
- 无法唯一确定命中哪池 → underdetermined=true 并说明。
只输出 JSON：{"hits": "<pool名或null>", "underdetermined": false, "reason": ""}"""


# CFG 原子判断：这条配置命令是否清零某些池的统计
_CFG_SYS = """你判断 APV sdns 设备上【这一条配置命令】是否会把某些 sdns pool 的
统计命中数(show statistics 的 Hit)清零/重置。注入了【此前配置】。
- 重新设置选池方法/重建池/复位统计等 → 列出被清零的池名。
- 仅新增 service/ip 等不影响既有池计数 → resets=[]。
只输出 JSON：{"resets": ["<池名>", ...]}"""


def run_arm_b(model, events: list[dict], out) -> dict:
    ledger: dict[str, int] = {}
    cfg_lines: list[str] = []
    buckets = {"correct": 0, "wrong": 0, "abstain": 0}
    for e in events:
        state = f"当前配置:\n" + ("\n".join(cfg_lines) or "(空)") + \
                f"\n各池当前累计命中: {json.dumps(ledger, ensure_ascii=False)}"
        if e["kind"] == "CFG":
            cfg_lines.append(e["cmd"])
            try:
                d, _ = _ask(model, _CFG_SYS, f"{state}\n\n【这条配置命令】{e['cmd']}")
                for p in d.get("resets") or []:
                    if p in ledger:
                        ledger[p] = 0
            except Exception:  # noqa: BLE001
                pass
        elif e["kind"] == "REQ":
            try:
                d, _ = _ask(model, _REQ_SYS, f"{state}\n\n【这一次操作】{e['who']}: {e['cmd']}")
                if not d.get("underdetermined"):
                    hp = d.get("hits")
                    if hp:
                        ledger[hp] = ledger.get(hp, 0) + 1
            except Exception:  # noqa: BLE001
                pass
        elif e["kind"] == "SHOW":
            pred = ledger.get(e["pool"], 0)
            hv = e["human_hit"]
            if hv is None:
                continue
            bucket = "correct" if pred == hv else "wrong"
            buckets[bucket] += 1
            rec = {"row": e["row"], "pool": e["pool"], "human_hit": hv,
                   "pred_ledger": pred, "bucket": bucket,
                   "ledger_snapshot": dict(ledger)}
            out.write(json.dumps(rec, ensure_ascii=False) + "\n"); out.flush()
            flag = "✓" if bucket == "correct" else "✗错"
            print(f"  [{flag}] row{e['row']} pool {e['pool']} 真人={hv} 状态机预测={pred}")
    return buckets


def main():
    model = _model()
    out = open(ROOT / "runtime/logs/falsify_statemachine.jsonl", "w")
    tot = {"correct": 0, "wrong": 0, "abstain": 0}
    for fn in ["sdns_method/sdns_method.xlsx", "sdns_DynamicDomain_MX/mx_record.xlsx"]:
        print(f"\n=== Arm B 重放 {fn} ===")
        ev = parse_events(CORPUS / fn)
        b = run_arm_b(model, ev, out)
        print(f"  小计: {b}")
        for k in tot:
            tot[k] += b[k]
    out.close()
    n = tot["correct"] + tot["wrong"]
    print(f"\n=== Arm B (状态机外置 + 逐事件原子LLM) {n}个SHOW点 ===")
    print(f"  ✓ 对真人:        {tot['correct']}")
    print(f"  ✗ 错(选错池累积): {tot['wrong']}")
    acc = tot["correct"] / n if n else 0
    print(f"  准确率: {acc:.0%}")
    print(f"  注:状态机只+1/归零,破守恒结构上不可能;错只来自LLM单次选错池(局部可定位)")
    print(f"  证据: runtime/logs/falsify_statemachine.jsonl")


if __name__ == "__main__":
    main()
