"""第二轮验证:事后裁判(v4 confidence_f)是否对"来源错"的值睁眼瞎。
挑 self-echo 重灾 case,跑 confidence_f,看它给 PASS/高分 → 坐实通病:
四版都靠事后裁判,裁判判'看着合不合理'判不出'来源对不对'。"""
import json, sys, re
from pathlib import Path
import openpyxl
ROOT = Path("/Users/jiangyongze/Library/CloudStorage/SynologyDrive-home/Project/InfoTest_Engine")
sys.path.insert(0, str(ROOT))
from dotenv import load_dotenv
load_dotenv(str(ROOT / "environment"), override=False)

# 重灾(self-echo多/match少) + 干净对照
CASES = {"203031754302210934": "重灾 0命中/11自配",
         "203031753342777976": "重灾 2命中/9自配",
         "203031754291994899": "重灾 1命中/8自配",
         "203031754291995045": "对照 6命中/0自配"}

def load_rows(aid):
    ws = openpyxl.load_workbook(ROOT / f"workspace/outputs/{aid}/case.xlsx", data_only=True).active
    rows = []
    for r in range(28, ws.max_row + 1):
        e,f,g = [str(ws.cell(r,c).value or "").strip() for c in (5,6,7)]
        if e or f or g:
            rows.append({"E":e,"F":f,"G":g})
    return rows

def main():
    from main.case_compiler.confidence_f import score_case
    from main.ist_core.agents._llm import build_agent_chat_model
    model = build_agent_chat_model()
    idx = json.load(open(ROOT/"runtime/logs/human_autoid_index.json"))
    print("=== v4裁判 confidence_f 对来源错的值判分 ===\n")
    for aid, tag in CASES.items():
        rows = load_rows(aid)
        try:
            r = score_case(rows, need_intent="", model=model)
        except Exception as e:
            print(f"[{aid}] {tag}: 异常 {e}"); continue
        dec = "PASS" if (not r.get("abstain") and (r.get("overall") or 0)>=0.5) else "CUT/abstain"
        print(f"[{tag}]")
        print(f"  overall={r.get('overall')} abstain={r.get('abstain')} → 裁判判: {dec}")
        print(f"  (该case断言值实际62%是自配回显/错值,真值率见common_disease)\n")

if __name__ == "__main__":
    main()
