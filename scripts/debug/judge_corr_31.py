"""闸2:confidence_f跑全31产物,算 score 与真值命中率的相关。
证明"裁判分数与真值不相关"是统计结论非4样本巧合。"""
import json, sys
from pathlib import Path
ROOT = Path("/Users/jiangyongze/Library/CloudStorage/SynologyDrive-home/Project/InfoTest_Engine")
sys.path.insert(0, str(ROOT))
from dotenv import load_dotenv
load_dotenv(str(ROOT/"environment"), override=False)
import scripts.debug.common_disease_31 as CD
import openpyxl, glob

def load_rows(aid):
    ws = openpyxl.load_workbook(ROOT/f"workspace/outputs/{aid}/case.xlsx", data_only=True).active
    rows=[]
    for r in range(28, ws.max_row+1):
        e,f,g=[str(ws.cell(r,c).value or "").strip() for c in (5,6,7)]
        if e or f or g: rows.append({"E":e,"F":f,"G":g})
    return rows

def main():
    from main.case_compiler.confidence_f import score_case
    from main.ist_core.agents._llm import build_agent_chat_model
    model = build_agent_chat_model()
    idx = json.load(open(ROOT/"runtime/logs/human_autoid_index.json"))
    prods = sorted(set(x.split("/")[-2] for x in glob.glob(str(ROOT/"workspace/outputs/2030*/case.xlsx"))))
    out = open(ROOT/"runtime/logs/judge_corr_31.jsonl","w")
    pairs=[]  # (match_rate, judge_score)
    for aid in prods:
        hf=idx.get(aid)
        if not hf: continue
        mcps,mcfg=CD._machine_block(aid); gcps,gcfg=CD._human_block(aid,hf)
        if not mcps or not gcps: continue
        gkeys=CD._gold_keys(gcps); cfgtoks=CD._toks(mcfg)
        nmatch=sum(1 for f,g in mcps if CD.classify(g,gkeys,cfgtoks)=="MATCH_GOLD")
        mrate=nmatch/len(mcps)
        try:
            r=score_case(load_rows(aid), need_intent="", model=model)
            score=r.get("overall") or 0.0; ab=r.get("abstain")
        except Exception as e:
            print(f"[{aid}] 异常{e}"); continue
        pairs.append((mrate,score))
        rec={"autoid":aid,"match_rate":round(mrate,2),"judge_score":score,"abstain":ab,"n_cp":len(mcps)}
        out.write(json.dumps(rec,ensure_ascii=False)+"\n"); out.flush()
        print(f"  {aid} 真值率={mrate:.0%} 裁判={score:.2f}{'(abs)' if ab else ''}")
    out.close()
    # 相关系数(纯numpy,无scipy依赖)
    import numpy as np
    mr=np.array([p[0] for p in pairs]); js=np.array([p[1] for p in pairs])
    if len(mr)>2 and mr.std()>0 and js.std()>0:
        pear=np.corrcoef(mr,js)[0,1]
        # spearman=rank的pearson
        rmr=mr.argsort().argsort(); rjs=js.argsort().argsort()
        spear=np.corrcoef(rmr,rjs)[0,1]
    else:
        pear=spear=float('nan')
    print(f"\n=== {len(pairs)}产物 裁判分vs真值命中率 ===")
    print(f"  Pearson r = {pear:.3f}   Spearman ρ = {spear:.3f}")
    print(f"  (接近0=不相关,坐实裁判信号是噪声;接近1=裁判有效)")
    print(f"  证据: runtime/logs/judge_corr_31.jsonl")

if __name__=="__main__":
    main()
