"""V3 上机验证 + <RUNTIME> 回填驱动（确定性，不经主 agent，免编排 churn、可报数）。

流程对应用户两条铁律：
  首跑(采集设备真实输出) → 抽 golden 值(受限 LLM,抽不出=CANNOT→留空,绝不猜)
  → qa_fill_runtime 锁死回填 → 复跑确认(回填断言现应 pass) → 报数。

设备串行（框架全局锁），用底层 FrameworkMCPClient 一条会话跑 + 取全量明细。
LLM 只做"从设备真实输出里抽出 <RUNTIME> 那段真实值"，绝不编。

用法：python -m scripts.debug.run_v3_verify_fill <case.xlsx> [build] [module]
"""
from __future__ import annotations

import sys
import time

import openpyxl

from main.langchain_env import langchain_load_dotenv_if_present

_EXTRACT_SYS = (
    "你在回填一个测试断言里**离线不可知的运行时值**。铁律：不许编——只能从设备真实输出里抽，"
    "抽不到/不确定就只回 CANNOT 这一个词。"
)
_EXTRACT_TMPL = """断言模式（xlsx G 列，<RUNTIME> 是待填槽位）：
{cur_g}

产生待匹配输出的命令（设备据它产出回显）：
{observe}

设备本次真实执行明细（ground truth）：
---
{detail}
---

任务：在明细里找到上面那条命令的真实输出，抽出应当填进 <RUNTIME> 这一段的**真实值**。
- 部分模式（前缀…<RUNTIME>）：只输出槽位那一段（如前缀 "Hits:" 后的真实数字 "42"）。
- 整值 <RUNTIME>：输出该断言要测的那个真实落点值（如本次真实解析出的 IP）。
只输出要填进 <RUNTIME> 的那段文本本身，不要任何解释、不要引号。抽不到/不确定就只回 CANNOT。"""


def _read_autoids(xlsx_path: str) -> list[str]:
    ws = openpyxl.load_workbook(xlsx_path, data_only=True).active
    ids = []
    for r in range(29, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if not a:
            continue
        a = str(a).strip()
        if a.startswith("999999"):
            break
        if a:
            ids.append(a)
    return ids


def _run_cases(client, module, build, autoids, max_s=180) -> dict[str, dict]:
    # max_s 600→180:崩溃/hang 的 case(unknown)会一直轮询到超时,600s×6=1h 是上次 verify
    # 2.4h 的大头。正常 case 框架 ~1-3min 内出结果;180s 足够,3min 不出基本是 hang,早断省时间。
    """串行跑一批 autoid，返回 {autoid: {verdict, task_id, detail}}（全量明细）。"""
    out = {}
    for aid in autoids:
        rec = {"verdict": "error", "task_id": "", "detail": ""}
        try:
            d = client.deliver(module, aid, str(_XLSX))
            if d.get("error"):
                rec["detail"] = f"deliver 失败: {d.get('error')}"
                out[aid] = rec; continue
            run = client.run_and_wait(module, aid, build, [aid], max_s=max_s)
            if run.get("busy"):
                rec["verdict"] = "busy"; rec["detail"] = run.get("message", "busy")
                out[aid] = rec; continue
            if run.get("error"):
                rec["detail"] = f"运行失败: {run.get('error')}"
                out[aid] = rec; continue
            rec["verdict"] = (run.get("results") or {}).get(aid) or run.get("result") or "unknown"
            rec["task_id"] = run.get("task_id", "")
            rec["detail"] = client.fetch_case_detail(aid, max_chars=6000)
        except Exception as e:  # noqa: BLE001
            rec["detail"] = f"异常: {e}"
        out[aid] = rec
        print(f"    [{aid}] verdict={rec['verdict']} ({len(rec['detail'])} chars detail)", flush=True)
    return out


def main():
    langchain_load_dotenv_if_present()
    if len(sys.argv) < 2:
        print("用法: run_v3_verify_fill <case.xlsx> [build] [module]")
        return 2
    global _XLSX
    from pathlib import Path
    _XLSX = Path(sys.argv[1]).resolve()
    if not _XLSX.is_file():
        print(f"xlsx 不存在: {_XLSX}"); return 2

    from main.case_compiler.config import get_config
    cfg = get_config()
    build = sys.argv[2] if len(sys.argv) > 2 else cfg.build
    module = sys.argv[3] if len(sys.argv) > 3 else cfg.staging_module

    from main.case_compiler.runtime_fill import list_runtime_slots, apply_fills
    from main.case_compiler.device_mcp_client import FrameworkMCPClient
    from main.case_compiler import verify_cache as vc
    from main.ist_core.agents import _llm

    root = Path(__file__).resolve().parents[2]
    mindmap_dir = _XLSX.parent
    autoids = _read_autoids(str(_XLSX))
    slots = list_runtime_slots(_XLSX)
    slot_autoids = sorted({s.autoid for s in slots})

    # ---- 上机 pass 缓存（铁律③：上机通过的不反复跑，只跑有问题的）----
    cache = vc.load_cache(mindmap_dir)
    rows_by = vc.case_rows_by_autoid(_XLSX)
    hash_by = {a: vc.case_content_hash(rows_by.get(a, [])) for a in autoids}
    cached_pass = [a for a in autoids if vc.is_cached_pass(cache, a, hash_by[a])]
    to_run = [a for a in autoids if a not in cached_pass]

    print(f"=== verify+fill: {_XLSX} ===", flush=True)
    print(f"build={build} module={module}  cases={len(autoids)}  "
          f"<RUNTIME>槽位={len(slots)}（分布于 {len(slot_autoids)} 个 case）", flush=True)
    print(f"上机缓存: 已通过跳过={len(cached_pass)}  本次需上机={len(to_run)}", flush=True)
    if not slots:
        print("没有 <RUNTIME> 待回填槽位——要么都是确定值，要么已回填完。", flush=True)

    t0 = time.time()
    model = _llm.build_agent_chat_model()

    # ---- 首跑：只跑非缓存(有问题/新/内容变)的 case ----
    print(f"\n[1/3] 首跑 {len(to_run)} 个需上机的 case（缓存通过的 {len(cached_pass)} 个跳过）…", flush=True)
    first = {}
    if to_run:
        with FrameworkMCPClient() as client:
            first = _run_cases(client, module, build, to_run)

    # ---- 抽 golden 值（受限 LLM，抽不出=CANNOT→留空）----
    print(f"\n[2/3] 从设备真实输出抽 golden 值（{len(slots)} 个槽位）…", flush=True)
    fills = []
    for s in slots:
        detail = (first.get(s.autoid) or {}).get("detail", "")
        val = ""
        # 护栏(不猜):只在 detail 确含该槽位的前序观测命令时才抽。崩溃 case 没专属日志会
        # 回退到别 case 的总日志——命令对不上就不抽,绝不从错日志里抽出错值。
        anchored = bool(detail and s.observe_cmd and s.observe_cmd in detail)
        if anchored:
            prompt = _EXTRACT_TMPL.format(cur_g=s.current_g, observe=s.observe_cmd or "(未知)",
                                          detail=detail[:5000])
            try:
                r = model.invoke([("system", _EXTRACT_SYS), ("user", prompt)])
                txt = str(getattr(r, "content", r)).strip()
                if txt and txt.upper() != "CANNOT":
                    val = txt
            except Exception as e:  # noqa: BLE001
                print(f"    抽值异常 {s.slot_id}: {e}", flush=True)
        if val:
            tag = f"'{val}'"
        elif not anchored:
            tag = "留空(detail 不含该观测命令,防抽错)"
        else:
            tag = "CANNOT→留空"
        print(f"    {s.slot_id}  [{s.current_g[:28]}] ← {tag}", flush=True)
        fills.append({"slot_id": s.slot_id, "runtime_value": val,
                      "evidence": (detail[:160] if val else "")})

    # ---- 锁死回填 ----
    task_meta = ";".join(f"{a}={(first.get(a) or {}).get('task_id','')}" for a in slot_autoids)[:120]
    res = apply_fills(_XLSX, fills, project_root=root, run_meta=f"build={build};{task_meta}")
    print(f"\n回填: {res.summary()}", flush=True)
    print(f"  filled={res.filled}", flush=True)
    print(f"  left_blank(如实留空)={res.left_blank}", flush=True)

    # ---- 复跑确认（被填的 case）----
    filled_autoids = sorted({sid.split('#')[0] for sid in res.filled})
    confirm = {}
    if filled_autoids:
        print(f"\n[3/3] 复跑确认（{len(filled_autoids)} 个被回填的 case）…", flush=True)
        with FrameworkMCPClient() as client:
            confirm = _run_cases(client, module, build, filled_autoids)

    # ---- 全量最终裁决：缓存通过→pass；被回填→复跑结果；其余→首跑 ----
    def _verdict(a):
        if a in cached_pass:
            return "pass"
        if a in confirm:
            return (confirm.get(a) or {}).get("verdict")
        return (first.get(a) or {}).get("verdict", "?")
    final = {a: _verdict(a) for a in autoids}
    remaining = len(list_runtime_slots(_XLSX))
    passed = [a for a, v in final.items() if v == "pass"]
    failed = [a for a, v in final.items() if v not in ("pass",)]

    # ---- 记缓存：本次新通过且无 <RUNTIME> 残留的 case（用回填后的内容哈希）----
    rows_after = vc.case_rows_by_autoid(_XLSX)
    newly_cached = 0
    for a in to_run:
        if final[a] == "pass":
            ra = rows_after.get(a, [])
            if not vc.has_unfilled_runtime(ra):
                vc.record_pass(cache, a, vc.case_content_hash(ra), build=build,
                               task_id=(confirm.get(a) or first.get(a) or {}).get("task_id", ""))
                newly_cached += 1
    vc.save_cache(mindmap_dir, cache)

    print(f"\n=== 汇总（耗时 {time.time()-t0:.0f}s）===", flush=True)
    print(f"成品: {_XLSX}", flush=True)
    print(f"case 总数={len(autoids)}  完整通过={len(passed)}  未通过={len(failed)}", flush=True)
    print(f"上机缓存: 本次跳过(已通过)={len(cached_pass)}  本次新记入={newly_cached}  缓存总通过={len([a for a in cache if cache[a].get('verdict')=='pass'])}", flush=True)
    print(f"回填: <RUNTIME>槽位={len(slots)} 填上={len(res.filled)} 留空(待人工补值)={remaining}", flush=True)
    if confirm:
        ok = [a for a in confirm if confirm[a]["verdict"] == "pass"]
        print(f"被回填 case 复跑通过: {len(ok)}/{len(confirm)} → {ok}", flush=True)
    if failed:
        print("--- 未通过 case（下次 verify 只重跑这些，已通过的不再上机）---", flush=True)
        for a in failed:
            had_slot = a in slot_autoids
            reason = "回填后仍fail/非确定值" if had_slot else "grounded断言真实失败/环境"
            print(f"  {a}: verdict={final[a]}  ({reason})", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
