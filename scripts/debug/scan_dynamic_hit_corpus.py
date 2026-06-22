"""扫 /tmp/real_cases 380 真人用例，找含「累计 Hit 动态序列」的——
就是清零规则归纳的多样本 oracle。Hit 单元格形如 'Hit:\\s+3' 或 'Hit: 3'。"""
import glob
import os
import re

import openpyxl

HIT = re.compile(r"Hit:\s*\\?s?\+?\s*(\d+)")
files = glob.glob("/tmp/real_cases/**/*.xlsx", recursive=True)
dyn = []
for f in files:
    try:
        ws = openpyxl.load_workbook(f, data_only=True).active
    except Exception:
        continue
    hits = []
    for r in range(1, ws.max_row + 1):
        g = str(ws.cell(r, 7).value or "")
        m = HIT.search(g)
        if m:
            hits.append(int(m.group(1)))
    if len(hits) >= 2 and len(set(hits)) >= 2:
        dyn.append((os.path.relpath(f, "/tmp/real_cases"), hits))

print(f"总 {len(files)} 真人 xlsx，含动态 Hit 序列(≥2点且有变化): {len(dyn)}")
print()
for f, h in sorted(dyn)[:25]:
    print(f"  {f}: {h[:14]}")
