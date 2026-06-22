"""扫真人用例的「动态路由」模式:同一 case 内多条 found <ip>，且 IP 有≥2种
(说明在验"每次请求命中不同池"的序列行为)。这类是 rr/wrr 的主流真人测法，
比累计 Hit 大得多——清零规则归纳的真正多样本 oracle。"""
import glob
import os
import re

import openpyxl

IP = re.compile(r"\b(\d{1,3}(?:\.\d{1,3}){3})\b")
ALGO = re.compile(r"\b(rr|wrr|sr|lc|lb|hi|hip|chi|method|轮询|加权)\b", re.I)
files = glob.glob("/tmp/real_cases/**/*.xlsx", recursive=True)
routing = []
for f in files:
    try:
        ws = openpyxl.load_workbook(f, data_only=True).active
    except Exception:
        continue
    found_ips, has_algo = [], False
    for r in range(1, ws.max_row + 1):
        E = str(ws.cell(r, 5).value or "")
        F = str(ws.cell(r, 6).value or "")
        G = str(ws.cell(r, 7).value or "")
        if ALGO.search(G) or ALGO.search(F):
            has_algo = True
        if "found" in F.lower():
            m = IP.search(G)
            if m:
                found_ips.append(m.group(1))
    uniq = set(found_ips)
    if len(found_ips) >= 3 and len(uniq) >= 2:
        routing.append((os.path.relpath(f, "/tmp/real_cases"),
                        len(found_ips), len(uniq), has_algo))

print(f"总 {len(files)} 真人 xlsx")
print(f"动态路由模式(≥3条found且≥2种IP): {len(routing)}")
print(f"  其中显式提到算法(rr/wrr/method等)的: {sum(1 for x in routing if x[3])}")
print()
for f, n, u, a in sorted(routing, key=lambda x: -x[1])[:20]:
    print(f"  {'[算法]' if a else '      '} {f}: {n}条found / {u}种IP")
