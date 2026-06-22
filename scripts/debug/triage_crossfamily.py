"""闸1:跨57族分诊验证。每族1 case,LLM对每断言点判来源(只看配置+请求,剥答案),
统计:①真实oracle来源跨族分布 ②分诊是否正确路由(能算族推对率) ③拒绝硬推率。
证明4路分诊覆盖380全样本族多样性,不只4族。"""
import json, sys, re
from pathlib import Path
ROOT = Path("/Users/jiangyongze/Library/CloudStorage/SynologyDrive-home/Project/InfoTest_Engine")
CORPUS = Path("/tmp/real_cases")
sys.path.insert(0, str(ROOT))
from dotenv import load_dotenv
load_dotenv(str(ROOT/"environment"), override=False)
import openpyxl
IP=re.compile(r"[0-9a-fA-F]+(?:[.:][0-9a-fA-F:]+)+"); NUM=re.compile(r"\d+")

_SYS="""你给一条 APV sdns 测试用例的每个【断言点】判定期望值的来源,能算的就算出来。
你只能看到用例的【配置命令+请求序列】,看不到设备输出,也看不到答案。
对每个断言点(标#序号),输出来源分类之一:
- "algorithm": 期望由通用选池算法(rr/wrr)决定,据配置顺序能算。
- "config_intent": 期望值就是某条配置命令里填的参数(monitor dst/link地址等),能从配置文本定位。
- "script_runtime": 期望是注入脚本运行时输出,编译期无法算。
- "precedent_needed": 需同类已验证先例才能定值。
- "underdetermined": 信息缺失/有歧义,无法唯一确定。
对algorithm/config_intent两类必须给predicted(推的期望值),其余null。绝不硬猜,拿不准归underdetermined。
只输出JSON: {"slots":[{"slot":0,"source":"config_intent","predicted":"172.16.35.231","why":"..."}]}"""

def extract(xlsx):
    ws=openpyxl.load_workbook(xlsx,data_only=True).active
    steps,golds=[],[]
    for r in range(28,ws.max_row+1):
        e,f,g=[str(ws.cell(r,c).value or "").strip() for c in (5,6,7)]
        if not(e or f or g): continue
        if e=="check_point" and "found" in f.lower():
            steps.append(f"  [断言点#{len(golds)}] 期望=____(待填)"); golds.append(g)
        else: steps.append(f"  {e}|{f}|{g[:80]}")
    return steps,golds

def _ask(model,steps,timeout=150):
    from langchain_core.messages import SystemMessage,HumanMessage
    import concurrent.futures as cf
    user="【用例配置+请求序列】\n"+"\n".join(steps)+"\n\n按要求对每个断言点输出JSON。"
    with cf.ThreadPoolExecutor(max_workers=1) as ex:
        resp=ex.submit(model.invoke,[SystemMessage(content=_SYS),HumanMessage(content=user)]).result(timeout=timeout)
    m=re.search(r"\{.*\}",str(resp.content),re.S)
    return json.loads(m.group(0)) if m else {"slots":[]}

def _norm(s):
    s=str(s); ips=IP.findall(s)
    return ips[-1] if ips else (NUM.findall(s)[-1] if NUM.findall(s) else s.strip())

def _match(pred, gold):
    """宽松匹配:IP/数字用_norm精确比;字符串答案做大小写无关的alnum包含比,
    治'status: NOERROR' vs 'NOERROR'这种标签前缀假错(gold带字段名,pred只给值)。"""
    if not pred: return False
    if _norm(pred)==_norm(gold): return True
    # gold无IP无数字=纯字符串答案(DNS状态码等):剥非字母数字后做包含
    if not IP.findall(str(gold)) and not NUM.findall(str(gold)):
        pa=re.sub(r"[^A-Za-z0-9]","",str(pred)).upper()
        ga=re.sub(r"[^A-Za-z0-9]","",str(gold)).upper()
        return bool(pa) and len(pa)>=3 and (pa in ga or ga in pa)
    return False


def main():
    from main.ist_core.agents._llm import build_agent_chat_model
    model=build_agent_chat_model()
    sample=json.load(open(ROOT/"runtime/logs/crossfamily_sample.json"))
    out=open(ROOT/"runtime/logs/triage_crossfamily.jsonl","w")
    from collections import Counter
    src_dist=Counter(); n_comp=n_comp_ok=n_defer=n_slots=0; n_case=0
    for rel in sample:
        xp=CORPUS/rel
        if not xp.exists(): continue
        steps,golds=extract(xp)
        if not golds: continue
        fam=rel.split("/")[-2]
        try: d=_ask(model,steps[:60])  # 截断防超长
        except Exception as e:
            print(f"[{fam}] 异常{str(e)[:40]}"); continue
        n_case+=1
        slots={s["slot"]:s for s in d.get("slots",[]) if "slot" in s}
        cc=co=cd=0; detail=[]
        for i,gold in enumerate(golds):
            s=slots.get(i,{}); src=s.get("source","missing"); src_dist[src]+=1; n_slots+=1
            hit=None
            if src in ("algorithm","config_intent"):
                n_comp+=1; cc+=1
                hit=_match(s.get("predicted"), gold)
                if hit: n_comp_ok+=1; co+=1
            elif src in ("script_runtime","precedent_needed","underdetermined"): n_defer+=1; cd+=1
            detail.append({"slot":i,"src":src,"gold":_norm(gold),
                           "pred":(_norm(s.get("predicted")) if s.get("predicted") else None),
                           "hit":hit})
        rec={"case":rel,"family":fam,"n_slots":len(golds),"n_computable":cc,"comp_ok":co,
             "n_defer":cd,"detail":detail}
        out.write(json.dumps(rec,ensure_ascii=False)+"\n"); out.flush()
        print(f"  [{fam[:24]:24}] 点{len(golds):2} 判能算{cc:2}→对{co:2} 判缓推{cd:2}")
    out.close()
    print(f"\n=== {n_case}族分诊({n_slots}断言点) ===")
    print("真实oracle来源分布:")
    for k,v in src_dist.most_common(): print(f"  {k:18}: {v:3} ({100*v/n_slots:.0f}%)")
    print(f"\n判'能算'{n_comp}点→推对{n_comp_ok} ({100*n_comp_ok/n_comp if n_comp else 0:.0f}%)")
    print(f"判'该缓推'{n_defer}点 ({100*n_defer/n_slots:.0f}%) ← 这些不硬推=避开self-echo")
    print(f"证据: runtime/logs/triage_crossfamily.jsonl")

if __name__=="__main__": main()
