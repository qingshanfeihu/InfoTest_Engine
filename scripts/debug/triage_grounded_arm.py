"""Arm A+grounding:在 triage_crossfamily 同样本上,给 LLM 注入产品手册/footprint
文法(命令各参数位语义),对照零-grounding 基线(comp_ok 36%)看抬升。
红线:footprint 返回的是命令文法(数据),LLM 自己推;无任何 per-case 规则/答案表。"""
import json, sys, re
from pathlib import Path
ROOT = Path("/Users/jiangyongze/Library/CloudStorage/SynologyDrive-home/Project/InfoTest_Engine")
CORPUS = Path("/tmp/real_cases")
sys.path.insert(0, str(ROOT))
from dotenv import load_dotenv
load_dotenv(str(ROOT / "environment"), override=False)
import openpyxl  # noqa: E402
# 复用基线臂的抽取/匹配/归一,保证唯一变量是"是否注入 grounding"
from scripts.debug.triage_crossfamily import extract, _match, _norm, _SYS  # noqa: E402

CMD_PREFIX = re.compile(r'^(?:no|clear|show)\s+')


def case_commands(xlsx):
    """从一条 case 的配置/请求行通用抽取命令前缀(前3 token),供 footprint 查文法。
    不挑命令、不按族分支——shell 噪声(cat/curl)footprint 查不中自然丢弃。"""
    ws = openpyxl.load_workbook(xlsx, data_only=True).active
    cmds = set()
    for r in range(28, ws.max_row + 1):
        e = str(ws.cell(r, 5).value or "").strip()
        g = str(ws.cell(r, 7).value or "").strip()
        if e in ("APV_0", "test_env") and g:
            for line in g.split("\n"):
                ln = CMD_PREFIX.sub("", line.strip().replace('"', ""))
                toks = ln.split()
                if len(toks) >= 2 and re.match(r"[a-z]", toks[0]):
                    cmds.add(" ".join(toks[:3]))
    return sorted(cmds)


def grounding_block(cmds, lookup, cap=12):
    """对每个命令前缀查 footprint,拼成 grounding 文本。命中即收,未命中跳过。"""
    chunks = []
    for c in cmds:
        try:
            r = lookup.invoke({"command": c})
        except Exception:  # noqa: BLE001
            continue
        if r and "未找到" not in r and "未命中" not in r[:20]:
            chunks.append(r.strip())
        if len(chunks) >= cap:
            break
    return "\n\n".join(chunks)


def ask_grounded(model, steps, ground, timeout=180):
    from langchain_core.messages import SystemMessage, HumanMessage
    import concurrent.futures as cf
    user = ("【该产品相关命令的 CLI 文法(权威,标明各参数位含义,用于定位配置里的期望值)】\n"
            + (ground or "(无命中文法)")
            + "\n\n【用例配置+请求序列】\n" + "\n".join(steps)
            + "\n\n按要求对每个断言点输出JSON。文法只帮你定位/解析,算法语义仍靠你自己推。")
    with cf.ThreadPoolExecutor(max_workers=1) as ex:
        resp = ex.submit(model.invoke,
                         [SystemMessage(content=_SYS), HumanMessage(content=user)]).result(timeout=timeout)
    m = re.search(r"\{.*\}", str(resp.content), re.S)
    return json.loads(m.group(0)) if m else {"slots": []}


def main():
    from main.ist_core.agents._llm import build_agent_chat_model
    from main.ist_core.tools.knowledge.footprint_lookup import qa_footprint_lookup
    model = build_agent_chat_model()
    sample = json.load(open(ROOT / "runtime/logs/crossfamily_sample.json"))
    # 基线明细:用于逐 case 对照抬升
    base = {}
    for l in open(ROOT / "runtime/logs/triage_crossfamily.jsonl"):
        try:
            r = json.loads(l); base[r["family"]] = r
        except Exception:  # noqa: BLE001
            pass
    out = open(ROOT / "runtime/logs/triage_grounded.jsonl", "w")
    only = sys.argv[1:]  # 可传族名子集快测
    g_comp = g_ok = b_comp = b_ok = n_case = 0
    for rel in sample:
        fam = rel.split("/")[-2]
        if only and fam not in only:
            continue
        xp = CORPUS / rel
        if not xp.exists():
            continue
        steps, golds = extract(xp)
        if not golds:
            continue
        ground = grounding_block(case_commands(xp), qa_footprint_lookup)
        try:
            d = ask_grounded(model, steps[:60], ground)
        except Exception as e:  # noqa: BLE001
            print(f"[{fam}] 异常{str(e)[:40]}"); continue
        n_case += 1
        slots = {s["slot"]: s for s in d.get("slots", []) if "slot" in s}
        cc = co = 0
        for i, gold in enumerate(golds):
            s = slots.get(i, {}); src = s.get("source", "missing")
            if src in ("algorithm", "config_intent"):
                cc += 1
                if _match(s.get("predicted"), gold):
                    co += 1
        brec = base.get(fam, {})
        bc, bo = brec.get("n_computable", 0), brec.get("comp_ok", 0)
        g_comp += cc; g_ok += co; b_comp += bc; b_ok += bo
        out.write(json.dumps({"family": fam, "ground_chars": len(ground),
                              "g_comp": cc, "g_ok": co, "b_comp": bc, "b_ok": bo},
                             ensure_ascii=False) + "\n"); out.flush()
        arrow = "↑" if co > bo else ("↓" if co < bo else "=")
        print(f"  [{fam[:22]:22}] grounding{len(ground):4}字 | 基线{bo:2}/{bc:2} → 接地{co:2}/{cc:2} {arrow}")
    out.close()
    print(f"\n=== {n_case}族 grounding 对照 ===")
    print(f"  基线(零手册)能算推对: {b_ok}/{b_comp} = {100*b_ok//max(b_comp,1)}%")
    print(f"  +footprint 文法 grounding: {g_ok}/{g_comp} = {100*g_ok//max(g_comp,1)}%")
    print(f"  证据: runtime/logs/triage_grounded.jsonl")


if __name__ == "__main__":
    main()
