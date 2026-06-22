"""最小样本:验证「oracle 来源分诊 + 能算的族编译期推期望值」。

对每个真人动态 case:
  1. 取 check_point/found 行的 G 作 gold,挖空 G 喂给 LLM(连同配置+请求序列)。
  2. LLM 对每个空:标 oracle 来源 ∈ {algorithm, config_intent, script_runtime,
     precedent_needed, underdetermined};能算的(algorithm/config_intent)现场推期望值。
  3. 比对:① 分诊是否把 epolicy 这类判成 script/precedent(不硬推) ②能算族推得准不准。

红线:LLM 只看本 case 配置+请求,不看任何设备输出/真人答案。推不出就标欠定,不硬猜。
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path("/Users/jiangyongze/Library/CloudStorage/SynologyDrive-home/Project/InfoTest_Engine")
CORPUS = ROOT / "knowledge" / "framework" / "mirror" if False else Path("/tmp/real_cases")

IP = re.compile(r"(\d{1,3}(?:\.\d{1,3}){3}|[0-9a-fA-F:]{2,}:[0-9a-fA-F:]+)")
NUM = re.compile(r"\d+")


def extract(xlsx: Path, max_slots: int = 18):
    """返回 (steps, golds)。steps 是含挖空标记的有序行;golds 是每个空的真值 G。
    max_slots: 大文件(多条 case 拼一起)只取前 N 个断言点,够验分诊+推值,防 prompt 爆。
    取到第 max_slots 个断言点所在行后停(含其后续上下文到下个断言点前)。"""
    ws = openpyxl.load_workbook(xlsx, data_only=True).active
    steps, golds = [], []
    for r in range(28, ws.max_row + 1):
        E = str(ws.cell(r, 5).value or "").strip()
        F = str(ws.cell(r, 6).value or "").strip()
        G = str(ws.cell(r, 7).value or "").strip()
        if not (E or F or G):
            continue
        if E == "check_point" and "found" in F.lower():
            if len(golds) >= max_slots:
                break
            slot = len(golds)
            golds.append(G)
            steps.append(f"  [断言点#{slot}] 期望=____(待你填+标来源)")
        else:
            steps.append(f"  {E}|{F}|{G[:80]}")
    return steps, golds


_SYS = """你给一条 APV sdns 测试用例的每个【断言点】判定期望值的来源,能算的就算出来。
你只能看到用例的【配置命令 + 请求序列】,看不到设备输出,也看不到答案。

对每个断言点(标 #序号),输出来源分类之一:
- "algorithm": 期望由通用选池算法(rr轮转/wrr加权)决定,你能据配置顺序算出。
- "config_intent": 期望值就是本用例某条配置命令里填的参数(如 monitor 的 dst 地址、
  link 地址),你能从配置文本直接定位。
- "script_runtime": 期望是注入脚本(如 epolicy TCL active_members)的运行时输出,
  编译期无法计算。
- "precedent_needed": 需要同类已验证先例才能定值,本用例信息不足。
- "underdetermined": 信息缺失/有歧义,无法唯一确定。

对 algorithm / config_intent 两类,**必须**给出 predicted(你推的期望值,如 IP 或数);
其余类 predicted 留 null。绝不硬猜——拿不准就归 underdetermined。

只输出 JSON: {"slots":[{"slot":0,"source":"config_intent","predicted":"172.16.35.231",
"why":"r38 monitor dns_v4 dst=231"}, ...]}"""


def _model():
    sys.path.insert(0, str(ROOT))
    from dotenv import load_dotenv
    load_dotenv(str(ROOT / "environment"), override=False)
    from main.ist_core.agents._llm import build_agent_chat_model
    return build_agent_chat_model()


def _ask(model, steps, timeout=150):
    from langchain_core.messages import SystemMessage, HumanMessage
    import concurrent.futures as cf
    user = "【用例配置+请求序列】\n" + "\n".join(steps) + "\n\n按要求对每个断言点输出 JSON。"
    msgs = [SystemMessage(content=_SYS), HumanMessage(content=user)]
    with cf.ThreadPoolExecutor(max_workers=1) as ex:
        resp = ex.submit(model.invoke, msgs).result(timeout=timeout)
    txt = str(resp.content)
    m = re.search(r"\{.*\}", txt, re.S)
    return json.loads(m.group(0)) if m else {"slots": []}


def _norm(s):
    """抽 G 里的关键标识(IP/末段数字)做宽松比对。"""
    s = str(s)
    ips = IP.findall(s)
    return ips[-1] if ips else (NUM.findall(s)[-1] if NUM.findall(s) else s.strip())


def _match(pred, gold) -> bool:
    """宽松判定:gold 的关键标识(最长 IP 或末位数)是否出现在 pred 里。
    剥端口噪声(pred='3ffd::ac10:23e7:53' 对 gold='3ffd::ac10:23e7' 算命中)。"""
    if pred is None:
        return False
    g, p = str(gold), str(pred)
    gips = IP.findall(g)
    if gips:  # gold 含 IP:取最长那个,看是否是 pred 的子串
        key = max(gips, key=len)
        return key in p
    gnums = NUM.findall(g)  # gold 是纯数(端口/计数):末位数 ∈ pred 的数集
    return bool(gnums) and gnums[-1] in NUM.findall(p)


def main():
    # 各族代表:轮转/参数回显/脚本/状态
    cases = sys.argv[1:] or [
        "smoke_test/sdns/sdns_method/sdns_method.xlsx",            # ①轮转
        "smoke_test/sdns/sdns_health_check_dns/dns_link_dst_addr.xlsx",  # ②参数回显
        "smoke_test/sdns/epolicy/epolicy_1.xlsx",                  # ③脚本
        "smoke_test/sdns/failaction/failaction_2.xlsx",            # ④状态
    ]
    model = _model()
    out = open(ROOT / "runtime/logs/triage_oracle.jsonl", "w")
    for rel in cases:
        xp = CORPUS / rel
        if not xp.exists():
            print(f"skip {rel}: 不存在"); continue
        steps, golds = extract(xp)
        try:
            d = _ask(model, steps)
        except Exception as e:
            print(f"[{rel.split('/')[-1]}] LLM异常: {e}"); continue
        slots = {s["slot"]: s for s in d.get("slots", []) if "slot" in s}
        # 统计:能算族(algorithm/config_intent)推对率;脚本/先例族是否被正确"不硬推"
        n_comp = n_comp_ok = n_defer = 0
        for i, gold in enumerate(golds):
            s = slots.get(i, {})
            src = s.get("source", "missing")
            if src in ("algorithm", "config_intent"):
                n_comp += 1
                if _match(s.get("predicted"), gold):
                    n_comp_ok += 1
            elif src in ("script_runtime", "precedent_needed", "underdetermined"):
                n_defer += 1
        fam = rel.split("/")[-2]
        rec = {"case": rel, "family": fam, "n_slots": len(golds),
               "n_computable": n_comp, "computable_correct": n_comp_ok,
               "n_deferred": n_defer,
               "detail": [{"slot": i, "gold": _norm(g), "gold_raw": g[:40],
                           "src": slots.get(i, {}).get("source"),
                           "pred_raw": slots.get(i, {}).get("predicted"),
                           "ok": _match(slots.get(i, {}).get("predicted"), g)}
                          for i, g in enumerate(golds)]}
        out.write(json.dumps(rec, ensure_ascii=False) + "\n"); out.flush()
        acc = f"{n_comp_ok}/{n_comp}" if n_comp else "—"
        print(f"[{fam:20}] 断言点{len(golds):2} | 判'能算'{n_comp:2}→推对{acc} | "
              f"判'该缓推'{n_defer:2}")
    out.close()
    print(f"\n证据落: runtime/logs/triage_oracle.jsonl")


if __name__ == "__main__":
    main()
