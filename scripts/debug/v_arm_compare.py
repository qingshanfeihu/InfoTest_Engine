"""V层弱断言判据·两臂对照实验(用现有31个draft + 重解析need_intent,直接调LLM)。

留证据决定走哪条:
- Arm-A 领域judge: 现有 confidence_f.score_case(喂领域料,LLM懂信安测试再判)。
- Arm-B 结构抽象: 两步——
    step1(懂领域) 把 need_intent 分解成"可观测验证点"列表,每点标 dynamic(序列/分布)还是 static。
    step2(不懂领域) 纯覆盖判定:给验证点列表 + 断言列表(值抽象成类型token),判每个点是否被某断言"可证伪地覆盖"。
覆盖率 = 被覆盖点数 / 总点数;动态点未覆盖是硬伤。

输出 runtime/logs/v_arm_compare.jsonl,每 case 一行,含两臂完整输入输出(证据)。
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))


def _load_rows(xlsx):
    import openpyxl
    ws = openpyxl.load_workbook(xlsx, data_only=True).active
    rows = []
    for r in range(29, ws.max_row + 1):
        E = str(ws.cell(r, 5).value or "")
        F = str(ws.cell(r, 6).value or "")
        G = str(ws.cell(r, 7).value or "")
        if E or F or G:
            rows.append({"E": E, "F": F, "G": G})
    return rows


def _need_map():
    from main.ist_core.tools.device.compile_prep import qa_compile_prep
    qa_compile_prep.invoke({"mindmap_path": "workspace/inputs/automatic_case/dongkl.txt",
                            "out_name": "_recon_intent"})
    m = json.load(open(ROOT / "workspace/outputs/_recon_intent/manifest.json"))
    out = {}
    for c in m["cases"]:
        si = c.get("step_intents", [])
        need = "; ".join(f"{s.get('desc','')}→{s.get('expected','')}"
                         for s in si if s.get("desc"))
        out[c["autoid"]] = {"title": c.get("title", ""), "need": need}
    return out


def _model():
    from dotenv import load_dotenv
    env_path = ROOT / "environment"
    if env_path.exists():
        load_dotenv(str(env_path), override=False)
    from main.ist_core.agents._llm import build_agent_chat_model
    return build_agent_chat_model()


def _ask(model, sys_prompt, user, timeout=120):
    from langchain_core.messages import SystemMessage, HumanMessage
    import concurrent.futures as cf
    msgs = [SystemMessage(content=sys_prompt), HumanMessage(content=user)]
    last = ""
    for _ in range(3):  # 瞬时空响应重试
        with cf.ThreadPoolExecutor(max_workers=1) as ex:
            resp = ex.submit(model.invoke, msgs).result(timeout=timeout)
        txt = str(resp.content)
        last = txt
        m = re.search(r"\{.*\}", txt, re.S)
        if m:
            try:
                return json.loads(m.group(0)), txt
            except json.JSONDecodeError:
                continue
    raise RuntimeError(f"LLM三次未返回可解析JSON,末次={last[:120]!r}")


# ---------- Arm-A: 现有领域 judge ----------
def arm_a(rows, need, model):
    from main.case_compiler.confidence_f import score_case
    r = score_case(rows, need_intent=need, model=model)
    return {"overall": r.get("overall"), "abstain": r.get("abstain"),
            "decision": "PASS" if (not r.get("abstain") and (r.get("overall") or 0) >= 0.5) else "CUT",
            "rows": [{"g": s.cp_g, "score": s.score} for s in r.get("rows", [])]}


# ---------- Arm-B step1: 需求→可观测验证点(懂领域) ----------
_DECOMP_SYS = """你是测试需求分析师。把一条测试需求分解成"可观测验证点"列表。
每个验证点是测试中一个必须被断言确认的、可观测的事实。标注它是:
- dynamic: 依赖运行时行为/序列/分布(如"第N次请求命中第N个pool"、"按权重轮转"),必须逐跳/计数验证。
- static: 一次性配置/显示确认(如"show显示算法为rr"、"配置存在")。
只输出JSON: {"points":[{"id":1,"desc":"...","kind":"dynamic|static"}]}。不要解释。"""


def arm_b_step1(need, model):
    data, raw = _ask(model, _DECOMP_SYS, f"需求:\n{need}")
    return data.get("points", []), raw


# ---------- Arm-B step2: 覆盖判定(不懂领域,纯蕴含) ----------
_COVER_SYS = """你是逻辑覆盖判定器。给你两个列表:
A) 验证点(每个有 id/desc/kind)
B) 断言(每条是"观测算子 + 期望模式",值已抽象成类型token如<ip>/<name>/<num>)
对每个验证点判断:B中是否存在某条断言,能在该点行为发生错误时失败(可证伪地覆盖它)。
关键:dynamic点要求断言能区分"不同次/不同对象"的结果(如逐次不同命中值、计数Hit:<num>);
若只有单次存在性断言(found <ip> 一次),不足以覆盖dynamic点。
只输出JSON: {"covered":[{"id":1,"by":"断言摘要或null","ok":true/false}]}。不解释。"""


def _abstract(g):
    g = re.sub(r"\d+\.\d+\.\d+\.\d+", "<ip>", g)
    g = re.sub(r"[0-9a-fA-F:]{2,}:[0-9a-fA-F:]+", "<ip6>", g)
    g = re.sub(r"\b\d+\b", "<num>", g)
    g = re.sub(r"[\w.]+\.com\b", "<name>", g)
    return g


def arm_b_step2(points, rows, model):
    cps = [{"op": r["F"], "pat": _abstract(r["G"])} for r in rows if r["E"] == "check_point"] \
        if rows and "E" in rows[0] else \
        [{"op": r.get("F"), "pat": _abstract(r.get("G", ""))} for r in rows]
    user = "A) 验证点:\n" + json.dumps(points, ensure_ascii=False) + \
           "\n\nB) 断言:\n" + json.dumps(cps, ensure_ascii=False)
    data, raw = _ask(model, _COVER_SYS, user)
    cov = data.get("covered", [])
    return cov, cps, raw


def arm_b(rows, need, model):
    # check_point 行(带E标记的用E,否则全传)
    cp_rows = [r for r in rows if r.get("E") == "check_point"] or rows
    pts, raw1 = arm_b_step1(need, model)
    cov, cps, raw2 = arm_b_step2(pts, cp_rows, model)
    dyn = [p for p in pts if p.get("kind") == "dynamic"]
    covmap = {c["id"]: c.get("ok") for c in cov}
    dyn_uncov = [p for p in dyn if not covmap.get(p["id"])]
    n_ok = sum(1 for c in cov if c.get("ok"))
    rate = n_ok / len(pts) if pts else 0.0
    # 决策:动态点有未覆盖=CUT;否则按覆盖率
    decision = "CUT" if dyn_uncov else ("PASS" if rate >= 0.7 else "CUT")
    return {"decision": decision, "cover_rate": round(rate, 2),
            "n_points": len(pts), "n_dynamic": len(dyn), "dyn_uncovered": len(dyn_uncov),
            "points": pts, "covered": cov, "abstracted_cps": cps}


def main():
    import glob
    reps = sys.argv[1:] or [
        "203031753342777976",  # rr动态序列
        "203031754287572672",  # show展示=static
        "203031754291994838",  # mid多IP无Hit
        "203031754291994899",  # weak单IP
        "203031754291994861",  # 另一个
    ]
    needs = _need_map()
    model = _model()
    out = open("runtime/logs/v_arm_compare.jsonl", "w")
    for aid in reps:
        xp = ROOT / f"workspace/outputs/{aid}/case.xlsx"
        if not xp.exists():
            print(f"skip {aid}: no xlsx"); continue
        rows = _load_rows(str(xp))
        need = needs.get(aid, {}).get("need", "")
        title = needs.get(aid, {}).get("title", "")
        try:
            a = arm_a(rows, need, model)
        except Exception as e:
            import traceback
            a = {"error": traceback.format_exc()[-300:]}
        try:
            b = arm_b(rows, need, model)
        except Exception as e:
            import traceback
            b = {"error": traceback.format_exc()[-300:]}
        rec = {"autoid": aid, "title": title, "need": need, "arm_a": a, "arm_b": b}
        out.write(json.dumps(rec, ensure_ascii=False) + "\n"); out.flush()
        print(f"[{aid}] {title[:22]}")
        print(f"   A(领域judge): {a.get('decision')} overall={a.get('overall')}")
        print(f"   B(结构抽象): {b.get('decision')} 覆盖率={b.get('cover_rate')} 动态点={b.get('n_dynamic')} 动态未覆盖={b.get('dyn_uncovered')}")
    out.close()
    print("\n证据落: runtime/logs/v_arm_compare.jsonl")


if __name__ == "__main__":
    main()
