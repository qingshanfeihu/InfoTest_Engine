"""yzg 重编后一次性静态审计:不上机,纯查产物结构。

- emit 覆盖率(26 个 autoid 各有无 case.xlsx)
- 重启簇(668000/15/30/44):是否真无 reboot、是否用 clear→config 恢复范式
- 全体:listener/dig 目标是否落在 env ★ 可达段(Fix2 是否生效)
- 14 个 ANSWER:0 长尾:基线完整性(sdns on/host/pool/绑定/listener)
- 破坏性命令扫描(应为 0)

用法: .venv/bin/python -m scripts.debug.yzg_postcompile_audit
"""
from __future__ import annotations

import glob
import os
import re

from main.langchain_env import langchain_load_dotenv_if_present


def _steps(path):
    from openpyxl import load_workbook
    if not path or not os.path.exists(path):
        return []
    ws = load_workbook(path, data_only=True).active
    out = []
    for r in ws.iter_rows(values_only=True):
        if str(r[0] or "").startswith("999"):
            break
        E = str(r[4] or "").strip() if len(r) > 4 else ""
        F = str(r[5] or "").strip() if len(r) > 5 else ""
        G = str(r[6] or "").strip() if len(r) > 6 else ""
        if E or F:
            out.append((E, F, G))
    return out


def _dir(sid):
    g = glob.glob(f"workspace/outputs/*{sid}")
    return g[0] + "/case.xlsx" if g else None


def main():
    langchain_load_dotenv_if_present()
    from main.ist_core.tools._shared.env_facts import get_env_facts
    F = get_env_facts()
    ok = set(F.listener_ips())
    blind = set(F.unreachable_lb_ips())

    # yzg 全 autoid(短 id)
    all_ids = ["655154", "655173", "655188", "655203", "655218", "655233", "655248",
               "655262", "655276", "655290", "667986", "668000", "668015", "668030",
               "668044", "668059", "676594", "676612", "676626", "676640", "676654",
               "676668", "681539", "681556", "681571", "681588"]
    reboot_cluster = ["668000", "668015", "668030", "668044"]
    answer0_14 = ["655188", "655203", "655218", "655262", "655276", "676594", "676654",
                  "676668", "676626", "676640", "681539", "681556", "655173", "667986"]

    print("=== 1. emit 覆盖率 ===")
    missing = [s for s in all_ids if not _dir(s)]
    print(f"  {len(all_ids)-len(missing)}/{len(all_ids)} 有产物;缺: {missing or '无'}")

    print("\n=== 2. 破坏性命令全扫(应 0)===")
    destr = []
    for s in all_ids:
        st = _steps(_dir(s))
        blob = " ".join(g for _, _, g in st)
        if re.search(r"\b(reboot|reload|shutdown|halt|poweroff)\b", blob, re.I):
            destr.append(s)
    print(f"  含破坏性命令: {destr or '无 ✓'}")

    print("\n=== 3. 重启簇:无 reboot + 用 clear/no→config 恢复 ===")
    for s in reboot_cluster:
        st = _steps(_dir(s))
        if not st:
            print(f"  {s}: 无产物")
            continue
        blob = " ".join(g for _, _, g in st).lower()
        has_reboot = bool(re.search(r"\breboot\b", blob))
        has_restore = "config memory" in blob or "config file" in blob or "config all" in blob or "config net" in blob
        has_clear = "clear" in blob or "no sdns" in blob
        print(f"  {s}: reboot={has_reboot!s:5s} 清(clear/no)={has_clear!s:5s} 恢复(config)={has_restore!s:5s}")

    print("\n=== 4. 全体 listener/dig 目标可达性(Fix2)===")
    bad_reach = []
    for s in all_ids:
        st = _steps(_dir(s))
        ips = set()
        for E, Fm, G in st:
            for m in re.finditer(r"(?:listener |@)(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})", G or ""):
                ips.add(m.group(1))
        bad = [ip for ip in ips if ip in blind]
        if bad:
            bad_reach.append((s, bad))
    print(f"  落在⚠够不着段的: {bad_reach or '无 ✓(全部 ★ 可达)'}")

    print("\n=== 5. 14 个 ANSWER:0 长尾:基线完整性 ===")
    for s in answer0_14:
        st = _steps(_dir(s))
        if not st:
            print(f"  {s}: 无产物")
            continue
        blob = " ".join(g for _, _, g in st).lower()
        chain = {
            "on": "sdns on" in blob,
            "host": "sdns host name" in blob,
            "svc": "sdns service ip" in blob,
            "pool": "sdns pool name" in blob,
            "绑定": "sdns host pool" in blob,
            "listen": "sdns listener" in blob,
        }
        miss = [k for k, v in chain.items() if not v]
        fwd = "forward" in blob
        print(f"  {s}: 缺={miss or '基线完整'}{' forward' if fwd else ''}")


if __name__ == "__main__":
    raise SystemExit(main())
