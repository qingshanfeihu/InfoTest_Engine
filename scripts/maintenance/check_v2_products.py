"""v2 产物静态检查（PLAN_footprint_v2_compile.md §2.4 verify 退化态 / §六 验收）。

本轮不上机（§〇）：对已产出的 case.xlsx 做**静态结构检查**，算论文 §5.6 静态对照指标：
- 完整性：excel 含的 autoid case 数 vs 脑图 autoid 数（治旧管线 41→1 丢失）；
- 结构约束（命题3.18，确定性可查）：命令∈allowlist、断言挂观测算子（非悬空）、IP 全可达；
- 不可达 IP 编造率。

不替代 grade（V 段语义覆盖度）；只查与意图无关的结构约束。

用法：
    .venv/bin/python -m scripts.maintenance.check_v2_products workspace/outputs/<脑图名>/case.xlsx [...]
    .venv/bin/python -m scripts.maintenance.check_v2_products --glob 'workspace/outputs/*/case.xlsx'
"""

from __future__ import annotations

import argparse
import glob as _glob
import logging
import sys
from pathlib import Path

logger = logging.getLogger("check_v2_products")


def _load_rows(xlsx_path: str) -> list[dict]:
    """读 case.xlsx 数据区为 [{A,E,F,G}...]，到哨兵止。"""
    import openpyxl
    ws = openpyxl.load_workbook(xlsx_path, data_only=True).active
    rows = []
    for r in range(29, ws.max_row + 1):
        A = ws.cell(r, 1).value
        if A and str(A).startswith("999999"):
            break
        rows.append({
            "A": str(A or "").strip(),
            "E": str(ws.cell(r, 5).value or "").strip(),
            "F": str(ws.cell(r, 6).value or "").strip(),
            "G": str(ws.cell(r, 7).value or "").strip(),
        })
    return rows


def _split_cases(rows: list[dict]) -> dict[str, list[dict]]:
    """按 A 列 autoid 切分成 {autoid: [steps]}。A 列只在 case 首行有值，后续步骤继承。"""
    cases: dict[str, list[dict]] = {}
    cur = None
    for row in rows:
        if row["A"]:
            cur = row["A"]
            cases.setdefault(cur, [])
        if cur is not None:
            cases[cur].append(row)
    return cases


def check_one(xlsx_path: str) -> dict:
    """对一个 case.xlsx 算静态指标。复用 structural_gate（命令 allowlist + 断言非悬空）
    与 env_facts（IP 可达），口径与 emit 结构门一致。"""
    from main.ist_core.tools.device.structural_gate import check_structural_constraints
    from main.ist_core.tools._shared.env_facts import get_env_facts

    rows = _load_rows(xlsx_path)
    cases = _split_cases(rows)
    facts = get_env_facts()

    n_cases = len(cases)
    dangling = 0
    cmd_violations = 0
    unreachable_ips: set[str] = set()
    n_assertions = 0

    for autoid, steps in cases.items():
        # 结构约束门（命令 allowlist + 断言非悬空）
        sres = check_structural_constraints(autoid, steps)
        for v in sres.violations:
            if v.code == "dangling_assertion":
                dangling += 1
            elif v.code == "cmd_not_in_allowlist":
                cmd_violations += 1
        # IP 可达
        for s in steps:
            if s["E"] == "check_point":
                n_assertions += 1
            if (s["E"].startswith("APV") and s["F"] in ("cmd_config", "cmds_config")) \
                    or s["E"] == "test_env":
                for ip in facts.unreachable_ipv4s(s["G"]):
                    unreachable_ips.add(ip)

    return {
        "path": xlsx_path,
        "n_cases": n_cases,
        "n_assertions": n_assertions,
        "dangling_assertions": dangling,
        "cmd_not_in_allowlist": cmd_violations,
        "unreachable_ips": sorted(unreachable_ips),
        "structurally_clean": dangling == 0 and cmd_violations == 0 and not unreachable_ips,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="v2 产物静态结构检查（§5.6 静态对照）")
    parser.add_argument("paths", nargs="*", help="case.xlsx 路径（可多个）")
    parser.add_argument("--glob", default="", help="按 glob 匹配多个 case.xlsx")
    args = parser.parse_args(argv)

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    paths = list(args.paths)
    if args.glob:
        paths += sorted(_glob.glob(args.glob))
    paths = [p for p in paths if Path(p).is_file()]
    if not paths:
        logger.error("无可检查的 case.xlsx（传路径或 --glob）")
        return 1

    print("=== v2 产物静态结构检查（命题3.18 / §5.6 静态对照）===")
    all_clean = True
    total_cases = 0
    for p in paths:
        try:
            r = check_one(p)
        except Exception as exc:  # noqa: BLE001
            print(f"\n{p}\n  ERROR: {exc}")
            all_clean = False
            continue
        total_cases += r["n_cases"]
        flag = "✅" if r["structurally_clean"] else "❌"
        print(f"\n{flag} {p}")
        print(f"   case 数: {r['n_cases']}  断言数: {r['n_assertions']}")
        print(f"   悬空断言: {r['dangling_assertions']}  "
              f"越界命令: {r['cmd_not_in_allowlist']}  "
              f"不可达 IP: {r['unreachable_ips'] or '无'}")
        if not r["structurally_clean"]:
            all_clean = False

    print(f"\n=== 汇总: {len(paths)} 个 excel，{total_cases} 个 case，"
          f"结构{'全部合格 ✅' if all_clean else '存在违规 ❌'} ===")
    return 0 if all_clean else 2


if __name__ == "__main__":
    sys.exit(main())
