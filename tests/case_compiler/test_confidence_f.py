"""grade 捕获感知(confidence_f)+ 防回归。

验证:捕获+比较关系断言(found(寄存器v1)、G空H=v1)在证据里被正确呈现成"关系断言",
而非旧的 found() 空括号(那会让判官误判悬空/弱);同时非捕获的字面量断言呈现逐字不变(防回归)。
免 LLM:build_judge_evidence 是纯拼装函数,确定性可断言。
"""
from __future__ import annotations

import openpyxl

from main.case_compiler.confidence_f import build_judge_evidence, link_assertion_to_config


def test_capture_assertion_shows_register():
    """T1:捕获点呈现寄存器 + 捕获源,绝不出现 found() 空括号。"""
    rows = [
        {"E": "test_env", "F": "routera", "G": "dig @172.16.32.70 foo.com A +short", "H": "v1"},
        {"E": "test_env", "F": "routera", "G": "dig @172.16.32.70 foo.com A +short", "H": ""},
        {"E": "check_point", "F": "found", "G": "", "H": "v1"},
    ]
    ev = build_judge_evidence(rows, need_intent="会话保持:连续同池")
    assert "寄存器 v1" in ev                                   # 寄存器引用被呈现
    assert "捕获自" in ev                                       # 捕获源被溯源
    assert "dig @172.16.32.70 foo.com A +short" in ev          # 捕获自的真实命令
    assert "跨观测关系" in ev                                   # 标成关系断言
    assert "found()" not in ev                                  # 致命的空括号呈现绝不出现


def test_literal_assertion_unchanged():
    """T2 防回归:字面量断言(G有值、H空)呈现逐字 = 旧实现,无"寄存器"字样。"""
    rows = [
        {"E": "APV_0", "F": "cmd_config", "G": "show sdns host", "H": ""},
        {"E": "check_point", "F": "found", "G": "some.domain.com", "H": ""},
    ]
    ev = build_judge_evidence(rows, need_intent="x")
    assert "  [0] 断言: found(some.domain.com)" in ev           # 旧格式逐字
    assert "寄存器" not in ev                                    # 非捕获断言不沾"寄存器"措辞


def test_mixed_capture_and_literal():
    """T3:捕获点走新呈现、字面量点走老呈现,互不污染。"""
    rows = [
        {"E": "test_env", "F": "routera", "G": "dig @172.16.32.70 foo.com A +short", "H": "v1"},
        {"E": "check_point", "F": "found", "G": "", "H": "v1"},          # 捕获关系
        {"E": "APV_0", "F": "cmd_config", "G": "show sdns host", "H": ""},
        {"E": "check_point", "F": "found", "G": "NOERROR", "H": ""},     # 字面量
    ]
    ev = build_judge_evidence(rows, need_intent="x")
    assert "寄存器 v1" in ev                                     # 捕获点新呈现
    assert "断言: found(NOERROR)" in ev                          # 字面量点老呈现
    # 捕获关系的语义说明只该出现一次(只对捕获点),不污染字面量点
    assert ev.count("跨观测关系/捕获比较") == 1


def test_not_found_capture_register():
    """not_found 捕获(异池)同样被识别为关系断言。"""
    rows = [
        {"E": "test_env", "F": "routera", "G": "dig @172.16.32.70 foo.com A +short", "H": "v1"},
        {"E": "test_env", "F": "routera", "G": "dig @172.16.32.70 foo.com A +short", "H": ""},
        {"E": "check_point", "F": "not_found", "G": "", "H": "v1"},
    ]
    ev = build_judge_evidence(rows, need_intent="无保持:每次换池")
    assert "not_found(寄存器 v1)" in ev
    assert "not_found()" not in ev


def test_link_carries_cp_h_and_capture_src():
    """link_assertion_to_config 给捕获 check_point 带上 cp_h + capture_src;非捕获为空。"""
    rows = [
        {"E": "test_env", "F": "routera", "G": "dig @ip foo A +short", "H": "v1"},
        {"E": "check_point", "F": "found", "G": "", "H": "v1"},
        {"E": "check_point", "F": "found", "G": "literal", "H": ""},
    ]
    links = link_assertion_to_config(rows)
    assert links[0]["cp_h"] == "v1"
    assert "dig @ip foo A +short" in links[0]["capture_src"]
    assert links[1]["cp_h"] == ""
    assert links[1]["capture_src"] == ""


def test_load_case_rows_reads_h(tmp_path):
    """T4:生产行加载器 _load_case_rows 必须读第 8 列(H),否则 grade 永远拿不到捕获变量。"""
    from main.ist_core.tools.device.precedent_tools import _load_case_rows
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(29, 1, "111")
    ws.cell(29, 5, "test_env"); ws.cell(29, 6, "routera")
    ws.cell(29, 7, "dig @ip foo A +short"); ws.cell(29, 8, "v1")
    ws.cell(30, 5, "check_point"); ws.cell(30, 6, "found")
    ws.cell(30, 7, ""); ws.cell(30, 8, "v1")
    p = tmp_path / "case.xlsx"
    wb.save(str(p))
    rows = _load_case_rows(str(p))
    assert rows[0]["H"] == "v1"
    assert rows[1]["H"] == "v1"
    assert rows[1]["E"] == "check_point"
