"""V3 approach A：确定性编译流水线 compile_pipeline。"""

from __future__ import annotations

import importlib
import json

# 工具名 compile_pipeline 与其所在模块同名 → device 包属性被工具对象遮蔽
# （`from device import compile_pipeline` 会拿到 StructuredTool 而非模块）；
# 用 importlib 取真模块以访问内部 helper / 做 monkeypatch。
CP = importlib.import_module("main.ist_core.tools.device.compile_pipeline")


def _case():
    return {
        "autoid": "111",
        "title": "sdns host rr算法",
        "group_path": ["A组", "算法测试"],
        "step_intents": [
            {"desc": "配置pool rr算法", "expected": ""},
            {"desc": "客户端发请求", "expected": "命中第一个pool"},
        ],
    }


def test_brief_template_has_five_elements_and_no_commands():
    b = CP._build_case_brief(_case(), product_version="10.5",
                             manual_glob="10.5_cli__part*.md")
    # 五要素都在
    assert "需求：autoid=111" in b
    assert "10.5" in b
    assert "A组 / 算法测试" in b
    assert "observe-then-assert" in b   # 规则
    assert "kb_footprint" in b    # 指路
    assert "只生成 draft" in b           # 边界
    # 零硬编码：不出现具体设备命令
    assert "sdns listener" not in b
    assert "命中第一个pool" in b          # 脑图期望如实带入


def test_version_guard():
    r = CP.compile_pipeline.invoke({"mindmap_path": "x.txt", "product_version": ""})
    assert r.startswith("error") and "product_version" in r
    r2 = CP.compile_pipeline.invoke({"mindmap_path": "", "product_version": "10.5"})
    assert r2.startswith("error") and "mindmap_path" in r2


def test_pipeline_per_case_no_barrier(monkeypatch):
    """per-case 流水线：每 case 独立 draft→grade，无屏障。mock execute_fork_skill + prep + merge。

    验证：draft 落 xlsx、grade 判 PASS、PASS 进 merge、无 escalated。
    """
    calls = []
    root = CP._project_root()
    outdir = root / "workspace" / "outputs" / "t_pipe"
    outdir.mkdir(parents=True, exist_ok=True)
    (outdir / "manifest.json").write_text(json.dumps({
        "cases": [_case(), {**_case(), "autoid": "222", "title": "case2"}],
        "groups": {},
    }, ensure_ascii=False), encoding="utf-8")

    class _Fake:
        def __init__(self, fn): self._fn = fn
        def invoke(self, d): return self._fn(d)

    def fake_prep(d):
        calls.append(("prep", d.get("mindmap_path")))
        return "manifest 已产"

    def fake_merge(d):
        cs = json.loads(d["cases_json"]); calls.append(("merge", len(cs)))
        return f"已产出 {len(cs)} case"

    # execute_fork_skill(skill, brief)：draft brief 含"需求："→落 xlsx 返回路径；
    # grade brief 含"xlsx_path="→返回 PASS。
    def fake_fork(skill, brief, tag="", summary_sink=None):
        if skill == "ist-compile-draft":
            # 从 brief 提 autoid
            aid = brief.split("autoid=")[1].split("，")[0].strip()
            calls.append(("draft", aid))
            cdir = root / "workspace" / "outputs" / aid
            cdir.mkdir(parents=True, exist_ok=True)
            _write_min_xlsx(cdir / "case.xlsx", aid)
            return f"xlsx: workspace/outputs/{aid}/case.xlsx"
        else:
            calls.append(("grade", skill))
            return "VERDICT: PASS 覆盖目标行为"

    prep_mod = importlib.import_module("main.ist_core.tools.device.compile_prep")
    import main.ist_core.tools.device.emit_xlsx_tool as em
    import main.ist_core.skills.loader as loader_mod
    monkeypatch.setattr(prep_mod, "compile_prep", _Fake(fake_prep))
    monkeypatch.setattr(em, "compile_emit_merged", _Fake(fake_merge))
    monkeypatch.setattr(loader_mod, "execute_fork_skill", fake_fork)

    res = CP._run_pipeline("dongkl.txt", "10.5", "t_pipe",
                           draft_skill="ist-compile-draft", grade_skill="ist-compile-grade")
    # 两 case 各 draft 一次 + grade 一次，全 PASS → merge 2
    assert ("prep", "dongkl.txt") in calls
    assert ("draft", "111") in calls and ("draft", "222") in calls
    assert calls.count(("grade", "ist-compile-grade")) == 2
    assert ("merge", 2) in calls
    assert sorted(res["done"]) == ["111", "222"]
    assert not res["escalated"]


def test_pipeline_observability_aggregates_llm_and_tool_calls(monkeypatch):
    """Phase 0：pipeline 聚合每 case draft/grade fork 的 LLM 往返 + 工具调用次数。

    mock 的 execute_fork_skill 经 summary_sink 回传可观测指标（模拟真 fork 的统计），
    验证 result['observability'] 汇总正确——这是「预检索是否真减少 LLM 调用/查找」的度量。
    """
    root = CP._project_root()
    outdir = root / "workspace" / "outputs" / "t_obs"
    outdir.mkdir(parents=True, exist_ok=True)
    (outdir / "manifest.json").write_text(json.dumps({
        "cases": [_case(), {**_case(), "autoid": "222", "title": "case2"}],
        "groups": {},
    }, ensure_ascii=False), encoding="utf-8")

    class _Fake:
        def __init__(self, fn): self._fn = fn
        def invoke(self, d): return self._fn(d)

    def fake_fork(skill, brief, tag="", summary_sink=None):
        if skill == "ist-compile-draft":
            aid = brief.split("autoid=")[1].split("，")[0].strip()
            cdir = root / "workspace" / "outputs" / aid
            cdir.mkdir(parents=True, exist_ok=True)
            _write_min_xlsx(cdir / "case.xlsx", aid)
            if summary_sink is not None:   # draft fork：5 轮 LLM + probe/footprint 各几次
                summary_sink.update({"ai_rounds": 5,
                                     "tool_calls": {"dev_probe": 2, "kb_footprint": 3}})
            return f"xlsx: {aid}"
        if summary_sink is not None:       # grade fork：2 轮 LLM + 1 次 compile_score
            summary_sink.update({"ai_rounds": 2, "tool_calls": {"compile_score": 1}})
        return "判定：PASS"

    prep_mod = importlib.import_module("main.ist_core.tools.device.compile_prep")
    import main.ist_core.tools.device.emit_xlsx_tool as em
    import main.ist_core.skills.loader as loader_mod
    monkeypatch.setattr(prep_mod, "compile_prep", _Fake(lambda d: "ok"))
    monkeypatch.setattr(em, "compile_emit_merged", _Fake(lambda d: "merged"))
    monkeypatch.setattr(loader_mod, "execute_fork_skill", fake_fork)

    res = CP._run_pipeline("x.txt", "10.5", "t_obs",
                           draft_skill="ist-compile-draft", grade_skill="ist-compile-grade")
    obs = res["observability"]["total"]
    # 2 case，各 draft(ai=5)+grade(ai=2) → draft_llm=10 grade_llm=4 合计=14
    assert obs["draft_llm_rounds"] == 10
    assert obs["grade_llm_rounds"] == 4
    assert obs["total_llm_rounds"] == 14
    # 工具：dev_probe 2×2=4, kb_footprint 3×2=6, compile_score 1×2=2
    assert obs["tool_calls"]["dev_probe"] == 4
    assert obs["tool_calls"]["kb_footprint"] == 6
    assert obs["tool_calls"]["compile_score"] == 2
    assert set(res["observability"]["per_case"].keys()) == {"111", "222"}
    # 渲染行：要削减的成本键(dev_probe/kb_footprint)排在普通工具前 + 进 phases
    line = CP._format_observability(obs)
    assert "dev_probe=4" in line and "kb_footprint=6" in line
    assert line.index("dev_probe") < line.index("compile_score")
    assert any("观测(LLM/查找成本" in p for p in res["phases"])


def test_pipeline_escalates_after_max_rounds(monkeypatch):
    """grade 持续 CUT → ≤N 轮后 escalate，不进 merge。"""
    root = CP._project_root()
    outdir = root / "workspace" / "outputs" / "t_esc"
    outdir.mkdir(parents=True, exist_ok=True)
    (outdir / "manifest.json").write_text(json.dumps({"cases": [_case()], "groups": {}},
                                                     ensure_ascii=False), encoding="utf-8")
    n_draft = {"n": 0}

    class _Fake:
        def __init__(self, fn): self._fn = fn
        def invoke(self, d): return self._fn(d)

    def fake_fork(skill, brief, tag="", summary_sink=None):
        if skill == "ist-compile-draft":
            n_draft["n"] += 1
            aid = brief.split("autoid=")[1].split("，")[0].strip()
            cdir = root / "workspace" / "outputs" / aid
            cdir.mkdir(parents=True, exist_ok=True)
            _write_min_xlsx(cdir / "case.xlsx", aid)
            return f"xlsx: {aid}"
        return "VERDICT: CUT 断言太弱，未覆盖动态行为"

    prep_mod = importlib.import_module("main.ist_core.tools.device.compile_prep")
    import main.ist_core.tools.device.emit_xlsx_tool as em
    import main.ist_core.skills.loader as loader_mod
    monkeypatch.setattr(prep_mod, "compile_prep", _Fake(lambda d: "ok"))
    monkeypatch.setattr(em, "compile_emit_merged", _Fake(lambda d: "merged"))
    monkeypatch.setattr(loader_mod, "execute_fork_skill", fake_fork)

    res = CP._run_pipeline("x.txt", "10.5", "t_esc",
                           draft_skill="ist-compile-draft", grade_skill="ist-compile-grade")
    assert res["done"] == []
    assert "111" in res["escalated"]
    assert n_draft["n"] == CP._MAX_REWORK_ROUNDS  # 重做到上限


def _setup_single_case(monkeypatch, out_name, fake_fork):
    """公共脚手架：单 case manifest + mock prep/merge/fork，返回 _run_pipeline 结果。"""
    root = CP._project_root()
    outdir = root / "workspace" / "outputs" / out_name
    outdir.mkdir(parents=True, exist_ok=True)
    (outdir / "manifest.json").write_text(json.dumps({"cases": [_case()], "groups": {}},
                                                     ensure_ascii=False), encoding="utf-8")

    class _Fake:
        def __init__(self, fn): self._fn = fn
        def invoke(self, d): return self._fn(d)

    prep_mod = importlib.import_module("main.ist_core.tools.device.compile_prep")
    import main.ist_core.tools.device.emit_xlsx_tool as em
    import main.ist_core.skills.loader as loader_mod
    monkeypatch.setattr(prep_mod, "compile_prep", _Fake(lambda d: "ok"))
    monkeypatch.setattr(em, "compile_emit_merged", _Fake(lambda d: "merged"))
    monkeypatch.setattr(loader_mod, "execute_fork_skill", fake_fork)
    return CP._run_pipeline("x.txt", "10.5", out_name,
                            draft_skill="ist-compile-draft", grade_skill="ist-compile-grade")


def test_draft_recursion_escalates_immediately_no_rework(monkeypatch):
    """Fix E：draft 返回 [recursion-limit] → 恰好 1 次 draft 尝试即 escalate，不做 3 轮等价重做。"""
    n_draft = {"n": 0}

    def fake_fork(skill, brief, tag="", summary_sink=None):
        if skill == "ist-compile-draft":
            n_draft["n"] += 1
            return ("ERROR: fork skill 'ist-compile-draft' execution failed: "
                    "[recursion-limit] Recursion limit of 200 reached")
        return "判定：PASS"

    res = _setup_single_case(monkeypatch, "t_recursion", fake_fork)
    assert res["done"] == []
    assert "111" in res["escalated"]
    assert n_draft["n"] == 1                      # 不再 _MAX_REWORK_ROUNDS 轮等价重做
    reason = res["escalated"]["111"]
    assert reason["rounds"][0]["verdict"] == "DRAFT_RECURSION"
    assert "递归" in reason["summary"]


def test_fork_wallclock_timeout_escalates_and_not_transient(monkeypatch):
    """Fix A：单 fork 超 _FORK_WALLCLOCK_S → 返回 [fork-wallclock]、escalate；该标记非 transient。"""
    import time as _t
    monkeypatch.setattr(CP, "_FORK_WALLCLOCK_S", 0.2)
    n_draft = {"n": 0}

    def fake_fork(skill, brief, tag="", summary_sink=None):
        if skill == "ist-compile-draft":
            n_draft["n"] += 1
            _t.sleep(1.0)                          # 超 0.2s 墙钟 → 看门狗放弃等待
            return "xlsx: never"
        return "判定：PASS"

    res = _setup_single_case(monkeypatch, "t_wallclock", fake_fork)
    assert res["done"] == []
    assert "111" in res["escalated"]
    assert n_draft["n"] == 1                      # wallclock 也立即 escalate，不重做
    # [fork-wallclock] 标记不应被判 transient（否则会被 4× 重试放大）
    from main.ist_core.resilience import is_transient_error
    assert not is_transient_error(f"ERROR: [fork-wallclock] 超 {int(CP._FORK_WALLCLOCK_S)}s")


def test_fork_run_token_propagates_into_watchdog_thread(monkeypatch):
    """Fix A 回归红线：copy_context 把 _current_run_token 带进看门狗线程，dev_probe single-flight 不破。"""
    from main.ist_core.tools.device import run_case
    seen = {}

    def fake_fork(skill, brief, tag="", summary_sink=None):
        seen[skill] = run_case._current_run_token.get()   # 看门狗线程内读 contextvar
        if skill == "ist-compile-draft":
            aid = brief.split("autoid=")[1].split("，")[0].strip()
            cdir = CP._project_root() / "workspace" / "outputs" / aid
            cdir.mkdir(parents=True, exist_ok=True)
            _write_min_xlsx(cdir / "case.xlsx", aid)
            return f"xlsx: {aid}"
        return "判定：PASS"

    res = _setup_single_case(monkeypatch, "t_token", fake_fork)
    assert res["done"] == ["111"]
    assert seen.get("ist-compile-draft", "").startswith("run-")   # 非 None、是本 run token


def test_transient_retry_capped_by_total_wallclock(monkeypatch):
    """Fix B：transient ERROR 在总墙钟超时即停退，不跑满 _TRANSIENT_RETRIES+1 次。"""
    monkeypatch.setattr(CP, "_FORK_TRANSIENT_WALLCLOCK_S", 0.0)   # 总墙钟立即到点
    monkeypatch.setattr(CP, "_TRANSIENT_BASE_SLEEP", 0.0)
    n_draft = {"n": 0}

    def fake_fork(skill, brief, tag="", summary_sink=None):
        if skill == "ist-compile-draft":
            n_draft["n"] += 1
            return "ERROR: fork skill 'ist-compile-draft' execution failed: Request timed out"
        return "判定：PASS"

    res = _setup_single_case(monkeypatch, "t_transcap", fake_fork)
    # 每轮 _fork_call 因总墙钟=0 只做 1 次尝试（不再 transient 重试 4 次）；DRAFT_FAIL 走满 N 轮重做。
    # 有 Fix B：N 轮 × 1 = N 次；无 Fix B：N 轮 × (4 重试+1) = N×5 次。断言落在前者。
    assert n_draft["n"] == CP._MAX_REWORK_ROUNDS
    assert n_draft["n"] < CP._MAX_REWORK_ROUNDS * (CP._TRANSIENT_RETRIES + 1)
    assert "111" in res["escalated"]


def _write_min_xlsx(path, autoid):
    """写一个最小的、_load_case_rows 能读的 case.xlsx（数据区从 29 行起，含 check_point）。"""
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active
    ws.cell(29, 1, autoid)
    # case 意图是 "rr算法"，配置区须真配 rr method（满足 B 层变体保真门，否则判"算法缺配"回流）
    ws.cell(29, 5, "APV_0"); ws.cell(29, 6, "cmds_config")
    ws.cell(29, 7, "sdns on\nsdns host method www.x.com rr")
    ws.cell(30, 5, "APV_0"); ws.cell(30, 6, "cmd_config"); ws.cell(30, 7, "show sdns listener")
    ws.cell(31, 5, "check_point"); ws.cell(31, 6, "found"); ws.cell(31, 7, "172.16.34.1")
    wb.save(path)


# --- grade 裁定解析：取最后裁定词，治"讨论工具CUT后自己判PASS"被误读 ---
def test_verdict_clean_pass():
    assert CP._parse_grade_verdict("## 审批结论：PASS 覆盖目标行为") is True


def test_verdict_clean_cut():
    assert CP._parse_grade_verdict("## 审批结论：CUT 断言太弱") is False


def test_verdict_discuss_cut_then_conclude_pass():
    """grade 先复核 confidence_score 的 CUT，再下 PASS 结论 → 应判 PASS（取末位）。"""
    out = ("compile_score 给 overall=0.2 / CUT，但我批判性看待——"
           "它孤立评估断言，忽略了序列语义。\n\n## 最终裁定：PASS")
    assert CP._parse_grade_verdict(out) is True


def test_verdict_discuss_pass_then_conclude_cut():
    """反向：先说某条像 PASS，最终结论 CUT → 应判 CUT。"""
    out = ("断言1看似 PASS，但整体未覆盖动态行为。\n\n## 最终裁定：CUT 需补命中断言")
    assert CP._parse_grade_verdict(out) is False


def test_verdict_error_is_not_pass():
    assert CP._parse_grade_verdict("ERROR: fork 超时") is False


def test_verdict_neither_word_is_not_pass():
    assert CP._parse_grade_verdict("（grade 没给明确裁定）") is False


def test_verdict_cut_then_trailing_pass_mention_is_cut():
    """fix B：CUT 结论后跟"…改成 X 才能 PASS"的重做意见——结构化标记须判 CUT，
    不被朴素 rfind 把末尾的 PASS 误读成通过、放弱断言进 merge。"""
    out = ("断言只 found 了一个域名，没覆盖计数行为。\n\n"
           "## 判定：CUT\n建议改成 found_times 计数断言，才能达到 PASS 标准。")
    assert CP._parse_grade_verdict(out) is False


def test_verdict_marker_pass_takes_precedence():
    """有结构化标记时以标记为准（即便后文又提了 CUT 词）。"""
    out = "## 判定：PASS\n（注意别再写成 CUT 式弱断言）"
    assert CP._parse_grade_verdict(out) is True


def test_load_case_rows_preserves_i_column_for_found_times(tmp_path):
    """fix A：_load_case_rows 必须读回第 9 列(I)，否则 merge 丢 found_times 次数 / input_var。"""
    import openpyxl
    from main.ist_core.tools.device.precedent_tools import _load_case_rows
    wb = openpyxl.Workbook(); ws = wb.active
    ws.cell(29, 1, "AID1")
    ws.cell(29, 5, "APV_0"); ws.cell(29, 6, "cmd_config"); ws.cell(29, 7, "show sdns listener")
    ws.cell(30, 5, "check_point"); ws.cell(30, 6, "found_times")
    ws.cell(30, 7, "sdns listener"); ws.cell(30, 9, "16")   # I 列=次数
    p = tmp_path / "case.xlsx"; wb.save(p)
    rows = _load_case_rows(str(p))
    ft = [r for r in rows if r.get("F") == "found_times"]
    assert ft and ft[0].get("I") == "16", f"I 列(found_times 次数)丢失: {rows}"


def test_emit_rejects_found_times_mandatory(tmp_path, monkeypatch):
    """found_times 在 emit 被**无条件**拒绝(A 层机械崩溃门,不受 strict_structural opt-in)。

    取代旧「fix A：合并保留 found_times 次数」——found_times 框架 xlsx 分派只传 2 参、缺 times →
    TypeError 崩整份文件、崩溃点后全 case unknown(dongkl 实测漏网 6 行崩了 31 个 case)。这是**机械
    可判、误判即真错**的语法错级问题(等同 SWE-agent linter),根本不该进 excel。
    (语义"某 claim 可不可证伪"归 verifiability 工具 + LLM,**不在此门**;见
    structural_gate._check_no_found_times。)
    """
    import json as _json
    from main.ist_core.tools.device.emit_xlsx_tool import compile_emit
    steps = [
        {"E": "APV_0", "F": "cmd_config", "G": "show sdns listener", "desc": "查看"},
        {"E": "check_point", "F": "found_times", "G": "sdns listener", "I": "16", "desc": "断言16条"},
    ]
    # strict_structural 默认 False——门仍应无条件拦下(证明与 opt-in 解耦)
    out = compile_emit.invoke({"autoid": "AIDX", "steps_json": _json.dumps(steps, ensure_ascii=False),
                               "init_commands": "sdns on", "out_name": "_pytest_ft_reject"})
    assert out.startswith("error"), out
    assert "found_times" in out, out
    assert ("崩整份文件" in out or "TypeError" in out or "不支持" in out), out   # 机械崩溃理由
    assert ("found" in out and "abs_found" in out), out                        # 指向合法替代形态
    import shutil
    shutil.rmtree(CP._project_root() / "workspace" / "outputs" / "_pytest_ft_reject", ignore_errors=True)


def test_emit_rejects_dangling_assertion_mandatory(tmp_path, monkeypatch):
    """悬空断言在 emit 被**无条件**拒绝(必崩门集合,不受 strict_structural opt-in)。

    实证(dongkl 778012 重编版):配置步 cmds_config 后直接 check_point、前面无任何 result 生产步 →
    框架 found(expect, result=None) 抛 TypeError 崩整份文件 → 第三轮上机 1 pass + 33 unknown。
    检查本身早已在 _check_dangling_assertions,但躲在 strict_structural(默认 False)后面——worker
    漏传参数即漏网。本测试钉死:**不传 strict_structural 也必须拦**。
    """
    import json as _json
    from main.ist_core.tools.device.emit_xlsx_tool import compile_emit
    steps = [
        {"E": "APV_0", "F": "cmds_config", "G": "sdns on\nsdns pool name pool_v4", "desc": "初始化配置"},
        {"E": "check_point", "F": "found", "G": "pool_v4", "desc": "验证配置生效(悬空:前面无观测步)"},
    ]
    out = compile_emit.invoke({"autoid": "AIDD", "steps_json": _json.dumps(steps, ensure_ascii=False),
                               "init_commands": "sdns on", "out_name": "_pytest_dangling_reject"})
    assert out.startswith("error"), out
    assert "dangling_assertion" in out, out
    assert ("崩溃整份文件" in out or "TypeError" in out), out   # 机械崩溃理由
    assert "观测步" in out, out                                  # 修法指引(紧前放不带 H 观测步)
    # 对照:断言紧前有 cmd_config 观测步(产 result 回显) → 应通过必崩门
    steps_ok = [
        {"E": "APV_0", "F": "cmds_config", "G": "sdns on\nsdns pool name pool_v4", "desc": "初始化配置"},
        {"E": "APV_0", "F": "cmd_config", "G": "show sdns pool", "desc": "观测:产生可检查回显"},
        {"E": "check_point", "F": "found", "G": "pool_v4", "desc": "验证配置生效"},
    ]
    out_ok = compile_emit.invoke({"autoid": "AIDD2", "steps_json": _json.dumps(steps_ok, ensure_ascii=False),
                                  "init_commands": "sdns on", "out_name": "_pytest_dangling_ok"})
    assert "dangling_assertion" not in out_ok, out_ok
    import shutil
    for d in ("_pytest_dangling_reject", "_pytest_dangling_ok"):
        shutil.rmtree(CP._project_root() / "workspace" / "outputs" / d, ignore_errors=True)


def test_emit_rejects_short_numeric_autoid():
    """autoid 短号格式门:纯数字但 <15 位=完整 autoid 的尾段缩写,拒绝(OBS-19)。

    实证:deepseek 重编时 main brief 用短号习惯 → worker `compile_emit(autoid='778012')` →
    短号烧进 xlsx ID 列 → 需求关联/框架报告 ID 断链。非纯数字 id(如测试用 'AIDX')不受限。
    """
    import json as _json
    from main.ist_core.tools.device.emit_xlsx_tool import compile_emit
    steps = [
        {"E": "APV_0", "F": "cmd_config", "G": "show sdns pool", "desc": "观测"},
        {"E": "check_point", "F": "found", "G": "pool_v4", "desc": "断言"},
    ]
    out = compile_emit.invoke({"autoid": "778012", "steps_json": _json.dumps(steps, ensure_ascii=False),
                               "init_commands": "sdns on", "out_name": "_pytest_shortid_reject"})
    assert out.startswith("error"), out
    assert "短号" in out and "18 位" in out, out
    # 完整 18 位数字 autoid 与字母混合测试 id 都放行(不触发本门)
    for ok_id in ("203031753342778012", "AIDX"):
        out_ok = compile_emit.invoke({"autoid": ok_id, "steps_json": _json.dumps(steps, ensure_ascii=False),
                                      "init_commands": "sdns on", "out_name": "_pytest_shortid_ok"})
        assert "短号" not in out_ok, out_ok
    import shutil
    for d in ("_pytest_shortid_reject", "_pytest_shortid_ok"):
        shutil.rmtree(CP._project_root() / "workspace" / "outputs" / d, ignore_errors=True)


def test_emit_rejects_manual_ip_cleanup_in_test_env():
    """test_env 变更触发机 IP(ip addr add/del)无条件拒绝(必崩门第三条,OBS-17 根因)。

    框架 ssh_server 对 add 自动记账(不去重)并在下一 case 开头 delete 恢复:
    - 自行 del → 恢复对已删 IP delete 报错 → 崩整份文件(dongkl_final 22 unknown 级联);
    - 跨 case 重复 add 同一 IP → 恢复第二条必失败 → RTNETLINK 残留污染后续 case 回显
      (第五轮 8/9、第七轮 7/8 成批假 fail)。
    触发机网络状态归框架管理;轮转按请求轮转,无需多源 IP。add 与 del 均拒。
    """
    import json as _json
    from main.ist_core.tools.device.emit_xlsx_tool import compile_emit
    for tag, cmd in (("del", "ip addr del 172.16.34.210/24 dev ens192"),
                     ("add", "ip addr add 172.16.34.210/24 dev ens192")):
        steps = [
            {"E": "test_env", "F": "routera", "G": cmd, "desc": "变更触发机IP"},
            {"E": "test_env", "F": "routera", "G": "dig @172.16.34.70 x.com", "desc": "触发"},
            {"E": "check_point", "F": "found", "G": "x", "desc": "断言"},
        ]
        out = compile_emit.invoke({"autoid": "AIDIP", "steps_json": _json.dumps(steps, ensure_ascii=False),
                                   "init_commands": "sdns on", "out_name": f"_pytest_ip{tag}_reject"})
        assert out.startswith("error"), (tag, out)
        assert "manual_ip_cleanup" in out and "假 fail" in out, (tag, out)
    # 普通 test_env(dig,不动 IP)不触发
    steps_ok = [
        {"E": "test_env", "F": "routera", "G": "dig @172.16.34.70 x.com", "desc": "触发"},
        {"E": "check_point", "F": "found", "G": "x", "desc": "断言"},
    ]
    out_ok = compile_emit.invoke({"autoid": "AIDIP2", "steps_json": _json.dumps(steps_ok, ensure_ascii=False),
                                  "init_commands": "sdns on", "out_name": "_pytest_ipok"})
    assert "manual_ip_cleanup" not in out_ok, out_ok
    import shutil
    for d in ("_pytest_ipdel_reject", "_pytest_ipadd_reject", "_pytest_ipok"):
        shutil.rmtree(CP._project_root() / "workspace" / "outputs" / d, ignore_errors=True)


def test_emit_rejects_empty_or_literal_backslash_n_commands():
    """命令载荷完整性门(必崩门第四条,第八轮实证):G 空/None 与字面 \\n 必拒。

    LLM 重编 steps_json 质量问题曾批量漏进 xlsx——"None" 被发给设备、多命令拼一行,
    11 个 case 全部被设备 ^ 拒。载荷形态机械可判,与命令内容无关。
    """
    import json as _json
    from main.ist_core.tools.device.emit_xlsx_tool import compile_emit
    base_cp = {"E": "check_point", "F": "found", "G": "x", "desc": "断言"}
    obs = {"E": "APV_0", "F": "cmd_config", "G": "show version", "desc": "观测"}
    # G=None(json null) 与字面 "None" → 拒；纯空串占位步 → 放行(框架零循环无害)
    out = compile_emit.invoke({"autoid": "AIDN", "steps_json": _json.dumps(
        [{"E": "APV_0", "F": "cmds_config", "G": None, "desc": "空配置"}, obs, base_cp]),
        "init_commands": "sdns on", "out_name": "_pytest_gnone"})
    assert out.startswith("error") and "empty_command_payload" in out, out
    out_lit = compile_emit.invoke({"autoid": "AIDN2", "steps_json": _json.dumps(
        [{"E": "APV_0", "F": "cmds_config", "G": "None", "desc": "字面None"}, obs, base_cp]),
        "init_commands": "sdns on", "out_name": "_pytest_gnone2"})
    assert "empty_command_payload" in out_lit, out_lit
    out_empty = compile_emit.invoke({"autoid": "AIDN3", "steps_json": _json.dumps(
        [{"E": "APV_0", "F": "cmds_config", "G": "", "desc": "空占位"}, obs, base_cp]),
        "init_commands": "sdns on", "out_name": "_pytest_gempty"})
    assert "empty_command_payload" not in out_empty, out_empty
    # G 含字面 \n（JSON 里双反斜杠 → 解析后是字面反斜杠+n）→ **自动纠正为真实换行**
    # (2026-07-04 V轮 token 取证:拒绝打回一轮 worker+grade 重做 ≈1-2M token,而命令语境
    # 字面 \n 只可能是想换行写错了,无损替换 + 返回注明即可,17 卷的重做轮直接消失)
    out2 = compile_emit.invoke({"autoid": "AIDBN", "steps_json":
        '[{"E":"APV_0","F":"cmds_config","G":"cmd a\\\\ncmd b","desc":"拼行"},'
        '{"E":"APV_0","F":"cmd_config","G":"show version","desc":"观测"},'
        '{"E":"check_point","F":"found","G":"x","desc":"断言"}]',
        "init_commands": "sdns on", "out_name": "_pytest_gbn"})
    assert "已产出" in out2 and "自动把" in out2, out2
    # 真实换行的多命令 → 放行
    out3 = compile_emit.invoke({"autoid": "AIDOK", "steps_json": _json.dumps(
        [{"E": "APV_0", "F": "cmds_config", "G": "cmd a\ncmd b", "desc": "多命令"}, obs, base_cp]),
        "init_commands": "sdns on", "out_name": "_pytest_gok"})
    assert "empty_command_payload" not in out3 and "literal_backslash_n" not in out3, out3
    import shutil
    for d in ("_pytest_gnone", "_pytest_gnone2", "_pytest_gempty", "_pytest_gbn", "_pytest_gok"):
        shutil.rmtree(CP._project_root() / "workspace" / "outputs" / d, ignore_errors=True)
