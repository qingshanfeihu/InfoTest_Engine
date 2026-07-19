# -*- coding: utf-8 -*-
"""#61 003 device-verified 根因:emit 期全角逗号 `，` → 半角 `,` 归一化(correct-by-construction)。

003 escalated 真凶(case.xlsx openpyxl 亲验 + dev_run_case 设备真错):worker 把 importKey/
importCert/importRootCA 的 data 字段写成全角逗号(`vh1， cert/...`)。框架 get_parameter 按**半角**
`,` 拆位置参(test_xlsx.py:307-311),全角 `，` 不被识别为分隔符 → 整串当一参 → `importKey() missing
keyfile` → TypeError → 整案 not_run。全角逗号在 dispatch G(命令/参数分隔位)永不合法(设备 CLI/文件
路径/vhost 名全 ASCII)→ 无歧义 auto-normalize(区别 #56 模糊动作名 hint-not-rewrite)。

归一化与字面 `\\n` 纠正同址(compile_emit init_rows 与各 gate 之前),覆盖 init_g + 非 check_point 步 G;
check_point 断言 pattern 匹配设备回显、保守不动。落 signal 可观测、非静默。
"""
from __future__ import annotations

import shutil

import openpyxl
import pytest

from main.ist_core.tools.device.emit_xlsx_tool import compile_emit
from main.ist_core.compile_engine_v8 import _shared as _sh
from main.ist_core.memory.footprint import signals

_A = "203099999999900361"


@pytest.fixture(autouse=True)
def _relax_and_clean(monkeypatch):
    # 隔离归一化于门噪声:命令存在性门关(SSL 命令是否入手册非本测关注)、provenance 可选
    monkeypatch.setenv("IST_COMMAND_EXISTENCE_GATE", "0")
    monkeypatch.setenv("IST_PROVENANCE_OPTIONAL", "1")
    shutil.rmtree(_sh.outputs_root() / _A, ignore_errors=True)
    yield
    shutil.rmtree(_sh.outputs_root() / _A, ignore_errors=True)


def _emit_and_load(steps, init=""):
    out = compile_emit.func(_A, steps=steps, init_commands=init)
    xp = _sh.outputs_root() / _A / "case.xlsx"
    assert xp.exists(), f"emit 未产出 xlsx:{out[:300]}"
    ws = openpyxl.load_workbook(xp, data_only=True).active
    return ws


def _dispatch_cells(ws):
    """dispatch 列(G=第7列)的字符串值——排除模板描述列(c4/c5 中文散文合法含全角逗号)。"""
    return [ws.cell(r, 7).value for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, 7).value, str)]


def test_fullwidth_comma_in_import_step_normalized():
    """003 直接形态:importKey/importCert data 的全角逗号 → 半角(否则 get_parameter 崩)。"""
    steps = [
        {"E": "APV_0", "F": "cmd_config", "G": "ssl host virtual vh1", "desc": "vhost"},
        {"E": "APV_0", "F": "importKey", "G": "vh1， cert/rsaca/1024rsa.key", "desc": "私钥(全角)"},
        {"E": "APV_0", "F": "importCert", "G": "vh1， cert/rsaca/1024rsa.crt", "desc": "证书(全角)"},
        {"E": "APV_0", "F": "cmd_config", "G": "show ssl certificate vh1 simple", "desc": "查看"},
        {"E": "check_point", "F": "found", "G": "Subject:", "desc": "断言"},
        {"E": "APV_0", "F": "cmd_config", "G": "clear ssl host vh1", "desc": "清理"},
    ]
    cells = _dispatch_cells(_emit_and_load(steps))
    assert any("vh1, cert/rsaca/1024rsa.key" == c for c in cells), "importKey G 全角逗号未归一化为半角"
    assert any("vh1, cert/rsaca/1024rsa.crt" == c for c in cells), "importCert G 全角逗号未归一化为半角"
    assert not any("，" in c for c in cells), "dispatch G 列不应残留全角逗号"


def test_fullwidth_comma_in_init_normalized():
    """init(cmds_config 文件级前置)的全角逗号同样归一化(init 各行也过 get_parameter)。"""
    # 复用过门的完整 case 结构(config-write + 断言 + teardown),init 带全角逗号
    steps = [
        {"E": "APV_0", "F": "cmd_config", "G": "ssl host virtual vh1", "desc": "vhost"},
        {"E": "APV_0", "F": "cmd_config", "G": "show ssl certificate vh1 simple", "desc": "查看"},
        {"E": "check_point", "F": "found", "G": "Subject:", "desc": "断言"},
        {"E": "APV_0", "F": "cmd_config", "G": "clear ssl host vh1", "desc": "清理"},
    ]
    ws = _emit_and_load(steps, init="ssl host virtual vh0， cert/x.key")
    # init 落 cmds_config 行(列位不固定),全表扫含 vh0 的单元确认归一化
    vh0_cells = [ws.cell(r, c).value
                 for r in range(1, ws.max_row + 1) for c in range(1, ws.max_column + 1)
                 if isinstance(ws.cell(r, c).value, str) and "vh0" in ws.cell(r, c).value]
    assert vh0_cells, "未找到 init(vh0)行"
    assert all("，" not in v for v in vh0_cells), f"init 全角逗号未归一化:{vh0_cells}"
    assert any("vh0, cert" in v for v in vh0_cells), "init 全角逗号未变半角"


def test_checkpoint_pattern_comma_preserved():
    """作用域边界:check_point 断言 pattern 的全角逗号**保守不动**(它匹配设备回显、非 dispatch 参数)。

    团队裁决「归一化 data 字段」——data=dispatch G(命令/参数),不含断言 pattern。锁边界防未来误扩。"""
    steps = [
        {"E": "APV_0", "F": "cmd_config", "G": "show ssl certificate vh1", "desc": "查看"},
        {"E": "check_point", "F": "found", "G": "CN=a，b", "desc": "断言(全角逗号,保守保留)"},
    ]
    cells = _dispatch_cells(_emit_and_load(steps))
    assert any("CN=a，b" == c for c in cells), "check_point pattern 的全角逗号不应被归一化(作用域边界)"


def test_healthy_input_no_change():
    """健康输入(无全角逗号)不改动:半角逗号原样保留,归一化恒等。"""
    steps = [
        {"E": "APV_0", "F": "importKey", "G": "vh1, cert/rsaca/1024rsa.key", "desc": "半角(正常)"},
        {"E": "APV_0", "F": "cmd_config", "G": "show ssl certificate vh1", "desc": "查看"},
        {"E": "check_point", "F": "found", "G": "Subject:", "desc": "断言"},
    ]
    cells = _dispatch_cells(_emit_and_load(steps))
    assert any("vh1, cert/rsaca/1024rsa.key" == c for c in cells), "半角逗号正常输入应原样保留"


def test_signal_registered_and_emitted(monkeypatch, tmp_path):
    """归一化落 signal(可观测非静默):`fullwidth_comma_normalized` ∈ SIGNALS 闭集且实际写出。"""
    assert "fullwidth_comma_normalized" in signals.SIGNALS, "信号未注册入闭集(会记 _unknown:)"
    sig_log = tmp_path / "sig.jsonl"
    monkeypatch.setattr(signals, "_LOG", sig_log)
    steps = [
        {"E": "APV_0", "F": "importKey", "G": "vh1， cert/rsaca/1024rsa.key", "desc": "全角"},
        {"E": "APV_0", "F": "cmd_config", "G": "show ssl certificate vh1", "desc": "查看"},
        {"E": "check_point", "F": "found", "G": "Subject:", "desc": "断言"},
    ]
    compile_emit.func(_A, steps=steps, init_commands="")
    assert sig_log.exists(), "signal 未落盘"
    import json
    recs = [json.loads(ln) for ln in sig_log.read_text(encoding="utf-8").splitlines() if ln.strip()]
    hit = [r for r in recs if r.get("signal") == "fullwidth_comma_normalized"]
    assert hit, f"未见 fullwidth_comma_normalized 信号(非 _unknown:):{[r.get('signal') for r in recs]}"
    # count 在 payload 下嵌套(emit_signal 把 **payload 收进 payload 键)
    assert hit[0].get("payload", {}).get("count", 0) >= 1 and hit[0].get("source") == "compile_emit"
