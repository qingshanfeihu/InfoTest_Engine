"""回填重放生命周期回归(2026-07-05 第三轮扫出的洞)。

洞:compile_runtime_fill 只写给定卷(通常合并卷),per-case 卷仍 <RUNTIME>;之后
任何 compile_emit_merged 重合并从 per-case 卷重建,**静默丢掉全部已填值**。
修复:fill 把成功回填按内容键(autoid+observe_cmd+原 G)记 sidecar,merged 后按内容重放。
守:①未变卷面重放命中恢复值;②重编改了前序观测→键变→安全跳过不猜;
③内容键含 observe_cmd(整值槽 G 全是 <RUNTIME> 不独特,靠观测命令锚定)。
"""

from __future__ import annotations

import json
import shutil
from pathlib import Path

import openpyxl
import pytest

from main.ist_core.tools.device.emit_xlsx_tool import compile_emit, compile_emit_merged
from main.ist_core.tools.device.runtime_fill_tools import compile_runtime_fill
from main.case_compiler.runtime_fill import list_runtime_slots, RUNTIME_PLACEHOLDER
from main.ist_core.compile_engine_v8 import _shared as _sh

_A = "203099999999900011"
_SUB = "rt_replay_ut"


def _dirs():
    return _sh.outputs_root() / _A, _sh.outputs_root() / _SUB


@pytest.fixture(autouse=True)
def _clean(monkeypatch):
    monkeypatch.setenv("IST_PROVENANCE_OPTIONAL", "1")
    for d in _dirs():
        shutil.rmtree(d, ignore_errors=True)
    yield
    for d in _dirs():
        shutil.rmtree(d, ignore_errors=True)


def _emit(observe="show sdns service status"):
    steps = [
        {"E": "APV_0", "F": "cmd_config", "G": "sdns listener 172.16.34.70 53"},
        {"E": "APV_0", "F": "cmd_config", "G": observe},
        {"E": "check_point", "F": "found", "G": RUNTIME_PLACEHOLDER},
    ]
    return compile_emit.func(_A, steps=steps, strict_structural=False, provenance={
        "autoid": _A, "steps": [
            {"layer": "G", "source": {"kind": "footprint", "ref": "x"}},
            {"layer": "G", "source": {"kind": "footprint", "ref": "y"}},
            {"layer": "V", "source": {"kind": "device_runtime", "ref": ""}}]})


def _merged():
    compile_emit_merged.func(autoids=[_A], out_name=_SUB)
    return _sh.outputs_root() / _SUB / "case.xlsx"


def _fill(merged):
    slots = list_runtime_slots(merged)
    assert slots, "合并卷应有 RUNTIME 槽"
    out = compile_runtime_fill.func(str(merged), fills_json=[
        {"slot_id": slots[0].slot_id, "runtime_value": "PTR-auto-42",
         "evidence": "设备 PTR-auto-42"}])
    assert "filled: ['" in out
    return merged.parent / "runtime_fills.json"


def test_replay_recovers_on_unchanged_remerge():
    _emit()
    merged = _merged()
    side = _fill(merged)
    assert side.is_file()
    rec = json.loads(side.read_text(encoding="utf-8"))[0]
    assert rec["observe_cmd"] == "show sdns service status"   # 键含观测命令
    res = compile_emit_merged.func(autoids=[_A], out_name=_SUB)
    assert "回填重放: 恢复 1/1" in res
    wb = openpyxl.load_workbook(str(merged), data_only=True)
    vals = [str(c.value or "") for row in wb.active.iter_rows() for c in row]
    wb.close()
    assert any("PTR-auto-42" in v for v in vals)


def test_replay_skips_when_observe_recompiled():
    _emit()
    merged = _merged()
    _fill(merged)
    # 重编:改前序观测命令 → 内容键变 → 重合并不匹配,安全跳过不猜
    _emit(observe="show sdns service status detail")
    res = compile_emit_merged.func(autoids=[_A], out_name=_SUB)
    assert "0/1" in res, f"重编观测后应报不匹配: {res[-160:]}"
    wb = openpyxl.load_workbook(str(merged), data_only=True)
    vals = [str(c.value or "") for row in wb.active.iter_rows() for c in row]
    wb.close()
    assert any(RUNTIME_PLACEHOLDER in v for v in vals)   # 仍是占位,没被旧值污染
    assert not any("PTR-auto-42" in v for v in vals)
