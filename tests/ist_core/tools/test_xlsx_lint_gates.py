"""成品卷 lint 门回归(2026-07-04 dongkl 闭环取证驱动)。

事故:orchestrator 用 run_python 直改 case.xlsx 绕过 compile_emit 的崩溃门,
直改版带"dig(H)后直接断言"形态,上机 result=None 抛 TypeError 崩整份 pytest
(39 秒截断、34 case 只跑 1 个,连续两轮)。修复:lint 挂到凭证(compile_emit lint)
与合并(emit_merged)的必经之路——任何来源的卷面都逃不过。
"""
from __future__ import annotations

import json
import shutil
from pathlib import Path

import openpyxl
import pytest

from main.ist_core.tools.device import compile_emit, compile_emit_merged
from main.ist_core.tools.device.structural_gate import lint_xlsx_case

AID = "203031750000000201"

_STEPS = [
    {"D": "配置基线", "E": "APV_0", "F": "cmds_config",
     "G": "sdns on\nsdns listener 172.16.34.70\nsdns host name t.com\nsdns service ip s1 172.16.35.213\nsdns pool name p1\nsdns pool service p1 s1\nsdns host pool t.com p1"},
    {"D": "触发", "E": "test_env", "F": "routera", "G": "dig @172.16.34.70 t.com"},
    {"D": "断言", "E": "check_point", "F": "found", "G": r"\b172\.16\.35\.213\b"},
]


@pytest.fixture()
def emitted_case():
    out = compile_emit.invoke({"autoid": AID, "steps": _STEPS, "out_name": AID})
    assert "已产出" in out
    xp = Path("workspace/outputs") / AID / "case.xlsx"
    yield xp
    shutil.rmtree(xp.parent, ignore_errors=True)


def _corrupt(xp: Path, mutate) -> None:
    wb = openpyxl.load_workbook(xp)
    mutate(wb.active)
    wb.save(xp)


def _find_row(ws, col: int, needle: str) -> int:
    for row in ws.iter_rows(min_row=2):
        if needle in str(row[col - 1].value or ""):
            return row[0].row
    raise AssertionError(f"找不到含 {needle!r} 的行")


def test_lint_clean_case_passes(emitted_case):
    res = lint_xlsx_case(emitted_case)
    assert res.ok, [v.code for v in res.violations]


def test_lint_catches_dangling_assertion(emitted_case):
    def mutate(ws):
        r = _find_row(ws, 7, "dig @")
        ws.cell(r, 8).value = "v1"  # dig 加 H → result 不更新 → 后续断言悬空
    _corrupt(emitted_case, mutate)
    res = lint_xlsx_case(emitted_case)
    assert any(v.code == "dangling_assertion" for v in res.violations)


def test_lint_catches_invalid_regex(emitted_case):
    def mutate(ws):
        r = _find_row(ws, 6, "found")
        ws.cell(r, 7).value = r"172\.16\.35\.213[^"
    _corrupt(emitted_case, mutate)
    res = lint_xlsx_case(emitted_case)
    assert any(v.code == "assertion_regex_invalid" for v in res.violations)


def test_lint_catches_short_mode_status_assertion(emitted_case):
    def mutate(ws):
        rd = _find_row(ws, 7, "dig @")
        ws.cell(rd, 7).value = str(ws.cell(rd, 7).value) + " +short"
        ra = _find_row(ws, 6, "found")
        ws.cell(ra, 7).value = r"status:\s*NOERROR"
    _corrupt(emitted_case, mutate)
    res = lint_xlsx_case(emitted_case)
    assert any(v.code == "short_mode_status_assertion" for v in res.violations)


def test_lint_catches_undefined_capture_ref(emitted_case):
    def mutate(ws):
        ra = _find_row(ws, 6, "found")
        ws.cell(ra, 8).value = "ghost_reg"
    _corrupt(emitted_case, mutate)
    res = lint_xlsx_case(emitted_case)
    assert any(v.code == "undefined_capture_ref" for v in res.violations)


def test_lint_catches_dns_label_over_63(emitted_case):
    def mutate(ws):
        rd = _find_row(ws, 7, "dig @")
        ws.cell(rd, 7).value = "dig @172.16.34.70 www." + "x" * 120 + ".com"
    _corrupt(emitted_case, mutate)
    res = lint_xlsx_case(emitted_case)
    assert any(v.code == "dns_label_over_63" for v in res.violations)


def test_emit_merged_rejects_lint_violation(emitted_case):
    # 直改场景全真模拟:emit 落的 lint 凭证新鲜,但卷面在凭证后被改坏、直改者又把凭证签名
    # 重签回当前 mtime → 合并的 lint 最后防线仍拦下(凭证门放行后 lint 是终卷必经)。
    def mutate(ws):
        r = _find_row(ws, 7, "dig @")
        ws.cell(r, 8).value = "v1"
    _corrupt(emitted_case, mutate)
    credp = emitted_case.parent / ".grade_credential.json"
    cred = json.loads(credp.read_text())
    cred["xlsx_mtime"] = emitted_case.stat().st_mtime  # 伪造签名(直改者能做到的极限)
    credp.write_text(json.dumps(cred))
    try:
        merged = compile_emit_merged.invoke({"autoids": [AID], "out_name": "_pytest_lint_merged"})
        assert merged.startswith("error") and "lint" in merged and AID in merged
    finally:
        shutil.rmtree(Path("workspace/outputs") / "_pytest_lint_merged", ignore_errors=True)


# ---- 分发白名单 + 直连槽/execute/变量注入(2026-07-05 框架源码取证轮) ----

from main.ist_core.tools.device.structural_gate import check_crash_gates_mandatory


def _codes(steps):
    return {v.code for v in check_crash_gates_mandatory(steps).violations}


def test_dispatch_unknown_e_rejected():
    # E 不在 devices 表 → 框架静默跳过整步(test_xlsx.py:281-287)——必拒
    steps = [{"E": "APV0", "F": "cmd_config", "G": "show version"},
             {"E": "check_point", "F": "found", "G": "x"}]
    assert "unknown_dispatch_target" in _codes(steps)


def test_dispatch_unknown_f_on_test_env_rejected_but_case_normalized():
    # F 不在 Env 方法闭集 → getattr AttributeError 崩整卷;大小写是可修形态(emit 归一)不拒
    bad = [{"E": "test_env", "F": "clienta", "G": "dig @1.2.3.4 t.com"},
           {"E": "check_point", "F": "found", "G": "x"}]
    assert "unknown_dispatch_method" in _codes(bad)
    cased = [{"E": "test_env", "F": "RouterA", "G": "dig @1.2.3.4 t.com"},
             {"E": "check_point", "F": "found", "G": "x"}]
    assert "unknown_dispatch_method" not in _codes(cased)


def test_host_slot_cmd_is_observation():
    # E=直连槽 F=cmd 返回回显(ssh_server.py:93-106)——后接断言不是悬空
    steps = [{"E": "server231", "F": "cmd", "G": "systemctl status named"},
             {"E": "check_point", "F": "found", "G": "active"}]
    codes = _codes(steps)
    assert "dangling_assertion" not in codes and "unknown_dispatch_target" not in codes


def test_host_slot_execute_observation_depends_on_action():
    # execute 动作有 return(访问)→ 观测;无 return(创建Mysql数据库用户)→ 断言悬空
    ok = [{"E": "server231", "F": "execute", "G": "访问：http://172.16.35.213/"},
          {"E": "check_point", "F": "found", "G": "200 OK"}]
    assert "dangling_assertion" not in _codes(ok)
    bad = [{"E": "server231", "F": "execute", "G": "创建Mysql数据库用户：u1，p1，1.2.3.4"},
           {"E": "check_point", "F": "found", "G": "OK"}]
    assert "dangling_assertion" in _codes(bad)


def test_host_slot_manual_ip_change_rejected():
    # 直连槽与 test_env 同走 ssh_server.cmd 记账/自动恢复——ip addr 变更同拒
    steps = [{"E": "routera", "F": "cmd", "G": "ip addr add 10.0.0.9/24 dev eth1"},
             {"E": "check_point", "F": "found", "G": "x"}]
    assert "manual_ip_cleanup" in _codes(steps)


def test_apv1_payload_integrity_now_covered():
    # 旧版载荷完整性门漏 APV_1——字面 \n 现在同拒
    steps = [{"E": "APV_1", "F": "cmds_config", "G": "sdns on\\nsdns listener 1.2.3.4"},
             {"E": "check_point", "F": "found", "G": "x"}]
    assert "literal_backslash_n" in _codes(steps)


def test_lint_i_column_injection_rules(emitted_case):
    # 非 check_point 步 I 引用未捕获变量 → 框架 raise NameError 崩整卷(test_xlsx.py:319-324)
    def bad_ref(ws):
        r = _find_row(ws, 7, "dig @")
        ws.cell(r, 9).value = "no_such_var"          # I 列
        ws.cell(r, 7).value = "dig @172.16.34.70 {} A"
    _corrupt(emitted_case, bad_ref)
    res = lint_xlsx_case(emitted_case)
    assert not res.ok and any(v.code == "undefined_capture_ref" for v in res.violations)


def test_lint_i_without_placeholder(emitted_case):
    # 带 I 但 G 无 {} → 注入静默不发生
    def no_ph(ws):
        r0 = _find_row(ws, 7, "sdns on")
        ws.cell(r0, 8).value = "v_cap"               # 前步 H 捕获,引用合法
        r = _find_row(ws, 7, "dig @")
        ws.cell(r, 9).value = "v_cap"                # I 引用但 G 无 {}
    _corrupt(emitted_case, no_ph)
    res = lint_xlsx_case(emitted_case)
    assert any(v.code == "injection_without_placeholder" for v in res.violations)


def test_lint_comma_splits_parameters(emitted_case):
    # 框架 get_parameter 按引号外逗号切参——test_env 步 G 裸逗号第二段错传 prompt
    def add_comma(ws):
        r = _find_row(ws, 7, "dig @")
        ws.cell(r, 7).value = "dig @172.16.34.70 t.com, extra"
    _corrupt(emitted_case, add_comma)
    res = lint_xlsx_case(emitted_case)
    assert any(v.code == "comma_splits_parameters" for v in res.violations)


def test_lint_comma_kwargs_and_quoted_allowed(emitted_case):
    # timeout= 具名段是框架惯用;引号内逗号不切——都放行
    def ok_commas(ws):
        r = _find_row(ws, 7, "dig @")
        ws.cell(r, 7).value = 'curl -H "a,b" http://t.com/, timeout=30'
    _corrupt(emitted_case, ok_commas)
    res = lint_xlsx_case(emitted_case)
    assert not any(v.code == "comma_splits_parameters" for v in res.violations)


def test_lint_autoid_row_needs_dispatch(emitted_case):
    # 框架 ifrun:autoid 行 E 列空 → 整 case 静默不跑
    def orphan_autoid(ws):
        r = _find_row(ws, 1, AID)
        # 把 autoid 挪到上一行独占(E 列空)——人工卷常见格式
        ws.insert_rows(r)
        ws.cell(r, 1).value = ws.cell(r + 1, 1).value
        ws.cell(r + 1, 1).value = None
    _corrupt(emitted_case, orphan_autoid)
    res = lint_xlsx_case(emitted_case)
    assert any(v.code == "autoid_row_not_runnable" for v in res.violations)


def test_execute_observation_parsed_from_mirror():
    # 观测性判定从 mirror 注册表源码解析(数据按引用流,不硬编码);抽查两侧已知事实
    from main.ist_core.tools.device.structural_gate import (
        _execute_returning_actions, _APV_ACTION_SRC, _CLIENT_ACTION_SRC)
    client = _execute_returning_actions(_CLIENT_ACTION_SRC)
    apv = _execute_returning_actions(_APV_ACTION_SRC)
    if not client or not apv:
        import pytest
        pytest.skip("框架 mirror 不在盘上")
    assert "访问" in client and "创建Mysql数据库用户" not in client
    assert "提取PTR自动生成名称" in apv and "等待健康检查up" not in apv
    # 设备侧动作后接断言:有 return 的观测/无 return 的悬空
    ok = [{"E": "APV_0", "F": "execute", "G": "提取PTR自动生成名称：x.com|8.8.8.8"},
          {"E": "check_point", "F": "found", "G": "x"}]
    assert "dangling_assertion" not in _codes(ok)
    bad = [{"E": "APV_0", "F": "execute", "G": "等待健康检查up：30"},
           {"E": "check_point", "F": "found", "G": "UP"}]
    assert "dangling_assertion" in _codes(bad)


def test_line_anchor_assertions_rejected():
    # 2026-07-06 588691:框架 found/not_found 无 MULTILINE,^ 锚恒假——必假门
    steps = [{"E": "APV_0", "F": "cmd_config", "G": "show sdns session persistence x"},
             {"E": "check_point", "F": "found", "G": "^www\\.zyq\\.com"}]
    assert "line_anchor_never_matches" in _codes(steps)
    steps2 = [{"E": "APV_0", "F": "cmd_config", "G": "show x"},
              {"E": "check_point", "F": "not_found", "G": "^entry"}]
    assert "line_anchor_never_matches" in _codes(steps2)
    # \n 前缀锚定与 abs_found 不受影响
    ok = [{"E": "APV_0", "F": "cmd_config", "G": "show x"},
          {"E": "check_point", "F": "found", "G": "\\nwww\\.zyq\\.com"},
          {"E": "check_point", "F": "abs_found", "G": "^literal"}]
    assert "line_anchor_never_matches" not in _codes(ok)


def test_dollar_anchor_assertions_rejected():
    # $/\Z 结尾锚:窗口末尾是提示符,无 MULTILINE 下结尾锚恒假(2026-07-06)
    base = [{"E": "test_env", "F": "clientc", "G": "dig @1.1.1.1 a.com A +short"}]
    assert "line_anchor_never_matches" in _codes(base + [
        {"E": "check_point", "F": "found", "G": "NOERROR$"}])
    assert "line_anchor_never_matches" in _codes(base + [
        {"E": "check_point", "F": "not_found", "G": "entry\\Z"}])
    # 字面 \$(转义美元)不是锚,放行
    assert "line_anchor_never_matches" not in _codes(base + [
        {"E": "check_point", "F": "found", "G": "price \\$"}])


def test_assertion_matches_command_echo_rejected():
    # 模式命中命令原文=恒真/恒fail(窗口含回显;2026-07-06 588691 round1 同族)
    dig = {"E": "test_env", "F": "clientc", "G": "dig @172.16.34.70 www.a.com A +short"}
    # found 查询目标 IP → 回显必中,恒真假 PASS
    assert "assertion_matches_command_echo" in _codes([
        dig, {"E": "check_point", "F": "found", "G": "\\b172\\.16\\.34\\.70\\b"}])
    # not_found 命令关键字 → 恒 fail
    assert "assertion_matches_command_echo" in _codes([
        dig, {"E": "check_point", "F": "not_found", "G": "www\\.a\\.com"}])
    # 数据形态断言不误杀
    assert "assertion_matches_command_echo" not in _codes([
        dig, {"E": "check_point", "F": "found", "G": "\\b172\\.16\\.35\\.21\\b"}])
    # I 路径:寄存器捕获步的命令原文同判
    assert "assertion_matches_command_echo" in _codes([
        {"E": "test_env", "F": "clientc", "G": "dig @1.2.3.4 b.com", "H": "v1"},
        dig,
        {"E": "check_point", "F": "abs_found", "G": "1.2.3.4", "I": "v1"}])
    # 期望取寄存器(H 非空)是运行时值,不判
    assert "assertion_matches_command_echo" not in _codes([
        {"E": "test_env", "F": "clientc", "G": "dig @1.2.3.4 b.com", "H": "v1"},
        dig, {"E": "check_point", "F": "found", "G": "", "H": "v1"}])


def test_zero_assertion_case_rejected():
    # 零 check_point → close() success==0 → 恒 FAIL(check_point.py:126)
    assert "no_assertion_in_case" in _codes([
        {"E": "APV_0", "F": "cmd_config", "G": "show version"}])
    assert "no_assertion_in_case" not in _codes([
        {"E": "test_env", "F": "clientc", "G": "dig @1.1.1.1 a.com"},
        {"E": "check_point", "F": "found", "G": "NOERROR"}])


def test_injection_format_crash_rejected():
    # I 注入时 G 经 str.format(单值):裸大括号/命名占位/多占位全崩(test_xlsx.py:309)
    cap = {"E": "test_env", "F": "clientc", "G": "dig @1.1.1.1 a.com +short", "H": "v1"}
    assert "injection_format_crash" in _codes([cap,
        {"E": "test_env", "F": "clientc", "G": "curl -d '{\"a\":1}' http://x/ {}", "I": "v1"}])
    assert "injection_format_crash" in _codes([cap,
        {"E": "test_env", "F": "clientc", "G": "ping {} && ping {}", "I": "v1"}])
    assert "injection_format_crash" in _codes([cap,
        {"E": "test_env", "F": "clientc", "G": "ping {ip}", "I": "v1"}])
    assert "injection_format_crash" not in _codes([cap,
        {"E": "test_env", "F": "clientc", "G": "ping {}", "I": "v1"}])


def test_register_shadows_framework_name_rejected():
    # H 撞框架执行器名字空间(locals()[H]=值 覆盖框架状态)——闭集 ast 解析 mirror
    dig = {"E": "test_env", "F": "clientc", "G": "dig @1.1.1.1 a.com +short"}
    assert "register_shadows_framework_name" in _codes([
        {**dig, "H": "result"},
        dig, {"E": "check_point", "F": "found", "G": "NOERROR"}])
    assert "register_shadows_framework_name" in _codes([
        {**dig, "H": "value"},
        dig, {"E": "check_point", "F": "found", "G": "NOERROR"}])
    assert "register_shadows_framework_name" not in _codes([
        {**dig, "H": "v1"},
        dig, {"E": "check_point", "F": "found", "G": "NOERROR"}])


def test_cmd_config_multiline_rejected():
    # cmd_config 的 G 含换行:执行层 replace 删换行→命令无分隔粘连(test_xlsx.py:307)
    assert "cmd_config_multiline" in _codes([
        {"E": "APV_0", "F": "cmd_config", "G": "slb real r1\nslb group g1"},
        {"E": "check_point", "F": "found", "G": "r1"}])
    # cmds_config 多行合法;cmd_config 尾随换行无害
    assert "cmd_config_multiline" not in _codes([
        {"E": "APV_0", "F": "cmds_config", "G": "slb real r1\nslb group g1"},
        {"E": "APV_0", "F": "cmd_config", "G": "show version\n"},
        {"E": "check_point", "F": "found", "G": "r1"}])


def test_inline_multiline_flag_exempts_anchor_gate():
    # (?m) 内联 flag 开 MULTILINE 后行锚合法——正确修法不被必崩门误杀(评审建议)
    base = [{"E": "test_env", "F": "clientc", "G": "dig @1.1.1.1 a.com A"}]
    assert "line_anchor_never_matches" not in _codes(base + [
        {"E": "check_point", "F": "found", "G": "(?m)^www\\.a\\.com"}])
    assert "line_anchor_never_matches" not in _codes(base + [
        {"E": "check_point", "F": "found", "G": "(?im)entry$"}])


def test_repeated_manual_index_format_is_legal():
    # {0} {0} 重复手动索引对单值 format 合法;{} 与 {0} 混用才崩(评审建议)
    cap = {"E": "test_env", "F": "clientc", "G": "dig @1.1.1.1 a.com +short", "H": "v1"}
    assert "injection_format_crash" not in _codes([cap,
        {"E": "test_env", "F": "clientc", "G": "ping {0} && arp -a | grep {0}", "I": "v1"}])
    assert "injection_format_crash" in _codes([cap,
        {"E": "test_env", "F": "clientc", "G": "ping {} && ping {0}", "I": "v1"}])


def test_reserved_names_narrowed_to_executor_frame():
    # 闭集只收执行帧名字:辅助函数局部名(如 get_parameter 的 m)不误杀
    from main.ist_core.tools.device.structural_gate import _framework_reserved_names
    ns = _framework_reserved_names()
    assert "result" in ns and "value" in ns      # 执行函数局部
    assert "v1" not in ns and "ip1" not in ns    # 常规寄存器名
