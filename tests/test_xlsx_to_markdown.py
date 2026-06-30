"""Tests for main.xlsx_to_markdown — xlsx/xls → GFM markdown conversion."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from main.xlsx_to_markdown import (
    _cell_to_str,
    _row_to_md,
    _sheet_to_markdown,
    convert_xlsx_to_markdown,
    write_markdown,
)


def _create_xlsx(path: Path, sheets: dict[str, list[list]]) -> Path:
    wb = openpyxl.Workbook()
    first = True
    for name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(str(path))
    return path


class TestCellToStr:
    def test_none(self):
        assert _cell_to_str(None) == ""

    def test_string(self):
        assert _cell_to_str("hello") == "hello"

    def test_number(self):
        assert _cell_to_str(42) == "42"

    def test_pipe_escaped(self):
        assert _cell_to_str("a|b") == "a\\|b"

    def test_newline_replaced(self):
        assert _cell_to_str("a\nb") == "a<br>b"

    def test_crlf_replaced(self):
        assert _cell_to_str("a\r\nb") == "a<br>b"

    def test_cr_replaced(self):
        assert _cell_to_str("a\rb") == "a<br>b"

    def test_whitespace_stripped(self):
        assert _cell_to_str("  hello  ") == "hello"


class TestRowToMd:
    def test_basic(self):
        assert _row_to_md(["A", "B", "C"]) == "| A | B | C |"

    def test_with_none(self):
        assert _row_to_md(["A", None, "C"]) == "| A |  | C |"


class TestSheetToMarkdown:
    def test_basic_sheet(self, tmp_path: Path):
        p = _create_xlsx(tmp_path / "test.xlsx", {
            "Sheet1": [["Name", "Value"], ["foo", 1], ["bar", 2]],
        })
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        ws = wb["Sheet1"]
        md = _sheet_to_markdown(ws)
        wb.close()

        assert "## Sheet1" in md
        assert "| Name | Value |" in md
        assert "| --- | --- |" in md
        assert "| foo | 1 |" in md
        assert "| bar | 2 |" in md

    def test_empty_sheet(self, tmp_path: Path):
        p = _create_xlsx(tmp_path / "empty.xlsx", {"Empty": []})
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        ws = wb["Empty"]
        md = _sheet_to_markdown(ws)
        wb.close()

        assert "## Empty" in md
        assert "空 sheet" in md

    def test_trailing_none_trimmed(self, tmp_path: Path):
        p = _create_xlsx(tmp_path / "trail.xlsx", {
            "S1": [["A", "B", None, None], ["x", "y"]],
        })
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        ws = wb["S1"]
        md = _sheet_to_markdown(ws)
        wb.close()

        assert "| A | B |" in md


class TestConvertXlsxToMarkdown:
    def test_full_conversion(self, tmp_path: Path):
        p = _create_xlsx(tmp_path / "data.xlsx", {
            "Users": [["Name", "Age"], ["Alice", 30]],
            "Scores": [["Subject", "Score"], ["Math", 95]],
        })
        md = convert_xlsx_to_markdown(p)

        assert "# data" in md
        assert "## Users" in md
        assert "## Scores" in md
        assert "| Alice | 30 |" in md
        assert "| Math | 95 |" in md

    def test_source_name_in_output(self, tmp_path: Path):
        p = _create_xlsx(tmp_path / "myfile.xlsx", {"S1": [["A"], [1]]})
        md = convert_xlsx_to_markdown(p)
        assert "myfile.xlsx" in md


class TestWriteMarkdown:
    def test_writes_file(self, tmp_path: Path):
        xlsx = _create_xlsx(tmp_path / "input.xlsx", {"S1": [["Col"], ["Val"]]})
        out = tmp_path / "output" / "result.md"
        result = write_markdown(xlsx, out)
        assert result == out
        assert out.exists()
        content = out.read_text(encoding="utf-8")
        assert "## S1" in content

    def test_creates_parent_dirs(self, tmp_path: Path):
        xlsx = _create_xlsx(tmp_path / "input.xlsx", {"S1": [["A"], [1]]})
        out = tmp_path / "deep" / "nested" / "dir" / "result.md"
        write_markdown(xlsx, out)
        assert out.exists()
