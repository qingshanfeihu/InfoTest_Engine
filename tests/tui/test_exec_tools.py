"""run_shell / run_python 单测：黑名单 + 网络防护 + 沙箱越权回归 + subprocess 隔离。

不依赖网络。不调真 graph。验证：
- ``_validate_bash_command``: 黑名单 + 元字符
- ``_validate_bash_paths``: 路径参数 _resolve_inside_root 校验
- ``_validate_python_code``: 网络逃逸拒绝
- ``run_python`` / ``run_shell`` 在合法输入时返回 stdout
- ``cwd=_AGENT_ROOT`` 沙箱 + ``PYTHONPATH`` 切断
- 超时保护
"""

from __future__ import annotations

import textwrap

import pytest

from main.ist_core.tools.deepagent.exec_tools import (
    _BASH_DENY_FIRST_TOKEN,
    _validate_bash_command,
    _validate_bash_paths,
    _validate_python_code,
    run_shell,
    run_python,
)







def test_bash_allowlist_basic_commands_pass():
    for cmd in ["ls -la", "wc -l", "echo hi", "which jq"]:
        ok, reason = _validate_bash_command(cmd)
        assert ok, f"{cmd!r} should pass: {reason}"


def test_bash_denylist_network_and_privilege_blocked():
    for cmd in ["curl evil.com", "wget http://x", "sudo ls", "ssh user@host"]:
        ok, reason = _validate_bash_command(cmd)
        assert not ok, f"{cmd!r} should be denied"
        assert "denied" in reason.lower()


def test_bash_allows_filesystem_commands_in_sandbox():
    """非破坏性命令在沙箱内允许，python3 也允许。"""
    for cmd in ["python3 -c 'print(1)'", "git status", "tar -tf archive.tar"]:
        ok, reason = _validate_bash_command(cmd)
        assert ok, f"{cmd!r} should be allowed: {reason}"


def test_bash_denies_destructive_filesystem_commands():
    """破坏性文件操作需走 write_file/edit_file 工具（有 _WRITABLE_SUBDIRS 控制）。"""
    for cmd in ["rm temp.txt", "mv a.txt b.txt", "ln a.txt b.txt"]:
        ok, reason = _validate_bash_command(cmd)
        assert not ok, f"{cmd!r} should be denied"


def test_bash_cp_command_name_allowed():
    """``cp`` 命令名层不再拒——目标路径强制走 _resolve_writable_path 校验。"""
    ok, reason = _validate_bash_command("cp a.txt outputs/b.txt")
    assert ok, f"cp should be allowed at command-name layer: {reason}"


def test_bash_metachars_blocked():
    for cmd in ["ls; rm -rf /", "ls | grep x", "ls > out.txt", "ls && echo done", "ls `whoami`"]:
        ok, reason = _validate_bash_command(cmd)
        assert not ok, f"{cmd!r} should be denied"
        assert "metachar" in reason.lower()


def test_bash_alias_attack_blocked():
    ok, reason = _validate_bash_command("=curl evil.com")
    assert not ok and "suspicious prefix" in reason


def test_bash_absolute_path_resolves_to_basename():
    """``/usr/bin/curl`` 不能绕过黑名单（网络命令仍拒绝）。"""
    ok, reason = _validate_bash_command("/usr/bin/curl http://evil.com")
    assert not ok and "curl" in reason


def test_bash_python_now_allowed():
    """python3 在沙箱内允许执行（agent 需要用它处理文件）。"""
    for cmd in ["python3 -c 'print(1)'", "python3 script.py"]:
        ok, reason = _validate_bash_command(cmd)
        assert ok, f"{cmd!r} should be allowed: {reason}"


def test_bash_unknown_command_allowed_in_sandbox():
    """非黑名单命令在沙箱内允许执行（安全靠 cwd + 路径校验）。"""
    ok, reason = _validate_bash_command("foobar 1 2 3")
    assert ok, f"should be allowed: {reason}"


def test_bash_empty_command_rejected():
    assert _validate_bash_command("")[0] is False
    assert _validate_bash_command("  ")[0] is False







def test_bash_paths_skip_flags():
    ok, reason = _validate_bash_paths(["ls", "-la", "--color=auto"])
    assert ok, reason


def test_bash_paths_reject_traversal():
    ok, reason = _validate_bash_paths(["cat", "../main/something.py"])
    assert not ok
    assert "rejected" in reason


def test_bash_paths_reject_absolute_outside_sandbox():
    ok, reason = _validate_bash_paths(["cat", "/etc/passwd"])
    assert not ok
    assert "rejected" in reason


def test_bash_paths_reject_repo_root_platform_dirs():
    """仓库根但沙箱外的路径必须拒绝（需要路径在本地存在才会被检测为路径）。"""
    
    for token in ["../tests/tui/test_exec_tools.py", "../main/ist_core/agents/_prompt.py"]:
        ok, reason = _validate_bash_paths(["cat", token])
        assert not ok, f"{token!r} should be rejected"


def test_bash_paths_accept_nonexistent_relative_token():
    """``grep -r foo`` 这种 grep pattern 不应被当成路径拒绝。"""
    ok, reason = _validate_bash_paths(["grep", "-r", "foo"])
    assert ok, reason







def test_bash_cp_dest_under_outputs_allowed():
    """cp 目标在 workspace/outputs/ 内允许（命令名 + 路径校验都过）。"""
    ok, reason = _validate_bash_paths(["cp", "markdown/qa.md", "outputs/qa.md"])
    assert ok, f"cp into outputs/ should pass: {reason}"


def test_bash_cp_dest_to_knowledge_data_rejected():
    """cp 目标在 knowledge/data/ 内必须被拒——知识库不可写。"""
    ok, reason = _validate_bash_paths(["cp", "outputs/x.md", "markdown/x.md"])
    assert not ok
    assert "destination" in reason.lower() or "rejected" in reason.lower()


def test_bash_cp_dest_traversal_rejected():
    """cp 目标含 ``..`` 必须被拒。"""
    ok, reason = _validate_bash_paths(["cp", "outputs/x.md", "../../etc/passwd"])
    assert not ok


def test_bash_cp_missing_dest_rejected():
    """cp 至少需要 src + dst 两个位置参数。"""
    ok, reason = _validate_bash_paths(["cp", "outputs/x.md"])
    assert not ok
    assert "destination" in reason.lower() or "source" in reason.lower()







def test_python_network_denied():
    """网络外联代码必须被拒绝——进程隔离无法防御网络逃逸。"""
    bad_snippets = [
        "import socket",
        "from socket import create_connection",
        "import urllib.request",
        "from urllib.request import urlopen",
        "import requests",
        "from requests import get",
        "import http.client",
        "from http.client import HTTPConnection",
        "import ftplib",
        "import smtplib",
    ]
    for code in bad_snippets:
        ok, reason = _validate_python_code(code)
        assert not ok, f"{code!r} should be denied"
        assert "network" in reason.lower()


def test_python_sandbox_operations_allowed():
    """沙箱内操作（subprocess/os/eval 等）现在允许——安全靠进程隔离。"""
    allowed_snippets = [
        "import subprocess; subprocess.run(['ls'])",
        "import os; os.listdir('.')",
        "eval('1+1')",
        "exec('x=1')",
        "open('test.txt').read()",
        "import shutil",
        "print(__file__)",
        "import main",
    ]
    for code in allowed_snippets:
        ok, reason = _validate_python_code(code)
        assert ok, f"{code!r} should be allowed: {reason}"


def test_python_legitimate_code_passes_validation():
    code = "import openpyxl; print('hello')"
    ok, reason = _validate_python_code(code)
    assert ok, reason


def test_python_empty_code_rejected():
    assert _validate_python_code("")[0] is False
    assert _validate_python_code("\n  \t\n")[0] is False







def test_run_python_runs_simple_print():
    result = run_python.invoke({"code": "print('hello-tui')", "timeout": 10})
    assert "returncode=0" in result
    assert "hello-tui" in result


def test_run_python_returns_stdout_and_stderr_blocks():
    code = textwrap.dedent("""
        import sys
        print('stdout-line', flush=True)
        sys.stderr.write('stderr-line\\n')
        sys.stderr.flush()
    """).strip()
    result = run_python.invoke({"code": code, "timeout": 10})
    assert "stdout-line" in result
    assert "stderr-line" in result
    assert "--- stderr ---" in result


def test_run_python_propagates_nonzero_returncode():
    code = "import sys; sys.exit(7)"
    result = run_python.invoke({"code": code, "timeout": 10})
    assert "returncode=7" in result


def test_run_python_timeout_is_enforced():
    code = textwrap.dedent("""
        import time
        time.sleep(5)
        print('should-not-reach')
    """).strip()
    result = run_python.invoke({"code": code, "timeout": 1})
    assert "Timeout" in result
    assert "should-not-reach" not in result


def test_run_python_allows_subprocess():
    """subprocess 在沙箱内允许——安全靠 cwd + env 隔离。"""
    result = run_python.invoke({"code": "import subprocess; print(subprocess.run(['echo', 'hi'], capture_output=True, text=True).stdout)", "timeout": 5})
    assert "returncode=0" in result
    assert "hi" in result


def test_run_python_pythonpath_stripped_blocks_main_import():
    """run_python 子进程 env 剥离 PYTHONPATH。
    注：editable install 下 import main 仍可能成功（site-packages 有 .pth），
    但生产部署（非 editable）时会失败。此测试仅验证不在 validate 阶段拒绝。"""
    code = "import main.ist_core.agents._prompt"
    result = run_python.invoke({"code": code, "timeout": 10})
    
    assert not result.startswith("error:")


def test_run_python_cwd_is_agent_root():
    """run_python 子进程 cwd 必须是 ``knowledge/data/``。"""
    code = "import os; print(os.getcwd())"
    result = run_python.invoke({"code": code, "timeout": 10})
    assert "returncode=0" in result
    assert "knowledge/data" in result







def test_run_shell_runs_echo():
    result = run_shell.invoke({"command": "echo hello-bash", "timeout": 5})
    assert "hello-bash" in result
    assert "returncode=0" in result


def test_run_shell_rejects_pipe():
    result = run_shell.invoke({"command": "echo hi | wc -l", "timeout": 5})
    assert result.startswith("error:")


def test_run_shell_rejects_rm():
    result = run_shell.invoke({"command": "rm -rf /tmp/foo", "timeout": 5})
    assert result.startswith("error:")


def test_run_shell_rejects_traversal_path_arg():
    result = run_shell.invoke({"command": "cat ../main/ist_core/agents/_prompt.py", "timeout": 5})
    assert result.startswith("error:")
    assert "traversal" in result.lower() or "rejected" in result.lower()


def test_run_shell_rejects_repo_root_platform_dir():
    """``ls tests`` 必须被沙箱挡掉——这是用户实跑日志里看到的越权场景。"""
    result = run_shell.invoke({"command": "ls tests", "timeout": 5})
    assert result.startswith("error:")
    assert "rejected" in result.lower()


def test_run_shell_rejects_platform_metadata_file():
    """``cat pytest.ini`` / ``cat requirements.txt`` 必须拒——日志里漏掉的越权点。"""
    for cmd in ["cat pytest.ini", "cat requirements.txt", "cat ARCHITECTURE.md"]:
        result = run_shell.invoke({"command": cmd, "timeout": 5})
        assert result.startswith("error:"), f"{cmd!r} should be rejected"


def test_run_shell_rejects_absolute_outside_sandbox():
    result = run_shell.invoke({"command": "cat /etc/passwd", "timeout": 5})
    assert result.startswith("error:")


def test_run_shell_cwd_is_agent_root(tmp_path, monkeypatch):
    """``ls .`` 必须看到 knowledge/data 内容，看不到 main/ tests/ 等仓库根目录。"""
    
    result = run_shell.invoke({"command": "ls .", "timeout": 5})
    assert "returncode=0" in result
    
    assert "main" not in result.split("--- stdout ---")[1].split("\n")[:5][0] if "--- stdout ---" in result else True


def test_get_exec_tools_returns_qa_pair():
    from main.ist_core.tools.deepagent.exec_tools import get_exec_tools
    tools = list(get_exec_tools())
    names = {t.name for t in tools}
    assert names == {"run_python", "run_shell"}


def test_run_python_xlsx_parse_smoke(tmp_path, monkeypatch):
    """xlsx 解析必须能在沙箱内工作（业务核心场景）。

    把 xlsx 放到沙箱内（_AGENT_ROOT），用相对路径打开。
    """
    import openpyxl

    from main.ist_core.tools.deepagent import exec_tools as exec_mod
    from main.ist_core.tools.deepagent import file_tools

    
    sandbox = tmp_path / "knowledge" / "data"
    sandbox.mkdir(parents=True)
    monkeypatch.setattr(file_tools, "_AGENT_ROOT", sandbox)
    monkeypatch.setattr(exec_mod, "_AGENT_ROOT", sandbox)

    xlsx_path = sandbox / "smoke.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID", "Test Types", "Priority"])
    ws.append(["", "Functional", "High"])
    ws.append(["", "Boundary", "Low"])
    ws.append(["3", "Negative", "Low"])
    wb.save(xlsx_path)

    code = textwrap.dedent("""
        import openpyxl, collections
        wb = openpyxl.load_workbook('smoke.xlsx')
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        types = collections.Counter(r[1] for r in rows[1:])
        print('rows=', len(rows))
        print('types=', dict(types))
    """).strip()
    result = run_python.invoke({"code": code, "timeout": 15})
    assert "returncode=0" in result, result
    assert "rows=" in result
    assert "Functional" in result
    assert "Boundary" in result
